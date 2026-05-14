#!/usr/bin/env python3
"""
Google Maps Detail Re-Scraper
Re-searches each city with different category terms and clicks into
every result to get website URLs, PLZ, categories, etc.
Much faster than enriching one company at a time.
"""

import json, os, re, sys, time, random, subprocess, signal
from datetime import datetime

CHROME_PATH = os.path.expanduser("~/.cache/ms-playwright/chromium-1223/chrome-linux64/chrome")
RESULTS_FILE = "/home/z/my-project/gmaps_travel_results.json"
PROGRESS_FILE = "/home/z/my-project/gmaps_detail_progress.json"
EXCEL_OUTPUT = "/home/z/my-project/download/Google_Maps_Firmen_Erlangen_50km.xlsx"
GITHUB_TOKEN = "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"
PROJECT_DIR = "/home/z/my-project"

# Category-based search terms to cover all business types
CATEGORY_TERMS = [
    "IT", "Software", "Ingenieurbüro", "Consulting", "Marketing",
    "Steuerberater", "Rechtsanwalt", "Architekt", "Handwerk",
    "Elektro", "Metallbau", "Maschinenbau", "Baufirma",
    "Kfz", "Immobilien", "Versicherung", "Logistik",
    "Apotheke", "Arzt", "Friseur", "Bäckerei", "Gastronomie",
]

# Major cities to re-scrape (ordered by number of companies needing websites)
PRIORITY_CITIES = [
    "Erlangen", "Nürnberg", "Fürth", "Bamberg",
    "Herzogenaurach", "Schwabach", "Forchheim",
    "Lauf an der Pegnitz", "Hersbruck", "Neustadt an der Aisch",
    "Eckental", "Ebermannstadt", "Roth", "Altdorf bei Nürnberg",
    "Zirndorf", "Strullendorf", "Höchstadt an der Aisch",
]


def load_json(path, default=None):
    try:
        with open(path) as f:
            return json.load(f)
    except:
        return default if default is not None else {}

def save_json(path, data):
    with open(path, 'w') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def git_push(message):
    try:
        os.chdir(PROJECT_DIR)
        subprocess.run(["git", "add", "-A"], capture_output=True, timeout=30)
        r = subprocess.run(["git", "commit", "-m", message], capture_output=True, timeout=30)
        if r.returncode != 0 and "nothing to commit" in r.stdout.decode():
            return
        subprocess.run(
            ["git", "push", f"https://{GITHUB_TOKEN}@github.com/BirdNest055/Heyjobs.git", "main"],
            capture_output=True, timeout=60
        )
        print(f"  [GIT] {message}", flush=True)
    except Exception as e:
        print(f"  [GIT-ERR] {e}", flush=True)


def extract_detail(page, town, search_term):
    result = {
        "name": "", "rating": "", "review_count": "", "category": "",
        "address": "", "plz": "", "city": "", "phone": "", "website": "",
        "lat": "", "lon": "", "gmaps_url": "",
        "search_town": town, "search_term": search_term,
        "scraped_at": datetime.now().isoformat(),
        "source": "google_maps_detail",
    }
    try:
        page.wait_for_selector('[role="main"]', timeout=6000)
    except:
        pass
    time.sleep(0.3)

    try:
        el = page.query_selector('h1.DUwDvf, h1.fontHeadlineLarge')
        if el: result["name"] = el.inner_text().strip()
    except: pass

    try:
        el = page.query_selector('div.F7nice')
        if el:
            se = el.query_selector('span[aria-label]')
            if se:
                a = se.get_attribute('aria-label') or ''
                m = re.search(r'([\d,]+)\s*Sterne', a)
                if m: result["rating"] = m.group(1).replace(',', '.')
    except: pass

    try:
        el = page.query_selector('button[aria-label*="Bewertungen"]')
        if el:
            a = el.get_attribute('aria-label') or ''
            m = re.search(r'([\d.]+)\s*Bewertungen', a)
            if m: result["review_count"] = m.group(1)
    except: pass

    try:
        el = page.query_selector('button[jsaction*="category"]')
        if el: result["category"] = el.inner_text().strip()
    except: pass

    # Website - most important field
    try:
        el = page.query_selector('a[data-item-id="authority"]')
        if el:
            href = el.get_attribute('href') or ''
            if href and 'google' not in href.lower():
                result["website"] = href
    except: pass
    if not result["website"]:
        try:
            el = page.query_selector('a[aria-label*="Website"]')
            if el:
                href = el.get_attribute('href') or ''
                if href and 'google' not in href.lower() and 'gstatic' not in href.lower() and not href.startswith('tel:'):
                    result["website"] = href
        except: pass

    try:
        for el in page.query_selector_all('[data-item-id]'):
            iid = el.get_attribute('data-item-id') or ''
            if 'address' in iid and not result["address"]:
                text = re.sub(r'[\ue000-\uefff]', '', el.inner_text()).strip()
                text = re.sub(r'\s+', ' ', text).strip()
                result["address"] = text
                m = re.search(r'(\d{5})', text)
                if m: result["plz"] = m.group(1)
                m = re.search(r'\d{5}\s+([A-Za-zäöüßÄÖÜ\-\.]+(?:\s+[A-Za-zäöüßÄÖÜ\-\.]+)*)', text)
                if m: result["city"] = m.group(1).strip()
            elif 'phone:tel:' in iid and not result["phone"]:
                m = re.search(r'phone:tel:(.+)', iid)
                if m: result["phone"] = m.group(1)
    except: pass

    try:
        url = page.url
        m = re.search(r'@(-?[\d.]+),(-?[\d.]+)', url)
        if m:
            result["lat"] = m.group(1)
            result["lon"] = m.group(2)
        result["gmaps_url"] = url
    except: pass

    return result


def search_and_extract_all(page, town, search_term, max_scroll=15):
    """Search and click into every result to get full detail data"""
    companies = []
    seen_names = set()

    query = f"{town} {search_term}"
    url = f"https://www.google.com/maps/search/{query.replace(' ', '+')}"

    try:
        page.goto(url, timeout=20000, wait_until="domcontentloaded")
    except:
        try:
            page.goto(url, timeout=15000)
        except:
            print(f"  [WARN] Cannot load: {query}", flush=True)
            return companies

    time.sleep(random.uniform(3, 5))

    # Cookie consent
    try:
        accept = page.query_selector('button[aria-label*="Alle ablehnen"], button[aria-label*="Reject"]')
        if accept:
            accept.click()
            time.sleep(1)
    except:
        pass

    # Scroll to load all results
    try:
        scroll_panel = page.query_selector('div[role="feed"]')
        if scroll_panel:
            last_count = 0
            same_rounds = 0
            for i in range(max_scroll):
                page.evaluate('(el) => el.scrollTop = el.scrollHeight', scroll_panel)
                time.sleep(random.uniform(0.8, 1.3))
                items = page.query_selector_all('a[href*="/maps/place/"]')
                if len(items) == last_count:
                    same_rounds += 1
                    if same_rounds >= 3:
                        break
                else:
                    same_rounds = 0
                last_count = len(items)
    except:
        pass

    # Collect unique links
    seen_hrefs = set()
    unique_items = []
    for item in page.query_selector_all('a[href*="/maps/place/"]'):
        href = item.get_attribute('href') or ''
        if href and href not in seen_hrefs:
            seen_hrefs.add(href)
            unique_items.append((item, href))

    print(f"  [{query}] {len(unique_items)} results", flush=True)

    for idx, (item, href) in enumerate(unique_items):
        try:
            name = ""
            try:
                ne = item.query_selector('.fontHeadlineSmall, .qBF1Pd')
                if ne: name = ne.inner_text().strip()
            except: pass
            if not name:
                try: name = item.get_attribute('aria-label') or ''
                except: pass

            if name in seen_names:
                continue
            seen_names.add(name)

            # Click into detail page
            item.click()
            time.sleep(random.uniform(1.5, 2.5))

            detail = extract_detail(page, town, search_term)
            if detail.get("name"):
                companies.append(detail)
                w = detail.get('website', '')[:30]
                print(f"    [{idx+1}/{len(unique_items)}] {detail['name'][:35]} | web={w} | plz={detail.get('plz','-')}", flush=True)
            elif name:
                detail["name"] = name
                companies.append(detail)

            # Go back to results
            try:
                page.go_back(timeout=8000)
                time.sleep(random.uniform(1, 2))
            except:
                page.goto(url, timeout=15000)
                time.sleep(2)

        except Exception as e:
            print(f"    [{idx+1}] Err: {str(e)[:40]}", flush=True)
            try:
                page.goto(url, timeout=15000)
                time.sleep(2)
            except:
                pass

    return companies


def create_excel(companies):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import CellIsRule
    except:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Google Maps Firmen"

    headers = [
        "#", "Firmenname", "Bewertung", "Anzahl Bewertungen", "Kategorie",
        "Adresse", "PLZ", "Ort", "Telefon", "Website",
        "Breitengrad", "Längengrad", "Google Maps Link",
        "Suchort", "Suchbegriff", "Quelle", "Scraped At"
    ]

    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    tb = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.border = hf, hfill, tb

    for idx, comp in enumerate(companies, 1):
        rd = [idx, comp.get("name",""), comp.get("rating",""), comp.get("review_count",""),
              comp.get("category",""), comp.get("address",""), comp.get("plz",""),
              comp.get("city",""), comp.get("phone",""), comp.get("website",""),
              comp.get("lat",""), comp.get("lon",""), comp.get("gmaps_url",""),
              comp.get("search_town",""), comp.get("search_term",""),
              comp.get("source",""), comp.get("scraped_at","")]
        for col, v in enumerate(rd, 1):
            ws.cell(row=idx+1, column=col, value=v).border = tb

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(companies)+1}"
    widths = [5,40,10,12,30,45,8,25,22,45,12,12,55,25,20,22,22]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i+1)].width = w

    gf = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ws.conditional_formatting.add(f"J2:J{len(companies)+1}", CellIsRule(operator="notEqual", formula=['""'], fill=gf))

    ws2 = wb.create_sheet("Statistiken")
    stats = [
        ("Gesamtzahl Firmen", len(companies)),
        ("Mit Website", sum(1 for c in companies if c.get("website"))),
        ("Ohne Website", sum(1 for c in companies if not c.get("website"))),
        ("Mit Telefon", sum(1 for c in companies if c.get("phone"))),
        ("Mit PLZ", sum(1 for c in companies if c.get("plz"))),
        ("Mit Kategorie", sum(1 for c in companies if c.get("category"))),
        ("Detail-basiert", sum(1 for c in companies if c.get("source")=="google_maps_detail")),
    ]
    for r, (s, v) in enumerate(stats, 1):
        ws2.cell(row=r, column=1, value=s).font = Font(bold=True)
        ws2.cell(row=r, column=2, value=v)
    ws2.column_dimensions['A'].width = 25

    wb.save(EXCEL_OUTPUT)
    print(f"  [EXCEL] {len(companies)} saved", flush=True)


def main():
    print("=== GMaps Detail Re-Scraper ===", flush=True)

    # Kill leftover
    subprocess.run(["pkill", "-9", "Xvfb"], capture_output=True)
    subprocess.run(["pkill", "-9", "chromium"], capture_output=True)
    time.sleep(2)

    # Load data
    progress = load_json(PROGRESS_FILE, {"completed_searches": [], "cities_done": []})
    travel_results = load_json(RESULTS_FILE, [])

    all_companies = {}
    for c in travel_results:
        key = f"{c.get('name','').lower().strip()}|{c.get('city', c.get('search_town','')).lower().strip()}"
        if key not in all_companies:
            all_companies[key] = c
        else:
            ec = all_companies[key]
            for k, v in c.items():
                if v and not ec.get(k):
                    ec[k] = v

    print(f"Loaded {len(all_companies)} companies", flush=True)

    # Determine which city+term combos to search
    cities_done = set(progress.get("cities_done", []))
    cities_to_do = [c for c in PRIORITY_CITIES if c not in cities_done]
    print(f"Cities to process: {len(cities_to_do)}", flush=True)

    # Start Xvfb
    xvfb_proc = subprocess.Popen(
        ['Xvfb', ':99', '-screen', '0', '1920x1080x24', '-ac', '+extension', 'GLX', '+render', '-noreset'],
        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL
    )
    time.sleep(3)
    os.environ['DISPLAY'] = ':99'
    print("Xvfb started", flush=True)

    from playwright.sync_api import sync_playwright

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=False,
            executable_path=CHROME_PATH,
            args=['--no-sandbox', '--disable-gpu', '--disable-dev-shm-usage',
                  '--disable-blink-features=AutomationControlled', '--lang=de-DE']
        )
        context = browser.new_context(
            viewport={"width": 1920, "height": 1080},
            locale="de-DE",
            user_agent="Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
        )
        page = context.new_page()
        print("Browser ready", flush=True)

        push_counter = 0
        total_new = 0

        for city in cities_to_do:
            print(f"\n=== {city} ===", flush=True)

            for term in CATEGORY_TERMS:
                skey = f"{city}|{term}"
                if skey in progress.get("completed_searches", []):
                    continue

                try:
                    companies = search_and_extract_all(page, city, term)
                except Exception as e:
                    print(f"  [ERR] {city} {term}: {e}", flush=True)
                    try:
                        page.goto('about:blank', timeout=5000)
                    except:
                        pass
                    progress.setdefault("completed_searches", []).append(skey)
                    save_json(PROGRESS_FILE, progress)
                    continue

                new_count = 0
                updated_count = 0
                for c in companies:
                    key = f"{c.get('name','').lower().strip()}|{c.get('city', c.get('search_town','')).lower().strip()}"
                    if key not in all_companies:
                        all_companies[key] = c
                        new_count += 1
                    else:
                        # Update missing fields (especially website!)
                        ec = all_companies[key]
                        updated = False
                        if c.get("website") and "google" not in c["website"].lower() and not ec.get("website"):
                            ec["website"] = c["website"]
                            updated = True
                        if c.get("category") and not ec.get("category"):
                            ec["category"] = c["category"]
                            updated = True
                        if c.get("plz") and not ec.get("plz"):
                            ec["plz"] = c["plz"]
                            updated = True
                        if c.get("address") and not ec.get("address"):
                            ec["address"] = c["address"]
                            updated = True
                        if c.get("phone") and not ec.get("phone"):
                            ec["phone"] = c["phone"]
                            updated = True
                        if c.get("city") and not ec.get("city"):
                            ec["city"] = c["city"]
                            updated = True
                        if c.get("lat") and not ec.get("lat"):
                            ec["lat"] = c["lat"]
                        updated = True
                        if c.get("lon") and not ec.get("lon"):
                            ec["lon"] = c["lon"]
                            updated = True
                        if updated:
                            ec["source"] = "google_maps_detail"
                            ec["enriched_at"] = datetime.now().isoformat()
                            updated_count += 1

                progress.setdefault("completed_searches", []).append(skey)
                push_counter += new_count + updated_count
                total_new += new_count

                print(f"  New: {new_count} | Updated: {updated_count} | Total: {len(all_companies)} | Push: {push_counter}", flush=True)

                save_json(RESULTS_FILE, list(all_companies.values()))
                save_json(PROGRESS_FILE, progress)

                if push_counter >= 10:
                    create_excel(list(all_companies.values()))
                    git_push(f"GMaps Detail: {city} {term} | +{push_counter} | {len(all_companies)} total")
                    push_counter = 0

                time.sleep(random.uniform(1, 2))

            cities_done.add(city)
            progress["cities_done"] = list(cities_done)
            save_json(PROGRESS_FILE, progress)
            print(f"  City {city} done!", flush=True)

        # Final save
        create_excel(list(all_companies.values()))
        git_push(f"GMaps Detail COMPLETE | {len(all_companies)} total | {total_new} new")

        browser.close()

    try:
        os.kill(xvfb_proc.pid, signal.SIGTERM)
    except: pass

    print(f"\n=== DONE | {len(all_companies)} companies | {total_new} new ===", flush=True)


if __name__ == "__main__":
    main()
