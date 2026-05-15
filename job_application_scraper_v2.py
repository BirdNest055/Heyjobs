#!/usr/bin/env python3
"""
Job Application Scraper v2 - Improved career page detection + job extraction.
Visits company websites, finds career pages, extracts junior IT positions,
generates customized application emails, and saves to Excel.
"""

import asyncio
import json
import os
import re
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import urljoin, urlparse, urlunparse

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# ── CONFIG ────────────────────────────────────────────────────────────────────
EXCEL_OUT = "/home/z/my-project/download/Bewerbungen_Junior_IT.xlsx"
PROGRESS_FILE = "/home/z/my-project/job_scraper_progress_v2.json"
LOG_FILE = "/home/z/my-project/job_scraper_v2_log.txt"
GITHUB_REPO = "/home/z/my-project"
DISPLAY = ":99"
PUSH_EVERY = 10
MAX_COMPANIES = 500
TIMEOUT_PER_COMPANY = 120
MAX_JOBS_PER_COMPANY = 5

# ── EMAIL TEMPLATE ────────────────────────────────────────────────────────────
EMAIL_TEMPLATE = """Sehr geehrte{anrede_suffix} {anrede_target},

ich sende Ihnen hiermit meine Bewerbungsunterlagen für die Stelle als {stelle} in Vollzeit.

Ich schließe meine Ausbildung zum Fachinformatiker für Systemintegration bald ab. Das Fachgespräch findet am 11. Juni 2026 statt, ab diesem Zeitpunkt stehe ich für eine Einstellung zur Verfügung.

Während meiner Ausbildung hatte ich Berührungspunkte mit verschiedenen IT-Bereichen: Systemintegration, Webentwicklung und Datenbankentwicklung.

In meiner letzten Praxisphase habe ich mich mit Azure beschäftigt, was direkt in mein IHK-Abschlussprojekt geflossen ist: die automatisierte Anwendungsbereitstellung via App Attach in Azure Virtual Desktop, von der Planung bis zur Dokumentation eigenständig umgesetzt.

Ich arbeite selbstständig und kommuniziere offen. Das haben mir auch meine Betreuer in den Praxisphasen zurückgemeldet. Einen Führerschein der Klasse B besitze ich aktuell nicht, bin aber bereit, diesen zeitnah zu erwerben.

Meine vollständigen Unterlagen finden Sie im Anhang. Ich freue mich auf ein persönliches Gespräch.

Mit freundlichen Grüßen,"""


def log(msg):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line, flush=True)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")


def load_progress():
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"processed": [], "jobs": [], "last_index": 0}


def save_progress(progress):
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump(progress, f, ensure_ascii=False, indent=2)


def git_push(msg):
    try:
        subprocess.run(["git", "add", "-A"], cwd=GITHUB_REPO, capture_output=True, timeout=30)
        result = subprocess.run(["git", "commit", "-m", msg], cwd=GITHUB_REPO, capture_output=True, timeout=30)
        result = subprocess.run(["git", "push", "origin", "main"], cwd=GITHUB_REPO,
                                capture_output=True, text=True, timeout=60)
        if result.returncode == 0:
            log(f"✅ Pushed: {msg}")
        else:
            log(f"⚠️ Push failed: {result.stderr[:200]}")
    except Exception as e:
        log(f"⚠️ Git error: {e}")


def normalize_url(url):
    """Clean up a URL."""
    if not url:
        return ""
    url = url.strip()
    if not url.startswith(("http://", "https://")):
        url = "https://" + url
    # Remove trailing slashes
    return url.rstrip("/")


def get_base_url(url):
    """Extract base URL (scheme + netloc)."""
    parsed = urlparse(url)
    return f"{parsed.scheme}://{parsed.netloc}"


def load_companies():
    """Load and merge company data from all sources, sorted by IT relevance."""
    companies = {}

    # 1. Google Maps data (11k+)
    wb1 = openpyxl.load_workbook("/home/z/my-project/download/Google_Maps_Firmen_Erlangen_50km.xlsx")
    ws1 = wb1["Google Maps Firmen"]
    for r in range(2, ws1.max_row + 1):
        name = (ws1.cell(r, 2).value or "").strip()
        if not name:
            continue
        cat = ws1.cell(r, 5).value or ""
        adresse = ws1.cell(r, 6).value or ""
        plz = ws1.cell(r, 7).value or ""
        ort = ws1.cell(r, 8).value or ""
        telefon = ws1.cell(r, 9).value or ""
        website = ws1.cell(r, 10).value or ""

        if name not in companies:
            companies[name] = {
                "name": name, "kategorie": cat, "adresse": adresse,
                "plz": plz, "ort": ort, "telefon": telefon, "website": website,
                "email": "", "ansprechpartner": ""
            }
        elif website and not companies[name]["website"]:
            companies[name]["website"] = website

    # 2. Enriched data (websites + emails)
    wb2 = openpyxl.load_workbook("/home/z/my-project/download/Firmen_Bamberg_Erlangen_Nuernberg_mit_Website_HTML.xlsx")
    ws2 = wb2["Firmen mit Website & HTML"]
    for r in range(2, ws2.max_row + 1):
        name = (ws2.cell(r, 2).value or "").strip()
        if not name:
            continue
        web_found = ws2.cell(r, 10).value or ""
        email = ws2.cell(r, 11).value or ""

        if name in companies:
            if web_found and not companies[name]["website"]:
                companies[name]["website"] = web_found
            if email and not companies[name]["email"]:
                companies[name]["email"] = email
        else:
            companies[name] = {
                "name": name, "kategorie": ws2.cell(r, 4).value or "",
                "adresse": ws2.cell(r, 5).value or "",
                "plz": ws2.cell(r, 7).value or "",
                "ort": ws2.cell(r, 8).value or "",
                "telefon": ws2.cell(r, 9).value or "",
                "website": web_found, "email": email,
                "ansprechpartner": ""
            }

    # 3. Arbeitsagentur jobs for confirmed IT employers
    wb3 = openpyxl.load_workbook("/home/z/my-project/download/Junior_IT_Jobs_Bamberg_Erlangen_Nuernberg.xlsx")
    ws3 = wb3["Junior IT-Jobs"]
    aa_jobs = {}
    for r in range(2, ws3.max_row + 1):
        employer = (ws3.cell(r, 22).value or "").strip()
        job_title = ws3.cell(r, 2).value or ""
        job_url = ws3.cell(r, 27).value or ""
        job_ort = ws3.cell(r, 18).value or ""
        email_col = None
        for c in range(1, ws3.max_column + 1):
            if ws3.cell(1, c).value and "mail" in str(ws3.cell(1, c).value).lower():
                email_col = c
        aa_email = ws3.cell(r, email_col).value if email_col else ""
        if employer:
            if employer not in aa_jobs:
                aa_jobs[employer] = []
            aa_jobs[employer].append({
                "title": job_title, "url": job_url, "ort": job_ort, "email": aa_email or ""
            })

    # Score for IT relevance
    it_keywords = [
        'it ', 'it-', 'software', 'data ', 'daten', 'technolog', 'digital',
        'systemhaus', 'informatik', 'cyber', 'cloud', 'tech', 'computer',
        'netzwerk', 'consulting', 'entwicklung', 'systemintegrat',
        'admin', 'devops', 'sre', 'platform', 'engineer',
        'sap', 'oracle', 'microsoft', 'linux', 'server',
        'service', 'lösung', 'solution', 'automatisier',
        'maschinenbau', 'industrie', 'elektro', 'energie',
        'bank', 'versicher', 'finanz', 'klinik', 'university',
        'universität', 'forschung', 'research', 'gruppe', 'holding',
        'telekommunikat', 'logistik', 'versand', 'handel'
    ]

    scored = []
    for name, data in companies.items():
        if not data["website"]:
            continue
        score = 0
        text = f"{name} {data['kategorie']}".lower()
        for kw in it_keywords:
            if kw in text:
                score += 1
        if name in aa_jobs:
            score += 5
        scored.append((score, name, data))

    scored.sort(key=lambda x: -x[0])

    result = []
    for score, name, data in scored:
        data["relevance_score"] = score
        data["arbeitsagentur_jobs"] = aa_jobs.get(name, [])
        result.append(data)

    log(f"Loaded {len(result)} companies with websites (sorted by IT relevance)")
    return result


def customize_email(stelle, firma, ansprechpartner=""):
    """Generate customized email text from template."""
    if ansprechpartner and ansprechpartner.lower() not in ["n/a", "unbekannt", "", "damen und herren"]:
        # Determine gender
        parts = ansprechpartner.strip().split()
        first_name = parts[-1] if parts else ""  # Last word is often first name in German

        female_names = ['anna', 'maria', 'sabine', 'monika', 'katrin', 'sandra',
                        'nicole', 'stephanie', 'julia', 'laura', 'sarah', 'lena',
                        'christine', 'barbara', 'heike', 'petra', 'susanne', 'angela',
                        'michaela', 'silvia', 'daniela', 'elke', 'birgit', 'ute',
                        'claudia', 'tanja', 'jennifer', 'lisa', 'marie', 'kathrin']

        if any(t in ansprechpartner.lower() for t in ['frau', 'ms.', 'mrs.']):
            anrede_suffix = ""
            anrede_target = f"Frau {ansprechpartner.replace('Frau', '').replace('Mrs.', '').replace('Ms.', '').strip()}"
        elif any(t in ansprechpartner.lower() for t in ['herr', 'hr.', 'mr.']):
            anrede_suffix = "r"
            anrede_target = f"Herr {ansprechpartner.replace('Herr', '').replace('Hr.', '').replace('Mr.', '').strip()}"
        elif first_name.lower() in female_names:
            anrede_suffix = ""
            anrede_target = f"Frau {ansprechpartner}"
        else:
            anrede_suffix = "r"
            anrede_target = f"Herr {ansprechpartner}"
    else:
        anrede_suffix = "r"
        anrede_target = "Damen und Herren"

    return EMAIL_TEMPLATE.format(
        anrede_suffix=anrede_suffix,
        anrede_target=anrede_target,
        stelle=stelle
    )


def extract_emails_from_text(text):
    """Find all email addresses in text."""
    if not text:
        return []
    pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
    return list(set(re.findall(pattern, text)))


def is_it_job_title(text):
    """Check if text looks like an IT job title."""
    text_lower = text.lower()
    it_patterns = [
        'it-', 'it ', 'informatik', 'software', 'systemintegration',
        'systemadmin', 'netzwerk', 'devops', 'cloud', 'data engineer',
        'data scien', 'datenbank', 'cyber', 'security', 'entwickler',
        'developer', 'administrator', 'linux', 'sap ', 'azure', 'aws ',
        'frontend', 'backend', 'fullstack', 'full-stack', 'sre ',
        'platform engineer', 'infrastruktur', 'infrastructure',
        'it-support', 'helpdesk', 'it-consultant', 'it berater',
        'it manager', 'it specialist', 'system engineer', 'network',
        'automation', 'automatisier', 'test', 'qa ', 'quality',
        'scrum master', 'product owner', 'agile', 'desktop',
        'application', 'anwendungs', 'digital', 'webentwickler',
        'web developer', 'mobile', 'app entwickler', 'dev ',
        'ux ', 'ui ', 'user experience', 'user interface',
        'machine learning', 'ai ', 'artificial intelligence',
        'künstliche intelligenz', 'projektmanag', 'project manag',
    ]
    return any(p in text_lower for p in it_patterns)


def is_junior_title(text):
    """Check if text indicates junior-level position."""
    text_lower = text.lower()
    junior_patterns = [
        'junior', 'trainee', 'entry', 'graduate', 'start',
        'berufseinsteig', 'jun.', 'young professional',
        'werkstudent', 'praktikant', 'azubi', 'ausbildung',
        '1st level', 'first level', 'level 1', 'beginner',
    ]
    # Also accept positions that don't specify level (often open to juniors)
    senior_patterns = ['senior', 'lead', 'manager', 'director', 'head of',
                       'vp ', 'chief', 'c-level', 'principal', 'staff']
    has_junior = any(p in text_lower for p in junior_patterns)
    has_senior = any(p in text_lower for p in senior_patterns)
    if has_junior:
        return "explicit_junior"
    elif has_senior:
        return "senior"
    else:
        return "open_level"


def clean_job_title(raw_text, max_len=150):
    """Extract a clean job title from raw element text."""
    # Take the first meaningful line
    lines = [l.strip() for l in raw_text.split("\n") if l.strip()]
    if not lines:
        return raw_text[:max_len]

    # Often the title is the first or second line
    for line in lines[:3]:
        line = line.strip(" →|•►▪▶↳-–—")
        if 10 <= len(line) <= max_len:
            return line
    return lines[0][:max_len]


async def scrape_company_jobs(page, company):
    """Visit company website, find career pages, extract IT jobs."""
    url = normalize_url(company["website"])
    if not url:
        return [], "", ""

    base_url = get_base_url(url)
    jobs_found = []
    contact_email = company.get("email", "")
    ansprechpartner = company.get("ansprechpartner", "")

    # ── Step 1: Visit main page, collect career links ──
    try:
        log(f"  🌐 Main page: {url}")
        resp = await page.goto(url, wait_until="domcontentloaded", timeout=25000)
        if not resp or resp.status >= 500:
            log(f"  ⚠️ HTTP error for main page")
            return [], contact_email, ansprechpartner
    except Exception as e:
        log(f"  ⚠️ Failed to load main page: {str(e)[:80]}")
        return [], contact_email, ansprechpartner

    await page.wait_for_timeout(2000)

    # Collect career page links from main page
    career_links = []

    try:
        links = await page.evaluate("""() => {
            const results = [];
            document.querySelectorAll('a[href]').forEach(a => {
                const text = (a.textContent || '').trim().toLowerCase();
                const href = (a.getAttribute('href') || '').toLowerCase();
                const keywords = ['karriere', 'jobs', 'stellen', 'career', 'bewer', 
                                   'offene', 'vacanc', 'position', 'ausschreib',
                                   'unser team', 'arbeit bei', 'work with', 'join us',
                                   'work at', 'now hiring'];
                const isCareer = keywords.some(kw => text.includes(kw) || href.includes(kw));
                if (isCareer && a.href) {
                    results.push({href: a.href, text: (a.textContent || '').trim().substring(0, 100)});
                }
            });
            return results;
        }""")

        for link in links[:10]:
            href = link["href"]
            # Only same domain or subdomain
            if urlparse(href).netloc.replace("www.", "") == urlparse(base_url).netloc.replace("www.", ""):
                career_links.append(href)
    except Exception as e:
        log(f"  ⚠️ Link extraction error: {str(e)[:60]}")

    # Add standard career paths
    standard_paths = [
        "/karriere", "/jobs", "/stellenangebote", "/career", "/careers",
        "/de/karriere", "/de/jobs", "/de/career", "/en/career",
        "/offene-stellen", "/bewerbung", "/vacancies", "/unser-team",
        "/wir-haben-platz", "/work-with-us", "/join-us",
    ]
    for path in standard_paths:
        full = base_url + path
        if full not in career_links:
            career_links.append(full)

    log(f"  🔗 {len(career_links)} career links to check")

    # ── Step 2: Visit career pages and look for IT jobs ──
    visited = set()
    for career_url in career_links[:8]:
        if career_url in visited:
            continue
        visited.add(career_url)

        try:
            log(f"  📄 Checking: {career_url}")
            resp = await page.goto(career_url, wait_until="domcontentloaded", timeout=20000)
            if not resp or resp.status >= 400:
                continue
            await page.wait_for_timeout(2500)
        except:
            continue

        # Get full page text
        try:
            page_text = await page.inner_text("body")
        except:
            continue

        # Quick check: does this page have IT-related content at all?
        if not is_it_job_title(page_text[:3000]):
            continue

        # Extract job listings from the page
        try:
            # Method 1: Look for structured job listing elements
            job_items = await page.evaluate("""() => {
                const jobs = [];
                // Common selectors for job listings
                const selectors = [
                    'article', '.job', '.vacancy', '.position', '.stellenanzeige',
                    '.job-listing', '.job-item', '.job-entry', '.vacancy-item',
                    '.career-item', '.position-item', '[data-job]', '[data-vacancy]',
                    '.list-item', '.card', '.teaser', '.entry',
                    'li', 'tr', '.row'
                ];
                
                for (const sel of selectors) {
                    const elements = document.querySelectorAll(sel);
                    for (const el of elements) {
                        const text = (el.textContent || '').trim();
                        if (text.length < 15 || text.length > 2000) continue;
                        
                        const textLower = text.toLowerCase();
                        // Check for IT keywords
                        const itKws = ['it-', 'it ', 'informatik', 'software', 'systemintegration',
                                       'entwickler', 'developer', 'administrator', 'admin', 'devops',
                                       'cloud', 'netzwerk', 'network', 'data ', 'linux', 'sap',
                                       'azure', 'aws', 'security', 'cyber', 'frontend', 'backend',
                                       'fullstack', 'infrastruktur', 'support', 'consultant',
                                       'engineer', 'architekt', 'automatisier', 'test', 'digital',
                                       'webentwickl', 'datenbank', 'scrum', 'projektmanag'];
                        const hasIT = itKws.some(kw => textLower.includes(kw));
                        if (!hasIT) continue;
                        
                        // Get the title (first meaningful line)
                        const lines = text.split('\\n').map(l => l.trim()).filter(l => l.length > 3);
                        let title = lines[0] || '';
                        if (title.length > 150) title = title.substring(0, 150);
                        
                        // Get link
                        const link = el.closest('a') || el.querySelector('a');
                        const href = link ? link.getAttribute('href') : '';
                        
                        // Avoid duplicates
                        const key = title.toLowerCase().substring(0, 50);
                        if (jobs.some(j => j.key === key)) continue;
                        
                        jobs.push({
                            title: title,
                            href: href || '',
                            text: text.substring(0, 500),
                            key: key
                        });
                        
                        if (jobs.length >= 15) break;
                    }
                    if (jobs.length >= 15) break;
                }
                return jobs;
            }""")

            for item in job_items:
                title = item["title"].strip()
                if len(title) < 5:
                    continue

                # Determine if junior
                junior_status = is_junior_title(title)
                if junior_status == "senior":
                    continue  # Skip senior positions

                # Build full URL
                job_url = ""
                if item["href"]:
                    job_url = urljoin(career_url, item["href"])

                jobs_found.append({
                    "title": title,
                    "url": job_url,
                    "description": item["text"][:500],
                    "is_junior": junior_status == "explicit_junior"
                })

                if len(jobs_found) >= MAX_JOBS_PER_COMPANY:
                    break

        except Exception as e:
            log(f"  ⚠️ Job extraction error: {str(e)[:60]}")

        # If we found jobs, also check for contact info on this page
        if jobs_found:
            # Extract emails
            if not contact_email:
                emails = extract_emails_from_text(page_text)
                for em in emails:
                    if not any(g in em.lower() for g in ['noreply', 'no-reply', 'spam', 'example.com', 'test']):
                        contact_email = em
                        break
                if not contact_email and emails:
                    contact_email = emails[0]

            # Try to find Ansprechpartner
            if not ansprechpartner:
                for line in page_text.split("\n"):
                    line = line.strip()
                    if any(t in line.lower() for t in ['herr', 'frau']) and \
                       any(t in line.lower() for t in ['kontakt', 'bewer', 'personal', 'hr', 'ansprech']):
                        ansprechpartner = line[:80]
                        break

            break  # Found jobs, no need to check more career pages

    # ── Step 3: If no jobs found on career pages, check Impressum for contact ──
    if not contact_email:
        for imp_path in ["/impressum", "/kontakt", "/contact"]:
            try:
                imp_url = base_url + imp_path
                resp = await page.goto(imp_url, wait_until="domcontentloaded", timeout=12000)
                if resp and resp.status < 400:
                    await page.wait_for_timeout(1500)
                    imp_text = await page.inner_text("body")
                    emails = extract_emails_from_text(imp_text)
                    for em in emails:
                        if not any(g in em.lower() for g in ['noreply', 'no-reply', 'spam']):
                            contact_email = em
                            break
                    if not contact_email and emails:
                        contact_email = emails[0]

                    # Ansprechpartner from Impressum
                    if not ansprechpartner:
                        for line in imp_text.split("\n"):
                            line = line.strip()
                            if any(t in line.lower() for t in ['herr', 'frau']):
                                if any(t in line.lower() for t in ['personal', 'hr', 'bewer', 'kontakt', 'ansprech']):
                                    ansprechpartner = line[:80]
                                    break
                    if contact_email:
                        break
            except:
                continue

    # ── Step 4: Also check Arbeitsagentur data for this company ──
    aa_jobs = company.get("arbeitsagentur_jobs", [])
    for aa_job in aa_jobs:
        title = aa_job["title"]
        junior_status = is_junior_title(title)
        if junior_status == "senior":
            continue

        # Check if we already have this job
        existing_titles = {j["title"].lower()[:40] for j in jobs_found}
        if title.lower()[:40] in existing_titles:
            continue

        jobs_found.append({
            "title": title,
            "url": aa_job.get("url", ""),
            "description": f"Quelle: Arbeitsagentur",
            "is_junior": junior_status == "explicit_junior"
        })

        if aa_job.get("email") and not contact_email:
            contact_email = aa_job["email"]

    # Deduplicate
    seen = set()
    unique_jobs = []
    for j in jobs_found:
        key = j["title"].lower().strip()[:50]
        if key not in seen:
            seen.add(key)
            j["email"] = contact_email
            j["ansprechpartner"] = ansprechpartner
            unique_jobs.append(j)
        if len(unique_jobs) >= MAX_JOBS_PER_COMPANY:
            break

    return unique_jobs, contact_email, ansprechpartner


async def main():
    log("=" * 60)
    log("Job Application Scraper v2 - Starting")
    log("=" * 60)

    companies = load_companies()
    progress = load_progress()
    processed_names = set(progress["processed"])
    jobs_all = progress["jobs"]
    start_idx = progress.get("last_index", 0)

    log(f"Already processed: {len(processed_names)}")
    log(f"Jobs found so far: {len(jobs_all)}")

    # Filter to unprocessed
    to_process = [(i, c) for i, c in enumerate(companies)
                  if c["name"] not in processed_names and c["website"]]
    log(f"Companies to process: {len(to_process)}")

    if not to_process:
        log("Nothing to process!")
        save_excel(jobs_all)
        return

    to_process = to_process[:MAX_COMPANIES]

    # Launch browser
    from playwright.async_api import async_playwright

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=False,
            args=[
                "--no-sandbox",
                "--disable-dev-shm-usage",
                "--window-size=1920,1080",
                "--disable-blink-features=AutomationControlled",
            ]
        )

        context = await browser.new_context(
            viewport={"width": 1920, "height": 1080},
            user_agent="Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
            locale="de-DE",
        )

        page = await context.new_page()
        page.set_default_timeout(25000)
        page.set_default_navigation_timeout(25000)

        processed_count = 0
        total_new_jobs = 0

        for idx, (orig_idx, company) in enumerate(to_process):
            company_name = company["name"]
            log(f"\n[{idx+1}/{len(to_process)}] {company_name} (score: {company.get('relevance_score', 0)})")

            try:
                jobs, email, ansprech = await asyncio.wait_for(
                    scrape_company_jobs(page, company),
                    timeout=TIMEOUT_PER_COMPANY
                )
            except asyncio.TimeoutError:
                log(f"  ⏰ Timeout for {company_name}")
                jobs, email, ansprech = [], "", ""
            except Exception as e:
                log(f"  ❌ Error: {str(e)[:80]}")
                jobs, email, ansprech = [], "", ""

            if jobs:
                log(f"  ✅ Found {len(jobs)} job(s)!")
                for job in jobs:
                    email_text = customize_email(
                        stelle=job["title"],
                        firma=company_name,
                        ansprechpartner=job.get("ansprechpartner", ansprech)
                    )

                    adresse = company.get("adresse", "")
                    if company.get("plz"):
                        adresse += f", {company['plz']}"
                    if company.get("ort"):
                        adresse += f" {company['ort']}"
                    adresse = adresse.strip(", ")

                    job_entry = {
                        "firma": company_name,
                        "stelle": job["title"],
                        "stellenbeschreibung": job.get("description", ""),
                        "email": job.get("email", email or company.get("email", "")),
                        "ansprechpartner": job.get("ansprechpartner", ansprech),
                        "bewerbung_geschickt": "",
                        "letzter_kontakt": "",
                        "status": "Offen",
                        "website": company["website"],
                        "adresse": adresse,
                        "email_text": email_text,
                        "job_url": job.get("url", ""),
                        "is_junior": job.get("is_junior", False),
                        "scraped_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    }
                    jobs_all.append(job_entry)
                    total_new_jobs += 1
            else:
                log(f"  ℹ️ No IT jobs found")

            processed_names.add(company_name)
            processed_count += 1

            # Save progress
            progress["processed"] = list(processed_names)
            progress["jobs"] = jobs_all
            progress["last_index"] = orig_idx
            save_progress(progress)

            # Push every 10
            if processed_count % PUSH_EVERY == 0:
                log(f"\n📦 Checkpoint: {processed_count} processed, {len(jobs_all)} jobs")
                save_excel(jobs_all)
                git_push(f"Job Scraper v2: {processed_count} companies | {len(jobs_all)} jobs")

            await page.wait_for_timeout(800)

        await browser.close()

    # Final save
    log(f"\n{'=' * 60}")
    log(f"COMPLETE! {processed_count} companies, {total_new_jobs} new jobs")
    log(f"Total jobs: {len(jobs_all)}")

    save_excel(jobs_all)
    git_push(f"Job Scraper v2 COMPLETE: {processed_count} co | {len(jobs_all)} jobs | {datetime.now().strftime('%m-%d %H:%M')}")

    progress["last_run"] = datetime.now().isoformat()
    save_progress(progress)


def save_excel(jobs):
    """Generate the application-ready Excel file."""
    wb = openpyxl.Workbook()

    # ── Main Sheet ──
    ws = wb.active
    ws.title = "Bewerbungen"

    headers = [
        "Firma", "Stelle", "Stellenbeschreibung", "Email", "Ansprechpartner",
        "Bewerbung geschickt am", "Letzter Kontakt am", "Status", "Website",
        "Adresse", "Email Text", "Job-URL", "Junior?", "Scraped At"
    ]

    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, job in enumerate(jobs, 2):
        values = [
            job.get("firma", ""),
            job.get("stelle", ""),
            job.get("stellenbeschreibung", ""),
            job.get("email", ""),
            job.get("ansprechpartner", ""),
            job.get("bewerbung_geschickt", ""),
            job.get("letzter_kontakt", ""),
            job.get("status", "Offen"),
            job.get("website", ""),
            job.get("adresse", ""),
            job.get("email_text", ""),
            job.get("job_url", ""),
            "Ja" if job.get("is_junior") else "Nein",
            job.get("scraped_at", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    col_widths = {
        1: 30, 2: 40, 3: 50, 4: 30, 5: 22,
        6: 18, 7: 18, 8: 14, 9: 35, 10: 35,
        11: 70, 12: 40, 13: 10, 14: 16
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(jobs) + 1}"

    # Status dropdown
    status_dv = DataValidation(
        type="list",
        formula1='"Offen,Beworben,Interview eingeplant,Interview absolviert,Zusage,Absage,Kontaktiert,Wiedervorlage,Kein Interesse"',
        allow_blank=True
    )
    status_dv.error = "Bitte wählen Sie einen gültigen Status"
    status_dv.errorTitle = "Ungültiger Status"
    ws.add_data_validation(status_dv)
    for row in range(2, len(jobs) + 2):
        status_dv.add(ws.cell(row=row, column=8))

    # Conditional formatting
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    orange_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")

    last_row = len(jobs) + 1
    ws.conditional_formatting.add(f"H2:H{last_row}",
        CellIsRule(operator="equal", formula=['"Zusage"'], fill=green_fill))
    ws.conditional_formatting.add(f"H2:H{last_row}",
        CellIsRule(operator="equal", formula=['"Absage"'], fill=red_fill))
    ws.conditional_formatting.add(f"H2:H{last_row}",
        CellIsRule(operator="equal", formula=['"Beworben"'], fill=blue_fill))
    ws.conditional_formatting.add(f"H2:H{last_row}",
        CellIsRule(operator="equal", formula=['"Interview eingeplant"'], fill=yellow_fill))
    ws.conditional_formatting.add(f"H2:H{last_row}",
        CellIsRule(operator="equal", formula=['"Interview absolviert"'], fill=orange_fill))
    ws.conditional_formatting.add(f"M2:M{last_row}",
        CellIsRule(operator="equal", formula=['"Ja"'], fill=green_fill))

    # ── Email Templates Sheet ──
    ws_emails = wb.create_sheet("Email Vorlagen")
    email_headers = ["Firma", "Stelle", "Email an", "Betreff", "Email Text (fertig zum Kopieren)"]
    for col, h in enumerate(email_headers, 1):
        ws_emails.cell(1, column=col, value=h).font = Font(bold=True, size=11)

    for row_idx, job in enumerate(jobs, 2):
        ws_emails.cell(row_idx, column=1, value=job.get("firma", ""))
        ws_emails.cell(row_idx, column=2, value=job.get("stelle", ""))
        ws_emails.cell(row_idx, column=3, value=job.get("email", ""))
        betreff = f"Bewerbung als {job.get('stelle', 'Fachinformatiker Systemintegration')} - {job.get('firma', '')}"
        ws_emails.cell(row_idx, column=4, value=betreff)
        ws_emails.cell(row_idx, column=5, value=job.get("email_text", ""))

    ws_emails.column_dimensions["A"].width = 30
    ws_emails.column_dimensions["B"].width = 40
    ws_emails.column_dimensions["C"].width = 30
    ws_emails.column_dimensions["D"].width = 55
    ws_emails.column_dimensions["E"].width = 80
    ws_emails.freeze_panes = "A2"

    # ── Statistics Sheet ──
    ws_stats = wb.create_sheet("Statistiken")

    stats = [
        ("Bewerbungen Junior IT - Statistiken", ""),
        ("Stand", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("", ""),
        ("Gesamt Jobs gefunden", len(jobs)),
        ("Explizit Junior-Jobs", sum(1 for j in jobs if j.get("is_junior"))),
        ("Mit Email-Adresse", sum(1 for j in jobs if j.get("email"))),
        ("Mit Ansprechpartner", sum(1 for j in jobs if j.get("ansprechpartner"))),
        ("", ""),
        ("Nach Status:", ""),
    ]

    status_counts = {}
    for j in jobs:
        s = j.get("status", "Offen")
        status_counts[s] = status_counts.get(s, 0) + 1
    for status, count in sorted(status_counts.items()):
        stats.append((f"  {status}", count))

    stats.extend([("", ""), ("Nach Ort:", "")])
    ort_counts = {}
    for j in jobs:
        addr = j.get("adresse", "").lower()
        matched = False
        for city in ["Bamberg", "Erlangen", "Nürnberg", "Fürth", "Würzburg",
                      "Bayreuth", "Schweinfurt", "Coburg", "Hof", "Amberg",
                      "Regensburg", "Ingolstadt", "Herzogenaurach", "Forchheim",
                      "Bamberg", "Erlangen"]:
            if city.lower() in addr:
                ort_counts[city] = ort_counts.get(city, 0) + 1
                matched = True
                break
        if not matched:
            ort_counts["Sonstige"] = ort_counts.get("Sonstige", 0) + 1
    for ort, count in sorted(ort_counts.items(), key=lambda x: -x[1]):
        stats.append((f"  {ort}", count))

    for row_idx, (label, value) in enumerate(stats, 1):
        c1 = ws_stats.cell(row=row_idx, column=1, value=label)
        c2 = ws_stats.cell(row=row_idx, column=2, value=value)
        if row_idx == 1 or "Nach" in str(label):
            c1.font = Font(bold=True, size=12 if row_idx == 1 else 11)

    ws_stats.column_dimensions["A"].width = 35
    ws_stats.column_dimensions["B"].width = 15

    wb.save(EXCEL_OUT)
    log(f"📊 Excel saved: {EXCEL_OUT} ({len(jobs)} jobs)")


if __name__ == "__main__":
    os.environ["DISPLAY"] = DISPLAY
    asyncio.run(main())
