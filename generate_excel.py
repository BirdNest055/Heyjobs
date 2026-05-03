#!/usr/bin/env python3
"""Generate professional Excel file from Arbeitsagentur employer data"""
import json
import sys
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

# Import base styles
XLSX_SKILL_DIR = "/home/z/my-project/skills/xlsx"
sys.path.insert(0, XLSX_SKILL_DIR)
sys.path.insert(0, os.path.join(XLSX_SKILL_DIR, "templates"))

from templates.base import (
    PRIMARY, PRIMARY_LIGHT, SECONDARY, NEUTRAL_900, NEUTRAL_600,
    NEUTRAL_200, NEUTRAL_100, NEUTRAL_0, HEADER_TEXT,
    FONT_NAME, HEADER_BOLD,
    font_title, font_header, font_subheader, font_body, font_caption,
    fill_header, fill_total, fill_data_row,
    border_header, border_total,
    align_title, align_header, align_text, align_number,
    setup_sheet, style_header_row, style_data_row, style_total_row, auto_fit_columns,
)

# Load data
INPUT_FILE = '/home/z/my-project/download/arbeitsagentur_employers.json'
OUTPUT_FILE = '/home/z/my-project/download/Arbeitgeber_Arbeitsagentur_Bewerbungskontakte.xlsx'

with open(INPUT_FILE, 'r', encoding='utf-8') as f:
    raw_data = json.load(f)

# Clean data
data = [e for e in raw_data if e is not None and e.get('Name')]

print(f"Loaded {len(data)} employers")

# ========== Create workbook ==========
wb = Workbook()

# ========== Sheet 1: Arbeitgeberverzeichnis ==========
ws1 = wb.active
ws1.title = "Arbeitgeberverzeichnis"

# Columns: Arbeitgeber_ID, Name, Straße, PLZ, Stadt, Bundesland, Telefon, E-Mail, Website, Branche, Quelle, Bewerbungsstatus
columns = ['Arbeitgeber_ID', 'Name', 'Straße', 'PLZ', 'Stadt', 'Bundesland', 'Telefon', 'E-Mail', 'Website', 'Branche', 'Quelle', 'Bewerbungsstatus']
col_widths = [16, 35, 25, 8, 22, 22, 18, 30, 35, 30, 16, 18]

# Title
setup_sheet(ws1, title="Arbeitgeberverzeichnis - Arbeitsagentur (IT/Telekommunikation)", last_col=len(columns)+1)

# Header row (row 4)
header_row = 4
for col_idx, col_name in enumerate(columns, start=2):
    cell = ws1.cell(row=header_row, column=col_idx, value=col_name)
style_header_row(ws1, header_row, 2, len(columns)+1)

# Data rows
for row_idx, emp in enumerate(data):
    excel_row = header_row + 1 + row_idx
    row_data = [
        emp.get('Arbeitgeber_ID', ''),
        emp.get('Name', ''),
        emp.get('Straße', ''),
        emp.get('PLZ', ''),
        emp.get('Stadt', ''),
        emp.get('Bundesland', ''),
        emp.get('Telefon', ''),
        emp.get('E-Mail', ''),
        emp.get('Website', ''),
        emp.get('Branche', ''),
        emp.get('Quelle', ''),
        emp.get('Bewerbungsstatus', 'Offen'),
    ]
    for col_idx, value in enumerate(row_data, start=2):
        cell = ws1.cell(row=excel_row, column=col_idx, value=value or '')
    
    style_data_row(ws1, excel_row, 2, len(columns)+1, row_idx)
    
    # Make Website clickable
    website = emp.get('Website', '')
    if website and website.startswith('http'):
        cell = ws1.cell(row=excel_row, column=10)  # Website column
        cell.hyperlink = website
        cell.font = Font(name=FONT_NAME, size=11, color="0563C1", underline="single")

# Set column widths
ws1.column_dimensions['A'].width = 3  # margin
for i, width in enumerate(col_widths, start=2):
    ws1.column_dimensions[get_column_letter(i)].width = width

# Freeze panes
ws1.freeze_panes = 'B5'

# AutoFilter
ws1.auto_filter.ref = f"B4:{get_column_letter(len(columns)+1)}{header_row + len(data)}"

# Data validation for Bewerbungsstatus (column 13 = M)
status_validation = DataValidation(
    type="list",
    formula1='"Offen,Beworben,Interview,Angebot,Abgelehnt,Zurückgezogen"',
    allow_blank=True
)
status_validation.error = "Bitte wählen Sie einen gültigen Status"
status_validation.errorTitle = "Ungültiger Status"
ws1.add_data_validation(status_validation)
status_col = columns.index('Bewerbungsstatus') + 2
for row in range(header_row + 1, header_row + 1 + len(data)):
    status_validation.add(ws1.cell(row=row, column=status_col))

# ========== Sheet 2: Statistik Bundesland ==========
ws2 = wb.create_sheet("Statistik Bundesland")
setup_sheet(ws2, title="Statistik nach Bundesland", last_col=5)

# Count employers per Bundesland
bl_counts = {}
for emp in data:
    bl = emp.get('Bundesland', 'Unbekannt')
    if bl:
        bl_counts[bl] = bl_counts.get(bl, 0) + 1

# Sort by count descending
bl_sorted = sorted(bl_counts.items(), key=lambda x: -x[1])

# Headers
headers2 = ['Bundesland', 'Anzahl Arbeitgeber', 'Anteil (%)']
for col_idx, h in enumerate(headers2, start=2):
    ws2.cell(row=4, column=col_idx, value=h)
style_header_row(ws2, 4, 2, 4)

total = len(data)
for row_idx, (bl, count) in enumerate(bl_sorted):
    excel_row = 5 + row_idx
    ws2.cell(row=excel_row, column=2, value=bl)
    ws2.cell(row=excel_row, column=3, value=count)
    ws2.cell(row=excel_row, column=4, value=round(count / total * 100, 1) if total else 0)
    style_data_row(ws2, excel_row, 2, 4, row_idx)

# Total row
total_row = 5 + len(bl_sorted)
ws2.cell(row=total_row, column=2, value='Gesamt')
ws2.cell(row=total_row, column=3, value=total)
ws2.cell(row=total_row, column=4, value=100.0)
style_total_row(ws2, total_row, 2, 4)

ws2.column_dimensions['A'].width = 3
ws2.column_dimensions['B'].width = 28
ws2.column_dimensions['C'].width = 20
ws2.column_dimensions['D'].width = 14
ws2.freeze_panes = 'B5'

# ========== Sheet 3: Statistik Branche ==========
ws3 = wb.create_sheet("Statistik Branche")
setup_sheet(ws3, title="Statistik nach Branche", last_col=5)

branche_counts = {}
for emp in data:
    br = emp.get('Branche', 'Unbekannt')
    if br:
        branche_counts[br] = branche_counts.get(br, 0) + 1

br_sorted = sorted(branche_counts.items(), key=lambda x: -x[1])

headers3 = ['Branche', 'Anzahl Arbeitgeber', 'Anteil (%)']
for col_idx, h in enumerate(headers3, start=2):
    ws3.cell(row=4, column=col_idx, value=h)
style_header_row(ws3, 4, 2, 4)

for row_idx, (br, count) in enumerate(br_sorted):
    excel_row = 5 + row_idx
    ws3.cell(row=excel_row, column=2, value=br)
    ws3.cell(row=excel_row, column=3, value=count)
    ws3.cell(row=excel_row, column=4, value=round(count / total * 100, 1) if total else 0)
    style_data_row(ws3, excel_row, 2, 4, row_idx)

ws3.column_dimensions['A'].width = 3
ws3.column_dimensions['B'].width = 35
ws3.column_dimensions['C'].width = 20
ws3.column_dimensions['D'].width = 14
ws3.freeze_panes = 'B5'

# ========== Sheet 4: Bewerbungstracker ==========
ws4 = wb.create_sheet("Bewerbungstracker")
setup_sheet(ws4, title="Bewerbungstracker", last_col=10)

tracker_headers = ['Arbeitgeber', 'Stadt', 'Datum beworben', 'Kanal', 'Status', 'Antwort erhalten', 'Interview-Datum', 'Angebot', 'Notizen', 'Wiedervorlage']
for col_idx, h in enumerate(tracker_headers, start=2):
    ws4.cell(row=4, column=col_idx, value=h)
style_header_row(ws4, 4, 2, 11)

# Pre-fill employer names
for row_idx, emp in enumerate(data[:200]):  # First 200 for the tracker
    excel_row = 5 + row_idx
    ws4.cell(row=excel_row, column=2, value=emp.get('Name', ''))
    ws4.cell(row=excel_row, column=3, value=emp.get('Stadt', ''))
    style_data_row(ws4, excel_row, 2, 11, row_idx)

ws4.column_dimensions['A'].width = 3
for i, w in enumerate([30, 18, 14, 14, 14, 14, 14, 14, 25, 14], start=2):
    ws4.column_dimensions[get_column_letter(i)].width = w
ws4.freeze_panes = 'B5'

# ========== Sheet 5: Hinweise ==========
ws5 = wb.create_sheet("Hinweise")
setup_sheet(ws5, title="Hinweise zur Nutzung", last_col=6)

notes = [
    ("Datenquelle", "Alle Daten wurden von der Arbeitsagentur Jobbörse (arbeitsagentur.de) über die öffentliche REST API gesammelt."),
    ("Branche", "Branche 11: IT, Computer, Telekommunikation - entspricht dem Filter auf der Arbeitsagentur-Website."),
    ("Arbeitgeber-ID", "Format: DE-PLZ-NNNN (z.B. DE-90402-1234). Eindeutige Identifikation pro Arbeitgeber und Standort."),
    ("Kontaktdaten", "E-Mail, Telefon und Website sind teilweise nicht verfügbar, da die Arbeitsagentur diese hinter einem CAPTCHA schützt. Websites wurden teilweise per Web-Suche ergänzt."),
    ("Bewerbungsstatus", "Dropdown-Filter verfügbar: Offen, Beworben, Interview, Angebot, Abgelehnt, Zurückgezogen"),
    ("Bewerbungstracker", "Im Tab 'Bewerbungstracker' können Sie Ihren Bewerbungsfortschritt nachverfolgen."),
    ("Standortdaten", "Die Daten umfassen Arbeitgeber aus ganz Deutschland mit Schwerpunkt auf IT/Telekommunikation."),
    ("Deduplizierung", "Arbeitgeber mit dem gleichen Namen und der gleichen PLZ wurden zusammengeführt. Unternehmen mit mehreren Standorten erscheinen mehrfach."),
    ("Letzte Aktualisierung", datetime.now().strftime('%d.%m.%Y %H:%M')),
    ("Anzahl Arbeitgeber", f"{len(data)} eindeutige Arbeitgeber-Standort-Kombinationen"),
    ("API-Endpunkt", "rest.arbeitsagentur.de/jobboerse/jobsuche-service/pc/v6/jobs"),
]

for row_idx, (title, content) in enumerate(notes):
    excel_row = 4 + row_idx * 2
    cell_title = ws5.cell(row=excel_row, column=2, value=title)
    cell_title.font = font_subheader()
    cell_title.alignment = align_text()
    
    cell_content = ws5.cell(row=excel_row + 1, column=2, value=content)
    cell_content.font = font_body()
    cell_content.alignment = align_text()
    ws5.merge_cells(start_row=excel_row + 1, start_column=2, end_row=excel_row + 1, end_column=6)

ws5.column_dimensions['A'].width = 3
ws5.column_dimensions['B'].width = 18
ws5.column_dimensions['C'].width = 20
ws5.column_dimensions['D'].width = 20
ws5.column_dimensions['E'].width = 20
ws5.column_dimensions['F'].width = 20

# ========== Save ==========
wb.properties.creator = "Z.ai"
wb.properties.title = "Arbeitgeber Arbeitsagentur - IT/Telekommunikation"
wb.properties.subject = f"{len(data)} Arbeitgeber aus der Arbeitsagentur Jobbörse"
wb.properties.created = datetime.now()

wb.save(OUTPUT_FILE)
print(f"\nExcel file saved: {OUTPUT_FILE}")
print(f"Total employers: {len(data)}")
print(f"Sheets: Arbeitgeberverzeichnis, Statistik Bundesland, Statistik Branche, Bewerbungstracker, Hinweise")
