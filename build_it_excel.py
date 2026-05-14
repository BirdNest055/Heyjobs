#!/usr/bin/env python3
"""
Build IT-Arbeitgeber Excel from available data (AA API + Google Maps).
No HTTP requests - just data merging. Fast and reliable.
"""
import json, os, re, sys, sqlite3
from datetime import datetime
import openpyxl

EXCEL_IN = "/home/z/my-project/download/Google_Maps_Firmen_Erlangen_50km.xlsx"
EXCEL_OUT = "/home/z/my-project/download/IT_Arbeitgeber_Bamberg_Erlangen_Nuernberg.xlsx"
IT_JOBS_MAP = "/home/z/my-project/download/it_jobs_map.json"
DB_PATH = "/home/z/my-project/download/it_employers.db"

TARGET_CITIES = {"Bamberg", "Erlangen", "Nürnberg"}

def normalize_name(name):
    name = name.lower().strip()
    for s in ['gmbh & co. kg','gmbh & co kg','gmbh co. kg','gmbh co kg','gmbh & co.','ag & co. kg','ag & co kg','gmbh','ag','kg','ohg','eg','se','e.k.','ek','e.v.','ev','mbh','ug','gbr']:
        name = re.sub(re.escape(s), '', name, flags=re.IGNORECASE)
    name = re.sub(r'[^a-z0-9äöüß]', '', name)
    return name

def main():
    print("=== Build IT-Arbeitgeber Excel ===")
    
    # Load IT jobs
    with open(IT_JOBS_MAP) as f:
        it_jobs = json.load(f)
    print(f"IT-Jobs: {len(it_jobs)} Arbeitgeber")
    
    # Load Google Maps
    wb = openpyxl.load_workbook(EXCEL_IN)
    ws = wb['Google Maps Firmen']
    employers = []
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        ort = str(row[6] or '').strip()
        name = str(row[1] or '').strip()
        if not name or ort not in TARGET_CITIES: continue
        norm = normalize_name(name)
        if norm in seen or len(norm) < 3: continue
        seen.add(norm)
        employers.append({
            'name': name, 'ort': ort,
            'plz': str(row[5] or '').strip(),
            'adresse': str(row[4] or '').strip(),
            'telefon': str(row[7] or '').strip(),
            'website_flag': str(row[8] or '').strip(),
            'kategorie': str(row[3] or '').strip(),
            'bewertung': row[2],
            'lat': row[9], 'lng': row[10],
            'gmaps_link': str(row[11] or '').strip(),
        })
    print(f"Google Maps: {len(employers)} Arbeitgeber")
    
    # Match and build data
    results = []
    for emp in employers:
        norm = normalize_name(emp['name'])
        it_data = None; score = 0.0; matched_aa_name = ''
        
        if norm in it_jobs:
            it_data = it_jobs[norm]
            score = 50
            matched_aa_name = it_data.get('firma_original', '')
        else:
            for it_norm, it_d in it_jobs.items():
                if len(norm) > 4 and len(it_norm) > 4 and (norm[:6] in it_norm or it_norm[:6] in norm):
                    it_data = it_d; score = 40
                    matched_aa_name = it_d.get('firma_original', '')
                    break
        
        name_lower = emp['name'].lower()
        kat_lower = (emp.get('kategorie') or '').lower()
        it_name = any(w in name_lower for w in ['it ','it-','software','tech','digital','data','cloud','cyber','computer','informatik','system','consulting','automation','iot','embedded','telekommunikation'])
        it_kat = any(w in kat_lower for w in ['software','it','informatik','computer','technologie','digital','internet','telekommunikation'])
        if it_name: score += 20
        if it_kat: score += 15
        
        result = {
            'firmenname': emp['name'],
            'firmenname_normalisiert': norm,
            'ort': emp['ort'],
            'plz': emp['plz'],
            'strasse': emp['adresse'],
            'telefon': emp.get('telefon', ''),
            'email': '',
            'website_url': '',
            'website_html_path': '',
            'website_html_hash': '',
            'karriere_url': '',
            'karriere_html_path': '',
            'impressum_url': '',
            'impressum_html_path': '',
            'hat_it_jobs': 1 if it_data else 0,
            'it_job_titel': '',
            'it_job_anzahl': 0,
            'it_job_quelle': '',
            'it_job_details': '',
            'it_keywords_gefunden': '',
            'arbeitsagentur_match': matched_aa_name,
            'kategorie_google': emp.get('kategorie', ''),
            'bewertung_google': emp.get('bewertung'),
            'lat': emp.get('lat'),
            'lng': emp.get('lng'),
            'google_maps_link': emp.get('gmaps_link', ''),
            'handelsregister': '',
            'geschaeftsfuehrer': '',
            'it_abteilung_vermutung': 1 if score >= 30 else 0,
            'it_relevanz_score': min(score, 100),
            'scrape_datum': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'scrape_status': 'phase1_complete',
            'letzte_pruefung': datetime.now().strftime('%Y-%m-%d'),
        }
        
        if it_data:
            jobs = it_data.get('jobs', [])
            result['it_job_anzahl'] = len(jobs)
            result['it_job_titel'] = "; ".join(j.get('titel', '') for j in jobs[:10])
            result['it_job_quelle'] = 'Arbeitsagentur'
            result['it_job_details'] = json.dumps([{
                'titel': j.get('titel', ''),
                'beruf': j.get('hauptberuf', ''),
                'refnr': j.get('referenznummer', ''),
                'ort': j.get('ort', ''),
                'plz': j.get('plz', ''),
                'vollzeit': j.get('vollzeit', False),
                'unbefristet': j.get('unbefristet', False),
            } for j in jobs[:20]], ensure_ascii=False)
            # Extract unique IT keywords from job titles
            all_titles = ' '.join(j.get('titel', '').lower() for j in jobs)
            it_kws_found = [kw for kw in ['software','entwickler','developer','it-','data','cloud','devops','administrator','sap','scrum','docker','frontend','backend','informatik','consultant','security','linux','network','python','java','typescript','agile','tester','support','engineer'] if kw in all_titles]
            result['it_keywords_gefunden'] = ", ".join(sorted(set(it_kws_found)))
        
        results.append(result)
    
    # Sort: IT jobs first, then by score
    results.sort(key=lambda x: (-x['hat_it_jobs'], -x['it_relevanz_score'], x['firmenname']))
    
    # Build Excel
    wb_out = openpyxl.Workbook()
    
    # Sheet 1: All IT-relevant employers
    ws1 = wb_out.active
    ws1.title = "IT Arbeitgeber"
    
    columns = ['firmenname','firmenname_normalisiert','ort','plz','strasse','telefon',
               'email','website_url','website_html_path','karriere_url','impressum_url',
               'hat_it_jobs','it_job_titel','it_job_anzahl','it_job_quelle',
               'it_keywords_gefunden','arbeitsagentur_match','kategorie_google',
               'bewertung_google','lat','lng','google_maps_link',
               'handelsregister','geschaeftsfuehrer','it_abteilung_vermutung',
               'it_relevanz_score','scrape_datum','scrape_status','letzte_pruefung']
    
    headers_de = ['Firmenname','Firmenname (normalisiert)','Ort','PLZ','Straße','Telefon',
                  'E-Mail','Website URL','Website HTML Pfad','Karriere-URL','Impressum URL',
                  'IT-Jobs vorhanden','IT-Job Titel','Anzahl IT-Jobs','IT-Job Quelle',
                  'IT-Keywords gefunden','Arbeitsagentur Match','Kategorie (Google)',
                  'Bewertung (Google)','Breitengrad','Längengrad','Google Maps Link',
                  'Handelsregister','Geschäftsführer','IT-Abteilung vermutet',
                  'IT-Relevanz Score','Scrape-Datum','Scrape-Status','Letzte Prüfung']
    
    ws1.append(headers_de)
    for r in results:
        ws1.append([r.get(c, '') for c in columns])
    
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)
    
    # Sheet 2: Only employers with confirmed IT jobs
    ws2 = wb_out.create_sheet("Nur IT-Jobs bestätigt")
    ws2.append(headers_de)
    it_only = [r for r in results if r.get('hat_it_jobs')]
    for r in it_only:
        ws2.append([r.get(c, '') for c in columns])
    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)
    
    # Sheet 3: Statistics
    ws3 = wb_out.create_sheet("Statistiken")
    ws3.append(["Metrik", "Wert"])
    stats = [
        ("Gesamt Arbeitgeber", len(results)),
        ("Mit IT-Jobs (bestätigt)", len(it_only)),
        ("IT-Abteilung vermutet", sum(1 for r in results if r.get('it_abteilung_vermutung'))),
        ("Ø IT-Relevanz Score", round(sum(r['it_relevanz_score'] for r in results) / len(results), 1) if results else 0),
        ("", ""),
        ("Nach Ort:", ""),
    ]
    for city in TARGET_CITIES:
        city_emps = [r for r in results if r['ort'] == city]
        city_it = [r for r in city_emps if r.get('hat_it_jobs')]
        stats.append((f"{city} - Gesamt", len(city_emps)))
        stats.append((f"{city} - Mit IT-Jobs", len(city_it)))
    stats.append(("", ""))
    stats.append(("Top IT-Arbeitgeber:", ""))
    for r in it_only[:10]:
        stats.append((r['firmenname'], f"{r['it_job_anzahl']} IT-Jobs (Score: {r['it_relevanz_score']:.0f})"))
    
    for s in stats:
        ws3.append(s)
    
    # Sheet 4: Per-city breakdown
    ws4 = wb_out.create_sheet("Nach Ort")
    ws4.append(["Ort", "Gesamt", "IT-Jobs bestätigt", "IT-Abt. vermutet", "Ø Score", "Top Arbeitgeber"])
    for city in TARGET_CITIES:
        city_emps = [r for r in results if r['ort'] == city]
        city_it = sum(1 for r in city_emps if r.get('hat_it_jobs'))
        city_vmt = sum(1 for r in city_emps if r.get('it_abteilung_vermutung'))
        avg_score = round(sum(r['it_relevanz_score'] for r in city_emps) / len(city_emps), 1) if city_emps else 0
        top = max(city_emps, key=lambda x: x['it_relevanz_score'])['firmenname'] if city_emps else ''
        ws4.append([city, len(city_emps), city_it, city_vmt, avg_score, top])
    
    # Sheet 5: AA API job details (raw)
    ws5 = wb_out.create_sheet("IT-Job Details")
    ws5.append(["Firmenname (Google Maps)", "Firmenname (Arbeitsagentur)", "Ort", "Anzahl Jobs", "Job-Titel", "Berufe", "Referenznummern"])
    for r in it_only:
        aa_name = r.get('arbeitsagentur_match', '')
        try:
            details = json.loads(r.get('it_job_details', '[]'))
            titels = "; ".join(d.get('titel', '') for d in details[:10])
            berufe = "; ".join(sorted(set(d.get('beruf', '') for d in details if d.get('beruf'))))
            refs = "; ".join(d.get('refnr', '') for d in details[:5])
        except:
            titels = r.get('it_job_titel', '')
            berufe = ''
            refs = ''
        ws5.append([r['firmenname'], aa_name, r['ort'], r['it_job_anzahl'], titels, berufe, refs])
    
    wb_out.save(EXCEL_OUT)
    print(f"Excel saved: {EXCEL_OUT}")
    
    # Also save to SQLite for further processing
    conn = sqlite3.connect(DB_PATH)
    conn.executescript("""CREATE TABLE IF NOT EXISTS employers (
        id INTEGER PRIMARY KEY AUTOINCREMENT, firmenname TEXT NOT NULL,
        firmenname_normalisiert TEXT, ort TEXT, plz TEXT, strasse TEXT,
        telefon TEXT, email TEXT, website_url TEXT, website_html_path TEXT,
        website_html_hash TEXT, karriere_url TEXT, karriere_html_path TEXT,
        impressum_url TEXT, impressum_html_path TEXT, hat_it_jobs INTEGER DEFAULT 0,
        it_job_titel TEXT, it_job_anzahl INTEGER DEFAULT 0, it_job_quelle TEXT,
        it_job_details TEXT, it_keywords_gefunden TEXT,
        arbeitsagentur_match TEXT, kategorie_google TEXT,
        bewertung_google REAL, lat REAL, lng REAL, google_maps_link TEXT,
        handelsregister TEXT, geschaeftsfuehrer TEXT,
        it_abteilung_vermutung INTEGER DEFAULT 0, it_relevanz_score REAL DEFAULT 0.0,
        scrape_datum TEXT, scrape_status TEXT, letzte_pruefung TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP, updated_at TEXT DEFAULT CURRENT_TIMESTAMP
    ); CREATE INDEX IF NOT EXISTS idx_ort ON employers(ort);
    CREATE INDEX IF NOT EXISTS idx_hat_it ON employers(hat_it_jobs);
    CREATE INDEX IF NOT EXISTS idx_score ON employers(it_relevanz_score);""")
    conn.commit()
    
    cols = ['firmenname','firmenname_normalisiert','ort','plz','strasse',
            'telefon','email','website_url','website_html_path','website_html_hash',
            'karriere_url','karriere_html_path','impressum_url','impressum_html_path',
            'hat_it_jobs','it_job_titel','it_job_anzahl','it_job_quelle',
            'it_job_details','it_keywords_gefunden','kategorie_google',
            'bewertung_google','lat','lng','google_maps_link',
            'handelsregister','geschaeftsfuehrer','it_abteilung_vermutung',
            'it_relevanz_score','scrape_datum','scrape_status','letzte_pruefung']
    
    for r in results:
        vals = [r.get(c,'') for c in cols]
        conn.execute(f"INSERT OR REPLACE INTO employers ({','.join(cols)}) VALUES ({','.join(['?']*len(cols))})", vals)
    conn.commit()
    conn.close()
    print(f"SQLite saved: {DB_PATH}")
    
    # Summary
    print(f"\n{'='*50}")
    print(f"ERGEBNISSE:")
    print(f"  Gesamt: {len(results)} Arbeitgeber")
    print(f"  IT-Jobs bestätigt: {len(it_only)}")
    print(f"  IT-Abteilung vermutet: {sum(1 for r in results if r.get('it_abteilung_vermutung'))}")
    print(f"  Ø IT-Relevanz: {round(sum(r['it_relevanz_score'] for r in results)/len(results),1)}")
    for city in TARGET_CITIES:
        c = sum(1 for r in results if r['ort'] == city)
        ci = sum(1 for r in results if r['ort'] == city and r.get('hat_it_jobs'))
        print(f"  {city}: {c} Arbeitgeber, {ci} mit IT-Jobs")
    print(f"{'='*50}")

if __name__ == '__main__':
    main()
