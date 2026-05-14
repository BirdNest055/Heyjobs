#!/usr/bin/env python3
"""
High-Quality Junior IT-Jobs Scraper v4
=======================================
- CloakBrowser (stealth Chromium) for anti-detection
- Direct website navigation to find career pages
- DuckDuckGo search engine enrichment
- JSON-LD + structured + text-based job extraction
- Commits every 10 entries to GitHub
- Resumable with progress tracking
"""

import hashlib
import json
import os
import re
import subprocess
import sys
import time
import traceback
from datetime import datetime
from pathlib import Path
from urllib.parse import urljoin, urlparse, quote, unquote
from collections import Counter

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG
# ============================================================
EXCEL_EMPLOYERS = "/home/z/my-project/download/Firmen_Bamberg_Erlangen_Nuernberg_mit_Website_HTML.xlsx"
EXCEL_OUT = "/home/z/my-project/download/Junior_IT_Jobs_Quellennah.xlsx"
CAREER_HTML_DIR = "/home/z/my-project/download/career_pages_v3"
JOB_HTML_DIR = "/home/z/my-project/download/job_detail_v3"
PROGRESS_FILE = "/home/z/my-project/download/scrape_v3_progress.json"
GIT_REPO = "/home/z/my-project"
GITHUB_TOKEN = "GIT_TOKEN_REDACTED"

COMMIT_EVERY = 10
HOME_TIMEOUT = 12000  # Shorter timeouts for faster processing
CAREER_TIMEOUT = 15000
DETAIL_TIMEOUT = 10000
DDG_TIMEOUT = 10000

os.makedirs(CAREER_HTML_DIR, exist_ok=True)
os.makedirs(JOB_HTML_DIR, exist_ok=True)

# ============================================================
# KEYWORDS (same as v3)
# ============================================================
JUNIOR_KEYWORDS = [
    'junior', 'entry level', 'einsteiger', 'anfänger',
    'trainee', 'praktikant', 'intern', 'werkstudent', 'working student',
    'absolvent', 'graduate', 'berufseinsteiger', 'young professional',
    'associate', 'duales studium', 'ausbildung', 'apprentice', 'azubi',
    'nachwuchs', 'career starter',
]

SENIOR_EXCLUDE = [
    'senior', 'lead', 'principal', 'head of', 'director',
    'leitung', 'chef', 'geschäftsführer', 'vorstand', 'prof.',
]

IT_KEYWORDS = [
    'software', 'entwickler', 'developer', 'it-', 'informatik',
    'data', 'devops', 'cloud', 'cyber', 'security', 'sicherheit',
    'system', 'admin', 'administrator', 'netzwerk', 'network',
    'frontend', 'backend', 'fullstack', 'full-stack',
    'python', 'java', 'javascript', 'typescript', 'react', 'angular', 'vue',
    'sap', 'linux', 'sql', 'database', 'datenbank', 'api',
    'machine learning', 'ki', 'künstliche intelligenz', 'ai',
    'scrum', 'agile', 'testing', 'qa', 'automation',
    'consultant', 'analyst', 'engineer', 'architekt', 'architect',
    'programmierung', 'programming', 'coding', 'code',
    'web', 'mobile', 'app', 'anwendung',
    'infrastructure', 'infrastruktur', 'platform', 'plattform',
    'support', 'helpdesk', 'service desk',
    'erp', 'crm', 'dynamics', 'sharepoint',
    'digitalisierung', 'technologie', 'technology', 'digital',
    'fachinformatiker', 'it-systemelektroniker', 'it-kaufmann',
    'elektrotechnik', 'elektronik',
]

CAREER_URL_PATTERNS = [
    '/karriere', '/career', '/careers', '/jobs', '/stellenangebote',
    '/stellen', '/offene-stellen', '/vacancies', '/job',
    '/de/karriere', '/en/career', '/de/career', '/en/careers',
    '/de/jobs', '/en/jobs', '/unternehmen/karriere',
    '/ueber-uns/karriere', '/about/careers',
    '/jobboerse', '/recruiting', '/bewerbung',
    '/studenten', '/absolventen', '/graduates',
    '/early-career', '/einsteiger', '/berufseinsteiger',
    '/trainee', '/praktikum', '/werkstudent', '/ausbildung',
]


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def sanitize_filename(name, max_len=50):
    name = re.sub(r'[^\w\s.-]', '_', name)
    name = re.sub(r'_{2,}', '_', name)
    return name.strip('_ .')[:max_len]

def is_junior_job(title, description=""):
    text = f"{title} {description}".lower()
    for kw in SENIOR_EXCLUDE:
        if kw in text:
            return False, None
    for kw in JUNIOR_KEYWORDS:
        if kw in text:
            return True, kw
    exp_match = re.search(r'(\d+)\s*(?:jahr|year|jahre|years)\s*(?:berufs)?erfahrung', text)
    if exp_match:
        years = int(exp_match.group(1))
        if years <= 2:
            return True, f"erfahrung_{years}jahre"
    return False, None

def is_it_job(title, description=""):
    text = f"{title} {description}".lower()
    for kw in IT_KEYWORDS:
        if kw in text:
            return True, kw
    return False, None

def is_senior_job(title):
    text = title.lower()
    return any(kw in text for kw in SENIOR_EXCLUDE)

def extract_technologies(text):
    tech_keywords = [
        'python', 'java', 'javascript', 'typescript', 'c#', 'c++',
        'ruby', 'go', 'rust', 'kotlin', 'swift', 'scala', 'php',
        'react', 'angular', 'vue', 'svelte', 'next.js',
        'node.js', 'django', 'flask', 'spring', '.net',
        'docker', 'kubernetes', 'k8s', 'aws', 'azure', 'gcp',
        'terraform', 'ansible', 'jenkins', 'gitlab', 'github',
        'sql', 'mysql', 'postgresql', 'mongodb', 'redis',
        'kafka', 'graphql', 'rest', 'api',
        'sap', 'salesforce', 'dynamics', 'sharepoint',
        'linux', 'windows', 'html', 'css', 'tailwind',
        'agile', 'scrum', 'kanban', 'devops', 'ci/cd',
        'machine learning', 'deep learning', 'tensorflow', 'pytorch',
        'nlp', 'computer vision', 'data science',
        'jira', 'confluence', 'cybersecurity', 'pentest',
        'powershell', 'bash', 'shell', 'git',
    ]
    found = []
    text_lower = text.lower()
    for tech in tech_keywords:
        if re.search(r'\b' + re.escape(tech) + r'\b', text_lower):
            found.append(tech)
    return list(set(found))

def classify_job_category(title, description=""):
    text = f"{title} {description}".lower()
    categories = {
        'Softwareentwicklung': ['softwareentwickler', 'software developer', 'software engineer', 'entwickler', 'developer', 'programmierer', 'frontend', 'backend', 'fullstack', 'anwendungsentwicklung'],
        'Data Science & Analytics': ['data scientist', 'data analyst', 'data engineer', 'data analytics', 'business intelligence', 'machine learning', 'ki-entwickler'],
        'IT-Consulting': ['it consultant', 'it-consultant', 'it berater', 'sap consultant', 'technischer berater'],
        'Systemadministration': ['systemadministrator', 'sysadmin', 'admin', 'netzwerkadministrator', 'systemintegration'],
        'DevOps & Cloud': ['devops', 'cloud engineer', 'cloud architect', 'platform engineer', 'sre'],
        'Cyber Security': ['security', 'cybersecurity', 'it-sicherheit', 'pentest', 'soc analyst'],
        'IT-Support': ['it support', 'helpdesk', 'service desk', 'first level', 'second level'],
        'IT-Projektmanagement': ['project manager', 'projektmanager', 'scrum master', 'product owner'],
        'QA & Testing': ['quality assurance', 'qa', 'tester', 'test engineer'],
        'UX/UI Design': ['ux', 'ui', 'user experience', 'user interface'],
        'IT-Ausbildung & Studium': ['ausbildung', 'duales studium', 'trainee', 'praktikum', 'werkstudent', 'fachinformatiker'],
    }
    for category, keywords in categories.items():
        for kw in keywords:
            if kw in text:
                return category
    return 'IT (Sonstige)'

def extract_employment_type(title, description=""):
    text = f"{title} {description}".lower()
    if any(kw in text for kw in ['werkstudent', 'working student']):
        return 'Werkstudent'
    if any(kw in text for kw in ['praktikant', 'praktikum', 'intern']):
        return 'Praktikum'
    if any(kw in text for kw in ['ausbildung', 'apprentice', 'azubi']):
        return 'Ausbildung'
    if any(kw in text for kw in ['duales studium']):
        return 'Duales Studium'
    if any(kw in text for kw in ['teilzeit', 'part-time']):
        return 'Teilzeit'
    return 'Vollzeit'

def extract_remote_option(title, description=""):
    text = f"{title} {description}".lower()
    if any(kw in text for kw in ['fully remote', '100% remote']):
        return 'Remote'
    if any(kw in text for kw in ['remote', 'homeoffice', 'home-office']):
        return 'Remote möglich'
    if any(kw in text for kw in ['hybrid', 'teilweise remote']):
        return 'Hybrid'
    return 'Nicht angegeben'

def clean_city(city_raw):
    if not city_raw:
        return ''
    city = city_raw.strip()
    for suffix in [', Mittelfranken', ', Oberfranken', ', Bayern', ', Deutschland']:
        city = city.replace(suffix, '')
    return city.strip()

def classify_region(city):
    if not city:
        return 'Unbekannt'
    c = city.lower()
    if any(x in c for x in ['nürnberg', 'nuernberg', 'fürth', 'fuerth', 'erlangen', 'schwabach']):
        return 'Metropolregion Nürnberg (Kern)'
    if any(x in c for x in ['bamberg', 'forchheim', 'coburg', 'bayreuth']):
        return 'Oberfranken'
    return 'Weitere Regionen'

def classify_junior_type(title, junior_kw=''):
    text = title.lower()
    if 'ausbildung' in text or junior_kw == 'ausbildung': return 'Ausbildung'
    if 'duales studium' in text: return 'Duales Studium'
    if 'werkstudent' in text: return 'Werkstudent'
    if 'praktikum' in text or 'praktikant' in text: return 'Praktikum'
    if 'trainee' in text: return 'Trainee'
    if 'junior' in text: return 'Junior'
    if 'associate' in text: return 'Associate'
    if 'einsteiger' in text or 'berufseinsteiger' in text: return 'Berufseinsteiger'
    if 'absolvent' in text: return 'Absolvent'
    return 'Junior-fähig'

def classify_age(veroeffdatum):
    if not veroeffdatum:
        return 'Unbekannt', ''
    try:
        dt = datetime.strptime(str(veroeffdatum)[:10], '%Y-%m-%d')
        days = (datetime.now() - dt).days
        if days <= 7: return 'Frisch (≤ 7 Tage)', days
        elif days <= 30: return 'Aktuell (8-30 Tage)', days
        elif days <= 90: return 'Älter (31-90 Tage)', days
        else: return 'Veraltet (> 90 Tage)', days
    except:
        return 'Unbekannt', ''


# ============================================================
# EXCEL MANAGER
# ============================================================

class JuniorJobsExcel:
    HEADERS = [
        'Job-ID', 'Job-Titel', 'Job-Kategorie', 'Erfahrungslevel', 'Junior-Typ', 'Explizit Junior',
        'Beschäftigungsart', 'Remote-Option', 'Technologien', 'Ausschreibungstext', 'Job-URL',
        'Stadt', 'Region', 'PLZ', 'Straße',
        'Arbeitgeber', 'Arbeitgeber-Website', 'Karriere-URL', 'Branche',
        'Quelle', 'Quelle-Typ', 'Gefunden über', 'Arbeitsagentur RefNr',
        'Junior-Keyword Match', 'IT-Keyword Match', 'Ist Junior-Job', 'Ist IT-Job',
        'Veröffentlicht am', 'Veröffentlicht vor (Tage)', 'Alter der Ausschreibung',
        'Scrape-Datum', 'Scrape-Status', 'Arbeitgeber-ID',
        'Karriere-HTML Pfad', 'Job-HTML Pfad',
    ]

    def __init__(self, filepath):
        self.filepath = filepath
        if os.path.exists(filepath):
            self.wb = openpyxl.load_workbook(filepath)
            self.ws = self.wb.active
            self.job_counter = self.ws.max_row - 1
            self.existing_urls = set()
            self.existing_titles = {}
            for r in range(2, self.ws.max_row + 1):
                url = self.ws.cell(r, 11).value
                title = self.ws.cell(r, 2).value
                emp = self.ws.cell(r, 16).value
                if url: self.existing_urls.add(url)
                if title and emp:
                    self.existing_titles[f"{str(title).lower().strip()}_{str(emp).lower().strip()}"] = True
        else:
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active
            self.ws.title = "Junior IT-Jobs (Quellennah)"
            self._write_headers()
            self.job_counter = 0
            self.existing_urls = set()
            self.existing_titles = {}

    def _write_headers(self):
        header_font = Font(bold=True, color="FFFFFF", size=11, name='Calibri')
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        for col, header in enumerate(self.HEADERS, 1):
            cell = self.ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        widths = [10,45,22,14,18,14, 16,16,35,60,50, 18,28,8,25, 38,50,50,22, 16,18,30,22, 22,18,12,12, 14,18,18,18,14,16,50,50]
        for i, w in enumerate(widths):
            if i < len(self.HEADERS):
                self.ws.column_dimensions[get_column_letter(i+1)].width = w

    def is_duplicate(self, job_data):
        url = job_data.get('job_url', '')
        title = job_data.get('job_title', '')
        employer = job_data.get('employer_name', '')
        if url and url in self.existing_urls:
            return True
        key = f"{title.lower().strip()}_{employer.lower().strip()}"
        return key in self.existing_titles

    def add_job(self, job_data):
        if self.is_duplicate(job_data):
            return False
        self.job_counter += 1
        row = self.ws.max_row + 1
        title = job_data.get('job_title', '')
        description = job_data.get('description', '')
        junior_kw = job_data.get('junior_keyword_match', '')
        veroeff = job_data.get('date_posted', '')
        age_label, days_ago = classify_age(veroeff)

        values = [
            f'Q-{self.job_counter:05d}',
            title,
            classify_job_category(title, description),
            'Junior',
            classify_junior_type(title, junior_kw),
            'Ja' if junior_kw and junior_kw != 'no_senior_keyword' else 'Nein',
            extract_employment_type(title, description),
            extract_remote_option(title, description),
            ', '.join(extract_technologies(f"{title} {description}")),
            (description or '')[:1000],
            job_data.get('job_url', ''),
            clean_city(job_data.get('city', '')),
            classify_region(job_data.get('city', '')),
            job_data.get('plz', ''),
            job_data.get('street', ''),
            job_data.get('employer_name', ''),
            job_data.get('website_url', ''),
            job_data.get('career_url', ''),
            job_data.get('industry', ''),
            job_data.get('source', 'Website'),
            job_data.get('source_type', 'Website direkt'),
            job_data.get('found_via', ''),
            job_data.get('refnr', ''),
            junior_kw,
            job_data.get('it_keyword_match', ''),
            'Ja',
            'Ja',
            veroeff,
            days_ago if days_ago != '' else '',
            age_label,
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            job_data.get('scrape_status', 'completed'),
            job_data.get('employer_id', ''),
            job_data.get('career_html_path', ''),
            job_data.get('job_html_path', ''),
        ]
        for col, value in enumerate(values, 1):
            cell = self.ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.font = Font(name='Calibri', size=10)

        url = job_data.get('job_url', '')
        if url: self.existing_urls.add(url)
        key = f"{title.lower().strip()}_{job_data.get('employer_name', '').lower().strip()}"
        self.existing_titles[key] = True
        return True

    def save(self):
        self.wb.save(self.filepath)

    def get_job_count(self):
        return self.job_counter


# ============================================================
# PROGRESS + GIT
# ============================================================

def load_progress():
    if os.path.exists(PROGRESS_FILE):
        return json.load(open(PROGRESS_FILE, encoding='utf-8'))
    return {'completed': [], 'failed': [], 'last_index': 0, 'stats': {}}

def save_progress(progress):
    with open(PROGRESS_FILE, 'w', encoding='utf-8') as f:
        json.dump(progress, f, ensure_ascii=False, indent=1)

def git_commit_and_push(excel, count_since_last):
    try:
        excel.save()
        os.chdir(GIT_REPO)
        subprocess.run(['git', 'add', '-A'], capture_output=True, timeout=30)
        msg = f"Junior IT-Jobs quellennah - {excel.get_job_count()} Jobs (+{count_since_last})"
        result = subprocess.run(['git', 'commit', '-m', msg], capture_output=True, timeout=30)
        if result.returncode == 0:
            subprocess.run(
                ['git', 'push', f'https://{GITHUB_TOKEN}@github.com/BirdNest055/Heyjobs.git', 'main'],
                capture_output=True, timeout=60
            )
            print(f"  [GIT] Pushed ({excel.get_job_count()} jobs, +{count_since_last})")
            return True
    except Exception as e:
        print(f"  [GIT] Error: {str(e)[:80]}")
    return False


# ============================================================
# CLOAKBROWSER SCRAPER
# ============================================================

class CloakScraper:
    def __init__(self, excel):
        self.excel = excel
        self.jobs_since_commit = 0
        self.browser = None
        self.context = None
        self.stats = Counter()

    def start_browser(self):
        from playwright.sync_api import sync_playwright
        self._pw = sync_playwright().start()
        self.browser = self._pw.chromium.launch(
            headless=True,
            args=['--disable-blink-features=AutomationControlled', '--no-sandbox'],
        )
        self.context = self.browser.new_context(
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
            viewport={'width': 1920, 'height': 1080},
            locale='de-DE',
            timezone_id='Europe/Berlin',
        )
        print("  [BROWSER] Playwright gestartet")

    def stop_browser(self):
        for obj in [self.context, self.browser]:
            if obj:
                try: obj.close()
                except: pass
        if hasattr(self, '_pw') and self._pw:
            try: self._pw.stop()
            except: pass
        self.browser = None
        self.context = None
        self._pw = None

    def _dismiss_cookies(self, page):
        for sel in ['button:has-text("Akzeptieren")', 'button:has-text("Alle akzeptieren")',
                    'button:has-text("Accept")', 'button:has-text("Accept All")',
                    '#onetrust-accept-btn-handler', '#ccc-notify-accept',
                    'button:has-text("OK")', 'button:has-text("Verstanden")',
                    'button:has-text("Zustimmen")', 'button:has-text("Einverstanden")']:
            try:
                btn = page.query_selector(sel)
                if btn and btn.is_visible():
                    btn.click()
                    time.sleep(0.3)
                    return
            except:
                continue

    def _safe_page(self, url, timeout=HOME_TIMEOUT):
        """Create a new page and navigate with error handling."""
        page = self.context.new_page()
        try:
            page.goto(url, timeout=timeout, wait_until='domcontentloaded')
            time.sleep(1.5)
            self._dismiss_cookies(page)
            return page
        except Exception as e:
            try: page.close()
            except: pass
            return None

    def search_duckduckgo(self, query, max_results=5):
        results = []
        page = None
        try:
            page = self._safe_page(f'https://html.duckduckgo.com/html/?q={quote(query)}', DDG_TIMEOUT)
            if not page:
                return []
            time.sleep(1)
            links = page.query_selector_all('.result__a')
            for link in links[:max_results]:
                href = link.get_attribute('href') or ''
                text = (link.inner_text() or '').strip()
                if href and 'uddg=' in href:
                    actual_url = unquote(href.split('uddg=')[1].split('&')[0])
                elif href:
                    actual_url = href
                else:
                    continue
                results.append({'title': text, 'url': actual_url})
        except:
            pass
        finally:
            if page:
                try: page.close()
                except: pass
        return results

    def find_career_page(self, employer_name, base_url):
        """Find career page - returns (url, method) or (None, None)."""
        # Step 1: Try homepage
        page = self._safe_page(base_url, HOME_TIMEOUT)
        if not page:
            return None, None

        try:
            # Look for career links
            for text in ['Karriere', 'Career', 'Careers', 'Jobs', 'Stellenangebote',
                         'Stellen', 'Offene Stellen', 'Bewerbung', 'Join us',
                         'Wir stellen ein', 'Ausbildung', 'Werkstudent', 'Trainee']:
                try:
                    elems = page.query_selector_all(f'a:has-text("{text}")')
                    for elem in elems:
                        href = elem.get_attribute('href')
                        if href:
                            full_url = urljoin(base_url, href)
                            url_lower = full_url.lower()
                            if any(kw in url_lower for kw in ['karriere', 'career', 'job', 'stelle', 'vacanc', 'recruit', 'bewerb', 'ausbildung', 'trainee', 'werkstudent']):
                                page.close()
                                return full_url, 'website_navigation'
                except:
                    continue

            # Check nav/footer
            for area in ['nav a', 'header a', 'footer a']:
                try:
                    links = page.query_selector_all(area)
                    for link in links:
                        link_text = (link.inner_text() or '').strip().lower()
                        href = link.get_attribute('href')
                        if href and any(kw in link_text for kw in ['karriere', 'career', 'jobs', 'stellen', 'bewerbung']):
                            full_url = urljoin(base_url, href)
                            page.close()
                            return full_url, 'nav_footer'
                except:
                    continue

            # Try URL patterns
            parsed = urlparse(base_url)
            base_domain = f"{parsed.scheme}://{parsed.netloc}"
            for pattern in CAREER_URL_PATTERNS[:8]:  # Only try most common patterns
                try:
                    test_url = base_domain + pattern
                    response = page.goto(test_url, timeout=6000, wait_until='domcontentloaded')
                    if response and response.status < 400:
                        time.sleep(0.5)
                        content = (page.content() or '').lower()
                        if any(kw in content for kw in ['job', 'stelle', 'karriere', 'career', 'bewerb']):
                            page.close()
                            return test_url, 'url_pattern'
                except:
                    continue
        except:
            pass
        finally:
            try: page.close()
            except: pass

        # Step 2: DuckDuckGo search
        ddg_queries = [
            f'{employer_name} Karriere Jobs',
            f'{employer_name} Stellenangebote',
        ]
        for query in ddg_queries:
            results = self.search_duckduckgo(query, max_results=5)
            parsed_base = urlparse(base_url).netloc.lower().replace('www.', '')
            for res in results:
                res_domain = urlparse(res['url']).netloc.lower().replace('www.', '')
                url_lower = res['url'].lower()
                if parsed_base in res_domain:
                    if any(kw in url_lower for kw in ['karriere', 'career', 'job', 'stelle', 'vacanc', 'recruit', 'bewerb']):
                        return res['url'], 'ddg_search'

        return None, None

    def extract_jobs_from_career_page(self, career_url, employer_name):
        """Extract job listings from career page. Returns (jobs_list, html_path)."""
        page = self._safe_page(career_url, CAREER_TIMEOUT)
        if not page:
            return [], None

        try:
            # Scroll to load content
            for pos in [0.5, 1.0]:
                try:
                    page.evaluate(f'window.scrollTo(0, document.body.scrollHeight * {pos})')
                    time.sleep(0.3)
                except:
                    pass

            # Save HTML
            career_html = page.content()
            career_filename = sanitize_filename(employer_name) + '_career.html'
            career_filepath = os.path.join(CAREER_HTML_DIR, career_filename)
            with open(career_filepath, 'w', encoding='utf-8') as f:
                f.write(career_html)

            jobs = []

            # Strategy 1: JSON-LD
            try:
                json_ld_data = page.evaluate('''() => {
                    const results = [];
                    const scripts = document.querySelectorAll('script[type="application/ld+json"]');
                    for (const script of scripts) {
                        try {
                            const data = JSON.parse(script.textContent);
                            const items = Array.isArray(data) ? data : [data];
                            for (const item of items) {
                                if (item['@type'] === 'JobPosting') {
                                    results.push({title: item.title||'', url: item.url||'', description: (item.description||'').substring(0,2000), datePosted: item.datePosted||'', employmentType: item.employmentType||''});
                                }
                                if (item['@graph']) {
                                    for (const g of item['@graph']) {
                                        if (g['@type'] === 'JobPosting') {
                                            results.push({title: g.title||'', url: g.url||'', description: (g.description||'').substring(0,2000), datePosted: g.datePosted||'', employmentType: g.employmentType||''});
                                        }
                                    }
                                }
                            }
                        } catch(e) {}
                    }
                    return results;
                }''')
                for ld in json_ld_data:
                    if ld.get('title'):
                        jobs.append({
                            'title': ld['title'][:300],
                            'url': ld.get('url', ''),
                            'description': ld.get('description', ''),
                            'source': 'Website (JSON-LD)',
                            'date_posted': ld.get('datePosted', ''),
                        })
            except:
                pass

            # Strategy 2: Text-based extraction from page content
            # STRICT: Only extract lines that look like actual job titles
            # Must have BOTH a job indicator AND an IT keyword
            JOB_TITLE_INDICATORS = [
                'junior', 'senior', 'lead', 'manager', 'engineer', 'entwickler',
                'analyst', 'consultant', 'berater', 'administrator', 'admin',
                'specialist', 'expert', 'architect', 'architekt', 'designer',
                'developer', 'programmierer', 'praktikant', 'trainee',
                'werkstudent', 'ausbildung', 'duales studium', 'absolvent',
                'berufseinsteiger', 'einsteiger', '(m/w/d)', '(w/m/d)',
                'fachinformatiker', 'it-', 'support', 'helpdesk',
                'stelle', 'position', 'vacancy', 'referenz',
            ]
            try:
                page_text = page.evaluate('document.body.innerText')
                if page_text:
                    lines = [l.strip() for l in page_text.split('\n') if l.strip() and 15 < len(l.strip()) < 250]
                    for line in lines:
                        ll = line.lower()
                        has_job_indicator = any(kw in ll for kw in JOB_TITLE_INDICATORS)
                        has_it = any(kw in ll for kw in IT_KEYWORDS)
                        is_sen = any(kw in ll for kw in SENIOR_EXCLUDE)
                        # Must have BOTH job indicator AND IT keyword
                        if has_job_indicator and has_it and not is_sen:
                            jobs.append({
                                'title': line[:300],
                                'url': career_url,
                                'source': 'Website (Text)',
                            })
            except:
                pass

            # Strategy 3: Link extraction
            # STRICT: Links must have BOTH a job keyword AND IT relevance
            try:
                links = page.query_selector_all('a[href]')
                for link in links:
                    text = (link.inner_text() or '').strip()
                    href = link.get_attribute('href') or ''
                    if text and 15 < len(text) < 250:
                        tl = text.lower()
                        has_it = any(kw in tl for kw in IT_KEYWORDS)
                        has_job = any(kw in tl for kw in ['junior', 'trainee', 'werkstudent', 'praktikum', 'ausbildung', 'duales studium', 'absolvent', 'berufseinsteiger', 'einsteiger', 'stelle', 'position', '(m/w/d)', '(w/m/d)', 'fachinformatiker', 'entwickler', 'developer', 'engineer', 'consultant', 'admin', 'support', 'analyst'])
                        is_sen = any(kw in tl for kw in SENIOR_EXCLUDE)
                        # Must have IT keyword AND job keyword
                        if has_it and has_job and not is_sen:
                            job_url = urljoin(career_url, href) if href else career_url
                            jobs.append({
                                'title': text[:300],
                                'url': job_url,
                                'source': 'Website (Link)',
                            })
            except:
                pass

            # Deduplicate by title
            seen_titles = set()
            unique = []
            for job in jobs:
                title_key = job.get('title', '').lower().strip()[:80]
                if title_key not in seen_titles:
                    seen_titles.add(title_key)
                    unique.append(job)

            return unique, career_filepath

        except:
            return [], None
        finally:
            try: page.close()
            except: pass

    def visit_job_detail(self, job_url):
        """Quick visit to job detail for more info."""
        if not job_url or job_url.endswith(('.pdf', '.png', '.jpg')):
            return {}
        page = None
        try:
            page = self._safe_page(job_url, DETAIL_TIMEOUT)
            if not page:
                return {}
            
            result = {}
            
            # Save HTML
            content = page.content()
            if len(content) > 500:
                import random
                fname = f"job_{random.randint(10000,99999)}.html"
                fpath = os.path.join(JOB_HTML_DIR, fname)
                with open(fpath, 'w', encoding='utf-8') as f:
                    f.write(content)
                result['job_html_path'] = fpath
            
            # Extract description
            try:
                main = page.query_selector('main, article, .content, .job-detail, [class*="description"]')
                desc = (main.inner_text() or '').strip() if main else (page.evaluate('document.body.innerText') or '')[:3000]
                if desc:
                    result['description'] = desc[:3000]
            except:
                pass

            # JSON-LD on detail page
            try:
                ld = page.evaluate('''() => {
                    const scripts = document.querySelectorAll('script[type="application/ld+json"]');
                    for (const script of scripts) {
                        try {
                            const data = JSON.parse(script.textContent);
                            const item = Array.isArray(data) ? data[0] : data;
                            if (item['@type'] === 'JobPosting') return JSON.stringify(item);
                        } catch(e) {}
                    }
                    return null;
                }''')
                if ld:
                    ld_data = json.loads(ld)
                    result['date_posted'] = ld_data.get('datePosted', '')
                    result['description'] = ld_data.get('description', result.get('description', ''))[:3000]
            except:
                pass

            return result
        except:
            return {}
        finally:
            if page:
                try: page.close()
                except: pass

    def scrape_employer(self, employer_data):
        name = employer_data.get('name', '')
        website = employer_data.get('website', '')
        city = employer_data.get('city', '')
        plz = employer_data.get('plz', '')
        industry = employer_data.get('industry', '')

        if not website or website.lower() in ['ja', 'nein', 'n/a', '-']:
            return []
        if not website.startswith('http'):
            website = 'https://' + website

        print(f"  {name[:45]}", end='', flush=True)

        # Find career page
        career_url, method = self.find_career_page(name, website)
        
        if not career_url:
            print(f" -> keine Karriere-Seite")
            self.stats['no_career'] += 1
            return []
        
        self.stats['career_found'] += 1
        print(f" -> Karriere ({method[:8]})", end='', flush=True)

        # Extract jobs
        raw_jobs, career_html_path = self.extract_jobs_from_career_page(career_url, name)
        
        if not raw_jobs:
            print(f" -> 0 Jobs")
            self.stats['no_jobs'] += 1
            return []

        # Filter for junior IT
        junior_it_jobs = []
        for job in raw_jobs:
            title = job.get('title', '')
            desc = job.get('description', '')
            is_j, junior_kw = is_junior_job(title, desc)
            is_i, it_kw = is_it_job(title, desc)
            is_sen = is_senior_job(title)
            
            if is_i and not is_sen:
                job['junior_keyword_match'] = junior_kw if is_j else 'no_senior_keyword'
                job['it_keyword_match'] = it_kw
                junior_it_jobs.append(job)

        # Visit detail pages for a few
        for job in junior_it_jobs[:3]:
            if job.get('url') and job['url'] != career_url:
                detail = self.visit_job_detail(job['url'])
                for k, v in detail.items():
                    if v and not job.get(k):
                        job[k] = v

        # Add to Excel
        added = 0
        source_type = 'Website direkt' if method == 'website_navigation' or method == 'nav_footer' or method == 'url_pattern' else f'Suchmaschine ({method[:3]})'
        
        for job in junior_it_jobs:
            job_data = {
                'job_title': job.get('title', ''),
                'description': job.get('description', '')[:5000],
                'job_url': job.get('url', ''),
                'employer_name': name,
                'website_url': website,
                'career_url': career_url,
                'city': job.get('city', '') or city,
                'plz': job.get('plz', '') or plz,
                'industry': industry,
                'source': job.get('source', 'Website'),
                'source_type': source_type,
                'found_via': f'{method}: {name}',
                'junior_keyword_match': job.get('junior_keyword_match', ''),
                'it_keyword_match': job.get('it_keyword_match', ''),
                'career_html_path': career_html_path or '',
                'job_html_path': job.get('job_html_path', ''),
                'date_posted': job.get('date_posted', ''),
                'scrape_status': 'completed',
                'employer_id': hashlib.md5(name.encode()).hexdigest()[:12],
            }
            
            if self.excel.add_job(job_data):
                added += 1
                self.jobs_since_commit += 1

        self.stats['jobs_added'] += added
        print(f" -> {len(junior_it_jobs)} Junior-IT ({added} neu)")

        # Commit check
        if self.jobs_since_commit >= COMMIT_EVERY:
            if git_commit_and_push(self.excel, self.jobs_since_commit):
                self.jobs_since_commit = 0

        return junior_it_jobs


# ============================================================
# MAIN
# ============================================================

def main():
    print("=" * 70)
    print("Junior IT-Jobs Scraper v4 - QUELLENNAH")
    print("CloakBrowser + DuckDuckGo + Website-Navigation")
    print("=" * 70)

    # Load employers
    print("\n[1] Lade Arbeitgeber...")
    wb_emp = openpyxl.load_workbook(EXCEL_EMPLOYERS)
    ws_emp = wb_emp['Firmen mit Website & HTML']

    employers = []
    for r in range(2, ws_emp.max_row + 1):
        name = ws_emp.cell(r, 2).value or ''
        website_gm = ws_emp.cell(r, 9).value or ''
        website_found = ws_emp.cell(r, 10).value or ''
        ort = ws_emp.cell(r, 7).value or ''
        plz = ws_emp.cell(r, 6).value or ''
        kategorie = ws_emp.cell(r, 4).value or ''
        # Only use actual URLs, not 'Ja'/'Nein' flags
        website = ''
        if website_found and website_found.startswith('http'):
            website = website_found
        elif website_gm and website_gm.startswith('http'):
            website = website_gm
        if name and website:
            employers.append({'name': name, 'website': website, 'city': ort, 'plz': plz, 'industry': kategorie})

    # Priority: IT employers first
    it_kw = ['it', 'software', 'data', 'digital', 'tech', 'system', 'informatik',
             'computer', 'cloud', 'devops', 'cyber', 'entwickler', 'consulting',
             'telekommunikation', 'elektronik', 'automatisierung', 'engineering']
    it_employers = [e for e in employers if any(kw in (e['name'].lower() + ' ' + e['industry'].lower()) for kw in it_kw)]
    other_employers = [e for e in employers if e not in it_employers]
    priority_employers = it_employers + other_employers

    print(f"  {len(employers)} Arbeitgeber ({len(it_employers)} IT-Priorität)")

    # Load progress
    progress = load_progress()
    completed_names = set(progress.get('completed', []))
    start_index = progress.get('last_index', 0)
    print(f"  Bereits verarbeitet: {len(completed_names)}")

    # Init Excel
    excel = JuniorJobsExcel(EXCEL_OUT)
    print(f"  {excel.get_job_count()} existierende Jobs")

    # Start scraper
    scraper = CloakScraper(excel)
    scraper.start_browser()

    try:
        for idx, employer in enumerate(priority_employers):
            if idx < start_index:
                continue
            if employer['name'] in completed_names:
                continue

            try:
                print(f"\n[{idx+1}/{len(priority_employers)}] ", end='', flush=True)
                scraper.scrape_employer(employer)
                completed_names.add(employer['name'])
            except Exception as e:
                print(f" [✗] ERROR: {str(e)[:80]}", flush=True)
                traceback.print_exc()
                # Restart browser on error
                try: scraper.stop_browser()
                except: pass
                time.sleep(3)
                try:
                    scraper.start_browser()
                except Exception as e2:
                    print(f" [✗] Browser restart failed: {e2}", flush=True)
                    break

            # Save progress every 3 employers
            if idx % 3 == 0:
                progress['completed'] = list(completed_names)
                progress['last_index'] = idx + 1
                progress['stats'] = dict(scraper.stats)
                save_progress(progress)
                excel.save()

            time.sleep(0.5)

    except KeyboardInterrupt:
        print("\n[!] Abgebrochen")
    finally:
        scraper.stop_browser()
        excel.save()
        if scraper.jobs_since_commit > 0:
            git_commit_and_push(excel, scraper.jobs_since_commit)
        
        progress['completed'] = list(completed_names)
        progress['stats'] = dict(scraper.stats)
        save_progress(progress)
        
        print(f"\n{'='*70}")
        print(f"ERGEBNIS: {excel.get_job_count()} Junior IT-Jobs (quellennah)")
        print(f"Verarbeitet: {len(completed_names)} | Stats: {dict(scraper.stats)}")
        print(f"{'='*70}")


if __name__ == '__main__':
    main()
