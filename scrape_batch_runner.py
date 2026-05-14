#!/usr/bin/env python3
"""Batch runner that processes each employer in a separate subprocess."""
import json, subprocess, sys, time, os
from pathlib import Path

EXCEL_EMPLOYERS = "/home/z/my-project/download/Firmen_Bamberg_Erlangen_Nuernberg_mit_Website_HTML.xlsx"
PROGRESS_FILE = "/home/z/my-project/download/scrape_v3_progress.json"
SCRAPE_SCRIPT = "/home/z/my-project/scrape_subprocess.py"
GIT_REPO = "/home/z/my-project"
GITHUB_TOKEN = "GIT_TOKEN_REDACTED"

import openpyxl

def load_progress():
    if os.path.exists(PROGRESS_FILE):
        return json.load(open(PROGRESS_FILE, encoding='utf-8'))
    return {'completed': [], 'failed': [], 'last_index': 0, 'stats': {}}

def save_progress(progress):
    with open(PROGRESS_FILE, 'w', encoding='utf-8') as f:
        json.dump(progress, f, ensure_ascii=False, indent=1)

def git_commit():
    try:
        os.chdir(GIT_REPO)
        subprocess.run(['git', 'add', '-A'], capture_output=True, timeout=30)
        result = subprocess.run(['git', 'commit', '-m', 'Junior IT-Jobs quellennah update'], capture_output=True, timeout=30)
        if result.returncode == 0:
            subprocess.run(
                ['git', 'push', f'https://{GITHUB_TOKEN}@github.com/BirdNest055/Heyjobs.git', 'main'],
                capture_output=True, timeout=60
            )
            print("  [GIT] Pushed!", flush=True)
    except:
        pass

def main():
    print("Junior IT-Jobs Batch Runner", flush=True)
    print("=" * 50, flush=True)
    
    # Load employers
    wb = openpyxl.load_workbook(EXCEL_EMPLOYERS)
    ws = wb['Firmen mit Website & HTML']
    
    employers = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 2).value or ''
        website_gm = ws.cell(r, 9).value or ''
        website_found = ws.cell(r, 10).value or ''
        ort = ws.cell(r, 7).value or ''
        plz = ws.cell(r, 6).value or ''
        kategorie = ws.cell(r, 4).value or ''
        website = ''
        if website_found and website_found.startswith('http'):
            website = website_found
        elif website_gm and website_gm.startswith('http'):
            website = website_gm
        if name and website:
            employers.append({'name': name, 'website': website, 'city': ort, 'plz': plz, 'industry': kategorie})
    
    # IT companies first
    it_kw = ['it', 'software', 'data', 'digital', 'tech', 'system', 'informatik',
             'computer', 'cloud', 'devops', 'entwickler', 'consulting', 'engineering']
    it = [e for e in employers if any(kw in (e['name'] + ' ' + e['industry']).lower() for kw in it_kw)]
    others = [e for e in employers if e not in it]
    employers = it + others
    
    print(f"{len(employers)} employers ({len(it)} IT priority)", flush=True)
    
    progress = load_progress()
    completed = set(progress.get('completed', []))
    failed = set(progress.get('failed', []))
    
    print(f"Already completed: {len(completed)}, failed: {len(failed)}", flush=True)
    
    for idx, emp in enumerate(employers):
        if emp['name'] in completed or emp['name'] in failed:
            continue
        
        print(f"\n[{idx+1}/{len(employers)}] {emp['name'][:45]}", flush=True)
        
        # Run in subprocess with 60s timeout
        try:
            emp_json = json.dumps(emp, ensure_ascii=False)
            result = subprocess.run(
                ['python3', '-u', SCRAPE_SCRIPT],
                input=emp_json,
                capture_output=True,
                text=True,
                timeout=60,
                env={**os.environ, 'DISPLAY': ':99', 'PYTHONUNBUFFERED': '1'},
            )
            
            output = result.stdout.strip()
            error = result.stderr.strip()
            
            if result.returncode == 0:
                print(f"  {output}", flush=True)
                completed.add(emp['name'])
            else:
                print(f"  FAILED (exit {result.returncode}): {output[:100]}", flush=True)
                if error:
                    print(f"  stderr: {error[:100]}", flush=True)
                failed.add(emp['name'])
                
        except subprocess.TimeoutExpired:
            print(f"  TIMEOUT (60s)", flush=True)
            failed.add(emp['name'])
        except Exception as e:
            print(f"  ERROR: {str(e)[:80]}", flush=True)
            failed.add(emp['name'])
        
        # Save progress every 5
        if idx % 5 == 0:
            progress['completed'] = list(completed)
            progress['failed'] = list(failed)
            progress['last_index'] = idx + 1
            save_progress(progress)
            git_commit()
        
        time.sleep(1)
    
    # Final commit
    progress['completed'] = list(completed)
    progress['failed'] = list(failed)
    save_progress(progress)
    git_commit()
    
    print(f"\n{'='*50}", flush=True)
    print(f"DONE! Completed: {len(completed)}, Failed: {len(failed)}", flush=True)

if __name__ == '__main__':
    main()
