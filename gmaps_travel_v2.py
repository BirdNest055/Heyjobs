#!/usr/bin/env python3
"""
Google Maps Travel Scraper v2 - Production
Travels through Google Maps clicking into detail pages for rich data.
"""

import json, os, re, sys, time, random, subprocess, signal, traceback
from datetime import datetime
from pathlib import Path

# Config
CHROME_PATH = os.path.expanduser("~/.cache/ms-playwright/chromium-1223/chrome-linux64/chrome")
RESULTS_FILE = "/home/z/my-project/gmaps_travel_results.json"
PROGRESS_FILE = "/home/z/my-project/gmaps_travel_progress.json"
EXISTING_RESULTS = "/home/z/my-project/gmaps_erlangen_results.json"
TOWNS_FILE = "/home/z/my-project/gmaps_erlangen_towns.json"
OLD_PROGRESS = "/home/z/my-project/gmaps_erlangen_progress.json"
EXCEL_OUTPUT = "/home/z/my-project/download/Google_Maps_Firmen_Erlangen_50km.xlsx"
GITHUB_TOKEN = "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"
PROJECT_DIR = "/home/z/my-project"

SEARCH_TERMS = {
    1: ["Firma", "Unternehmen", "GmbH", "IT Firma", "Software", "Softwareentwicklung",
        "Handwerk", "Industrie", "Handel", "Dienstleistung", "Gastronomie",
        "Baufirma", "Kfz Werkstatt", "Steuerberater", "Rechtsanwalt",
        "Versicherung", "Immobilien", "Marketing Agentur", "Logistik",
        "Elektro", "Metallbau", "Maschinenbau", "Kunststoff", "Druckerei",
        "Apotheke", "Arztpraxis", "Friseur", "Bäckerei",
        "Ingenieurbüro", "Consulting", "Architekt", "Autohaus",
        "Bank", "Bildung", "Gesundheit", "Pharma"],
    2: ["Firma", "Unternehmen", "GmbH", "IT", "Handwerk", "Dienstleistung",
        "Industrie", "Handel", "Gastronomie", "Steuerberater", "Elektro"],
    3: ["Firma", "Unternehmen", "GmbH", "Handwerk", "Dienstleistung", "Handel", "IT"],
    4: ["Firma", "Unternehmen", "Gewerbe", "Handwerk", "Dienstleistung"],
}


def load_json(path, default=None):
    try:
        with open(path) as f:
            return json.load(f)
    except:
        return default if default is not None else {}

def save_json(path, data):
    with open(path, 'w') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def git_push(message):
    try:
        os.chdir(PROJECT_DIR)
        subprocess.run(["git", "add", "-A"], capture_output=True, timeout=30)
        r = subprocess.run(["git", "commit", "-m", message], capture_output=True, timeout=30)
        if r.returncode != 0 and "nothing to commit" in r.stdout.decode():
            return
        subprocess.run(
            ["git", "push", f"https://{GITHUB_TOKEN}@github.com/BirdNest055/Heyjobs.git", "main"],
            capture_output=True, timeout=60
        )
        print(f"  [GIT] {message}", flush=True)
    except Exception as e:
        print(f"  [GIT-ERR] {e}", flush=True)


def extract_detail(page, town, search_term):
    """Extract all details from a company's Google Maps detail page"""
    result = {
        "name": "", "rating": "", "review_count": "", "category": "",
        "address": "", "plz": "", "city": "", "phone": "", "website": "",
        "lat": "", "lon": "", "gmaps_url": "",
        "search_town": town, "search_term": search_term,
        "scraped_at": datetime.now().isoformat(),
        "source": "google_maps_detail",
    }
    try:
        page.wait_for_selector('[role="main"]', timeout=6000)
    except:
        pass
    time.sleep(0.5)

    # Name
    try:
        el = page.query_selector('h1.DUwDvf, h1.fontHeadlineLarge')
        if el:
            result["name"] = el.inner_text().strip()
    except:
        pass

    # Rating
    try:
        el = page.query_selector('div.F7nice')
        if el:
            se = el.query_selector('span[aria-label]')
            if se:
                a = se.get_attribute('aria-label') or ''
                m = re.search(r'([\d,]+)\s*Sterne', a)
                if m:
                    result["rating"] = m.group(1).replace(',', '.')
    except:
        pass

    # Reviews
    try:
        el = page.query_selector('button[aria-label*="Bewertungen"]')
        if el:
            a = el.get_attribute('aria-label') or ''
            m = re.search(r'([\d.]+)\s*Bewertungen', a)
            if m:
                result["review_count"] = m.group(1)
    except:
        pass

    # Category
    try:
        el = page.query_selector('button[jsaction*="category"]')
        if el:
            result["category"] = el.inner_text().strip()
    except:
        pass

    # Website
    try:
        el = page.query_selector('a[data-item-id="authority"]')
        if el:
            href = el.get_attribute('href') or ''
            if href and 'google' not in href.lower():
                result["website"] = href
    except:
        pass
    if not result["website"]:
        try:
            el = page.query_selector('a[aria-label*="Website"]')
            if el:
                href = el.get_attribute('href') or ''
                if href and 'google' not in href.lower() and 'gstatic' not in href.lower() and not href.startswith('tel:'):
                    result["website"] = href
        except:
            pass

    # Address, Phone from data-item-id
    try:
        for el in page.query_selector_all('[data-item-id]'):
            iid = el.get_attribute('data-item-id') or ''
            if 'address' in iid and not result["address"]:
                text = re.sub(r'[\ue000-\uefff]', '', el.inner_text()).strip()
                text = re.sub(r'\s+', ' ', text).strip()
                result["address"] = text
                m = re.search(r'(\d{5})', text)
                if m:
                    result["plz"] = m.group(1)
                m = re.search(r'\d{5}\s+([A-Za-zäöüßÄÖÜ\-\.]+(?:\s+[A-Za-zäöüßÄÖÜ\-\.]+)*)', text)
                if m:
                    result["city"] = m.group(1).strip()
            elif 'phone:tel:' in iid and not result["phone"]:
                m = re.search(r'phone:tel:(.+)', iid)
                if m:
                    result["phone"] = m.group(1)
    except:
        pass

    # Coordinates
    try:
        url = page.url
        m = re.search(r'@(-?[\d.]+),(-?[\d.]+)', url)
        if m:
            result["lat"] = m.group(1)
            result["lon"] = m.group(2)
        result["gmaps_url"] = url
    except:
        pass

    return result


def search_and_extract(page, town, search_term, max_scroll=12):
    """Search Google Maps, click into results, extract details"""
    companies = []
    seen_names = set()

    query = f"{town} {search_term}"
    url = f"https://www.google.com/maps/search/{query.replace(' ', '+')}"

    try:
        page.goto(url, timeout=20000, wait_until="domcontentloaded")
    except:
        try:
            page.goto(url, timeout=15000)
        except:
            print(f"  [WARN] Cannot load: {query}", flush=True)
            return companies

    time.sleep(random.uniform(3, 5))

    # Cookie consent
    try:
        accept = page.query_selector('button[aria-label*="Alle ablehnen"], button[aria-label*="Reject"]')
        if accept:
            accept.click()
            time.sleep(1)
    except:
        pass

    # Scroll to load results
    try:
        scroll_panel = page.query_selector('div[role="feed"]')
        if scroll_panel:
            last_count = 0
            same_rounds = 0
            for i in range(max_scroll):
                page.evaluate('(el) => el.scrollTop = el.scrollHeight', scroll_panel)
                time.sleep(random.uniform(0.8, 1.5))
                items = page.query_selector_all('a[href*="/maps/place/"]')
                if len(items) == last_count:
                    same_rounds += 1
                    if same_rounds >= 3:
                        break
                else:
                    same_rounds = 0
                last_count = len(items)
    except:
        pass

    # Collect unique result links
    seen_hrefs = set()
    unique_items = []
    for item in page.query_selector_all('a[href*="/maps/place/"]'):
        href = item.get_attribute('href') or ''
        if href and href not in seen_hrefs:
            seen_hrefs.add(href)
            unique_items.append((item, href))

    print(f"  [{query}] {len(unique_items)} unique results", flush=True)

    for idx, (item, href) in enumerate(unique_items):
        try:
            # Get name from link
            name = ""
            try:
                ne = item.query_selector('.fontHeadlineSmall, .qBF1Pd')
                if ne:
                    name = ne.inner_text().strip()
            except:
                pass
            if not name:
                try:
                    name = item.get_attribute('aria-label') or ''
                except:
                    pass

            if name in seen_names:
                continue
            seen_names.add(name)

            # Click into detail
            item.click()
            time.sleep(random.uniform(2, 3.5))

            detail = extract_detail(page, town, search_term)
            if detail.get("name"):
                companies.append(detail)
                w = detail.get('website', '-')
                p = detail.get('plz', '-')
                c = detail.get('category', '-')
                print(f"    [{idx+1}/{len(unique_items)}] {detail['name'][:40]} | web={w[:30]} | plz={p} | cat={c}", flush=True)
            elif name:
                detail["name"] = name
                companies.append(detail)

            # Go back to results
            try:
                page.go_back(timeout=8000)
                time.sleep(random.uniform(1.5, 2.5))
            except:
                page.goto(url, timeout=15000)
                time.sleep(2)

        except Exception as e:
            print(f"    [{idx+1}] Error: {e}", flush=True)
            try:
                page.goto(url, timeout=15000)
                time.sleep(2)
            except:
                pass

    return companies


def create_excel(companies):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import CellIsRule
    except:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Google Maps Firmen"

    headers = [
        "#", "Firmenname", "Bewertung", "Anzahl Bewertungen", "Kategorie",
        "Adresse", "PLZ", "Ort", "Telefon", "Website",
        "Breitengrad", "Längengrad", "Google Maps Link",
        "Suchort", "Suchbegriff", "Quelle", "Scraped At"
    ]

    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    tb = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = hf, hfill, ha, tb

    for idx, comp in enumerate(companies, 1):
        rd = [idx, comp.get("name",""), comp.get("rating",""), comp.get("review_count",""),
              comp.get("category",""), comp.get("address",""), comp.get("plz",""),
              comp.get("city",""), comp.get("phone",""), comp.get("website",""),
              comp.get("lat",""), comp.get("lon",""), comp.get("gmaps_url",""),
              comp.get("search_town",""), comp.get("search_term",""),
              comp.get("source",""), comp.get("scraped_at","")]
        for col, v in enumerate(rd, 1):
            ws.cell(row=idx+1, column=col, value=v).border = tb

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(companies)+1}"
    widths = [5,40,10,12,30,45,8,25,22,45,12,12,55,25,20,22,22]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i+1)].width = w

    gf = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ws.conditional_formatting.add(f"J2:J{len(companies)+1}", CellIsRule(operator="notEqual", formula=['""'], fill=gf))

    # Stats
    ws2 = wb.create_sheet("Statistiken")
    stats = [
        ("Gesamtzahl Firmen", len(companies)),
        ("Mit Website", sum(1 for c in companies if c.get("website"))),
        ("Ohne Website", sum(1 for c in companies if not c.get("website"))),
        ("Mit Telefon", sum(1 for c in companies if c.get("phone"))),
        ("Mit PLZ", sum(1 for c in companies if c.get("plz"))),
        ("Mit Adresse", sum(1 for c in companies if c.get("address"))),
        ("Mit Kategorie", sum(1 for c in companies if c.get("category"))),
        ("Detail-basiert", sum(1 for c in companies if c.get("source")=="google_maps_detail")),
    ]
    for r, (s, v) in enumerate(stats, 1):
        ws2.cell(row=r, column=1, value=s).font = Font(bold=True)
        ws2.cell(row=r, column=2, value=v)
    ws2.column_dimensions['A'].width = 25

    # Top Orte
    ws3 = wb.create_sheet("Top Orte")
    cc = {}
    for c in companies:
        city = c.get("city", c.get("search_town", "?"))
        cc[city] = cc.get(city, 0) + 1
    ws3.cell(row=1, column=1, value="Ort").font = Font(bold=True)
    ws3.cell(row=1, column=2, value="Anzahl").font = Font(bold=True)
    for r, (city, cnt) in enumerate(sorted(cc.items(), key=lambda x: -x[1]), 2):
        ws3.cell(row=r, column=1, value=city)
        ws3.cell(row=r, column=2, value=cnt)
    ws3.column_dimensions['A'].width = 35

    # Top Kategorien
    ws4 = wb.create_sheet("Top Kategorien")
    catc = {}
    for c in companies:
        cat = c.get("category", "Ohne")
        if cat:
            catc[cat] = catc.get(cat, 0) + 1
    ws4.cell(row=1, column=1, value="Kategorie").font = Font(bold=True)
    ws4.cell(row=1, column=2, value="Anzahl").font = Font(bold=True)
    for r, (cat, cnt) in enumerate(sorted(catc.items(), key=lambda x: -x[1]), 2):
        ws4.cell(row=r, column=1, value=cat)
        ws4.cell(row=r, column=2, value=cnt)
    ws4.column_dimensions['A'].width = 40

    wb.save(EXCEL_OUTPUT)
    print(f"  [EXCEL] {len(companies)} companies saved", flush=True)


def main():
    print("=== Starting Google Maps Travel Scraper v2 ===", flush=True)

    # Load data
    existing = load_json(EXISTING_RESULTS, [])
    progress = load_json(PROGRESS_FILE, {
        "completed_searches": [],
        "enriched_companies": [],
        "phase": "search_new_towns",
        "last_enrich_idx": 0,
    })
    towns = load_json(TOWNS_FILE, [])
    old_progress = load_json(OLD_PROGRESS, {"completed_searches": []})

    travel_results = load_json(RESULTS_FILE, [])
    all_companies = {}

    # Merge existing + travel results
    for c in existing:
        key = f"{c.get('name','').lower().strip()}|{c.get('city', c.get('search_town','')).lower().strip()}"
        if key not in all_companies:
            all_companies[key] = c
        else:
            for k, v in c.items():
                if v and not all_companies[key].get(k):
                    all_companies[key][k] = v

    for c in travel_results:
        key = f"{c.get('name','').lower().strip()}|{c.get('city', c.get('search_town','')).lower().strip()}"
        if key not in all_companies:
            all_companies[key] = c
        else:
            ec = all_companies[key]
            for k, v in c.items():
                if v and (not ec.get(k) or k in ("website","category","plz","address","source","scraped_at")):
                    ec[k] = v

    print(f"Merged: {len(existing)} existing + {len(travel_results)} travel = {len(all_companies)} total", flush=True)

    # Find missing towns
    searched = set()
    for s in old_progress.get("completed_searches", []):
        searched.add(s.split("|")[0])
    for s in progress.get("completed_searches", []):
        searched.add(s.split("|")[0])

    town_names = [t["name"] if isinstance(t, dict) else t for t in towns]
    missing = [t for t in town_names if t not in searched]
    print(f"Missing towns ({len(missing)}): {missing}", flush=True)

    # Start Xvfb in-process
    xvfb_proc = subprocess.Popen(
        ['Xvfb', ':99', '-screen', '0', '1920x1080x24', '-ac', '+extension', 'GLX', '+render', '-noreset'],
        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL
    )
    time.sleep(3)
    os.environ['DISPLAY'] = ':99'

    if xvfb_proc.poll() is not None:
        print("FATAL: Xvfb failed to start!", flush=True)
        sys.exit(1)
    print("Xvfb started", flush=True)

    from playwright.sync_api import sync_playwright

    with sync_playwright() as p:
        print("Launching browser...", flush=True)
        browser = p.chromium.launch(
            headless=False,
            executable_path=CHROME_PATH,
            args=['--no-sandbox', '--disable-gpu', '--disable-dev-shm-usage',
                  '--disable-blink-features=AutomationControlled', '--lang=de-DE']
        )
        context = browser.new_context(
            viewport={"width": 1920, "height": 1080},
            locale="de-DE",
            user_agent="Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
        )
        page = context.new_page()
        print("Browser ready!", flush=True)

        push_counter = 0

        # Phase 1: Search missing towns
        if progress.get("phase", "search_new_towns") == "search_new_towns":
            print(f"\n=== PHASE 1: {len(missing)} missing towns ===", flush=True)

            for town_name in missing:
                town_info = None
                for t in towns:
                    tn = t["name"] if isinstance(t, dict) else t
                    if tn == town_name:
                        town_info = t
                        break

                tier = town_info.get("population_tier", 4) if isinstance(town_info, dict) else 4
                terms = SEARCH_TERMS.get(tier, SEARCH_TERMS[4])

                print(f"\n--- {town_name} (tier {tier}, {len(terms)} terms) ---", flush=True)

                for term in terms:
                    skey = f"{town_name}|{term}"
                    if skey in progress.get("completed_searches", []):
                        continue

                    try:
                        companies = search_and_extract(page, town_name, term)
                    except Exception as e:
                        print(f"  [ERROR] {town_name} {term}: {e}", flush=True)
                        traceback.print_exc()
                        try:
                            page.goto('about:blank', timeout=5000)
                        except:
                            pass
                        continue

                    new_count = 0
                    for c in companies:
                        key = f"{c.get('name','').lower().strip()}|{c.get('city', c.get('search_town','')).lower().strip()}"
                        if key not in all_companies:
                            all_companies[key] = c
                            new_count += 1
                        else:
                            ec = all_companies[key]
                            for k, v in c.items():
                                if v and not ec.get(k):
                                    ec[k] = v

                    progress.setdefault("completed_searches", []).append(skey)
                    push_counter += new_count

                    print(f"  New: {new_count} | Total: {len(all_companies)} | Push counter: {push_counter}", flush=True)

                    save_json(RESULTS_FILE, list(all_companies.values()))
                    save_json(PROGRESS_FILE, progress)

                    if push_counter >= 10:
                        create_excel(list(all_companies.values()))
                        git_push(f"GMaps Travel: {town_name} +{push_counter} | {len(all_companies)} total")
                        push_counter = 0

                    time.sleep(random.uniform(1, 2))

            progress["phase"] = "enrich_existing"
            save_json(PROGRESS_FILE, progress)
            create_excel(list(all_companies.values()))
            git_push(f"GMaps Phase 1 DONE | {len(all_companies)} total")
            print(f"\n=== PHASE 1 DONE | {len(all_companies)} total ===", flush=True)

        # Phase 2: Enrich existing companies
        if progress.get("phase") == "enrich_existing":
            print(f"\n=== PHASE 2: Enriching with website URLs ===", flush=True)

            need = []
            for key, c in all_companies.items():
                hw = c.get("has_website", False)
                hu = bool(c.get("website") and "google" not in c.get("website","").lower())
                hp = bool(c.get("plz"))
                hc = bool(c.get("category"))
                if hw and not hu:
                    need.append((1, key, c))
                elif not hp:
                    need.append((2, key, c))
                elif not hc and hw:
                    need.append((3, key, c))

            need.sort(key=lambda x: x[0])
            p1 = sum(1 for p,_,_ in need if p==1)
            p2 = sum(1 for p,_,_ in need if p==2)
            p3 = sum(1 for p,_,_ in need if p==3)
            print(f"Enrichment needed: {len(need)} (P1 web:{p1} P2 plz:{p2} P3 cat:{p3})", flush=True)

            enriched = set(progress.get("enriched_companies", []))
            start = progress.get("last_enrich_idx", 0)
            done = 0

            for i in range(start, len(need)):
                pri, key, comp = need[i]
                name = comp.get("name", "")
                town = comp.get("city", comp.get("search_town", ""))

                if key in enriched:
                    continue

                query = f"{name} {town}"
                url = f"https://www.google.com/maps/search/{query.replace(' ', '+')}"

                try:
                    page.goto(url, timeout=15000, wait_until="domcontentloaded")
                    time.sleep(random.uniform(3, 5))

                    try:
                        accept = page.query_selector('button[aria-label*="Alle ablehnen"]')
                        if accept:
                            accept.click()
                            time.sleep(0.5)
                    except:
                        pass

                    first = page.query_selector('a[href*="/maps/place/"]')
                    if first:
                        first.click()
                        time.sleep(random.uniform(2, 3))
                        detail = extract_detail(page, town, "enrichment")

                        if detail:
                            updated = False
                            if detail.get("website") and "google" not in detail["website"].lower() and not comp.get("website"):
                                comp["website"] = detail["website"]
                                updated = True
                            if detail.get("phone") and not comp.get("phone"):
                                comp["phone"] = detail["phone"]
                                updated = True
                            if detail.get("address") and not comp.get("address"):
                                comp["address"] = detail["address"]
                                updated = True
                            if detail.get("plz") and not comp.get("plz"):
                                comp["plz"] = detail["plz"]
                                updated = True
                            if detail.get("category") and not comp.get("category"):
                                comp["category"] = detail["category"]
                                updated = True
                            if detail.get("city") and not comp.get("city"):
                                comp["city"] = detail["city"]
                                updated = True
                            if detail.get("lat"):
                                comp["lat"] = detail["lat"]
                            if detail.get("lon"):
                                comp["lon"] = detail["lon"]
                            if updated:
                                comp["source"] = "google_maps_detail"
                                comp["enriched_at"] = datetime.now().isoformat()
                            if comp.get("website"):
                                print(f"  [E] {name[:35]}: web={comp['website'][:40]}", flush=True)

                except Exception as e:
                    pass

                enriched.add(key)
                done += 1
                push_counter += 1

                if done % 10 == 0:
                    save_json(RESULTS_FILE, list(all_companies.values()))
                    progress["enriched_companies"] = list(enriched)[-500:]
                    progress["last_enrich_idx"] = i + 1
                    save_json(PROGRESS_FILE, progress)
                    create_excel(list(all_companies.values()))
                    git_push(f"GMaps Enrich: {done} done | {len(all_companies)} total")
                    push_counter = 0

                time.sleep(random.uniform(0.5, 1.5))

            save_json(RESULTS_FILE, list(all_companies.values()))
            progress["enriched_companies"] = list(enriched)[-500:]
            progress["last_enrich_idx"] = len(need)
            save_json(PROGRESS_FILE, progress)
            create_excel(list(all_companies.values()))
            git_push(f"GMaps COMPLETE | {len(all_companies)} total")

        browser.close()

    try:
        os.kill(xvfb_proc.pid, signal.SIGTERM)
    except:
        pass

    print(f"\n=== ALL DONE | {len(all_companies)} companies ===", flush=True)


if __name__ == "__main__":
    main()
