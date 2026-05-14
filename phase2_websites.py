#!/usr/bin/env python3
"""
Phase 2: Visit employer websites, find career pages, download HTML
Uses Playwright to browse like a human would.
Processes employers in parallel and commits every 10 entries.
"""
import asyncio, json, os, re, subprocess, hashlib
from datetime import datetime
from urllib.parse import urljoin, urlparse
from playwright.async_api import async_playwright
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

EXCEL_OUT = "/home/z/my-project/download/Junior_IT_Jobs_Bamberg_Erlangen_Nuernberg.xlsx"
JSON_IN = "/home/z/my-project/download/Junior_IT_Jobs_Bamberg_Erlangen_Nuernberg_raw.json"
CAREER_HTML_DIR = "/home/z/my-project/download/career_pages_html"
JOB_HTML_DIR = "/home/z/my-project/download/junior_jobs_html"
GIT_REPO = "/home/z/my-project"
GITHUB_TOKEN = "GIT_TOKEN_REDACTED"
COMMIT_EVERY = 10
MAX_CONCURRENT = 4

# Also load the existing employer data for website lookup
EMPLOYER_EXCEL = "/home/z/my-project/download/IT_Arbeitgeber_Bamberg_Erlangen_Nuernberg.xlsx"
FULL_EMPLOYER_EXCEL = "/home/z/my-project/download/Firmen_Bamberg_Erlangen_Nuernberg_mit_Website_HTML.xlsx"


def sanitize_filename(name, max_len=60):
    name = re.sub(r'[^\w\s.-]', '_', name)
    name = re.sub(r'_{2,}', '_', name)
    return name.strip('_')[:max_len]


def load_employer_websites():
    """Load employer -> website mapping from existing Excel files."""
    websites = {}  # normalized employer name -> website URL
    
    # From IT employer Excel
    try:
        wb = openpyxl.load_workbook(EMPLOYER_EXCEL)
        ws = wb['Nur IT-Jobs']
        for r in range(2, ws.max_row + 1):
            name = ws.cell(r, 2).value or ''
            website = ws.cell(r, 9).value
            if name and website:
                norm = re.sub(r'\s+', '', name.lower())[:15]
                if norm not in websites:
                    websites[norm] = {
                        'url': website,
                        'name': name,
                        'email': ws.cell(r, 8).value or '',
                        'phone': ws.cell(r, 7).value or '',
                        'city': ws.cell(r, 4).value or '',
                        'industry': ws.cell(r, 23).value or '',
                    }
        wb.close()
    except Exception as e:
        print(f"  Warning: Could not load IT employer Excel: {e}")
    
    # From full B/ER/N Excel
    try:
        wb2 = openpyxl.load_workbook(FULL_EMPLOYER_EXCEL)
        ws2 = wb2['Firmen mit Website & HTML']
        for r in range(2, ws2.max_row + 1):
            name = ws2.cell(r, 2).value or ''
            website = ws2.cell(r, 11).value  # Website (gefunden)
            if name and website:
                norm = re.sub(r'\s+', '', name.lower())[:15]
                if norm not in websites:
                    websites[norm] = {
                        'url': website,
                        'name': name,
                        'email': ws2.cell(r, 11).value or '',  # E-Mail column
                        'phone': ws2.cell(r, 8).value or '',
                        'city': ws2.cell(r, 7).value or '',
                        'industry': ws2.cell(r, 4).value or '',
                    }
        wb2.close()
    except Exception as e:
        print(f"  Warning: Could not load full employer Excel: {e}")
    
    return websites


def match_employer(aa_name, website_map):
    """Try to match an AA employer name to our website database."""
    # Normalize and try different approaches
    norm = re.sub(r'\s+', '', aa_name.lower())
    
    # Direct match
    for key, val in website_map.items():
        if key in norm or norm[:15] in key:
            return val
    
    # Try without common suffixes
    clean = re.sub(r'(gmbh|ag|ek|ug|kg|gbr|ohg|mbh|co|kgaa).*$', '', norm)
    if len(clean) >= 5:
        for key, val in website_map.items():
            clean_key = re.sub(r'(gmbh|ag|ek|ug|kg|gbr|ohg|mbh|co|kgaa).*$', '', key)
            if clean[:10] in clean_key or clean_key[:10] in clean:
                return val
    
    return None


CAREER_LINK_TEXTS = [
    'karriere', 'career', 'jobs', 'stellenangebote', 'stellen',
    'offene stellen', 'vacancies', 'careers', 'jobbörse',
    'recruiting', 'bewerbung', 'talent', 'opportunities',
    'join us', 'join', 'wir suchen', 'wir stellen ein',
    'offene positionen', 'open positions', 'arbeiten bei',
]

CAREER_URL_PATTERNS = [
    '/karriere', '/career', '/careers', '/jobs', '/stellenangebote',
    '/stellen', '/offene-stellen', '/vacancies', '/job',
    '/recruiting', '/bewerbung', '/talent', '/join-us',
    '/de/karriere', '/en/career', '/de/career', '/en/careers',
]


async def visit_employer(context, employer_name, website_url, semaphore, results):
    """Visit an employer website and find career pages."""
    async with semaphore:
        page = None
        try:
            page = await context.new_page()
            await page.set_extra_http_headers({'Accept-Language': 'de-DE,de;q=0.9,en-US;q=0.8'})
            
            result = {
                'employer_name': employer_name,
                'website_url': website_url,
                'career_url': None,
                'career_html_path': None,
                'job_urls_found': [],
                'error': None,
            }
            
            # Visit homepage
            try:
                await page.goto(website_url, timeout=15000, wait_until='domcontentloaded')
                await asyncio.sleep(1.5)
            except Exception as e:
                result['error'] = f"Homepage timeout: {str(e)[:50]}"
                results.append(result)
                return result
            
            # Dismiss cookie consent
            try:
                for sel in ['button:has-text("Akzeptieren")', 'button:has-text("Accept")',
                           'button:has-text("Alle akzeptieren")', 'button:has-text("Accept All")',
                           'button:has-text("OK")', 'button:has-text("Verstanden")',
                           '#onetrust-accept-btn-handler', '.cc-btn']:
                    try:
                        btn = await page.query_selector(sel)
                        if btn:
                            await btn.click()
                            await asyncio.sleep(0.3)
                            break
                    except:
                        continue
            except:
                pass
            
            # Find career link
            career_url = None
            
            # Strategy 1: Look for career links on page
            for sel in ['a:has-text("Karriere")', 'a:has-text("Career")',
                       'a:has-text("Jobs")', 'a:has-text("Stellenangebote")',
                       'a:has-text("Stellen")', 'a:has-text("Careers")',
                       'a:has-text("Wir suchen")', 'a:has-text("Bewerbung")',
                       'a:has-text("Join us")', 'a:has-text("Recruiting")']:
                try:
                    elems = await page.query_selector_all(sel)
                    for elem in elems:
                        href = await elem.get_attribute('href')
                        if href:
                            full_url = urljoin(website_url, href)
                            career_url = full_url
                            break
                except:
                    continue
                if career_url:
                    break
            
            # Strategy 2: Check nav and footer links
            if not career_url:
                try:
                    for sel in ['nav a', 'header a', 'footer a', '.nav a', '.menu a']:
                        links = await page.query_selector_all(sel)
                        for link in links:
                            text = (await link.inner_text()).strip().lower()
                            href = await link.get_attribute('href')
                            if href and any(kw in text for kw in CAREER_LINK_TEXTS):
                                career_url = urljoin(website_url, href)
                                break
                        if career_url:
                            break
                except:
                    pass
            
            # Strategy 3: Try common URL patterns
            if not career_url:
                parsed = urlparse(website_url)
                base = f"{parsed.scheme}://{parsed.netloc}"
                for pattern in CAREER_URL_PATTERNS:
                    try:
                        test_url = base + pattern
                        resp = await page.goto(test_url, timeout=6000, wait_until='domcontentloaded')
                        if resp and resp.status < 400:
                            content = await page.content()
                            if any(kw in content.lower() for kw in ['job', 'stelle', 'karriere', 'career', 'position', 'bewerb']):
                                career_url = test_url
                                break
                    except:
                        continue
                # Go back to homepage
                try:
                    await page.goto(website_url, timeout=10000, wait_until='domcontentloaded')
                except:
                    pass
            
            result['career_url'] = career_url
            
            # Visit career page and save HTML
            if career_url:
                try:
                    await page.goto(career_url, timeout=15000, wait_until='domcontentloaded')
                    await asyncio.sleep(2)
                    # Scroll to load JS content
                    await page.evaluate('window.scrollTo(0, document.body.scrollHeight/2)')
                    await asyncio.sleep(0.5)
                    await page.evaluate('window.scrollTo(0, document.body.scrollHeight)')
                    await asyncio.sleep(0.5)
                    
                    career_html = await page.content()
                    filename = sanitize_filename(employer_name) + '_career.html'
                    filepath = os.path.join(CAREER_HTML_DIR, filename)
                    with open(filepath, 'w', encoding='utf-8') as f:
                        f.write(career_html)
                    result['career_html_path'] = filepath
                    
                    # Find job links on career page
                    all_links = await page.query_selector_all('a')
                    for link in all_links:
                        text = (await link.inner_text()).strip()
                        href = await link.get_attribute('href')
                        if text and 8 < len(text) < 200 and href:
                            tl = text.lower()
                            is_it = any(kw in tl for kw in ['software','entwickler','developer','it-','data','devops','cloud','system','admin','frontend','backend','fullstack','python','java','sap','linux','sql','consultant','engineer','support','helpdesk','architekt','scrum','project','projekt','test','qa','ux','ui','design','informatik','fachinformatiker'])
                            if is_it:
                                result['job_urls_found'].append({
                                    'title': text[:150],
                                    'url': urljoin(career_url, href)
                                })
                    
                except Exception as e:
                    result['error'] = f"Career page error: {str(e)[:50]}"
            
            print(f"  {'[✓]' if career_url else '[-]'} {employer_name[:40]} | career: {career_url[:50] if career_url else 'none'} | jobs: {len(result['job_urls_found'])}")
            results.append(result)
            return result
            
        except Exception as e:
            result = {
                'employer_name': employer_name,
                'website_url': website_url,
                'career_url': None,
                'career_html_path': None,
                'job_urls_found': [],
                'error': str(e)[:80],
            }
            results.append(result)
            return result
        finally:
            if page:
                try:
                    await page.close()
                except:
                    pass


async def main():
    print("="*60)
    print("Phase 2: Visit employer websites for career pages")
    print("="*60)
    
    # Load jobs data
    with open(JSON_IN, 'r', encoding='utf-8') as f:
        jobs = json.load(f)
    print(f"  Loaded {len(jobs)} jobs from Phase 1")
    
    # Load employer website mapping
    print("\n  Loading employer website database...")
    website_map = load_employer_websites()
    print(f"  Found {len(website_map)} employer websites in database")
    
    # Get unique employers from jobs
    employers = {}
    for job in jobs:
        emp_name = job.get('employer_name', '')
        if not emp_name:
            continue
        norm = re.sub(r'\s+', '', emp_name.lower())[:20]
        if norm not in employers:
            # Try to find website
            match = match_employer(emp_name, website_map)
            employers[norm] = {
                'name': emp_name,
                'website_url': match['url'] if match else None,
                'email': match.get('email', '') if match else '',
                'phone': match.get('phone', '') if match else '',
                'city': match.get('city', '') if match else job.get('city', ''),
                'industry': match.get('industry', '') if match else '',
                'job_count': 0,
            }
        employers[norm]['job_count'] += 1
    
    employers_with_website = {k: v for k, v in employers.items() if v.get('website_url')}
    print(f"  Unique employers: {len(employers)}")
    print(f"  With website: {len(employers_with_website)}")
    print(f"  Without website: {len(employers) - len(employers_with_website)}")
    
    # Visit websites
    print(f"\n  Visiting {len(employers_with_website)} employer websites...")
    
    os.makedirs(CAREER_HTML_DIR, exist_ok=True)
    os.makedirs(JOB_HTML_DIR, exist_ok=True)
    
    results = []
    semaphore = asyncio.Semaphore(MAX_CONCURRENT)
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=['--no-sandbox', '--disable-dev-shm-usage', '--disable-gpu']
        )
        context = await browser.new_context(
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36',
            viewport={'width': 1920, 'height': 1080},
            locale='de-DE',
        )
        
        # Process in batches
        emp_list = list(employers_with_website.values())
        batch_size = 8
        
        for i in range(0, len(emp_list), batch_size):
            batch = emp_list[i:i+batch_size]
            print(f"\n--- Batch {i//batch_size + 1}/{(len(emp_list) + batch_size - 1)//batch_size} ---")
            
            tasks = [visit_employer(context, emp['name'], emp['website_url'], semaphore, results) for emp in batch]
            await asyncio.gather(*tasks, return_exceptions=True)
            
            # Commit every 10 entries
            if len(results) >= COMMIT_EVERY:
                _update_excel_and_commit(jobs, employers, results, employers_with_website)
                results_since = len(results)
        
        await browser.close()
    
    # Final update
    _update_excel_and_commit(jobs, employers, results, employers_with_website, final=True)
    
    # Stats
    career_found = len([r for r in results if r.get('career_url')])
    jobs_found = sum(len(r.get('job_urls_found', [])) for r in results)
    errors = len([r for r in results if r.get('error')])
    
    print(f"\n{'='*60}")
    print(f"Phase 2 DONE!")
    print(f"  Employers visited: {len(results)}")
    print(f"  Career pages found: {career_found}")
    print(f"  Job links found: {jobs_found}")
    print(f"  Errors: {errors}")
    print(f"{'='*60}")


def _update_excel_and_commit(jobs, employers, visit_results, employers_with_website, final=False):
    """Update the Excel with website visit results and commit."""
    # Build lookup from visit results
    career_lookup = {}
    for r in visit_results:
        emp_name = r.get('employer_name', '')
        norm = re.sub(r'\s+', '', emp_name.lower())[:20]
        career_lookup[norm] = r
    
    # Load current Excel and update
    try:
        wb = openpyxl.load_workbook(EXCEL_OUT)
        ws = wb['Junior IT-Jobs']
        
        # Update rows with career URL and HTML path
        updated = 0
        for row in range(2, ws.max_row + 1):
            emp_name = ws.cell(row, 11).value  # Arbeitgeber column
            if not emp_name:
                continue
            norm = re.sub(r'\s+', '', emp_name.lower())[:20]
            if norm in career_lookup:
                result = career_lookup[norm]
                if result.get('career_url'):
                    ws.cell(row, 20, result['career_url'])  # Karriere-URL
                if result.get('career_html_path'):
                    ws.cell(row, 22, result['career_html_path'])  # Karriere-HTML Pfad
                if result.get('website_url'):
                    ws.cell(row, 21, result['website_url'])  # Website-URL
                # Add job URLs from website
                if result.get('job_urls_found'):
                    job_titles = [j['title'] for j in result['job_urls_found'][:5]]
                    current_desc = ws.cell(row, 9).value or ''
                    if not current_desc:
                        ws.cell(row, 9, '; '.join(job_titles)[:500])
                # Add employer info
                emp_data = employers_with_website.get(norm, {})
                if emp_data.get('email') and not ws.cell(row, 17).value:
                    ws.cell(row, 17, emp_data['email'])
                if emp_data.get('phone') and not ws.cell(row, 16).value:
                    ws.cell(row, 16, emp_data['phone'])
                if emp_data.get('industry') and not ws.cell(row, 18).value:
                    ws.cell(row, 18, emp_data['industry'])
                updated += 1
        
        wb.save(EXCEL_OUT)
        print(f"  Excel updated: {updated} rows enriched with website data")
        
    except Exception as e:
        print(f"  Excel update error: {e}")
    
    # Git commit
    if final or len(visit_results) >= COMMIT_EVERY:
        try:
            os.chdir(GIT_REPO)
            subprocess.run(['git', 'add', '-A'], capture_output=True, timeout=30)
            r = subprocess.run(['git', 'commit', '-m', f'Phase 2: Website enrichment ({len(visit_results)} employers visited)'], capture_output=True, timeout=30)
            if r.returncode == 0:
                subprocess.run(['git', 'push', f'https://{GITHUB_TOKEN}@github.com/BirdNest055/Heyjobs.git', 'main', '--force-with-lease'], capture_output=True, timeout=60)
                print(f"  Git: Pushed!")
        except Exception as e:
            print(f"  Git error: {e}")


if __name__ == '__main__':
    asyncio.run(main())
