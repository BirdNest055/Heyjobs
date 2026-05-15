#!/usr/bin/env python3
"""
Job Application Scraper - Visits company websites to find junior IT positions,
extracts contact details, and generates customized application emails.

Output: Bewerbungen_Junior_IT.xlsx with all application-ready data.
Pushes to GitHub every 10 processed companies.
"""

import asyncio
import json
import os
import re
import hashlib
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import urljoin, urlparse

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── CONFIG ────────────────────────────────────────────────────────────────────
EXCEL_OUT = "/home/z/my-project/download/Bewerbungen_Junior_IT.xlsx"
PROGRESS_FILE = "/home/z/my-project/job_scraper_progress.json"
LOG_FILE = "/home/z/my-project/job_scraper_log.txt"
GITHUB_REPO = "/home/z/my-project"
DISPLAY = ":99"
PUSH_EVERY = 10
MAX_COMPANIES = 300  # Safety limit per run; increase as needed
TIMEOUT_PER_COMPANY = 90  # seconds
MAX_JOBS_PER_COMPANY = 5

# ── EMAIL TEMPLATE ────────────────────────────────────────────────────────────
EMAIL_TEMPLATE = """Sehr geehrte{anrede} {ansprechpartner},

Ich sende Ihnen hiermit meine Bewerbungsunterlagen für die Stelle als {stelle} in Vollzeit.

Ich schließe meine Ausbildung zum Fachinformatiker für Systemintegration bald ab. Das Fachgespräch findet am 11. Juni 2026 statt, ab diesem Zeitpunkt stehe ich für eine Einstellung zur Verfügung.

Während meiner Ausbildung hatte ich Berührungspunkte mit verschiedenen IT-Bereichen: Systemintegration, Webentwicklung und Datenbankentwicklung.

In meiner letzten Praxisphase habe ich mich mit Azure beschäftigt, was direkt in mein IHK-Abschlussprojekt geflossen ist: die automatisierte Anwendungsbereitstellung via App Attach in Azure Virtual Desktop, von der Planung bis zur Dokumentation eigenständig umgesetzt.

Ich arbeite selbstständig und kommuniziere offen. Das haben mir auch meine Betreuer in den Praxisphasen zurückgemeldet. Einen Führerschein der Klasse B besitze ich aktuell nicht, bin aber bereit, diesen zeitnah zu erwerben.

Meine vollständigen Unterlagen finden Sie im Anhang. Ich freue mich auf ein persönliches Gespräch.

Mit freundlichen Grüßen,"""


def log(msg):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line, flush=True)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")


def load_progress():
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"processed": [], "jobs": [], "last_index": 0}


def save_progress(progress):
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump(progress, f, ensure_ascii=False, indent=2)


def git_push(msg):
    try:
        subprocess.run(["git", "add", "-A"], cwd=GITHUB_REPO, capture_output=True, timeout=30)
        subprocess.run(["git", "commit", "-m", msg], cwd=GITHUB_REPO, capture_output=True, timeout=30)
        result = subprocess.run(["git", "push", "origin", "main"], cwd=GITHUB_REPO,
                                capture_output=True, text=True, timeout=60)
        if result.returncode == 0:
            log(f"✅ Pushed: {msg}")
        else:
            log(f"⚠️ Push failed: {result.stderr[:200]}")
    except Exception as e:
        log(f"⚠️ Git error: {e}")


def load_companies():
    """Load and merge company data from all sources."""
    companies = {}  # name -> data dict

    # 1. Load from Google Maps (11k+ companies)
    wb1 = openpyxl.load_workbook("/home/z/my-project/download/Google_Maps_Firmen_Erlangen_50km.xlsx")
    ws1 = wb1["Google Maps Firmen"]
    for r in range(2, ws1.max_row + 1):
        name = (ws1.cell(r, 2).value or "").strip()
        if not name:
            continue
        cat = ws1.cell(r, 5).value or ""
        adresse = ws1.cell(r, 6).value or ""
        plz = ws1.cell(r, 7).value or ""
        ort = ws1.cell(r, 8).value or ""
        telefon = ws1.cell(r, 9).value or ""
        website = ws1.cell(r, 10).value or ""

        if name not in companies:
            companies[name] = {
                "name": name, "kategorie": cat, "adresse": adresse,
                "plz": plz, "ort": ort, "telefon": telefon, "website": website,
                "email": "", "ansprechpartner": "", "quelle": "google_maps"
            }
        elif website and not companies[name]["website"]:
            companies[name]["website"] = website

    # 2. Enrich with website/HTML data
    wb2 = openpyxl.load_workbook("/home/z/my-project/download/Firmen_Bamberg_Erlangen_Nuernberg_mit_Website_HTML.xlsx")
    ws2 = wb2["Firmen mit Website & HTML"]
    for r in range(2, ws2.max_row + 1):
        name = (ws2.cell(r, 2).value or "").strip()
        if not name:
            continue
        web_found = ws2.cell(r, 10).value or ""
        email = ws2.cell(r, 11).value or ""
        html_file = ws2.cell(r, 12).value or ""

        if name in companies:
            if web_found and not companies[name]["website"]:
                companies[name]["website"] = web_found
            if email and not companies[name]["email"]:
                companies[name]["email"] = email
            companies[name]["html_file"] = html_file
        else:
            companies[name] = {
                "name": name, "kategorie": ws2.cell(r, 4).value or "",
                "adresse": ws2.cell(r, 5).value or "",
                "plz": ws2.cell(r, 7).value or "",
                "ort": ws2.cell(r, 8).value or "",
                "telefon": ws2.cell(r, 9).value or "",
                "website": web_found, "email": email,
                "html_file": html_file, "ansprechpartner": "",
                "quelle": "enriched"
            }

    # 3. Load from Arbeitsagentur jobs for extra employer info
    wb3 = openpyxl.load_workbook("/home/z/my-project/download/Junior_IT_Jobs_Bamberg_Erlangen_Nuernberg.xlsx")
    ws3 = wb3["Junior IT-Jobs"]
    arbeitsagentur_jobs = {}
    for r in range(2, ws3.max_row + 1):
        employer = (ws3.cell(r, 22).value or "").strip()  # Arbeitgeber
        job_title = ws3.cell(r, 2).value or ""
        job_url = ws3.cell(r, 27).value or ""
        job_ort = ws3.cell(r, 18).value or ""
        if employer:
            if employer not in arbeitsagentur_jobs:
                arbeitsagentur_jobs[employer] = []
            arbeitsagentur_jobs[employer].append({
                "title": job_title, "url": job_url, "ort": job_ort
            })

    # Filter to companies with websites that are IT-related OR all with websites
    # Priority: IT companies with websites first
    it_keywords = [
        'it ', 'it-', 'software', 'data ', 'daten', 'technolog', 'digital',
        'systemhaus', 'informatik', 'cyber', 'cloud', 'tech', 'computer',
        'netzwerk', 'consulting', 'entwicklung', 'systemintegrat',
        'admin', 'devops', 'sre', 'platform', 'engineer',
        'sap', 'oracle', 'microsoft', 'linux', 'server',
        'service', 'lösung', 'solution', 'automatisier',
        'maschinenbau', 'industrie', 'elektro', 'energie',
        'bank', 'versicher', 'finanz', 'klinik', 'university',
        'universität', 'forschung', 'research', 'gruppe', 'holding'
    ]

    # Score companies for IT relevance
    scored = []
    for name, data in companies.items():
        if not data["website"]:
            continue
        score = 0
        text = f"{name} {data['kategorie']}".lower()
        for kw in it_keywords:
            if kw in text:
                score += 1
        # Also check if they appear in Arbeitsagentur IT jobs
        if name in arbeitsagentur_jobs:
            score += 5  # High priority - already confirmed they hire IT
        scored.append((score, name, data))

    # Sort by score descending
    scored.sort(key=lambda x: -x[0])

    result = []
    for score, name, data in scored:
        data["relevance_score"] = score
        data["arbeitsagentur_jobs"] = arbeitsagentur_jobs.get(name, [])
        result.append(data)

    log(f"Loaded {len(result)} companies with websites (sorted by IT relevance)")
    log(f"  - {sum(1 for s, _, _ in scored if s >= 5)} confirmed IT employers (from Arbeitsagentur)")
    log(f"  - {sum(1 for s, _, _ in scored if 1 <= s < 5)} likely IT employers")
    log(f"  - {sum(1 for s, _, _ in scored if s == 0)} other employers with websites")

    return result


def customize_email(stelle, firma, ansprechpartner="", email=""):
    """Generate customized email text from template."""
    # Determine salutation
    if ansprechpartner and ansprechpartner.lower() not in ["n/a", "unbekannt", ""]:
        # Try to determine gender from name (best effort)
        parts = ansprechpartner.strip().split()
        first_name = parts[0] if parts else ""
        # Common female first names in German
        female_names = ['anna', 'maria', 'sabine', 'monika', 'katrin', 'sandra',
                        'nicole', 'stephanie', 'julia', 'laura', 'sarah', 'lena',
                        'christine', 'barbara', 'heike', 'petra', 'susanne', 'angela',
                        'michaela', 'silvia', 'daniela', 'elke', 'birgit', 'ute']
        if first_name.lower() in female_names:
            anrede = "Frau"
        elif any(t in ansprechpartner.lower() for t in ['frau', 'ms.', 'mrs.']):
            anrede = "Frau"
        elif any(t in ansprechpartner.lower() for t in ['herr', 'hr.', 'mr.']):
            anrede = "Herr"
        else:
            anrede = "Herr/Frau"
        anrede_line = f" {anrede} {ansprechpartner}"
    else:
        anrede_line = ""

    text = EMAIL_TEMPLATE.format(
        anrede="" if anrede_line and anrede != "Herr/Frau" else "r",
        ansprechpartner=ansprechpartner if ansprechpartner else "Damen und Herren",
        stelle=stelle
    )

    # Fix the opening line based on ansprechpartner
    if ansprechpartner and ansprechpartner.lower() not in ["n/a", "unbekannt", ""]:
        # Determine gender for proper grammar
        if anrede_line:
            opening = f"Sehr geehrte{anrede_line},"
        else:
            opening = f"Sehr geehrte/r {ansprechpartner},"
    else:
        opening = "Sehr geehrte Damen und Herren,"

    # Replace the first line
    lines = text.split("\n")
    lines[0] = opening
    return "\n".join(lines)


async def scrape_company_jobs(page, company):
    """Visit a company's website and extract junior IT job listings."""
    url = company["website"]
    if not url:
        return []

    # Ensure URL has scheme
    if not url.startswith(("http://", "https://")):
        url = "https://" + url

    jobs_found = []

    # Common career page URLs to try
    career_paths = [
        "/karriere", "/jobs", "/stellenangebote", "/kariere",
        "/career", "/careers", "/stellen", "/offene-stellen",
        "/bewerbung", "/job", "/vacancies", "/vacancy",
        "/working-with-us", "/wir-haben-platz", "/team",
        "/de/karriere", "/de/jobs", "/de/career",
        "/en/career", "/en/jobs",
    ]

    try:
        # Step 1: Visit main page
        log(f"  🌐 Visiting {url}")
        try:
            resp = await page.goto(url, wait_until="domcontentloaded", timeout=20000)
            if not resp or resp.status >= 400:
                log(f"  ⚠️ HTTP {resp.status if resp else 'N/A'} for {url}")
                return []
        except Exception as e:
            log(f"  ⚠️ Failed to load {url}: {str(e)[:100]}")
            return []

        await page.wait_for_timeout(1500)

        # Step 2: Look for career/jobs links on the main page
        career_links = set()
        try:
            # Search for links containing career/job-related text
            link_elements = await page.query_selector_all("a[href]")
            for el in link_elements:
                try:
                    href = await el.get_attribute("href") or ""
                    text = await el.inner_text()
                    text_lower = (text or "").lower().strip()
                    href_lower = href.lower()

                    # Check if link text or href suggests career page
                    career_keywords = ['karriere', 'jobs', 'stellen', 'career', 'bewer',
                                       'offene', 'vacanc', 'position', 'stellenausschreib']
                    is_career_link = any(kw in text_lower or kw in href_lower for kw in career_keywords)

                    if is_career_link and href:
                        full_url = urljoin(url, href)
                        if urlparse(full_url).netloc == urlparse(url).netloc or \
                           urlparse(full_url).netloc.replace("www.", "") == urlparse(url).netloc.replace("www.", ""):
                            career_links.add(full_url)
                except:
                    continue
        except:
            pass

        # Also try standard career paths
        base_url = f"{urlparse(url).scheme}://{urlparse(url).netloc}"
        for path in career_paths:
            career_links.add(base_url + path)

        log(f"  🔗 Found {len(career_links)} potential career links")

        # Step 3: Visit career pages to find IT jobs
        visited = set()
        for career_url in list(career_links)[:5]:  # Limit to 5 career pages
            if career_url in visited:
                continue
            visited.add(career_url)

            try:
                resp = await page.goto(career_url, wait_until="domcontentloaded", timeout=15000)
                if not resp or resp.status >= 400:
                    continue
                await page.wait_for_timeout(1500)
            except:
                continue

            # Extract page content
            try:
                content = await page.content()
                page_text = await page.inner_text("body")
            except:
                continue

            # Look for IT-related job listings
            it_keywords = ['it-', 'it ', 'informatik', 'software', 'systemintegration',
                           'systemadministr', 'netzwerk', 'devops', 'cloud', 'data ',
                           'daten', 'cyber', 'security', 'entwickler', 'developer',
                           'admin', 'linux', 'sap', 'azure', 'aws', 'frontend',
                           'backend', 'fullstack', 'full-stack', 'sre', 'platform',
                           'infrastruktur', 'infrastructure', 'support', 'helpdesk']

            junior_keywords = ['junior', 'trainee', 'eingang', 'entry', 'graduate',
                               'start', 'berufseinsteig', 'jun.', 'young',
                               'praktikant', 'werkstudent', 'azubi', 'ausbildung']

            # Find job listing elements
            job_elements = await page.query_selector_all("a, div, li, article, section, tr")
            for el in job_elements:
                try:
                    text = await el.inner_text()
                    if not text or len(text) > 1000:
                        continue
                    text_lower = text.lower()

                    # Check if this looks like an IT job
                    has_it = any(kw in text_lower for kw in it_keywords)
                    if not has_it:
                        continue

                    # Check if it's junior-level or open to juniors
                    has_junior = any(kw in text_lower for kw in junior_keywords)
                    # Also accept if no level specified (might be open)
                    is_open_level = not any(kw in text_lower for kw in ['senior', 'lead', 'manager', 'director', 'head of'])

                    if has_junior or is_open_level:
                        # Try to extract job title
                        title = text.strip().split("\n")[0].strip()[:200]

                        # Try to get link
                        job_url = ""
                        tag = await el.evaluate("el => el.tagName").catch(lambda: "") if hasattr(el, 'evaluate') else ""
                        try:
                            href = await el.get_attribute("href")
                            if href:
                                job_url = urljoin(career_url, href)
                        except:
                            pass

                        # Try to find parent link
                        if not job_url:
                            try:
                                parent_link = await el.evaluate_handle(
                                    "el => el.closest('a')?.href || ''"
                                )
                                job_url = str(await parent_link.json_value()) if parent_link else ""
                                if job_url and not job_url.startswith("http"):
                                    job_url = urljoin(career_url, job_url)
                            except:
                                pass

                        if title and len(title) > 5:
                            jobs_found.append({
                                "title": title,
                                "url": job_url,
                                "description": text[:500],
                                "is_junior": has_junior
                            })
                except:
                    continue

            # If we found jobs, try to extract more detail from first few
            if jobs_found:
                break  # Found jobs on this career page, no need to check more

        # Step 4: Also look for Impressum to find contact email
        email = company.get("email", "")
        ansprechpartner = ""

        if not email:
            impressum_paths = ["/impressum", "/imprint", "/kontakt", "/contact", "/datenschutz"]
            for imp_path in impressum_paths[:2]:
                try:
                    imp_url = base_url + imp_path
                    resp = await page.goto(imp_url, wait_until="domcontentloaded", timeout=10000)
                    if resp and resp.status < 400:
                        await page.wait_for_timeout(1000)
                        imp_text = await page.inner_text("body")
                        # Find email
                        email_match = re.search(r'[\w.+-]+@[\w-]+\.[\w.-]+', imp_text)
                        if email_match:
                            found_email = email_match.group(0)
                            # Prefer non-generic emails
                            if not any(g in found_email.lower() for g in ['info@', 'webmaster', 'noreply', 'spam']):
                                email = found_email
                            elif not email:
                                email = found_email
                        # Find Ansprechpartner
                        lines = imp_text.split("\n")
                        for line in lines:
                            line = line.strip()
                            if any(t in line.lower() for t in ['herr', 'frau', 'hr.', 'fr.']):
                                if any(t in line.lower() for t in ['kontakt', 'ansprech', 'bewer', 'personal', 'hr']):
                                    ansprechpartner = line[:80]
                                    break
                        if email and ansprechpartner:
                            break
                except:
                    continue

        # Deduplicate and limit jobs
        seen_titles = set()
        unique_jobs = []
        for job in jobs_found:
            title_key = job["title"].lower().strip()[:50]
            if title_key not in seen_titles:
                seen_titles.add(title_key)
                job["email"] = email
                job["ansprechpartner"] = ansprechpartner
                unique_jobs.append(job)
            if len(unique_jobs) >= MAX_JOBS_PER_COMPANY:
                break

        return unique_jobs

    except Exception as e:
        log(f"  ❌ Error scraping {company['name']}: {str(e)[:100]}")
        return []


async def main():
    log("=" * 60)
    log("Job Application Scraper - Starting")
    log("=" * 60)

    # Load data
    companies = load_companies()
    progress = load_progress()

    processed_names = set(progress["processed"])
    jobs_all = progress["jobs"]
    start_idx = progress["last_index"]

    log(f"Already processed: {len(processed_names)} companies")
    log(f"Jobs found so far: {len(jobs_all)}")
    log(f"Starting from index: {start_idx}")

    # Filter to unprocessed companies
    to_process = [(i, c) for i, c in enumerate(companies)
                  if c["name"] not in processed_names and c["website"]]
    log(f"Companies to process: {len(to_process)}")

    if not to_process:
        log("Nothing to process!")
        return

    # Limit per run
    to_process = to_process[:MAX_COMPANIES]
    log(f"Processing up to {MAX_COMPANIES} this run")

    # Launch browser
    from playwright.async_api import async_playwright

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=False,
            args=[
                "--no-sandbox",
                "--disable-dev-shm-usage",
                "--window-size=1920,1080",
                "--disable-blink-features=AutomationControlled",
            ]
        )

        context = await browser.new_context(
            viewport={"width": 1920, "height": 1080},
            user_agent="Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
            locale="de-DE",
        )

        page = await context.new_page()

        # Set reasonable timeouts
        page.set_default_timeout(20000)
        page.set_default_navigation_timeout(20000)

        processed_count = 0
        total_new_jobs = 0

        for idx, (orig_idx, company) in enumerate(to_process):
            company_name = company["name"]
            log(f"\n[{idx+1}/{len(to_process)}] {company_name} (score: {company.get('relevance_score', 0)})")

            try:
                jobs = await asyncio.wait_for(
                    scrape_company_jobs(page, company),
                    timeout=TIMEOUT_PER_COMPANY
                )
            except asyncio.TimeoutError:
                log(f"  ⏰ Timeout for {company_name}")
                jobs = []
            except Exception as e:
                log(f"  ❌ Error: {str(e)[:100]}")
                jobs = []

            if jobs:
                log(f"  ✅ Found {len(jobs)} job(s)")
                for job in jobs:
                    # Customize email for this job
                    email_text = customize_email(
                        stelle=job["title"],
                        firma=company_name,
                        ansprechpartner=job.get("ansprechpartner", ""),
                        email=job.get("email", "")
                    )

                    adresse = f"{company.get('adresse', '')}"
                    if company.get('plz'):
                        adresse += f", {company['plz']}"
                    if company.get('ort'):
                        adresse += f" {company['ort']}"

                    job_entry = {
                        "firma": company_name,
                        "stelle": job["title"],
                        "stellenbeschreibung": job.get("description", "")[:500],
                        "email": job.get("email", company.get("email", "")),
                        "ansprechpartner": job.get("ansprechpartner", ""),
                        "bewerbung_geschickt": "",
                        "letzter_kontakt": "",
                        "status": "Offen",
                        "website": company["website"],
                        "adresse": adresse.strip(", "),
                        "email_text": email_text,
                        "job_url": job.get("url", ""),
                        "is_junior": job.get("is_junior", False),
                        "scraped_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    }
                    jobs_all.append(job_entry)
                    total_new_jobs += 1
            else:
                log(f"  ℹ️ No IT jobs found")
                # Still mark as processed

            processed_names.add(company_name)
            processed_count += 1

            # Save progress
            progress["processed"] = list(processed_names)
            progress["jobs"] = jobs_all
            progress["last_index"] = orig_idx
            save_progress(progress)

            # Push every 10 companies
            if processed_count % PUSH_EVERY == 0:
                log(f"\n📦 Progress checkpoint: {processed_count} companies, {len(jobs_all)} jobs")
                # Save intermediate Excel
                save_excel(jobs_all)
                git_push(f"Job Scraper: {processed_count} companies processed | {len(jobs_all)} jobs found")

            # Small delay between companies
            await page.wait_for_timeout(1000)

        await browser.close()

    # Final save
    log(f"\n{'=' * 60}")
    log(f"Done! Processed {processed_count} companies, found {total_new_jobs} new jobs")
    log(f"Total jobs in database: {len(jobs_all)}")

    save_excel(jobs_all)
    git_push(f"Job Scraper COMPLETE: {processed_count} companies | {len(jobs_all)} jobs | {datetime.now().strftime('%Y-%m-%d')}")

    progress["last_run"] = datetime.now().isoformat()
    save_progress(progress)


def save_excel(jobs):
    """Generate the application-ready Excel file."""
    wb = openpyxl.Workbook()

    # ── Main Sheet: Bewerbungen ──
    ws = wb.active
    ws.title = "Bewerbungen"

    headers = [
        "Firma", "Stelle", "Stellenbeschreibung", "Email", "Ansprechpartner",
        "Bewerbung geschickt am", "Letzter Kontakt am", "Status", "Website",
        "Adresse", "Email Text", "Job-URL", "Junior?", "Scraped At"
    ]

    # Header styling
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    for row_idx, job in enumerate(jobs, 2):
        values = [
            job.get("firma", ""),
            job.get("stelle", ""),
            job.get("stellenbeschreibung", ""),
            job.get("email", ""),
            job.get("ansprechpartner", ""),
            job.get("bewerbung_geschickt", ""),
            job.get("letzter_kontakt", ""),
            job.get("status", "Offen"),
            job.get("website", ""),
            job.get("adresse", ""),
            job.get("email_text", ""),
            job.get("job_url", ""),
            "Ja" if job.get("is_junior") else "Nein",
            job.get("scraped_at", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Column widths
    col_widths = {
        1: 30,   # Firma
        2: 35,   # Stelle
        3: 45,   # Stellenbeschreibung
        4: 30,   # Email
        5: 22,   # Ansprechpartner
        6: 18,   # Bewerbung geschickt
        7: 18,   # Letzter Kontakt
        8: 12,   # Status
        9: 35,   # Website
        10: 35,  # Adresse
        11: 60,  # Email Text
        12: 35,  # Job-URL
        13: 10,  # Junior?
        14: 16,  # Scraped At
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(jobs) + 1}"

    # Data validation for Status column
    status_dv = DataValidation(
        type="list",
        formula1='"Offen,Beworben,Interview eingeplant,Interview absolviert,Zusage,Absage,Kontaktiert,Wiedervorlage,Kein Interesse"',
        allow_blank=True
    )
    status_dv.error = "Bitte wählen Sie einen gültigen Status"
    status_dv.errorTitle = "Ungültiger Status"
    ws.add_data_validation(status_dv)
    for row in range(2, len(jobs) + 2):
        status_dv.add(ws.cell(row=row, column=8))

    # Conditional formatting for Status
    from openpyxl.formatting.rule import CellIsRule

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    ws.conditional_formatting.add(f"H2:H{len(jobs)+1}",
        CellIsRule(operator="equal", formula=['"Zusage"'], fill=green_fill))
    ws.conditional_formatting.add(f"H2:H{len(jobs)+1}",
        CellIsRule(operator="equal", formula=['"Absage"'], fill=red_fill))
    ws.conditional_formatting.add(f"H2:H{len(jobs)+1}",
        CellIsRule(operator="equal", formula=['"Beworben"'], fill=blue_fill))
    ws.conditional_formatting.add(f"H2:H{len(jobs)+1}",
        CellIsRule(operator="equal", formula=['"Interview eingeplant"'], fill=yellow_fill))
    ws.conditional_formatting.add(f"H2:H{len(jobs)+1}",
        CellIsRule(operator="equal", formula=['"Interview absolviert"'], fill=yellow_fill))

    # Junior? column conditional formatting
    ws.conditional_formatting.add(f"M2:M{len(jobs)+1}",
        CellIsRule(operator="equal", formula=['"Ja"'], fill=green_fill))

    # ── Statistics Sheet ──
    ws_stats = wb.create_sheet("Statistiken")

    stats_data = [
        ("Bewerbungen Junior IT - Statistiken", ""),
        ("", ""),
        ("Gesamt Jobs gefunden", len(jobs)),
        ("Junior-Jobs", sum(1 for j in jobs if j.get("is_junior"))),
        ("Mit Email", sum(1 for j in jobs if j.get("email"))),
        ("Mit Ansprechpartner", sum(1 for j in jobs if j.get("ansprechpartner"))),
        ("", ""),
        ("Nach Status:", ""),
    ]

    # Count by status
    status_counts = {}
    for j in jobs:
        s = j.get("status", "Offen")
        status_counts[s] = status_counts.get(s, 0) + 1
    for status, count in sorted(status_counts.items()):
        stats_data.append((f"  {status}", count))

    stats_data.extend([
        ("", ""),
        ("Nach Ort:", ""),
    ])

    # Count by ort
    ort_counts = {}
    for j in jobs:
        addr = j.get("adresse", "")
        for city in ["Bamberg", "Erlangen", "Nürnberg", "Fürth", "Würzburg"]:
            if city.lower() in addr.lower():
                ort_counts[city] = ort_counts.get(city, 0) + 1
                break
        else:
            ort_counts["Sonstige"] = ort_counts.get("Sonstige", 0) + 1
    for ort, count in sorted(ort_counts.items(), key=lambda x: -x[1]):
        stats_data.append((f"  {ort}", count))

    for row_idx, (label, value) in enumerate(stats_data, 1):
        ws_stats.cell(row=row_idx, column=1, value=label).font = Font(
            bold=(row_idx <= 1 or "Nach" in str(label)))
        ws_stats.cell(row=row_idx, column=2, value=value)

    ws_stats.column_dimensions["A"].width = 35
    ws_stats.column_dimensions["B"].width = 15

    # ── Email Templates Sheet (for quick copy) ──
    ws_emails = wb.create_sheet("Email Vorlagen")
    ws_emails.cell(1, 1, value="Firma").font = Font(bold=True)
    ws_emails.cell(1, 2, value="Stelle").font = Font(bold=True)
    ws_emails.cell(1, 3, value="Email an").font = Font(bold=True)
    ws_emails.cell(1, 4, value="Betreff").font = Font(bold=True)
    ws_emails.cell(1, 5, value="Email Text (fertig zum Kopieren)").font = Font(bold=True)

    for row_idx, job in enumerate(jobs, 2):
        ws_emails.cell(row=row_idx, column=1, value=job.get("firma", ""))
        ws_emails.cell(row=row_idx, column=2, value=job.get("stelle", ""))
        ws_emails.cell(row=row_idx, column=3, value=job.get("email", ""))
        betreff = f"Bewerbung als {job.get('stelle', 'Fachinformatiker Systemintegration')} - {job.get('firma', '')}"
        ws_emails.cell(row=row_idx, column=4, value=betreff)
        ws_emails.cell(row=row_idx, column=5, value=job.get("email_text", ""))

    ws_emails.column_dimensions["A"].width = 30
    ws_emails.column_dimensions["B"].width = 35
    ws_emails.column_dimensions["C"].width = 30
    ws_emails.column_dimensions["D"].width = 50
    ws_emails.column_dimensions["E"].width = 80

    wb.save(EXCEL_OUT)
    log(f"📊 Excel saved: {EXCEL_OUT} ({len(jobs)} jobs)")


if __name__ == "__main__":
    os.environ["DISPLAY"] = DISPLAY
    asyncio.run(main())
