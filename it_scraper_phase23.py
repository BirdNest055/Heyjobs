#!/usr/bin/env python3
"""
Phase 2+3: Match Google Maps employers with IT jobs, enrich with websites & HTML
"""
import json, os, re, sys, time, hashlib, sqlite3, subprocess
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
import requests
from bs4 import BeautifulSoup
import openpyxl

sys.stdout.reconfigure(line_buffering=True)

EXCEL_IN = "/home/z/my-project/download/Google_Maps_Firmen_Erlangen_50km.xlsx"
EXCEL_OUT = "/home/z/my-project/download/IT_Arbeitgeber_Bamberg_Erlangen_Nuernberg.xlsx"
HTML_DIR = "/home/z/my-project/download/it_employer_html"
DB_PATH = "/home/z/my-project/download/it_employers.db"
IT_JOBS_MAP = "/home/z/my-project/download/it_jobs_map.json"
PROGRESS_FILE = "/home/z/my-project/download/it_scraper_progress.json"
GIT_DIR = "/home/z/my-project"

TARGET_CITIES = {"Bamberg", "Erlangen", "Nürnberg"}
MAX_WORKERS = 10
COMMIT_EVERY = 10

HTTP_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "de-DE,de;q=0.9,en;q=0.5",
}

SESSION = requests.Session()
SESSION.headers.update(HTTP_HEADERS)
SESSION.verify = False

import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

def sanitize_filename(name, max_len=80):
    name = re.sub(r'[äöüßÄÖÜ]', lambda m: {'ä':'ae','ö':'oe','ü':'ue','ß':'ss','Ä':'Ae','Ö':'Oe','Ü':'Ue'}[m.group()], name)
    name = re.sub(r'[^a-zA-Z0-9._-]', '_', name)
    name = re.sub(r'_+', '_', name).strip('_')
    return name[:max_len]

def normalize_name(name):
    name = name.lower().strip()
    name = re.sub(r'\s+', ' ', name)
    for suffix in ['gmbh & co. kg','gmbh & co kg','gmbh co. kg','gmbh co kg','gmbh & co.','ag & co. kg','ag & co kg','gmbh','ag','kg','ohg','eg','se','e.k.','ek','e.v.','ev','mbh','ug','gbr','& co. ohg','co. ohg']:
        name = re.sub(re.escape(suffix), '', name, flags=re.IGNORECASE)
    name = re.sub(r'[^a-z0-9äöüß]', '', name)
    return name

def make_absolute_url(base_url, relative_url):
    if not relative_url:
        return None
    if relative_url.startswith('http'):
        return relative_url
    if relative_url.startswith('//'):
        return 'https:' + relative_url
    if relative_url.startswith('/'):
        from urllib.parse import urlparse
        parsed = urlparse(base_url)
        return f"{parsed.scheme}://{parsed.netloc}{relative_url}"
    return f"{base_url.rstrip('/')}/{relative_url.lstrip('/')}"

def find_website(name, city):
    """Try to find website by constructing likely URLs."""
    clean = re.sub(r'(GmbH|AG|KG|OHG|UG|e\.K\.|e\.V\.|mbH|Co\.\s*KG|& Co\.)', '', name, flags=re.IGNORECASE).strip()
    domain = re.sub(r'[äöüß]', lambda m: {'ä':'ae','ö':'oe','ü':'ue','ß':'ss'}[m.group()], clean)
    domain = re.sub(r'[^a-zA-Z0-9]', '', domain).lower()
    
    if len(domain) < 3:
        return None
    
    candidates = [
        f"https://www.{domain}.de",
        f"https://{domain}.de",
        f"https://www.{domain}.com",
    ]
    
    for url in candidates:
        try:
            resp = SESSION.get(url, timeout=8, allow_redirects=True)
            if resp.status_code == 200 and 'text/html' in resp.headers.get('content-type', ''):
                return url
        except:
            pass
    return None

def download_html(url):
    try:
        resp = SESSION.get(url, timeout=12, allow_redirects=True)
        if resp.status_code == 200:
            ct = resp.headers.get('content-type', '')
            if 'text/' in ct:
                return resp.text
    except:
        pass
    return None

def save_html_file(content, company_name, suffix=""):
    if not content:
        return None, None
    safe = sanitize_filename(company_name)
    filename = f"{safe}_{suffix}.html" if suffix else f"{safe}.html"
    filepath = os.path.join(HTML_DIR, filename)
    os.makedirs(HTML_DIR, exist_ok=True)
    with open(filepath, 'w', encoding='utf-8', errors='replace') as f:
        f.write(content)
    html_hash = hashlib.sha256(content.encode()).hexdigest()[:16]
    return filepath, html_hash

def extract_contact_info(html):
    info = {}
    imp_url = None
    kar_url = None
    if not html:
        return info, None, None
    
    emails = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', html)
    emails = [e for e in emails if not any(x in e.lower() for x in ['example.com','sentry','webpack','noreply','wixpress','github.com','googleapis'])]
    mailtos = re.findall(r'mailto:([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})', html)
    emails = list(dict.fromkeys(emails + mailtos))[:5]
    if emails:
        info['email'] = emails[0]
    
    gf = re.search(r'(?:Geschäftsführer|Geschaeftsfuehrer)[\s:]*([^<\.]{2,100})', html, re.IGNORECASE)
    if gf:
        t = BeautifulSoup(gf.group(1), 'html.parser').get_text(strip=True)
        if len(t) > 2: info['geschaeftsfuehrer'] = t[:200]
    
    hr = re.search(r'(?:Handelsregister|HRB)[\s:]*([^<\.]{2,100})', html, re.IGNORECASE)
    if hr:
        t = BeautifulSoup(hr.group(1), 'html.parser').get_text(strip=True)
        info['handelsregister'] = t[:200]
    
    soup = BeautifulSoup(html, 'html.parser')
    for a in soup.find_all('a', href=True):
        h = a['href'].lower()
        t = a.get_text(strip=True).lower()
        if not imp_url and ('impressum' in h or 'impressum' in t or 'imprint' in h):
            imp_url = a['href']
        if not kar_url and any(w in h for w in ['karriere','career','jobs','stellenangebote','bewerbung']):
            kar_url = a['href']
    
    return info, imp_url, kar_url

def check_website_for_it(html):
    if not html:
        return False, []
    text = html.lower()
    it_kws = [kw for kw in ['software','entwickler','developer','informatik','it-','data','cloud','devops','administrator','sap','linux','scrum','docker','frontend','backend'] if kw in text]
    career_kws = [kw for kw in ['karriere','career','stellenangebote','jobs','bewerbung'] if kw in text]
    has_it = bool(it_kws) and bool(career_kws)
    specific = ['entwickler m/w/d','developer m/w/d','it-jobs','softwareentwickler gesucht','devops engineer']
    if any(s in text for s in specific):
        has_it = True
    return has_it, it_kws[:10]

# ============================================================
# MAIN PHASE 2+3
# ============================================================

def main():
    print("=== PHASE 2+3: Match & Enrich ===")
    
    # Load IT jobs map
    with open(IT_JOBS_MAP) as f:
        it_jobs_data = json.load(f)
    print(f"IT-Jobs geladen: {len(it_jobs_data)} Arbeitgeber")
    
    # Load Google Maps employers
    wb = openpyxl.load_workbook(EXCEL_IN)
    ws = wb['Google Maps Firmen']
    employers = []
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        ort = str(row[6] or '').strip()
        name = str(row[1] or '').strip()
        if not name or ort not in TARGET_CITIES:
            continue
        norm = normalize_name(name)
        if norm in seen or len(norm) < 3:
            continue
        seen.add(norm)
        employers.append({
            'name': name, 'ort': ort,
            'plz': str(row[5] or '').strip(),
            'adresse': str(row[4] or '').strip(),
            'telefon': str(row[7] or '').strip(),
            'kategorie': str(row[3] or '').strip(),
            'bewertung': row[2],
            'lat': row[9], 'lng': row[10],
            'gmaps_link': str(row[11] or '').strip(),
        })
    print(f"Google Maps: {len(employers)} Arbeitgeber")
    
    # Match with IT jobs
    matched = []
    for emp in employers:
        norm = normalize_name(emp['name'])
        it_data = None
        score = 0.0
        
        # Direct match
        if norm in it_jobs_data:
            it_data = it_jobs_data[norm]
            score = 50
        else:
            # Fuzzy
            for it_norm, it_d in it_jobs_data.items():
                if len(norm) > 4 and len(it_norm) > 4:
                    if norm[:6] in it_norm or it_norm[:6] in norm:
                        it_data = it_d
                        score = 40
                        break
        
        # Name/category IT relevance
        name_lower = emp['name'].lower()
        kat_lower = (emp.get('kategorie') or '').lower()
        it_name = ['it ','it-','software','tech','digital','data','cloud','cyber','computer','informatik','system','consulting','automation','iot','embedded','telekommunikation']
        it_kat = ['software','it','informatik','computer','technologie','digital','internet','telekommunikation']
        if any(w in name_lower for w in it_name): score += 20
        if any(w in kat_lower for w in it_kat): score += 15
        
        matched.append({
            'employer': emp,
            'it_data': it_data,
            'it_score': min(score, 100),
        })
    
    it_matched = sum(1 for m in matched if m['it_data'])
    print(f"Gematcht: {it_matched} mit IT-Jobs, {len(matched)-it_matched} ohne")
    
    # Prioritize: IT-matched first, then by score
    matched.sort(key=lambda x: (-x['it_score'], x['employer']['name']))
    
    # Load progress
    progress_file = PROGRESS_FILE
    if os.path.exists(progress_file):
        with open(progress_file) as f:
            progress = json.load(f)
    else:
        progress = {"completed": []}
    
    completed = set(progress['completed'])
    remaining = [m for m in matched if m['employer']['name'] not in completed]
    print(f"Verbleibend: {len(remaining)} (bereits {len(completed)} erledigt)")
    
    # Init DB
    conn = sqlite3.connect(DB_PATH)
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS employers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            firmenname TEXT NOT NULL,
            firmenname_normalisiert TEXT,
            ort TEXT, plz TEXT, strasse TEXT, telefon TEXT, email TEXT,
            website_url TEXT, website_html_path TEXT, website_html_hash TEXT,
            karriere_url TEXT, karriere_html_path TEXT,
            impressum_url TEXT, impressum_html_path TEXT,
            hat_it_jobs INTEGER DEFAULT 0, it_job_titel TEXT,
            it_job_anzahl INTEGER DEFAULT 0, it_job_quelle TEXT,
            it_job_details TEXT, it_keywords_gefunden TEXT,
            kategorie_google TEXT, bewertung_google REAL,
            lat REAL, lng REAL, google_maps_link TEXT,
            handelsregister TEXT, geschaeftsfuehrer TEXT,
            it_abteilung_vermutung INTEGER DEFAULT 0,
            it_relevanz_score REAL DEFAULT 0.0,
            scrape_datum TEXT, scrape_status TEXT, letzte_pruefung TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_ort ON employers(ort);
        CREATE INDEX IF NOT EXISTS idx_hat_it ON employers(hat_it_jobs);
        CREATE INDEX IF NOT EXISTS idx_score ON employers(it_relevanz_score);
    """)
    conn.commit()
    os.makedirs(HTML_DIR, exist_ok=True)
    
    def process_one(m):
        emp = m['employer']
        it_data = m['it_data']
        
        result = {
            'firmenname': emp['name'], 'ort': emp['ort'], 'plz': emp['plz'],
            'strasse': emp['adresse'], 'telefon': emp.get('telefon', ''),
            'email': '', 'website_url': '', 'website_html_path': '',
            'website_html_hash': '', 'karriere_url': '', 'karriere_html_path': '',
            'impressum_url': '', 'impressum_html_path': '',
            'hat_it_jobs': 1 if it_data else 0, 'it_job_titel': '',
            'it_job_anzahl': 0, 'it_job_quelle': '', 'it_job_details': '',
            'it_keywords_gefunden': '',
            'kategorie_google': emp.get('kategorie', ''),
            'bewertung_google': emp.get('bewertung'),
            'lat': emp.get('lat'), 'lng': emp.get('lng'),
            'google_maps_link': emp.get('gmaps_link', ''),
            'handelsregister': '', 'geschaeftsfuehrer': '',
            'it_abteilung_vermutung': 1 if m['it_score'] >= 30 else 0,
            'it_relevanz_score': m['it_score'],
        }
        
        # IT jobs
        if it_data:
            jobs = it_data.get('jobs', [])
            result['it_job_anzahl'] = len(jobs)
            result['it_job_titel'] = "; ".join(j.get('titel','') for j in jobs[:10])
            result['it_job_quelle'] = 'Arbeitsagentur'
            result['it_job_details'] = json.dumps([{
                'titel': j.get('titel',''), 'beruf': j.get('hauptberuf',''),
                'refnr': j.get('referenznummer',''), 'ort': j.get('ort',''),
                'vollzeit': j.get('vollzeit', False),
                'unbefristet': j.get('unbefristet', False),
            } for j in jobs[:20]], ensure_ascii=False)
        
        # Find website
        website = find_website(emp['name'], emp['ort'])
        if website:
            result['website_url'] = website
            main_html = download_html(website)
            if main_html:
                path, h = save_html_file(main_html, emp['name'])
                result['website_html_path'] = path
                result['website_html_hash'] = h
                
                has_it, it_kws = check_website_for_it(main_html)
                if has_it:
                    if not result['hat_it_jobs']: result['hat_it_jobs'] = 1
                    if not result['it_job_quelle']: result['it_job_quelle'] = 'Website'
                    result['it_keywords_gefunden'] = ", ".join(it_kws)
                    result['it_relevanz_score'] = min(result['it_relevanz_score'] + 15, 100)
                
                contact, imp_url, kar_url = extract_contact_info(main_html)
                if contact.get('email'): result['email'] = contact['email']
                if contact.get('geschaeftsfuehrer'): result['geschaeftsfuehrer'] = contact['geschaeftsfuehrer']
                if contact.get('handelsregister'): result['handelsregister'] = contact['handelsregister']
                
                if imp_url:
                    imp_abs = make_absolute_url(website, imp_url)
                    if imp_abs:
                        result['impressum_url'] = imp_abs
                        imp_html = download_html(imp_abs)
                        if imp_html:
                            path, _ = save_html_file(imp_html, emp['name'], "impressum")
                            result['impressum_html_path'] = path
                            ic, _, _ = extract_contact_info(imp_html)
                            if ic.get('email') and not result['email']: result['email'] = ic['email']
                            if ic.get('geschaeftsfuehrer') and not result['geschaeftsfuehrer']: result['geschaeftsfuehrer'] = ic['geschaeftsfuehrer']
                            if ic.get('handelsregister') and not result['handelsregister']: result['handelsregister'] = ic['handelsregister']
                
                if kar_url:
                    kar_abs = make_absolute_url(website, kar_url)
                    if kar_abs:
                        result['karriere_url'] = kar_abs
                        kar_html = download_html(kar_abs)
                        if kar_html:
                            path, _ = save_html_file(kar_html, emp['name'], "karriere")
                            result['karriere_html_path'] = path
        
        result['scrape_status'] = 'completed'
        result['scrape_datum'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        result['letzte_pruefung'] = datetime.now().strftime('%Y-%m-%d')
        return result
    
    # Process with thread pool
    total = len(remaining)
    processed = 0
    it_count = 0
    batch_commit = 0
    
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(process_one, m): m for m in remaining}
        
        for future in as_completed(futures):
            m = futures[future]
            try:
                result = future.result()
            except Exception as e:
                result = {'firmenname': m['employer']['name'], 'scrape_status': f'error: {e}'}
            
            # Save to DB
            cols = ['firmenname','firmenname_normalisiert','ort','plz','strasse',
                    'telefon','email','website_url','website_html_path','website_html_hash',
                    'karriere_url','karriere_html_path','impressum_url','impressum_html_path',
                    'hat_it_jobs','it_job_titel','it_job_anzahl','it_job_quelle',
                    'it_job_details','it_keywords_gefunden','kategorie_google',
                    'bewertung_google','lat','lng','google_maps_link',
                    'handelsregister','geschaeftsfuehrer','it_abteilung_vermutung',
                    'it_relevanz_score','scrape_datum','scrape_status','letzte_pruefung']
            vals = [
                result.get('firmenname',''), normalize_name(result.get('firmenname','')),
                result.get('ort',''), result.get('plz',''), result.get('strasse',''),
                result.get('telefon',''), result.get('email',''), result.get('website_url',''),
                result.get('website_html_path',''), result.get('website_html_hash',''),
                result.get('karriere_url',''), result.get('karriere_html_path',''),
                result.get('impressum_url',''), result.get('impressum_html_path',''),
                result.get('hat_it_jobs',0), result.get('it_job_titel',''),
                result.get('it_job_anzahl',0), result.get('it_job_quelle',''),
                result.get('it_job_details',''), result.get('it_keywords_gefunden',''),
                result.get('kategorie_google',''), result.get('bewertung_google'),
                result.get('lat'), result.get('lng'), result.get('google_maps_link',''),
                result.get('handelsregister',''), result.get('geschaeftsfuehrer',''),
                result.get('it_abteilung_vermutung',0), result.get('it_relevanz_score',0.0),
                result.get('scrape_datum',''), result.get('scrape_status',''),
                result.get('letzte_pruefung',''),
            ]
            conn.execute(f"INSERT OR REPLACE INTO employers ({','.join(cols)}) VALUES ({','.join(['?']*len(cols))})", vals)
            conn.commit()
            
            processed += 1
            batch_commit += 1
            progress['completed'].append(result.get('firmenname', ''))
            if result.get('hat_it_jobs'): it_count += 1
            
            status = "✓IT" if result.get('hat_it_jobs') else "   "
            ws_url = result.get('website_url','')[:35] if result.get('website_url') else '-'
            score = result.get('it_relevanz_score', 0)
            print(f"  [{processed}/{total}] {status} Score:{score:5.1f} | {result.get('firmenname','')[:42]:42} | {ws_url}")
            
            if batch_commit >= COMMIT_EVERY:
                # Save progress
                with open(progress_file, 'w') as f:
                    json.dump(progress, f, ensure_ascii=False)
                
                # Git commit & push
                try:
                    subprocess.run(['git','add','-A'], cwd=GIT_DIR, capture_output=True, timeout=30)
                    msg = f"IT-Scraper: {processed}/{total} | {it_count} IT-Arbeitgeber ({datetime.now().strftime('%H:%M')})"
                    r = subprocess.run(['git','commit','-m',msg], cwd=GIT_DIR, capture_output=True, timeout=30)
                    if r.returncode == 0:
                        subprocess.run(['git','push','origin','main'], cwd=GIT_DIR, capture_output=True, timeout=60)
                        print(f"  >>> GIT PUSH OK ({it_count} IT-Arbeitgeber)")
                except Exception as e:
                    print(f"  GIT Error: {e}")
                batch_commit = 0
    
    # Save final progress
    with open(progress_file, 'w') as f:
        json.dump(progress, f, ensure_ascii=False)
    
    # Generate Excel
    print("\n=== EXCEL GENERIEREN ===")
    wb_out = openpyxl.Workbook()
    ws_main = wb_out.active
    ws_main.title = "IT Arbeitgeber"
    
    cursor = conn.execute("SELECT * FROM employers ORDER BY it_relevanz_score DESC, hat_it_jobs DESC, firmenname")
    columns = [desc[0] for desc in cursor.description]
    
    header_map = {
        'id':'ID','firmenname':'Firmenname','firmenname_normalisiert':'Firmenname (normalisiert)',
        'ort':'Ort','plz':'PLZ','strasse':'Straße','telefon':'Telefon','email':'E-Mail',
        'website_url':'Website URL','website_html_path':'Website HTML Pfad',
        'website_html_hash':'Website HTML Hash','karriere_url':'Karriere-URL',
        'karriere_html_path':'Karriere HTML Pfad','impressum_url':'Impressum URL',
        'impressum_html_path':'Impressum HTML Pfad','hat_it_jobs':'IT-Jobs vorhanden',
        'it_job_titel':'IT-Job Titel','it_job_anzahl':'Anzahl IT-Jobs',
        'it_job_quelle':'IT-Job Quelle','it_job_details':'IT-Job Details',
        'it_keywords_gefunden':'IT-Keywords gefunden','kategorie_google':'Kategorie (Google)',
        'bewertung_google':'Bewertung (Google)','lat':'Breitengrad','lng':'Längengrad',
        'google_maps_link':'Google Maps Link','handelsregister':'Handelsregister',
        'geschaeftsfuehrer':'Geschäftsführer','it_abteilung_vermutung':'IT-Abteilung vermutet',
        'it_relevanz_score':'IT-Relevanz Score','scrape_datum':'Scrape-Datum',
        'scrape_status':'Scrape-Status','letzte_pruefung':'Letzte Prüfung',
        'created_at':'Erstellt am','updated_at':'Aktualisiert am',
    }
    
    headers = [header_map.get(c, c) for c in columns]
    ws_main.append(headers)
    for row in cursor:
        ws_main.append(list(row))
    
    for col in ws_main.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws_main.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
    
    # Stats sheet
    ws2 = wb_out.create_sheet("Statistiken")
    stats = [
        ("Gesamt Arbeitgeber", conn.execute("SELECT COUNT(*) FROM employers").fetchone()[0]),
        ("Mit IT-Jobs", conn.execute("SELECT COUNT(*) FROM employers WHERE hat_it_jobs=1").fetchone()[0]),
        ("Mit Website", conn.execute("SELECT COUNT(*) FROM employers WHERE website_url != ''").fetchone()[0]),
        ("Mit E-Mail", conn.execute("SELECT COUNT(*) FROM employers WHERE email != ''").fetchone()[0]),
        ("IT-Abt. vermutet", conn.execute("SELECT COUNT(*) FROM employers WHERE it_abteilung_vermutung=1").fetchone()[0]),
        ("Ø IT-Relevanz", round(conn.execute("SELECT AVG(it_relevanz_score) FROM employers").fetchone()[0] or 0, 1)),
    ]
    for s in stats:
        ws2.append(s)
    
    # Per-city
    ws3 = wb_out.create_sheet("Nach Ort")
    ws3.append(["Ort","Gesamt","Mit IT-Jobs","Mit Website","Mit E-Mail","Ø IT-Relevanz"])
    for row in conn.execute("""
        SELECT ort, COUNT(*),
               SUM(CASE WHEN hat_it_jobs=1 THEN 1 ELSE 0 END),
               SUM(CASE WHEN website_url != '' THEN 1 ELSE 0 END),
               SUM(CASE WHEN email != '' THEN 1 ELSE 0 END),
               ROUND(AVG(it_relevanz_score), 1)
        FROM employers GROUP BY ort ORDER BY COUNT(*) DESC
    """):
        ws3.append(list(row))
    
    # Top IT employers
    ws4 = wb_out.create_sheet("Top IT Arbeitgeber")
    ws4.append(["Firmenname","Ort","IT-Jobs","Anzahl Jobs","IT-Score","Website","E-Mail"])
    for row in conn.execute("""
        SELECT firmenname, ort, hat_it_jobs, it_job_anzahl, it_relevanz_score, website_url, email
        FROM employers WHERE hat_it_jobs=1 ORDER BY it_job_anzahl DESC, it_relevanz_score DESC
    """):
        ws4.append(list(row))
    
    wb_out.save(EXCEL_OUT)
    print(f"Excel: {EXCEL_OUT}")
    
    # Final git
    try:
        subprocess.run(['git','add','-A'], cwd=GIT_DIR, capture_output=True, timeout=30)
        subprocess.run(['git','commit','-m',f"IT-Scraper FERTIG: {it_count} IT-Arbeitgeber"], cwd=GIT_DIR, capture_output=True, timeout=30)
        subprocess.run(['git','push','origin','main'], cwd=GIT_DIR, capture_output=True, timeout=60)
    except:
        pass
    
    conn.close()
    print(f"\nFERTIG! {it_count} IT-Arbeitgeber aus {total} gescannt")

if __name__ == '__main__':
    main()
