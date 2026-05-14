#!/usr/bin/env python3
"""
High-Quality Junior IT-Jobs Scraper v3
=======================================
- Uses CloakBrowser (stealth Chromium) with Xvfb headful mode
- Navigates employer websites like a human
- Uses DuckDuckGo + Google search to find career pages
- Extracts real, current job listings from source websites
- Filters for junior-level IT positions
- Commits every 10 entries to GitHub
"""

import asyncio
import hashlib
import json
import os
import re
import subprocess
import sys
import time
import traceback
from datetime import datetime
from pathlib import Path
from urllib.parse import urljoin, urlparse, quote, unquote

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG
# ============================================================
EXCEL_EMPLOYERS = "/home/z/my-project/download/Firmen_Bamberg_Erlangen_Nuernberg_mit_Website_HTML.xlsx"
EXCEL_OUT = "/home/z/my-project/download/Junior_IT_Jobs_Quellennah.xlsx"
CAREER_HTML_DIR = "/home/z/my-project/download/career_pages_v3"
JOB_HTML_DIR = "/home/z/my-project/download/job_detail_v3"
PROGRESS_FILE = "/home/z/my-project/download/scrape_v3_progress.json"
GIT_REPO = "/home/z/my-project"
GITHUB_TOKEN = "GIT_TOKEN_REDACTED"

COMMIT_EVERY = 10
MAX_CONCURRENT_BROWSERS = 2  # CloakBrowser instances
PAGE_TIMEOUT = 25000
JOB_DETAIL_TIMEOUT = 15000

os.makedirs(CAREER_HTML_DIR, exist_ok=True)
os.makedirs(JOB_HTML_DIR, exist_ok=True)

# ============================================================
# KEYWORDS
# ============================================================
JUNIOR_KEYWORDS = [
    'junior', 'entry level', 'einsteiger', 'anfänger',
    'trainee', 'praktikant', 'intern', 'werkstudent', 'working student',
    'absolvent', 'graduate', 'berufseinsteiger', 'young professional',
    'associate', 'graduate program', 'trainee program',
    'duales studium', 'ausbildung', 'apprentice', 'azubi',
    'nachwuchs', 'career starter', 'erstjob',
    '(m/w/d)',  # common in German job postings
]

SENIOR_EXCLUDE = [
    'senior', 'lead', 'principal', 'head of', 'director',
    'leitung', 'chef', 'geschäftsführer', 'vorstand',
    'professor', 'prof.',
]

IT_KEYWORDS = [
    'software', 'entwickler', 'developer', 'it-', 'informatik',
    'data', 'devops', 'cloud', 'cyber', 'security', 'sicherheit',
    'system', 'admin', 'administrator', 'netzwerk', 'network',
    'frontend', 'backend', 'fullstack', 'full-stack', 'full stack',
    'python', 'java', 'javascript', 'typescript', 'react', 'angular', 'vue',
    'sap', 'linux', 'sql', 'database', 'datenbank', 'api',
    'machine learning', 'ki', 'künstliche intelligenz', 'ai',
    'scrum', 'agile', 'testing', 'qa', 'automation',
    'consultant', 'analyst', 'engineer', 'architekt', 'architect',
    'programmierung', 'programming', 'coding', 'code',
    'web', 'mobile', 'app', 'anwendung',
    'infrastructure', 'infrastruktur', 'platform', 'plattform',
    'support', 'helpdesk', 'service desk', 'ticket',
    'erp', 'crm', 'dynamics', 'sharepoint',
    'digitalisierung', 'technologie', 'technology', 'digital',
    'fachinformatiker', 'it-systemelektroniker', 'it-kaufmann',
    'elektrotechnik', 'elektronik',
]

CAREER_LINK_TEXTS = [
    'karriere', 'career', 'careers', 'jobs', 'stellenangebote',
    'stellen', 'offene stellen', 'vacancies', 'vacancy',
    'jobbörse', 'recruiting', 'bewerbung', 'talent',
    'opportunities', 'join us', 'join', 'wir suchen',
    'wir stellen ein', 'offene positionen', 'open positions',
    'unser team', 'arbeiten bei', 'work at', 'mitarbeiter',
    'jobbörse', 'stellenmarkt', 'jobportal', 'werkstudent',
    'praktikum', 'ausbildung', 'trainee', 'absolventen',
    'berufseinsteiger', 'nachwuchs',
]

CAREER_URL_PATTERNS = [
    '/karriere', '/career', '/careers', '/jobs', '/stellenangebote',
    '/stellen', '/offene-stellen', '/vacancies', '/vacancy',
    '/job', '/jobboerse', '/recruiting', '/bewerbung', '/talent',
    '/opportunities', '/join-us', '/join', '/werde-teil',
    '/unsere-jobs', '/arbeiten-bei', '/offene-positionen',
    '/de/karriere', '/en/career', '/de/career', '/en/careers',
    '/de/jobs', '/en/jobs', '/unternehmen/karriere',
    '/ueber-uns/karriere', '/about/careers', '/about-us/careers',
    '/job-search', '/stellensuche', '/jobboerse',
    '/studenten', '/absolventen', '/graduates', '/students',
    '/early-career', '/einsteiger', '/berufseinsteiger',
    '/trainee', '/praktikum', '/werkstudent', '/ausbildung',
]


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def sanitize_filename(name, max_len=60):
    name = re.sub(r'[^\w\s.-]', '_', name)
    name = re.sub(r'_{2,}', '_', name)
    name = name.strip('_ .')
    return name[:max_len]


def is_junior_job(title, description=""):
    text = f"{title} {description}".lower()
    # Check senior exclude first
    for kw in SENIOR_EXCLUDE:
        if kw in text:
            return False, None
    # Check junior keywords
    for kw in JUNIOR_KEYWORDS:
        if kw in text:
            return True, kw
    # Check experience requirement
    exp_match = re.search(r'(\d+)\s*(?:jahr|year|jahre|years)\s*(?:berufs)?erfahrung', text)
    if exp_match:
        years = int(exp_match.group(1))
        if years <= 2:
            return True, f"erfahrung_{years}jahre"
    if re.search(r'ohne\s*erfahrung|no\s*experience|keine\s*erfahrung|erfahrung\s*nicht', text):
        return True, "no_experience"
    return False, None


def is_it_job(title, description=""):
    text = f"{title} {description}".lower()
    for kw in IT_KEYWORDS:
        if kw in text:
            return True, kw
    return False, None


def is_senior_job(title):
    text = title.lower()
    return any(kw in text for kw in SENIOR_EXCLUDE)


def extract_technologies(text):
    tech_keywords = [
        'python', 'java', 'javascript', 'typescript', 'c#', 'c++',
        'ruby', 'go', 'rust', 'kotlin', 'swift', 'scala', 'php',
        'react', 'angular', 'vue', 'vue.js', 'svelte', 'next.js',
        'node.js', 'nodejs', 'express', 'django', 'flask', 'spring',
        '.net', 'asp.net', 'docker', 'kubernetes', 'k8s', 'aws', 'azure', 'gcp',
        'terraform', 'ansible', 'jenkins', 'gitlab', 'github',
        'sql', 'mysql', 'postgresql', 'postgres', 'mongodb', 'redis',
        'kafka', 'rabbitmq', 'graphql', 'rest', 'api',
        'sap', 'salesforce', 'dynamics', 'sharepoint', 'servicenow',
        'linux', 'windows', 'macos', 'unix',
        'html', 'css', 'scss', 'sass', 'tailwind', 'bootstrap',
        'agile', 'scrum', 'kanban', 'devops', 'ci/cd',
        'machine learning', 'deep learning', 'tensorflow', 'pytorch',
        'nlp', 'computer vision', 'data science', 'big data',
        'jira', 'confluence', 'cybersecurity', 'pentest',
        'powershell', 'bash', 'shell', 'git',
        'azure devops', 'bitbucket', 'intellij', 'vscode',
    ]
    found = []
    text_lower = text.lower()
    for tech in tech_keywords:
        if re.search(r'\b' + re.escape(tech) + r'\b', text_lower):
            found.append(tech)
    return list(set(found))


def classify_job_category(title, description=""):
    text = f"{title} {description}".lower()
    categories = {
        'Softwareentwicklung': ['softwareentwickler', 'software developer', 'software engineer', 'entwickler', 'developer', 'programmierer', 'frontend', 'backend', 'fullstack', 'full-stack', 'anwendungsentwicklung'],
        'Data Science & Analytics': ['data scientist', 'data analyst', 'data engineer', 'data analytics', 'business intelligence', 'machine learning', 'ki-entwickler', 'ai developer'],
        'IT-Consulting': ['it consultant', 'it-consultant', 'it berater', 'sap consultant', 'technischer berater', 'digitalisierungsberater'],
        'Systemadministration': ['system administrator', 'systemadministrator', 'sysadmin', 'admin', 'netzwerkadministrator', 'linux admin', 'systemintegration'],
        'DevOps & Cloud': ['devops', 'cloud engineer', 'cloud architect', 'platform engineer', 'sre', 'infrastruktur'],
        'Cyber Security': ['security', 'cybersecurity', 'information security', 'it-sicherheit', 'pentest', 'soc analyst'],
        'IT-Support': ['it support', 'helpdesk', 'service desk', 'first level', 'second level', 'user support', '1st level', '2nd level'],
        'IT-Projektmanagement': ['project manager', 'projektmanager', 'scrum master', 'product owner'],
        'QA & Testing': ['quality assurance', 'qa', 'tester', 'test engineer', 'testautomatisierung'],
        'UX/UI Design': ['ux', 'ui', 'user experience', 'user interface', 'design', 'usability'],
        'Datenbankadministration': ['dba', 'datenbankadministrator', 'database administrator'],
        'IT-Ausbildung & Studium': ['ausbildung', 'duales studium', 'trainee', 'praktikum', 'werkstudent', 'fachinformatiker'],
    }
    for category, keywords in categories.items():
        for kw in keywords:
            if kw in text:
                return category
    return 'IT (Sonstige)'


def extract_employment_type(title, description=""):
    text = f"{title} {description}".lower()
    if any(kw in text for kw in ['werkstudent', 'working student', 'studentische hilfskraft']):
        return 'Werkstudent'
    if any(kw in text for kw in ['praktikant', 'praktikum', 'intern', 'internship']):
        return 'Praktikum'
    if any(kw in text for kw in ['ausbildung', 'apprentice', 'azubi']):
        return 'Ausbildung'
    if any(kw in text for kw in ['duales studium', 'dual student']):
        return 'Duales Studium'
    if any(kw in text for kw in ['teilzeit', 'part-time', 'part time']):
        return 'Teilzeit'
    if any(kw in text for kw in ['minijob', 'mini-job', 'geringfügig']):
        return 'Minijob'
    return 'Vollzeit'


def extract_remote_option(title, description=""):
    text = f"{title} {description}".lower()
    if any(kw in text for kw in ['fully remote', '100% remote', 'voll remote']):
        return 'Remote'
    if any(kw in text for kw in ['remote', 'homeoffice', 'home-office', 'home office', 'work from home']):
        return 'Remote möglich'
    if any(kw in text for kw in ['hybrid', 'teilweise remote']):
        return 'Hybrid'
    if any(kw in text for kw in ['vor ort', 'on-site', 'onsite']):
        return 'Vor Ort'
    return 'Nicht angegeben'


def clean_city(city_raw):
    if not city_raw:
        return ''
    city = city_raw.strip()
    for suffix in [', Mittelfranken', ', Oberfranken', ', Unterfranken',
                   ', Bayern', ', Deutschland']:
        city = city.replace(suffix, '')
    return city.strip()


def classify_region(city):
    if not city:
        return 'Unbekannt'
    c = city.lower()
    if any(x in c for x in ['nürnberg', 'nuernberg', 'fürth', 'fuerth', 'erlangen', 'schwabach']):
        return 'Metropolregion Nürnberg (Kern)'
    if any(x in c for x in ['bamberg', 'forchheim', 'coburg', 'bayreuth']):
        return 'Oberfranken'
    if any(x in c for x in ['ansbach', 'roth', 'weißenburg', 'gunzenhausen']):
        return 'Mittelfranken (Land)'
    return 'Weitere Regionen'


def classify_junior_type(title, junior_kw=''):
    text = title.lower()
    if 'ausbildung' in text or junior_kw == 'ausbildung':
        return 'Ausbildung'
    if 'duales studium' in text or junior_kw == 'duales studium':
        return 'Duales Studium'
    if 'werkstudent' in text or junior_kw == 'werkstudent':
        return 'Werkstudent'
    if 'praktikum' in text or 'praktikant' in text or junior_kw == 'praktikant':
        return 'Praktikum'
    if 'trainee' in text or junior_kw == 'trainee':
        return 'Trainee'
    if 'junior' in text or junior_kw == 'junior':
        return 'Junior'
    if 'associate' in text or junior_kw == 'associate':
        return 'Associate'
    if 'einsteiger' in text or 'berufseinsteiger' in text:
        return 'Berufseinsteiger'
    if 'absolvent' in text or junior_kw == 'absolvent':
        return 'Absolvent'
    return 'Junior-fähig'


def classify_age(veroeffdatum):
    if not veroeffdatum:
        return 'Unbekannt', ''
    try:
        dt = datetime.strptime(str(veroeffdatum)[:10], '%Y-%m-%d')
        days = (datetime.now() - dt).days
        if days <= 7:
            return 'Frisch (≤ 7 Tage)', days
        elif days <= 30:
            return 'Aktuell (8-30 Tage)', days
        elif days <= 90:
            return 'Älter (31-90 Tage)', days
        else:
            return 'Veraltet (> 90 Tage)', days
    except:
        return 'Unbekannt', ''


# ============================================================
# EXCEL MANAGER
# ============================================================

class JuniorJobsExcel:
    HEADERS = [
        # A: Job identification
        'Job-ID',                    # A
        'Job-Titel',                 # B
        'Job-Kategorie',             # C
        'Erfahrungslevel',           # D
        'Junior-Typ',                # E
        'Explizit Junior',           # F
        # G: Job details
        'Beschäftigungsart',         # G
        'Remote-Option',             # H
        'Technologien',              # I
        'Ausschreibungstext',        # J
        'Job-URL',                   # K
        # L: Location
        'Stadt',                     # L
        'Region',                    # M
        'PLZ',                       # N
        'Straße',                    # O
        # P: Employer
        'Arbeitgeber',               # P
        'Arbeitgeber-Website',       # Q
        'Karriere-URL',              # R
        'Branche',                   # S
        # T: Source
        'Quelle',                    # T
        'Quelle-Typ',                # U  (Website direkt / Suchmaschine / Arbeitsagentur)
        'Gefunden über',             # V  (which search query)
        'Arbeitsagentur RefNr',      # W
        # X: Classification
        'Junior-Keyword Match',      # X
        'IT-Keyword Match',          # Y
        'Ist Junior-Job',            # Z
        'Ist IT-Job',                # AA
        # AB: Meta
        'Veröffentlicht am',         # AB
        'Veröffentlicht vor (Tage)', # AC
        'Alter der Ausschreibung',   # AD
        'Scrape-Datum',              # AE
        'Scrape-Status',             # AF
        'Arbeitgeber-ID',            # AG
        'Karriere-HTML Pfad',        # AH
        'Job-HTML Pfad',             # AI
    ]

    def __init__(self, filepath):
        self.filepath = filepath
        if os.path.exists(filepath):
            self.wb = openpyxl.load_workbook(filepath)
            self.ws = self.wb.active
            self.job_counter = self.ws.max_row - 1
            # Check existing job IDs to avoid duplicates
            self.existing_ids = set()
            self.existing_urls = set()
            self.existing_titles = {}  # title_lower -> employer
            for r in range(2, self.ws.max_row + 1):
                jid = self.ws.cell(r, 1).value
                url = self.ws.cell(r, 11).value  # Job-URL column
                title = self.ws.cell(r, 2).value
                employer = self.ws.cell(r, 16).value
                if jid:
                    self.existing_ids.add(jid)
                if url:
                    self.existing_urls.add(url)
                if title and employer:
                    self.existing_titles[f"{str(title).lower().strip()}_{str(employer).lower().strip()}"] = True
        else:
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active
            self.ws.title = "Junior IT-Jobs (Quellennah)"
            self._write_headers()
            self.job_counter = 0
            self.existing_ids = set()
            self.existing_urls = set()
            self.existing_titles = {}

    def _write_headers(self):
        header_font = Font(bold=True, color="FFFFFF", size=11, name='Calibri')
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        for col, header in enumerate(self.HEADERS, 1):
            cell = self.ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        widths = [10,45,22,14,18,14, 16,16,35,60,50, 18,28,8,25, 38,50,50,22, 16,18,30,22, 22,18,12,12, 14,18,18,18,14,16,50,50]
        for i, w in enumerate(widths):
            if i < len(self.HEADERS):
                self.ws.column_dimensions[get_column_letter(i+1)].width = w

    def is_duplicate(self, job_data):
        """Check if job already exists."""
        url = job_data.get('job_url', '')
        title = job_data.get('job_title', '')
        employer = job_data.get('employer_name', '')
        
        if url and url in self.existing_urls:
            return True
        
        key = f"{title.lower().strip()}_{employer.lower().strip()}"
        if key in self.existing_titles:
            return True
        
        return False

    def add_job(self, job_data):
        if self.is_duplicate(job_data):
            return False

        self.job_counter += 1
        row = self.ws.max_row + 1

        title = job_data.get('job_title', '')
        description = job_data.get('description', '')
        junior_kw = job_data.get('junior_keyword_match', '')
        
        # Calculate derived fields
        veroeff = job_data.get('date_posted', '')
        age_label, days_ago = classify_age(veroeff)

        values = [
            f'Q-{self.job_counter:05d}',              # A: Job-ID
            title,                                      # B: Job-Titel
            classify_job_category(title, description),  # C: Job-Kategorie
            'Junior',                                   # D: Erfahrungslevel
            classify_junior_type(title, junior_kw),     # E: Junior-Typ
            'Ja' if junior_kw and junior_kw != 'no_senior_keyword' else 'Nein',  # F: Explizit Junior
            extract_employment_type(title, description), # G: Beschäftigungsart
            extract_remote_option(title, description),   # H: Remote-Option
            ', '.join(extract_technologies(f"{title} {description}")),  # I: Technologien
            (description or '')[:1000],                  # J: Ausschreibungstext
            job_data.get('job_url', ''),                 # K: Job-URL
            clean_city(job_data.get('city', '')),        # L: Stadt
            classify_region(job_data.get('city', '')),   # M: Region
            job_data.get('plz', ''),                     # N: PLZ
            job_data.get('street', ''),                  # O: Straße
            job_data.get('employer_name', ''),           # P: Arbeitgeber
            job_data.get('website_url', ''),             # Q: Arbeitgeber-Website
            job_data.get('career_url', ''),              # R: Karriere-URL
            job_data.get('industry', ''),                # S: Branche
            job_data.get('source', 'Website'),           # T: Quelle
            job_data.get('source_type', 'Website direkt'),  # U: Quelle-Typ
            job_data.get('found_via', ''),               # V: Gefunden über
            job_data.get('refnr', ''),                   # W: RefNr
            junior_kw,                                  # X: Junior-Keyword Match
            job_data.get('it_keyword_match', ''),        # Y: IT-Keyword Match
            'Ja',                                       # Z: Ist Junior-Job
            'Ja',                                       # AA: Ist IT-Job
            veroeff,                                    # AB: Veröffentlicht am
            days_ago if days_ago != '' else '',          # AC: Veröffentlicht vor
            age_label,                                  # AD: Alter der Ausschreibung
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),  # AE: Scrape-Datum
            job_data.get('scrape_status', 'completed'),  # AF: Scrape-Status
            job_data.get('employer_id', ''),             # AG: Arbeitgeber-ID
            job_data.get('career_html_path', ''),        # AH: Karriere-HTML
            job_data.get('job_html_path', ''),           # AI: Job-HTML
        ]

        for col, value in enumerate(values, 1):
            cell = self.ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.font = Font(name='Calibri', size=10)

        # Track for dedup
        url = job_data.get('job_url', '')
        if url:
            self.existing_urls.add(url)
        key = f"{title.lower().strip()}_{job_data.get('employer_name', '').lower().strip()}"
        self.existing_titles[key] = True

        return True

    def save(self):
        self.wb.save(self.filepath)

    def get_job_count(self):
        return self.job_counter


# ============================================================
# PROGRESS MANAGER
# ============================================================

def load_progress():
    if os.path.exists(PROGRESS_FILE):
        return json.load(open(PROGRESS_FILE, encoding='utf-8'))
    return {'completed': [], 'failed': [], 'last_index': 0}

def save_progress(progress):
    with open(PROGRESS_FILE, 'w', encoding='utf-8') as f:
        json.dump(progress, f, ensure_ascii=False, indent=1)


# ============================================================
# GIT HELPER
# ============================================================

def git_commit_and_push(excel, count_since_last):
    try:
        excel.save()
        os.chdir(GIT_REPO)
        subprocess.run(['git', 'add', '-A'], capture_output=True, timeout=30)
        msg = f"Junior IT-Jobs quellennah - {excel.get_job_count()} Jobs (+{count_since_last})"
        result = subprocess.run(['git', 'commit', '-m', msg], capture_output=True, timeout=30)
        if result.returncode == 0:
            push_result = subprocess.run(
                ['git', 'push', f'https://{GITHUB_TOKEN}@github.com/BirdNest055/Heyjobs.git', 'main'],
                capture_output=True, timeout=60
            )
            if push_result.returncode == 0:
                print(f"  [GIT] Pushed ({excel.get_job_count()} jobs, +{count_since_last})")
                return True
            else:
                print(f"  [GIT] Push failed: {push_result.stderr.decode()[:100]}")
        return False
    except Exception as e:
        print(f"  [GIT] Error: {str(e)[:80]}")
        return False


# ============================================================
# CLOAKBROWSER SCRAPER
# ============================================================

class CloakScraper:
    def __init__(self, excel):
        self.excel = excel
        self.jobs_since_commit = 0
        self.browser = None
        self.context = None
    
    def start_browser(self):
        """Launch CloakBrowser with stealth + human behavior."""
        from cloakbrowser import launch
        self.browser = launch(
            headless=True,
            args=[
                '--fingerprint=42',
                '--fingerprint-platform=windows',
                '--disable-blink-features=AutomationControlled',
            ],
        )
        self.context = self.browser.new_context(
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36',
            viewport={'width': 1920, 'height': 1080},
            locale='de-DE',
            timezone_id='Europe/Berlin',
        )
        print("  [BROWSER] CloakBrowser gestartet (Stealth + Windows Fingerprint)")
    
    def stop_browser(self):
        if self.context:
            try:
                self.context.close()
            except:
                pass
        if self.browser:
            try:
                self.browser.close()
            except:
                pass
        print("  [BROWSER] Geschlossen")
    
    def search_duckduckgo(self, query, max_results=5):
        """Search DuckDuckGo for career pages."""
        results = []
        try:
            page = self.context.new_page()
            try:
                search_url = f'https://html.duckduckgo.com/html/?q={quote(query)}'
                page.goto(search_url, timeout=PAGE_TIMEOUT, wait_until='domcontentloaded')
                time.sleep(2)
                
                # Parse results
                links = page.query_selector_all('.result__a')
                for link in links[:max_results]:
                    href = link.get_attribute('href')
                    text = link.inner_text()
                    if href:
                        # DDG redirects - extract actual URL
                        if 'uddg=' in href:
                            actual_url = unquote(href.split('uddg=')[1].split('&')[0])
                        else:
                            actual_url = href
                        results.append({
                            'title': text.strip(),
                            'url': actual_url,
                        })
            finally:
                page.close()
        except Exception as e:
            print(f"    [DDG] Search error: {str(e)[:60]}")
        return results
    
    def search_google(self, query, max_results=5):
        """Search Google for career pages."""
        results = []
        try:
            page = self.context.new_page()
            try:
                search_url = f'https://www.google.com/search?q={quote(query)}&hl=de&num=10'
                page.goto(search_url, timeout=PAGE_TIMEOUT, wait_until='domcontentloaded')
                time.sleep(3)
                
                # Try to dismiss consent
                for sel in ['button:has-text("Akzeptieren")', 'button:has-text("Accept all")',
                           '#L2AGLb', '#W0wltc', 'button:has-text("Alle akzeptieren")']:
                    try:
                        btn = page.query_selector(sel)
                        if btn:
                            btn.click()
                            time.sleep(1)
                            break
                    except:
                        continue
                
                # Parse results - multiple selector strategies
                selectors = ['div.g a', 'div[data-sokoban-container] a', 'a[href*="/url"]', '.tF2Cxc a']
                seen_urls = set()
                for selector in selectors:
                    links = page.query_selector_all(selector)
                    for link in links[:10]:
                        href = link.get_attribute('href')
                        if href and 'google' not in href and 'youtube' not in href:
                            if href not in seen_urls:
                                seen_urls.add(href)
                                text = link.inner_text().strip()
                                results.append({
                                    'title': text[:100],
                                    'url': href,
                                })
                                if len(results) >= max_results:
                                    break
                    if len(results) >= max_results:
                        break
            finally:
                page.close()
        except Exception as e:
            print(f"    [Google] Search error: {str(e)[:60]}")
        return results
    
    def find_career_page(self, employer_name, base_url):
        """Find career/jobs page on employer website."""
        page = None
        try:
            page = self.context.new_page()
            
            # Step 1: Visit homepage
            try:
                page.goto(base_url, timeout=PAGE_TIMEOUT, wait_until='domcontentloaded')
                time.sleep(2)
            except Exception as e:
                print(f"    Homepage nicht erreichbar: {str(e)[:60]}")
                return None, None
            
            # Dismiss cookie banners
            self._dismiss_cookies(page)
            
            # Step 2: Look for career links on homepage
            career_url = self._find_career_link_on_page(page, base_url)
            if career_url:
                page.close()
                return career_url, 'website_navigation'
            
            # Step 3: Try common URL patterns
            career_url = self._try_url_patterns(page, base_url)
            if career_url:
                page.close()
                return career_url, 'url_pattern'
            
            page.close()
            return None, None
            
        except Exception as e:
            if page:
                try:
                    page.close()
                except:
                    pass
            return None, None
    
    def _dismiss_cookies(self, page):
        """Dismiss cookie consent banners."""
        consent_selectors = [
            'button:has-text("Akzeptieren")', 'button:has-text("Accept")',
            'button:has-text("Alle akzeptieren")', 'button:has-text("Accept All")',
            'button:has-text("OK")', 'button:has-text("Verstanden")',
            'button:has-text("Zustimmen")', 'button:has-text("Einverstanden")',
            '#onetrust-accept-btn-handler', '.cc-btn', '#accept-cookie-notification',
            '[class*="cookie"] button:first-child', '[id*="cookie"] button:first-child',
            'a:has-text("Akzeptieren")', 'button[aria-label="Accept"]',
        ]
        for sel in consent_selectors:
            try:
                btn = page.query_selector(sel)
                if btn and btn.is_visible():
                    btn.click()
                    time.sleep(0.5)
                    return
            except:
                continue
    
    def _find_career_link_on_page(self, page, base_url):
        """Find career page link on the current page."""
        # Strategy 1: Direct text matching in links
        for text in CAREER_LINK_TEXTS:
            try:
                selector = f'a:has-text("{text}")'
                elements = page.query_selector_all(selector)
                for elem in elements:
                    href = elem.get_attribute('href')
                    if href:
                        full_url = urljoin(base_url, href)
                        url_lower = full_url.lower()
                        if any(kw in url_lower for kw in ['karriere', 'career', 'job', 'stelle', 'vacanc', 'recruit', 'talent', 'bewerbung', 'ausbildung', 'trainee', 'werkstudent', 'praktikum']):
                            return full_url
            except:
                continue
        
        # Strategy 2: Nav/header/footer links
        for area in ['nav a', 'header a', 'footer a', '.nav a', '.navigation a', '.menu a', '#menu a']:
            try:
                links = page.query_selector_all(area)
                for link in links:
                    link_text = (link.inner_text() or '').strip().lower()
                    href = link.get_attribute('href')
                    if href and any(kw in link_text for kw in CAREER_LINK_TEXTS):
                        full_url = urljoin(base_url, href)
                        return full_url
            except:
                continue
        
        # Strategy 3: Check for common career section elements
        try:
            all_links = page.query_selector_all('a[href]')
            for link in all_links:
                href = link.get_attribute('href') or ''
                href_lower = href.lower()
                if any(pattern in href_lower for pattern in ['/karriere', '/career', '/jobs', '/stellen', '/jobboerse']):
                    full_url = urljoin(base_url, href)
                    return full_url
        except:
            pass
        
        return None
    
    def _try_url_patterns(self, page, base_url):
        """Try common career URL patterns."""
        parsed = urlparse(base_url)
        base = f"{parsed.scheme}://{parsed.netloc}"
        
        for pattern in CAREER_URL_PATTERNS:
            try:
                test_url = base + pattern
                response = page.goto(test_url, timeout=8000, wait_until='domcontentloaded')
                if response and response.status < 400:
                    time.sleep(1)
                    content = page.content()
                    content_lower = content.lower()
                    if any(kw in content_lower for kw in ['job', 'stelle', 'karriere', 'career', 'vacanc', 'position', 'bewerb', 'ausbildung', 'trainee']):
                        return test_url
            except:
                continue
        
        # Go back to homepage
        try:
            page.goto(base_url, timeout=PAGE_TIMEOUT, wait_until='domcontentloaded')
        except:
            pass
        
        return None
    
    def extract_jobs_from_career_page(self, career_url, employer_name):
        """Extract job listings from a career page."""
        page = None
        jobs = []
        try:
            page = self.context.new_page()
            page.goto(career_url, timeout=PAGE_TIMEOUT, wait_until='domcontentloaded')
            time.sleep(2)
            
            # Dismiss cookies
            self._dismiss_cookies(page)
            
            # Scroll to load lazy content
            for scroll_pos in [0.25, 0.5, 0.75, 1.0]:
                try:
                    page.evaluate(f'window.scrollTo(0, document.body.scrollHeight * {scroll_pos})')
                    time.sleep(0.5)
                except:
                    pass
            time.sleep(1)
            
            # Save career page HTML
            career_html = page.content()
            career_filename = sanitize_filename(employer_name) + '_career.html'
            career_filepath = os.path.join(CAREER_HTML_DIR, career_filename)
            with open(career_filepath, 'w', encoding='utf-8') as f:
                f.write(career_html)
            
            # Strategy 1: JSON-LD JobPosting
            json_ld_jobs = self._extract_json_ld(page, employer_name)
            jobs.extend(json_ld_jobs)
            
            # Strategy 2: Structured job elements
            struct_jobs = self._extract_structured_jobs(page, career_url, employer_name)
            jobs.extend(struct_jobs)
            
            # Strategy 3: Link-based extraction (IT/junior related)
            link_jobs = self._extract_link_based(page, career_url, employer_name, career_html)
            jobs.extend(link_jobs)
            
            # Strategy 4: If there's a job search/filter, try it
            search_jobs = self._try_job_search(page, career_url, employer_name)
            jobs.extend(search_jobs)
            
            page.close()
            
            # Deduplicate
            seen = set()
            unique = []
            for job in jobs:
                key = (job.get('title', '')[:100], job.get('url', '')[:200])
                if key not in seen:
                    seen.add(key)
                    unique.append(job)
            
            return unique, career_filepath
            
        except Exception as e:
            if page:
                try:
                    page.close()
                except:
                    pass
            return [], None
    
    def _extract_json_ld(self, page, employer_name):
        """Extract JSON-LD JobPosting data."""
        jobs = []
        try:
            json_ld_data = page.evaluate('''() => {
                const results = [];
                const scripts = document.querySelectorAll('script[type="application/ld+json"]');
                for (const script of scripts) {
                    try {
                        const data = JSON.parse(script.textContent);
                        const items = Array.isArray(data) ? data : [data];
                        for (const item of items) {
                            if (item['@type'] === 'JobPosting') {
                                results.push({
                                    title: item.title || '',
                                    url: item.url || '',
                                    description: item.description || '',
                                    datePosted: item.datePosted || '',
                                    employmentType: item.employmentType || '',
                                    jobLocation: item.jobLocation || '',
                                    hiringOrganization: item.hiringOrganization || {},
                                    baseSalary: item.baseSalary || {},
                                });
                            }
                            // Check for @graph
                            if (item['@graph']) {
                                for (const g of item['@graph']) {
                                    if (g['@type'] === 'JobPosting') {
                                        results.push({
                                            title: g.title || '',
                                            url: g.url || '',
                                            description: g.description || '',
                                            datePosted: g.datePosted || '',
                                            employmentType: g.employmentType || '',
                                            jobLocation: g.jobLocation || '',
                                            hiringOrganization: g.hiringOrganization || {},
                                            baseSalary: g.baseSalary || {},
                                        });
                                    }
                                }
                            }
                        }
                    } catch(e) {}
                }
                return results;
            }''')
            
            for ld_job in json_ld_data:
                if ld_job.get('title'):
                    # Extract location
                    location = ld_job.get('jobLocation', {})
                    city = ''
                    if isinstance(location, dict):
                        addr = location.get('address', {})
                        if isinstance(addr, dict):
                            city = addr.get('addressLocality', '')
                    
                    # Extract salary
                    salary = ''
                    sal_data = ld_job.get('baseSalary', {})
                    if isinstance(sal_data, dict):
                        val = sal_data.get('value', {})
                        if isinstance(val, dict):
                            salary = f"{val.get('minValue', '')}-{val.get('maxValue', '')} {sal_data.get('currency', 'EUR')}"
                    
                    jobs.append({
                        'title': ld_job['title'][:300],
                        'url': ld_job.get('url', ''),
                        'description': ld_job.get('description', '')[:5000],
                        'source': 'Website (JSON-LD)',
                        'employer_name': employer_name,
                        'date_posted': ld_job.get('datePosted', ''),
                        'employment_detail': ld_job.get('employmentType', ''),
                        'city': city,
                        'salary_range': salary,
                    })
        except:
            pass
        return jobs
    
    def _extract_structured_jobs(self, page, career_url, employer_name):
        """Extract from structured job listing elements."""
        jobs = []
        
        job_selectors = [
            '.job-listing a', '.job-list a', '.vacancy a', '.position a',
            '.job-item a', '.job-entry a', '.job-card a', '.job a',
            '[class*="job"] a', '[class*="position"] a', '[class*="vacancy"] a',
            '[class*="stelle"] a', '.listing a', '.opening a',
            'article a', '.card a', 'table tbody tr a',
            '.hiring a', '.open-position a', '.teaser a',
            '[class*="listing"] a', '[class*="teaser"] a',
            '[data-testid*="job"] a', '[data-type="job"] a',
        ]
        
        for selector in job_selectors:
            try:
                elements = page.query_selector_all(selector)
                for elem in elements:
                    href = elem.get_attribute('href')
                    title = (elem.inner_text() or '').strip()
                    if title and 5 < len(title) < 300:
                        job_url = urljoin(career_url, href) if href else career_url
                        jobs.append({
                            'title': title[:300],
                            'url': job_url,
                            'source': 'Website',
                            'employer_name': employer_name,
                        })
            except:
                continue
        
        return jobs
    
    def _extract_link_based(self, page, career_url, employer_name, page_html):
        """Extract links with IT/junior related text from HTML."""
        jobs = []
        try:
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(page_html, 'lxml')
            for link in soup.find_all('a', href=True):
                text = link.get_text(strip=True)
                href = link['href']
                if text and 8 < len(text) < 250:
                    text_lower = text.lower()
                    has_it = any(kw in text_lower for kw in IT_KEYWORDS)
                    has_junior = any(kw in text_lower for kw in JUNIOR_KEYWORDS)
                    has_job = any(kw in text_lower for kw in ['job', 'stelle', 'position', 'vacanc'])
                    
                    if (has_it or has_junior or has_job) and not any(kw in text_lower for kw in SENIOR_EXCLUDE):
                        job_url = urljoin(career_url, href)
                        if not any(j.get('url') == job_url for j in jobs):
                            jobs.append({
                                'title': text[:300],
                                'url': job_url,
                                'source': 'Website (Link)',
                                'employer_name': employer_name,
                            })
        except:
            pass
        return jobs
    
    def _try_job_search(self, page, career_url, employer_name):
        """Try to use job search/filter on the page."""
        jobs = []
        try:
            # Look for search inputs
            search_inputs = page.query_selector_all('input[type="search"], input[placeholder*="Job"], input[placeholder*="Stelle"], input[placeholder*="Such"]')
            for inp in search_inputs[:1]:
                try:
                    inp.fill('IT')
                    time.sleep(0.5)
                    # Try to submit
                    form = inp.evaluate_handle('el => el.closest("form")')
                    if form:
                        submit_btn = form.as_element().query_selector('button[type="submit"], input[type="submit"]')
                        if submit_btn:
                            submit_btn.click()
                            time.sleep(2)
                            # Extract results
                            new_jobs = self._extract_structured_jobs(page, career_url, employer_name)
                            jobs.extend(new_jobs)
                except:
                    continue
        except:
            pass
        return jobs
    
    def visit_job_detail(self, job):
        """Visit job detail page for more information."""
        job_url = job.get('url', '')
        if not job_url or job_url.endswith(('.pdf', '.png', '.jpg')):
            return {}
        
        page = None
        try:
            page = self.context.new_page()
            page.goto(job_url, timeout=JOB_DETAIL_TIMEOUT, wait_until='domcontentloaded')
            time.sleep(1.5)
            
            self._dismiss_cookies(page)
            
            content = page.content()
            
            # Save HTML
            employer_name = sanitize_filename(job.get('employer_name', 'unknown'))[:30]
            title_snippet = sanitize_filename(job.get('title', 'untitled')[:30])
            job_filename = f"{employer_name}_{title_snippet}.html"
            job_filepath = os.path.join(JOB_HTML_DIR, job_filename)
            with open(job_filepath, 'w', encoding='utf-8') as f:
                f.write(content)
            
            result = {'job_html_path': job_filepath}
            
            # Extract description
            try:
                main_content = page.query_selector(
                    'main, article, .content, .job-detail, .job-description, '
                    '[class*="job"] [class*="description"], [class*="detail"], '
                    '#content, .entry-content, .posting, .vacancy-detail'
                )
                if main_content:
                    description = (main_content.inner_text() or '').strip()
                else:
                    description = (page.evaluate('document.body.innerText') or '')[:5000]
                
                if description:
                    result['description'] = description
            except:
                pass
            
            # Try JSON-LD on detail page
            try:
                ld = page.evaluate('''() => {
                    const scripts = document.querySelectorAll('script[type="application/ld+json"]');
                    for (const script of scripts) {
                        try {
                            const data = JSON.parse(script.textContent);
                            const item = Array.isArray(data) ? data[0] : data;
                            if (item['@type'] === 'JobPosting') {
                                return JSON.stringify(item);
                            }
                        } catch(e) {}
                    }
                    return null;
                }''')
                if ld:
                    ld_data = json.loads(ld)
                    if not result.get('description'):
                        result['description'] = ld_data.get('description', '')
                    result['employment_detail'] = ld_data.get('employmentType', '')
                    result['date_posted'] = ld_data.get('datePosted', '')
                    sal = ld_data.get('baseSalary', {})
                    if isinstance(sal, dict):
                        val = sal.get('value', {})
                        if isinstance(val, dict):
                            result['salary_range'] = f"{val.get('minValue', '')}-{val.get('maxValue', '')} {sal.get('currency', 'EUR')}"
            except:
                pass
            
            return result
            
        except:
            return {}
        finally:
            if page:
                try:
                    page.close()
                except:
                    pass
    
    def scrape_employer(self, employer_data):
        """Full scraping pipeline for one employer."""
        name = employer_data.get('name', '')
        website = employer_data.get('website', '')
        city = employer_data.get('city', '')
        plz = employer_data.get('plz', '')
        industry = employer_data.get('industry', '')
        
        if not website or website.lower() in ['ja', 'nein', 'n/a', '-']:
            return []
        
        # Fix common website URL issues
        if not website.startswith('http'):
            website = 'https://' + website
        
        print(f"\n{'='*60}")
        print(f"  SCRAPER: {name[:50]}")
        print(f"  Website: {website[:60]}")
        print(f"{'='*60}")
        
        all_jobs = []
        career_url = None
        found_via = ''
        source_type = ''
        
        # Step 1: Try to find career page directly on website
        print(f"  [1] Suche Karriere-Seite auf {website[:40]}...")
        career_url, method = self.find_career_page(name, website)
        
        if career_url:
            found_via = f"Website: {website}"
            source_type = 'Website direkt'
            print(f"  [✓] Karriere gefunden ({method}): {career_url[:60]}")
        else:
            # Step 2: Search DuckDuckGo
            print(f"  [2] Suche via DuckDuckGo...")
            ddg_queries = [
                f'{name} Karriere Jobs',
                f'{name} Stellenangebote site:{urlparse(website).netloc}',
                f'{name} career jobs',
            ]
            for query in ddg_queries:
                results = self.search_duckduckgo(query, max_results=5)
                for res in results:
                    url_lower = res['url'].lower()
                    if any(kw in url_lower for kw in ['karriere', 'career', 'job', 'stelle', 'vacanc', 'recruit']):
                        # Check it's the same employer domain
                        if urlparse(website).netloc.lower().replace('www.', '') in urlparse(res['url']).netloc.lower().replace('www.', ''):
                            career_url = res['url']
                            found_via = f"DuckDuckGo: {query}"
                            source_type = 'Suchmaschine (DDG)'
                            break
                if career_url:
                    break
            
            if not career_url:
                # Step 3: Search Google
                print(f"  [3] Suche via Google...")
                google_queries = [
                    f'{name} Karriere Stellenangebote',
                    f'{name} Jobs career site:{urlparse(website).netloc}',
                ]
                for query in google_queries:
                    results = self.search_google(query, max_results=5)
                    for res in results:
                        url_lower = res['url'].lower()
                        if any(kw in url_lower for kw in ['karriere', 'career', 'job', 'stelle', 'vacanc']):
                            if urlparse(website).netloc.lower().replace('www.', '') in urlparse(res['url']).netloc.lower().replace('www.', ''):
                                career_url = res['url']
                                found_via = f"Google: {query}"
                                source_type = 'Suchmaschine (Google)'
                                break
                    if career_url:
                        break
        
        if not career_url:
            print(f"  [✗] Keine Karriere-Seite gefunden")
            return []
        
        # Step 4: Extract jobs from career page
        print(f"  [4] Extrahiere Jobs von {career_url[:50]}...")
        raw_jobs, career_html_path = self.extract_jobs_from_career_page(career_url, name)
        
        if not raw_jobs:
            print(f"  [✗] Keine Jobs gefunden")
            return []
        
        print(f"  [i] {len(raw_jobs)} Roh-Jobs extrahiert")
        
        # Step 5: Filter for junior IT jobs
        junior_it_jobs = []
        for job in raw_jobs:
            title = job.get('title', '')
            desc = job.get('description', '')
            
            is_j, junior_kw = is_junior_job(title, desc)
            is_i, it_kw = is_it_job(title, desc)
            is_sen = is_senior_job(title)
            
            if is_i and not is_sen:
                # IT job, not senior - include
                if is_j:
                    job['is_junior'] = True
                    job['junior_keyword_match'] = junior_kw
                else:
                    job['is_junior'] = True
                    job['junior_keyword_match'] = 'no_senior_keyword'
                job['is_it'] = True
                job['it_keyword_match'] = it_kw
                junior_it_jobs.append(job)
        
        # Step 6: Visit detail pages for top candidates
        visited = 0
        for job in junior_it_jobs[:8]:
            if job.get('url') and job['url'] != career_url:
                detail = self.visit_job_detail(job)
                if detail:
                    for k, v in detail.items():
                        if v and not job.get(k):
                            job[k] = v
                    visited += 1
        
        if visited:
            print(f"  [i] {visited} Detail-Seiten besucht")
        
        # Step 7: Build job records
        for job in junior_it_jobs:
            job_data = {
                'job_title': job.get('title', ''),
                'description': job.get('description', '')[:5000],
                'job_url': job.get('url', ''),
                'employer_name': name,
                'website_url': website,
                'career_url': career_url,
                'city': job.get('city', '') or city,
                'plz': job.get('plz', '') or plz,
                'street': job.get('street', ''),
                'industry': industry,
                'source': job.get('source', 'Website'),
                'source_type': source_type,
                'found_via': found_via,
                'junior_keyword_match': job.get('junior_keyword_match', ''),
                'it_keyword_match': job.get('it_keyword_match', ''),
                'career_html_path': career_html_path or '',
                'job_html_path': job.get('job_html_path', ''),
                'date_posted': job.get('date_posted', ''),
                'scrape_status': 'completed',
                'employer_id': hashlib.md5(name.encode()).hexdigest()[:12],
                'salary_range': job.get('salary_range', ''),
            }
            
            added = self.excel.add_job(job_data)
            if added:
                all_jobs.append(job_data)
                self.jobs_since_commit += 1
                print(f"    [+] {job_data['job_title'][:50]}")
                
                # Commit check
                if self.jobs_since_commit >= COMMIT_EVERY:
                    if git_commit_and_push(self.excel, self.jobs_since_commit):
                        self.jobs_since_commit = 0
        
        if junior_it_jobs:
            print(f"  [✓✓] {len(junior_it_jobs)} Junior IT-Jobs bei {name[:40]}")
        
        return all_jobs


# ============================================================
# MAIN
# ============================================================

def main():
    print("=" * 70)
    print("Junior IT-Jobs Scraper v3 - QUENLLENNAH")
    print("CloakBrowser + DuckDuckGo + Google + Website-Navigation")
    print("=" * 70)
    
    # Load employers
    print("\n[1] Lade Arbeitgeber-Daten...")
    wb_emp = openpyxl.load_workbook(EXCEL_EMPLOYERS)
    ws_emp = wb_emp['Firmen mit Website & HTML']
    
    employers = []
    for r in range(2, ws_emp.max_row + 1):
        name = ws_emp.cell(r, 2).value or ''
        website_gm = ws_emp.cell(r, 9).value or ''
        website_found = ws_emp.cell(r, 10).value or ''
        ort = ws_emp.cell(r, 7).value or ''
        plz = ws_emp.cell(r, 6).value or ''
        kategorie = ws_emp.cell(r, 4).value or ''
        
        # Prefer found website, fall back to Google Maps website
        website = website_found if website_found and website_found.lower() not in ['ja', 'nein', 'n/a', '-'] else website_gm
        
        if name:
            employers.append({
                'name': name,
                'website': website,
                'city': ort,
                'plz': plz,
                'industry': kategorie,
            })
    
    print(f"  {len(employers)} Arbeitgeber geladen")
    
    # Filter for IT-related employers first (priority)
    it_kw = ['it', 'software', 'data', 'digital', 'tech', 'system', 'informatik', 
             'computer', 'cloud', 'devops', 'cyber', 'entwickler', 'consulting',
             'telekommunikation', 'elektronik', 'automatisierung', 'engineering']
    
    it_employers = [e for e in employers if any(kw in (e['name'].lower() + ' ' + e['industry'].lower()) for kw in it_kw)]
    other_employers = [e for e in employers if e not in it_employers]
    
    # Priority: IT employers first, then larger companies, then rest
    priority_employers = it_employers + other_employers
    print(f"  {len(it_employers)} IT-Prioritäts-Arbeitgeber")
    print(f"  {len(other_employers)} weitere Arbeitgeber")
    
    # Load progress
    progress = load_progress()
    completed_names = set(progress.get('completed', []))
    failed_names = set(progress.get('failed', []))
    start_index = progress.get('last_index', 0)
    
    print(f"  Bereits verarbeitet: {len(completed_names)} erfolgreich, {len(failed_names)} fehlgeschlagen")
    
    # Initialize Excel
    print("\n[2] Initialisiere Excel...")
    excel = JuniorJobsExcel(EXCEL_OUT)
    print(f"  {excel.get_job_count()} existierende Jobs")
    
    # Start CloakBrowser
    print("\n[3] Starte CloakBrowser...")
    scraper = CloakScraper(excel)
    scraper.start_browser()
    
    # Scrape loop
    print(f"\n[4] Starte Scraping ({len(priority_employers)} Arbeitgeber)...")
    
    try:
        for idx, employer in enumerate(priority_employers):
            if idx < start_index:
                continue
            
            emp_name = employer['name']
            
            # Skip already completed
            if emp_name in completed_names:
                continue
            
            try:
                print(f"\n--- [{idx+1}/{len(priority_employers)}] {emp_name[:50]} ---")
                
                jobs = scraper.scrape_employer(employer)
                
                if jobs:
                    completed_names.add(emp_name)
                else:
                    # Not failed - just no junior IT jobs
                    completed_names.add(emp_name)
                
            except Exception as e:
                failed_names.add(emp_name)
                print(f"  [✗] Fehler: {str(e)[:80]}")
                traceback.print_exc()
                
                # Restart browser on error
                try:
                    scraper.stop_browser()
                except:
                    pass
                time.sleep(2)
                scraper.start_browser()
            
            # Update progress
            progress['completed'] = list(completed_names)
            progress['failed'] = list(failed_names)
            progress['last_index'] = idx + 1
            save_progress(progress)
            
            # Small delay between employers
            time.sleep(1)
    
    except KeyboardInterrupt:
        print("\n\n[!] Abgebrochen durch Benutzer")
    
    finally:
        scraper.stop_browser()
        
        # Final save
        excel.save()
        
        # Final commit
        if scraper.jobs_since_commit > 0:
            git_commit_and_push(excel, scraper.jobs_since_commit)
        
        print(f"\n{'='*70}")
        print(f"ERGEBNIS: {excel.get_job_count()} Junior IT-Jobs (quellennah)")
        print(f"Verarbeitet: {len(completed_names)} Arbeitgeber")
        print(f"Fehlgeschlagen: {len(failed_names)} Arbeitgeber")
        print(f"{'='*70}")


if __name__ == '__main__':
    main()
