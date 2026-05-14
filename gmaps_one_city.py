#!/usr/bin/env python3
"""Scrape one city from Google Maps with detail extraction"""
import json, os, re, sys, time, random, subprocess, signal
from datetime import datetime

CITY = sys.argv[1] if len(sys.argv) > 1 else "Erlangen"
CHROME_PATH = os.path.expanduser("~/.cache/ms-playwright/chromium-1223/chrome-linux64/chrome")
RESULTS_FILE = "/home/z/my-project/gmaps_travel_results.json"
PROGRESS_FILE = "/home/z/my-project/gmaps_detail_progress.json"
EXCEL_OUTPUT = "/home/z/my-project/download/Google_Maps_Firmen_Erlangen_50km.xlsx"
GITHUB_TOKEN = "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"
PROJECT_DIR = "/home/z/my-project"

TERMS = [
    "IT", "Software", "Ingenieurbüro", "Consulting", "Marketing",
    "Steuerberater", "Rechtsanwalt", "Architekt", "Handwerk",
    "Elektro", "Metallbau", "Maschinenbau", "Baufirma",
    "Kfz", "Immobilien", "Versicherung", "Logistik",
    "Apotheke", "Arzt", "Friseur", "Bäckerei", "Gastronomie",
]

def load_json(p, d=None):
    try:
        with open(p) as f: return json.load(f)
    except: return d if d is not None else {}

def save_json(p, d):
    with open(p, 'w') as f: json.dump(d, f, ensure_ascii=False, indent=2)

def git_push(msg):
    try:
        os.chdir(PROJECT_DIR)
        subprocess.run(["git", "add", "-A"], capture_output=True, timeout=30)
        r = subprocess.run(["git", "commit", "-m", msg], capture_output=True, timeout=30)
        if r.returncode != 0 and "nothing to commit" in r.stdout.decode(): return
        subprocess.run(["git", "push", f"https://{GITHUB_TOKEN}@github.com/BirdNest055/Heyjobs.git", "main"], capture_output=True, timeout=60)
        print(f"  [GIT] {msg}", flush=True)
    except: pass

def extract_detail(page, town, term):
    r = {"name":"","rating":"","review_count":"","category":"","address":"","plz":"","city":"","phone":"","website":"","lat":"","lon":"","gmaps_url":"","search_town":town,"search_term":term,"scraped_at":datetime.now().isoformat(),"source":"google_maps_detail"}
    try: page.wait_for_selector('[role="main"]', timeout=6000)
    except: pass
    time.sleep(0.3)
    try:
        e = page.query_selector('h1.DUwDvf, h1.fontHeadlineLarge')
        if e: r["name"] = e.inner_text().strip()
    except: pass
    try:
        e = page.query_selector('div.F7nice')
        if e:
            s = e.query_selector('span[aria-label]')
            if s:
                a = s.get_attribute('aria-label') or ''
                m = re.search(r'([\d,]+)\s*Sterne', a)
                if m: r["rating"] = m.group(1).replace(',','.')
    except: pass
    try:
        e = page.query_selector('button[aria-label*="Bewertungen"]')
        if e:
            a = e.get_attribute('aria-label') or ''
            m = re.search(r'([\d.]+)\s*Bewertungen', a)
            if m: r["review_count"] = m.group(1)
    except: pass
    try:
        e = page.query_selector('button[jsaction*="category"]')
        if e: r["category"] = e.inner_text().strip()
    except: pass
    try:
        e = page.query_selector('a[data-item-id="authority"]')
        if e:
            h = e.get_attribute('href') or ''
            if h and 'google' not in h.lower(): r["website"] = h
    except: pass
    if not r["website"]:
        try:
            e = page.query_selector('a[aria-label*="Website"]')
            if e:
                h = e.get_attribute('href') or ''
                if h and 'google' not in h.lower() and 'gstatic' not in h.lower() and not h.startswith('tel:'): r["website"] = h
        except: pass
    try:
        for e in page.query_selector_all('[data-item-id]'):
            iid = e.get_attribute('data-item-id') or ''
            if 'address' in iid and not r["address"]:
                t = re.sub(r'[\ue000-\uefff]','',e.inner_text()).strip()
                t = re.sub(r'\s+',' ',t).strip()
                r["address"] = t
                m = re.search(r'(\d{5})',t)
                if m: r["plz"] = m.group(1)
                m = re.search(r'\d{5}\s+([A-Za-zäöüßÄÖÜ\-\.]+(?:\s+[A-Za-zäöüßÄÖÜ\-\.]+)*)',t)
                if m: r["city"] = m.group(1).strip()
            elif 'phone:tel:' in iid and not r["phone"]:
                m = re.search(r'phone:tel:(.+)',iid)
                if m: r["phone"] = m.group(1)
    except: pass
    try:
        u = page.url
        m = re.search(r'@(-?[\d.]+),(-?[\d.]+)',u)
        if m: r["lat"],r["lon"] = m.group(1),m.group(2)
        r["gmaps_url"] = u
    except: pass
    return r

def create_excel(companies):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import CellIsRule
    except: return
    wb = Workbook()
    ws = wb.active
    ws.title = "Google Maps Firmen"
    headers = ["#","Firmenname","Bewertung","Anzahl Bewertungen","Kategorie","Adresse","PLZ","Ort","Telefon","Website","Breitengrad","Längengrad","Google Maps Link","Suchort","Suchbegriff","Quelle","Scraped At"]
    hf = Font(bold=True,color="FFFFFF",size=11)
    hfill = PatternFill(start_color="2F5496",end_color="2F5496",fill_type="solid")
    tb = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
    for col,h in enumerate(headers,1):
        c = ws.cell(row=1,column=col,value=h); c.font,hf; c.fill,hfill; c.border,tb
    for idx,comp in enumerate(companies,1):
        rd = [idx,comp.get("name",""),comp.get("rating",""),comp.get("review_count",""),comp.get("category",""),comp.get("address",""),comp.get("plz",""),comp.get("city",""),comp.get("phone",""),comp.get("website",""),comp.get("lat",""),comp.get("lon",""),comp.get("gmaps_url",""),comp.get("search_town",""),comp.get("search_term",""),comp.get("source",""),comp.get("scraped_at","")]
        for col,v in enumerate(rd,1):
            ws.cell(row=idx+1,column=col,value=v).border = tb
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(companies)+1}"
    for i,w in enumerate([5,40,10,12,30,45,8,25,22,45,12,12,55,25,20,22,22]):
        ws.column_dimensions[get_column_letter(i+1)].width = w
    gf = PatternFill(start_color="C6EFCE",end_color="C6EFCE",fill_type="solid")
    ws.conditional_formatting.add(f"J2:J{len(companies)+1}",CellIsRule(operator="notEqual",formula=['""'],fill=gf))
    ws2 = wb.create_sheet("Statistiken")
    for r,(s,v) in enumerate([("Gesamtzahl",len(companies)),("Mit Website",sum(1 for c in companies if c.get("website"))),("Ohne Website",sum(1 for c in companies if not c.get("website"))),("Mit Telefon",sum(1 for c in companies if c.get("phone"))),("Mit PLZ",sum(1 for c in companies if c.get("plz"))),("Mit Kategorie",sum(1 for c in companies if c.get("category"))),("Detail-basiert",sum(1 for c in companies if c.get("source")=="google_maps_detail"))],1):
        ws2.cell(row=r,column=1,value=s).font=Font(bold=True); ws2.cell(row=r,column=2,value=v)
    ws2.column_dimensions['A'].width=25
    wb.save(EXCEL_OUTPUT)
    print(f"  [EXCEL] {len(companies)} saved",flush=True)

def main():
    print(f"=== Scraping {CITY} ===",flush=True)
    subprocess.run(["pkill","-9","Xvfb"],capture_output=True)
    subprocess.run(["rm","-f","/tmp/.X99-lock"],capture_output=True)
    subprocess.run(["rm","-f","/tmp/.X11-unix/X99"],capture_output=True)
    subprocess.run(["pkill","-9","chromium"],capture_output=True)
    time.sleep(2)

    progress = load_json(PROGRESS_FILE,{"completed_searches":[],"cities_done":[]})
    results = load_json(RESULTS_FILE,[])
    all_comp = {}
    for c in results:
        k = f"{c.get('name','').lower().strip()}|{c.get('city',c.get('search_town','')).lower().strip()}"
        if k not in all_comp: all_comp[k] = c
        else:
            for kk,vv in c.items():
                if vv and not all_comp[k].get(kk): all_comp[k][kk] = vv

    print(f"Loaded {len(all_comp)} companies",flush=True)

    xvfb = subprocess.Popen(['Xvfb',':99','-screen','0','1920x1080x24','-ac','+extension','GLX','+render','-noreset'],stdout=subprocess.DEVNULL,stderr=subprocess.DEVNULL)
    time.sleep(3)
    os.environ['DISPLAY'] = ':99'
    if xvfb.poll() is not None:
        print("FATAL: Xvfb failed!",flush=True); sys.exit(1)

    from playwright.sync_api import sync_playwright
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False,executable_path=CHROME_PATH,args=['--no-sandbox','--disable-gpu','--disable-dev-shm-usage','--disable-blink-features=AutomationControlled','--lang=de-DE'])
        ctx = browser.new_context(viewport={"width":1920,"height":1080},locale="de-DE",user_agent="Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36")
        page = ctx.new_page()
        print("Browser ready",flush=True)

        push_ctr = 0
        for term in TERMS:
            skey = f"{CITY}|{term}"
            if skey in progress.get("completed_searches",[]): continue

            query = f"{CITY} {term}"
            url = f"https://www.google.com/maps/search/{query.replace(' ','+')}"
            try:
                page.goto(url,timeout=20000,wait_until="domcontentloaded")
                time.sleep(random.uniform(3,5))
            except:
                print(f"  [WARN] Cannot load: {query}",flush=True)
                progress.setdefault("completed_searches",[]).append(skey)
                save_json(PROGRESS_FILE,progress)
                continue

            # Cookie consent
            try:
                a = page.query_selector('button[aria-label*="Alle ablehnen"],button[aria-label*="Reject"]')
                if a: a.click(); time.sleep(1)
            except: pass

            # Scroll to load all
            try:
                sp = page.query_selector('div[role="feed"]')
                if sp:
                    lc,sr = 0,0
                    for i in range(15):
                        page.evaluate('(el) => el.scrollTop = el.scrollHeight',sp)
                        time.sleep(random.uniform(0.8,1.3))
                        items = page.query_selector_all('a[href*="/maps/place/"]')
                        if len(items)==lc:
                            sr+=1
                            if sr>=3: break
                        else: sr=0
                        lc = len(items)
            except: pass

            # Collect unique links
            seen_h = set()
            uitems = []
            for item in page.query_selector_all('a[href*="/maps/place/"]'):
                h = item.get_attribute('href') or ''
                if h and h not in seen_h:
                    seen_h.add(h)
                    uitems.append((item,h))

            print(f"  [{query}] {len(uitems)} results",flush=True)

            # Click into each result
            for idx,(item,href) in enumerate(uitems):
                try:
                    name = ""
                    try:
                        ne = item.query_selector('.fontHeadlineSmall,.qBF1Pd')
                        if ne: name = ne.inner_text().strip()
                    except: pass
                    if not name:
                        try: name = item.get_attribute('aria-label') or ''
                        except: pass
                    if not name: continue

                    item.click()
                    time.sleep(random.uniform(1.5,2.5))

                    detail = extract_detail(page,CITY,term)
                    if not detail.get("name"): detail["name"] = name

                    # Add/update
                    key = f"{detail['name'].lower().strip()}|{detail.get('city',CITY).lower().strip()}"
                    if key not in all_comp:
                        all_comp[key] = detail
                        push_ctr += 1
                    else:
                        ec = all_comp[key]
                        updated = False
                        for k,v in detail.items():
                            if v and not ec.get(k):
                                ec[k] = v
                                updated = True
                        if updated:
                            ec["source"] = "google_maps_detail"
                            ec["enriched_at"] = datetime.now().isoformat()
                            push_ctr += 1

                    w = detail.get('website','')[:30]
                    print(f"    [{idx+1}/{len(uitems)}] {detail['name'][:35]} | web={w} | plz={detail.get('plz','-')}",flush=True)

                    try:
                        page.go_back(timeout=8000)
                        time.sleep(random.uniform(1,2))
                    except:
                        page.goto(url,timeout=15000)
                        time.sleep(2)

                except Exception as e:
                    print(f"    [{idx+1}] Err: {str(e)[:40]}",flush=True)
                    try:
                        page.goto(url,timeout=15000)
                        time.sleep(2)
                    except: pass

            progress.setdefault("completed_searches",[]).append(skey)
            save_json(RESULTS_FILE,list(all_comp.values()))
            save_json(PROGRESS_FILE,progress)

            if push_ctr >= 10:
                create_excel(list(all_comp.values()))
                git_push(f"GMaps: {CITY} {term} | +{push_ctr} | {len(all_comp)} total")
                push_ctr = 0

            time.sleep(random.uniform(1,2))

        progress.setdefault("cities_done",[]).append(CITY)
        save_json(PROGRESS_FILE,progress)
        create_excel(list(all_comp.values()))
        git_push(f"GMaps: {CITY} DONE | {len(all_comp)} total")

        browser.close()

    try: os.kill(xvfb.pid,signal.SIGTERM)
    except: pass
    print(f"=== {CITY} complete | {len(all_comp)} total ===",flush=True)

if __name__ == "__main__":
    main()
