#!/usr/bin/env python3
"""
Phase 2: Enrich IT employers with websites and HTML.
Uses very short timeouts and skips companies likely to fail.
Processes in batches of 10, commits after each batch.
"""
import json, os, re, sys, time, hashlib, sqlite3, subprocess
from datetime import datetime
import requests
from bs4 import BeautifulSoup

sys.stdout.reconfigure(line_buffering=True)

DB_PATH = "/home/z/my-project/download/it_employers.db"
HTML_DIR = "/home/z/my-project/download/it_employer_html"
GIT_DIR = "/home/z/my-project"
COMMIT_EVERY = 10

SESSION = requests.Session()
SESSION.headers.update({"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"})
SESSION.verify = False
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

def sanitize_filename(name, max_len=80):
    name = re.sub(r'[äöüßÄÖÜ]', lambda m: {'ä':'ae','ö':'oe','ü':'ue','ß':'ss','Ä':'Ae','Ö':'Oe','Ü':'Ue'}[m.group()], name)
    name = re.sub(r'[^a-zA-Z0-9._-]', '_', name)
    name = re.sub(r'_+', '_', name).strip('_')
    return name[:max_len]

def make_abs(base, rel):
    if not rel: return None
    if rel.startswith('http'): return rel
    if rel.startswith('//'): return 'https:' + rel
    if rel.startswith('/'):
        from urllib.parse import urlparse
        p = urlparse(base)
        return f"{p.scheme}://{p.netloc}{rel}"
    return f"{base.rstrip('/')}/{rel.lstrip('/')}"

def find_website(name):
    """Try to find website by constructing likely domain names."""
    clean = re.sub(r'(GmbH|AG|KG|OHG|UG|e\.K\.|e\.V\.|mbH|Co\.\s*KG|& Co\.|GmbH & Co\. KG|und Firma|Inh\.|e\.Kfr\.|GbR)', '', name, flags=re.IGNORECASE).strip()
    # Remove common words that wouldn't be in domain
    for word in ['und', 'the', 'de', 'am', 'im', 'an', 'auf', 'der', 'die', 'das', 'von', 'zu', 'bei']:
        clean = re.sub(r'\b' + word + r'\b', '', clean, flags=re.IGNORECASE).strip()
    domain = re.sub(r'[äöüß]', lambda m: {'ä':'ae','ö':'oe','ü':'ue','ß':'ss'}[m.group()], clean)
    domain = re.sub(r'[^a-zA-Z0-9]', '', domain).lower()
    if len(domain) < 3 or len(domain) > 40: return None
    
    for url in [f"https://www.{domain}.de", f"https://{domain}.de"]:
        try:
            r = SESSION.get(url, timeout=4, allow_redirects=True)
            if r.status_code == 200 and 'text/html' in r.headers.get('content-type', ''):
                return url
        except: pass
    return None

def download_html(url, timeout=6):
    try:
        r = SESSION.get(url, timeout=timeout, allow_redirects=True)
        if r.status_code == 200 and 'text/' in r.headers.get('content-type', ''):
            return r.text
    except: pass
    return None

def save_html(content, name, suffix=""):
    if not content: return None, None
    fn = f"{sanitize_filename(name)}_{suffix}.html" if suffix else f"{sanitize_filename(name)}.html"
    fp = os.path.join(HTML_DIR, fn)
    os.makedirs(HTML_DIR, exist_ok=True)
    with open(fp, 'w', encoding='utf-8', errors='replace') as f: f.write(content)
    return fp, hashlib.sha256(content.encode()).hexdigest()[:16]

def extract_info(html):
    info = {}; imp_url = None; kar_url = None
    if not html: return info, None, None
    emails = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', html)
    emails = [e for e in emails if not any(x in e.lower() for x in ['example.com','sentry','webpack','noreply','wixpress','github'])]
    emails = list(dict.fromkeys(emails))[:5]
    if emails: info['email'] = emails[0]
    gf = re.search(r'(?:Geschäftsführer)[\s:]*([^<\.]{2,100})', html, re.IGNORECASE)
    if gf:
        t = BeautifulSoup(gf.group(1), 'html.parser').get_text(strip=True)
        if len(t)>2: info['geschaeftsfuehrer'] = t[:200]
    hr = re.search(r'(?:Handelsregister|HRB)[\s:]*([^<\.]{2,100})', html, re.IGNORECASE)
    if hr:
        t = BeautifulSoup(hr.group(1), 'html.parser').get_text(strip=True)
        info['handelsregister'] = t[:200]
    soup = BeautifulSoup(html, 'html.parser')
    for a in soup.find_all('a', href=True):
        h = a['href'].lower(); t = a.get_text(strip=True).lower()
        if not imp_url and ('impressum' in h or 'impressum' in t): imp_url = a['href']
        if not kar_url and any(w in h for w in ['karriere','career','jobs','stellenangebote','bewerbung']): kar_url = a['href']
    return info, imp_url, kar_url

def main():
    print("=== Website Enrichment ===")
    conn = sqlite3.connect(DB_PATH)
    
    # Get employers that need website enrichment
    rows = conn.execute("""
        SELECT id, firmenname, ort, hat_it_jobs, it_relevanz_score 
        FROM employers 
        WHERE website_url = '' AND scrape_status = 'phase1_complete'
        ORDER BY it_relevanz_score DESC, hat_it_jobs DESC
    """).fetchall()
    
    print(f"Arbeitgeber ohne Website: {len(rows)}")
    
    total = len(rows)
    processed = 0; found = 0; batch_commit = 0
    
    for row in rows:
        emp_id, name, ort, has_it, score = row
        
        ws = find_website(name)
        updates = {'website_url': ws or '', 'scrape_status': 'enriched'}
        
        if ws:
            found += 1
            html = download_html(ws)
            if html:
                path, h = save_html(html, name)
                updates['website_html_path'] = path
                updates['website_html_hash'] = h
                
                contact, imp_url, kar_url = extract_info(html)
                if contact.get('email'): updates['email'] = contact['email']
                if contact.get('geschaeftsfuehrer'): updates['geschaeftsfuehrer'] = contact['geschaeftsfuehrer']
                if contact.get('handelsregister'): updates['handelsregister'] = contact['handelsregister']
                
                if imp_url:
                    imp_abs = make_abs(ws, imp_url)
                    if imp_abs:
                        updates['impressum_url'] = imp_abs
                        imp_html = download_html(imp_abs, timeout=5)
                        if imp_html:
                            p2, _ = save_html(imp_html, name, "impressum")
                            updates['impressum_html_path'] = p2
                            ic, _, _ = extract_info(imp_html)
                            if ic.get('email') and not updates.get('email'): updates['email'] = ic['email']
                            if ic.get('geschaeftsfuehrer') and not updates.get('geschaeftsfuehrer'): updates['geschaeftsfuehrer'] = ic['geschaeftsfuehrer']
                            if ic.get('handelsregister') and not updates.get('handelsregister'): updates['handelsregister'] = ic['handelsregister']
                
                if kar_url:
                    kar_abs = make_abs(ws, kar_url)
                    if kar_abs:
                        updates['karriere_url'] = kar_abs
                        kar_html = download_html(kar_abs, timeout=5)
                        if kar_html:
                            p3, _ = save_html(kar_html, name, "karriere")
                            updates['karriere_html_path'] = p3
        
        # Update DB
        set_clause = ", ".join(f"{k} = ?" for k in updates.keys())
        conn.execute(f"UPDATE employers SET {set_clause}, updated_at = CURRENT_TIMESTAMP WHERE id = ?", 
                     list(updates.values()) + [emp_id])
        conn.commit()
        
        processed += 1; batch_commit += 1
        status = f"✓{ws[:35]}" if ws else "  -"
        it_flag = "IT" if has_it else "  "
        print(f"  [{processed}/{total}] {it_flag} Sc:{score:5.1f} | {name[:40]:40} | {status}")
        
        if batch_commit >= COMMIT_EVERY:
            # Rebuild Excel from DB
            rebuild_excel(conn)
            
            # Git commit & push
            try:
                subprocess.run(['git','add','-A'], cwd=GIT_DIR, capture_output=True, timeout=30)
                msg = f"Website-Enrichment: {processed}/{total} | {found} Websites ({datetime.now().strftime('%H:%M')})"
                r = subprocess.run(['git','commit','-m',msg], cwd=GIT_DIR, capture_output=True, timeout=30)
                if r.returncode == 0:
                    subprocess.run(['git','push','origin','main'], cwd=GIT_DIR, capture_output=True, timeout=60)
                    print(f"  >>> GIT PUSH ({found} Websites gefunden) <<<")
            except Exception as e: print(f"  GIT: {e}")
            batch_commit = 0
    
    # Final rebuild & commit
    rebuild_excel(conn)
    try:
        subprocess.run(['git','add','-A'], cwd=GIT_DIR, capture_output=True, timeout=30)
        subprocess.run(['git','commit','-m',f"Website-Enrichment FERTIG: {found}/{total}"], cwd=GIT_DIR, capture_output=True, timeout=30)
        subprocess.run(['git','push','origin','main'], cwd=GIT_DIR, capture_output=True, timeout=60)
    except: pass
    
    conn.close()
    print(f"\nFERTIG! {found}/{total} Websites gefunden")

def rebuild_excel(conn):
    """Rebuild Excel from current DB state."""
    import openpyxl
    EXCEL_OUT = "/home/z/my-project/download/IT_Arbeitgeber_Bamberg_Erlangen_Nuernberg.xlsx"
    
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "IT Arbeitgeber"
    
    cursor = conn.execute("SELECT * FROM employers ORDER BY it_relevanz_score DESC, hat_it_jobs DESC, firmenname")
    columns = [d[0] for d in cursor.description]
    hm = {'id':'ID','firmenname':'Firmenname','firmenname_normalisiert':'Firmenname (norm.)',
          'ort':'Ort','plz':'PLZ','strasse':'Straße','telefon':'Telefon','email':'E-Mail',
          'website_url':'Website URL','website_html_path':'HTML Pfad','website_html_hash':'HTML Hash',
          'karriere_url':'Karriere-URL','karriere_html_path':'Karriere HTML',
          'impressum_url':'Impressum URL','impressum_html_path':'Impressum HTML',
          'hat_it_jobs':'IT-Jobs','it_job_titel':'IT-Job Titel','it_job_anzahl':'Anzahl IT-Jobs',
          'it_job_quelle':'IT-Job Quelle','it_job_details':'IT-Job Details',
          'it_keywords_gefunden':'IT-Keywords','arbeitsagentur_match':'AA Match',
          'kategorie_google':'Kategorie (Google)','bewertung_google':'Bewertung',
          'lat':'Breitengrad','lng':'Längengrad','google_maps_link':'Google Maps',
          'handelsregister':'Handelsregister','geschaeftsfuehrer':'Geschäftsführer',
          'it_abteilung_vermutung':'IT-Abt. vermutet','it_relevanz_score':'IT-Relevanz Score',
          'scrape_datum':'Scrape-Datum','scrape_status':'Scrape-Status',
          'letzte_pruefung':'Letzte Prüfung','created_at':'Erstellt','updated_at':'Aktualisiert'}
    ws1.append([hm.get(c,c) for c in columns])
    for row in cursor: ws1.append(list(row))
    for col in ws1.columns:
        ml = max(len(str(cell.value or '')) for cell in col)
        ws1.column_dimensions[col[0].column_letter].width = min(ml+2, 60)
    
    # Stats
    ws2 = wb.create_sheet("Statistiken")
    ws2.append(["Metrik","Wert"])
    ws2.append(["Gesamt",conn.execute("SELECT COUNT(*) FROM employers").fetchone()[0]])
    ws2.append(["IT-Jobs bestätigt",conn.execute("SELECT COUNT(*) FROM employers WHERE hat_it_jobs=1").fetchone()[0]])
    ws2.append(["Mit Website",conn.execute("SELECT COUNT(*) FROM employers WHERE website_url != ''").fetchone()[0]])
    ws2.append(["Mit E-Mail",conn.execute("SELECT COUNT(*) FROM employers WHERE email != ''").fetchone()[0]])
    ws2.append(["IT-Abt. vermutet",conn.execute("SELECT COUNT(*) FROM employers WHERE it_abteilung_vermutung=1").fetchone()[0]])
    
    # Per city
    ws3 = wb.create_sheet("Nach Ort")
    ws3.append(["Ort","Gesamt","IT-Jobs","Website","E-Mail","Ø Score"])
    for row in conn.execute("""SELECT ort,COUNT(*),SUM(CASE WHEN hat_it_jobs=1 THEN 1 ELSE 0 END),
        SUM(CASE WHEN website_url != '' THEN 1 ELSE 0 END),SUM(CASE WHEN email != '' THEN 1 ELSE 0 END),
        ROUND(AVG(it_relevanz_score),1) FROM employers GROUP BY ort ORDER BY COUNT(*) DESC"""):
        ws3.append(list(row))
    
    # IT jobs only
    ws4 = wb.create_sheet("Nur IT-Jobs bestätigt")
    cursor2 = conn.execute("SELECT * FROM employers WHERE hat_it_jobs=1 ORDER BY it_job_anzahl DESC, it_relevanz_score DESC")
    ws4.append([hm.get(c,c) for c in [d[0] for d in cursor2.description]])
    for row in cursor2: ws4.append(list(row))
    
    wb.save(EXCEL_OUT)

if __name__ == '__main__':
    main()
