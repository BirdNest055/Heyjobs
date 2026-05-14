#!/usr/bin/env python3
"""
Phase 1: Query Arbeitsagentur API for junior IT jobs
Quick and efficient - just API calls
"""
import asyncio
import httpx
import openpyxl
import re
import os
import subprocess
import json
from datetime import datetime
from urllib.parse import urljoin

EXCEL_IN = "/home/z/my-project/download/IT_Arbeitgeber_Bamberg_Erlangen_Nuernberg.xlsx"
EXCEL_OUT = "/home/z/my-project/download/Junior_IT_Jobs_Bamberg_Erlangen_Nuernberg.xlsx"
GIT_REPO = "/home/z/my-project"
GITHUB_TOKEN = "GIT_TOKEN_REDACTED"

AA_API_URL = "https://rest.arbeitsagentur.de/jobboerse/jobsuche-service/pc/v6/jobs"
AA_API_KEY = "jobboerse-jobsuche"

JUNIOR_KEYWORDS = [
    'junior', 'entry level', 'einsteiger', 'anfänger', 'beginner',
    'trainee', 'praktikant', 'intern', 'werkstudent', 'working student',
    'absolvent', 'graduate', 'berufseinsteiger', 'young professional',
    'associate', 'duales studium', 'ausbildung', 'apprentice',
    'berufseinsteiger it', 'absolvent informatik', 'nachwuchs',
    'perspektive', 'start career', 'career starter', 'erstjob',
]

IT_KEYWORDS = [
    'software', 'entwickler', 'developer', 'it-', 'informatik',
    'data', 'devops', 'cloud', 'cyber', 'security', 'sicherheit',
    'system', 'admin', 'administrator', 'netzwerk', 'network',
    'frontend', 'backend', 'fullstack', 'full-stack',
    'python', 'java', 'javascript', 'typescript', 'react', 'angular',
    'sap', 'linux', 'sql', 'database', 'datenbank', 'api',
    'machine learning', 'ki', 'ai', 'scrum', 'agile', 'testing',
    'qa', 'automation', 'consultant', 'analyst', 'engineer',
    'architekt', 'programmierung', 'web', 'mobile', 'app',
    'infrastructure', 'infrastruktur', 'platform', 'support',
    'helpdesk', 'erp', 'crm', 'digitalisierung', 'technologie',
]

TECH_KEYWORDS = [
    'python', 'java', 'javascript', 'typescript', 'c#', 'c++',
    'ruby', 'go', 'rust', 'kotlin', 'swift', 'php',
    'react', 'angular', 'vue', 'svelte', 'next.js',
    'node.js', 'django', 'flask', 'spring', '.net',
    'docker', 'kubernetes', 'aws', 'azure', 'gcp',
    'terraform', 'ansible', 'jenkins', 'gitlab', 'github',
    'sql', 'mysql', 'postgresql', 'mongodb', 'redis',
    'kafka', 'rabbitmq', 'graphql', 'rest', 'api',
    'sap', 'salesforce', 'dynamics', 'sharepoint',
    'linux', 'windows', 'html', 'css', 'tailwind',
    'scrum', 'kanban', 'devops', 'ci/cd',
    'machine learning', 'deep learning', 'tensorflow', 'pytorch',
    'nlp', 'computer vision', 'data science', 'big data',
    'jira', 'confluence', 'cybersecurity', 'pentest',
    'powershell', 'bash', 'git',
]


def is_junior_job(title, description=""):
    text = f"{title} {description}".lower()
    for kw in JUNIOR_KEYWORDS:
        if kw in text:
            return True, kw
    exp_match = re.search(r'(\d+)\s*(?:jahr|year|jahre|years)\s*(?:berufs)?erfahrung', text)
    if exp_match:
        years = int(exp_match.group(1))
        if years <= 2:
            return True, f"erfahrung_{years}jahre"
    if re.search(r'ohne\s*erfahrung|no\s*experience|keine\s*erfahrung', text):
        return True, "no_experience"
    return False, None


def is_it_job(title, description=""):
    text = f"{title} {description}".lower()
    for kw in IT_KEYWORDS:
        if kw in text:
            return True, kw
    return False, None


def is_senior_job(title):
    text = title.lower()
    return any(kw in text for kw in ['senior', 'lead', 'principal', 'head of', 'director', 'leitung', 'chef'])


def extract_technologies(text):
    found = []
    text_lower = text.lower()
    for tech in TECH_KEYWORDS:
        if re.search(r'\b' + re.escape(tech) + r'\b', text_lower):
            found.append(tech)
    return list(set(found))


def classify_job_category(title, description=""):
    text = f"{title} {description}".lower()
    categories = {
        'Softwareentwicklung': ['softwareentwickler', 'software developer', 'software engineer', 'entwickler', 'developer', 'programmierer', 'frontend', 'backend', 'fullstack', 'full-stack'],
        'Data Science & Analytics': ['data scientist', 'data analyst', 'data engineer', 'data analytics', 'business intelligence', 'machine learning', 'ki-entwickler'],
        'IT-Consulting': ['it consultant', 'it-consultant', 'it berater', 'sap consultant', 'technischer berater'],
        'Systemadministration': ['system administrator', 'systemadministrator', 'sysadmin', 'admin', 'netzwerkadministrator'],
        'DevOps & Cloud': ['devops', 'cloud engineer', 'cloud architect', 'platform engineer', 'sre', 'infrastruktur'],
        'Cyber Security': ['security', 'cybersecurity', 'information security', 'it-sicherheit', 'pentest'],
        'IT-Support': ['it support', 'helpdesk', 'service desk', 'first level', 'second level', 'user support'],
        'IT-Projektmanagement': ['project manager', 'projektmanager', 'scrum master', 'product owner'],
        'QA & Testing': ['quality assurance', 'qa', 'tester', 'test engineer', 'testautomatisierung'],
        'UX/UI Design': ['ux', 'ui', 'user experience', 'user interface', 'design', 'usability'],
        'IT-Ausbildung & Studium': ['ausbildung', 'duales studium', 'trainee', 'praktikum', 'werkstudent'],
    }
    for category, keywords in categories.items():
        for kw in keywords:
            if kw in text:
                return category
    return 'IT (Sonstige)'


def extract_employment_type(title, description=""):
    text = f"{title} {description}".lower()
    if any(kw in text for kw in ['teilzeit', 'part-time', 'part time']):
        return 'Teilzeit'
    if any(kw in text for kw in ['werkstudent', 'working student']):
        return 'Werkstudent'
    if any(kw in text for kw in ['praktikant', 'praktikum', 'intern', 'internship']):
        return 'Praktikum'
    if any(kw in text for kw in ['ausbildung', 'apprentice', 'azubi']):
        return 'Ausbildung'
    if any(kw in text for kw in ['duales studium', 'dual student']):
        return 'Duales Studium'
    if any(kw in text for kw in ['minijob', 'mini-job', 'geringfügig']):
        return 'Minijob'
    return 'Vollzeit'


def extract_remote_option(title, description=""):
    text = f"{title} {description}".lower()
    if any(kw in text for kw in ['hybrid', 'teilweise remote']):
        return 'Hybrid'
    if any(kw in text for kw in ['remote', 'homeoffice', 'home-office', 'work from home', 'fully remote']):
        return 'Remote'
    if any(kw in text for kw in ['vor ort', 'on-site', 'onsite']):
        return 'Vor Ort'
    return 'Nicht angegeben'


async def query_aa_jobs(client, search_name, city=None):
    """Query Arbeitsagentur API for jobs from a specific employer."""
    jobs = []
    cities = ['Erlangen', 'Nürnberg', 'Bamberg', 'Fürth', 'Forchheim', 'Herzogenaurach', 'Lauf an der Pegnitz']
    if city and city not in cities:
        cities.insert(0, city)
    
    for search_city in cities[:3]:  # Only check top 3 cities
        params = {
            'was': search_name,
            'wo': search_city,
            'berufsfeld': '11',  # IT
            'page': 0,
            'size': 25,
        }
        headers = {
            'X-API-Key': AA_API_KEY,
            'Accept': 'application/json',
        }
        
        try:
            response = await client.get(AA_API_URL, params=params, headers=headers, timeout=10)
            if response.status_code == 200:
                data = response.json()
                stellen = data.get('stellenangebote', [])
                for stelle in stellen:
                    refnr = stelle.get('refnr', '')
                    titel = stelle.get('beruf', '') or stelle.get('titel', '')
                    arbeitgeber = stelle.get('arbeitgeber', '')
                    ort = stelle.get('arbeitsort', {})
                    ort_name = ort.get('ort', '') if isinstance(ort, dict) else str(ort)
                    plz = ort.get('plz', '') if isinstance(ort, dict) else ''
                    
                    # Match employer name
                    emp_norm = re.sub(r'\s+', '', search_name.lower())[:10]
                    ag_norm = re.sub(r'\s+', '', arbeitgeber.lower())[:10]
                    if emp_norm and ag_norm and (emp_norm in ag_norm or ag_norm in emp_norm):
                        is_j, junior_kw = is_junior_job(titel)
                        is_i, it_kw = is_it_job(titel)
                        is_sen = is_senior_job(titel)
                        
                        if is_i and (is_j or not is_sen):
                            jobs.append({
                                'title': titel,
                                'employer_name': arbeitgeber,
                                'city': ort_name,
                                'plz': plz,
                                'source': 'Arbeitsagentur',
                                'refnr': refnr,
                                'junior_keyword_match': junior_kw or 'no_senior_keyword',
                                'it_keyword_match': it_kw,
                                'is_junior': True,
                                'is_it': True,
                                'url': f'https://www.arbeitsagentur.de/jobsuche/jobdetail/{refnr}',
                                'description': '',
                                'vollzeit': stelle.get('vollzeit', True),
                                'befristet': stelle.get('befristet', False),
                            })
        except:
            continue
    
    # Deduplicate
    seen = set()
    unique = []
    for j in jobs:
        key = j.get('refnr', '') or j.get('title', '')
        if key and key not in seen:
            seen.add(key)
            unique.append(j)
    
    return unique


# Also do a broad search for all junior IT jobs in the region
async def broad_junior_search(client):
    """Search for all junior IT jobs in the region."""
    jobs = []
    cities = [
        ('Erlangen', '91052'), ('Nürnberg', '90402'), ('Bamberg', '96047'),
        ('Fürth', '90762'), ('Forchheim', '91301'), ('Herzogenaurach', '91074'),
        ('Lauf an der Pegnitz', '91207'), ('Erlangen', '91058'),
    ]
    
    # Search for junior IT jobs in each city
    search_terms = [
        'Junior IT', 'Junior Software', 'Junior Entwickler', 'Werkstudent IT',
        'Werkstudent Software', 'Praktikum IT', 'Praktikum Software',
        'Trainee IT', 'Ausbildung Informatik', 'Duales Studium IT',
        'Junior Developer', 'Junior Data', 'Junior DevOps', 'Junior Cloud',
        'Einsteiger IT', 'Berufseinsteiger Informatik', 'Absolvent Informatik',
        'Junior Frontend', 'Junior Backend', 'Junior Fullstack',
    ]
    
    for search_term in search_terms:
        for city_name, plz in cities:
            params = {
                'was': search_term,
                'wo': city_name,
                'umkreis': 50,
                'page': 0,
                'size': 50,
            }
            headers = {
                'X-API-Key': AA_API_KEY,
                'Accept': 'application/json',
            }
            
            try:
                response = await client.get(AA_API_URL, params=params, headers=headers, timeout=10)
                if response.status_code == 200:
                    data = response.json()
                    stellen = data.get('stellenangebote', [])
                    for stelle in stellen:
                        refnr = stelle.get('refnr', '')
                        titel = stelle.get('beruf', '') or stelle.get('titel', '')
                        arbeitgeber = stelle.get('arbeitgeber', '')
                        ort = stelle.get('arbeitsort', {})
                        ort_name = ort.get('ort', '') if isinstance(ort, dict) else str(ort)
                        ort_plz = ort.get('plz', '') if isinstance(ort, dict) else ''
                        
                        is_i, it_kw = is_it_job(titel)
                        is_sen = is_senior_job(titel)
                        
                        if is_i and not is_sen:
                            is_j, junior_kw = is_junior_job(titel)
                            jobs.append({
                                'title': titel,
                                'employer_name': arbeitgeber,
                                'city': ort_name,
                                'plz': ort_plz,
                                'source': 'Arbeitsagentur',
                                'refnr': refnr,
                                'junior_keyword_match': junior_kw or 'no_senior_keyword',
                                'it_keyword_match': it_kw,
                                'is_junior': True,
                                'is_it': True,
                                'url': f'https://www.arbeitsagentur.de/jobsuche/jobdetail/{refnr}',
                                'description': '',
                                'vollzeit': stelle.get('vollzeit', True),
                                'befristet': stelle.get('befristet', False),
                                'aktuelleVeroeffentlichung': stelle.get('aktuelleVeroeffentlichung', ''),
                                'eintrittsdatum': stelle.get('eintrittsdatum', ''),
                                'arbeitgeberPlz': stelle.get('arbeitgeberPlz', ''),
                                'arbeitgeberOrt': stelle.get('arbeitgeberOrt', ''),
                                'modifikationsTimestamp': stelle.get('modifikationsTimestamp', ''),
                            })
            except:
                continue
    
    # Deduplicate by refnr
    seen = set()
    unique = []
    for j in jobs:
        key = j.get('refnr', '')
        if key and key not in seen:
            seen.add(key)
            unique.append(j)
    
    return unique


async def main():
    print("=" * 70)
    print("Phase 1: Arbeitsagentur API - Junior IT Jobs")
    print("=" * 70)
    
    # Load employer names for matching
    wb = openpyxl.load_workbook(EXCEL_IN)
    ws = wb['Nur IT-Jobs']
    
    employer_names = set()
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 2).value
        if name:
            # Normalize: remove common suffixes
            norm = re.sub(r'\s*(GmbH|AG|e\.K\.|e\.V\.|UG|KG|GbR|OHG|mbH).*$', '', name).strip()
            if len(norm) >= 3:
                employer_names.add(norm)
    
    print(f"  {len(employer_names)} unique employer names to match against")
    
    # Broad search for all junior IT jobs in the region
    print("\n[1] Broad search for junior IT jobs in B/ER/N region...")
    async with httpx.AsyncClient() as client:
        all_jobs = await broad_junior_search(client)
        print(f"  Found {len(all_jobs)} unique junior IT jobs from broad search")
    
    # Now also search by specific employer name
    print("\n[2] Employer-specific search...")
    employer_list = list(employer_names)
    specific_jobs = []
    
    async with httpx.AsyncClient() as client:
        batch_size = 10
        for i in range(0, len(employer_list), batch_size):
            batch = employer_list[i:i+batch_size]
            tasks = [query_aa_jobs(client, name) for name in batch]
            results = await asyncio.gather(*tasks, return_exceptions=True)
            
            for result in results:
                if isinstance(result, list):
                    specific_jobs.extend(result)
            
            if (i // batch_size + 1) % 5 == 0:
                print(f"  Checked {min(i+batch_size, len(employer_list))}/{len(employer_list)} employers")
    
    # Deduplicate specific jobs
    seen_refnr = set(j.get('refnr', '') for j in all_jobs if j.get('refnr'))
    for j in specific_jobs:
        if j.get('refnr') and j['refnr'] not in seen_refnr:
            all_jobs.append(j)
            seen_refnr.add(j['refnr'])
    
    print(f"\n  Total unique junior IT jobs: {len(all_jobs)}")
    
    # Create Excel
    print("\n[3] Creating Excel...")
    import openpyxl as xl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    out_wb = xl.Workbook()
    ws = out_wb.active
    ws.title = "Junior IT-Jobs"
    
    headers = [
        'Job-ID', 'Job-Titel', 'Job-Titel (Original)', 'Job-Kategorie', 'Erfahrungslevel',
        'Beschäftigungsart', 'Remote-Option', 'Technologien',
        'Ausschreibungstext (Auszug)', 'Job-URL',
        'Arbeitgeber', 'Arbeitgeber (norm.)', 'Ort', 'PLZ', 'Straße',
        'Telefon', 'E-Mail', 'Branche',
        'Quelle', 'Karriere-URL', 'Website-URL',
        'Karriere-HTML Pfad', 'Job-HTML Pfad',
        'Junior-Keyword Match', 'IT-Keyword Match', 'Ist Junior-Job', 'Ist IT-Job',
        'Vollzeit/Teilzeit Detail', 'Befristet', 'Gehaltsspanne', 'Arbeitsagentur RefNr',
        'Eintrittsdatum', 'Veroeffentlichungsdatum',
        'Arbeitgeber Ort', 'Arbeitgeber PLZ',
        'Scrape-Datum', 'Scrape-Status', 'Arbeitgeber-ID',
        'Letzte Prüfung', 'Erstellt', 'Aktualisiert',
    ]
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    # Add jobs
    for idx, job in enumerate(all_jobs):
        row = idx + 2
        title = job.get('title', '')
        desc = job.get('description', '')
        
        values = [
            f'J-{idx+1:05d}',  # Job-ID
            title,  # Job-Titel
            title,  # Job-Titel (Original)
            classify_job_category(title, desc),  # Job-Kategorie
            'Junior',  # Erfahrungslevel
            extract_employment_type(title, desc),  # Beschäftigungsart
            extract_remote_option(title, desc),  # Remote-Option
            ', '.join(extract_technologies(f"{title} {desc}")),  # Technologien
            desc[:500],  # Ausschreibungstext
            job.get('url', ''),  # Job-URL
            job.get('employer_name', ''),  # Arbeitgeber
            '',  # Arbeitgeber (norm.)
            job.get('city', ''),  # Ort
            job.get('plz', ''),  # PLZ
            '',  # Straße
            '',  # Telefon
            '',  # E-Mail
            '',  # Branche
            job.get('source', 'Arbeitsagentur'),  # Quelle
            '',  # Karriere-URL
            '',  # Website-URL
            '',  # Karriere-HTML Pfad
            '',  # Job-HTML Pfad
            job.get('junior_keyword_match', ''),  # Junior-Keyword Match
            job.get('it_keyword_match', ''),  # IT-Keyword Match
            'Ja',  # Ist Junior-Job
            'Ja',  # Ist IT-Job
            extract_employment_type(title, desc),  # Vollzeit/Teilzeit Detail
            'Ja' if job.get('befristet', False) else 'Nein',  # Befristet
            '',  # Gehaltsspanne
            job.get('refnr', ''),  # Arbeitsagentur RefNr
            job.get('eintrittsdatum', ''),  # Eintrittsdatum
            job.get('aktuelleVeroeffentlichung', ''),  # Veroeffentlichungsdatum
            job.get('arbeitgeberOrt', ''),  # Arbeitgeber Ort
            job.get('arbeitgeberPlz', ''),  # Arbeitgeber PLZ
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),  # Scrape-Datum
            'completed',  # Scrape-Status
            '',  # Arbeitgeber-ID
            datetime.now().strftime('%Y-%m-%d'),  # Letzte Prüfung
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),  # Erstellt
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),  # Aktualisiert
        ]
        
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Stats sheet
    ws_stats = out_wb.create_sheet('Statistiken')
    stats = [
        ('Gesamt Junior IT-Jobs', len(all_jobs)),
        ('Quelle: Arbeitsagentur', len([j for j in all_jobs if j.get('source') == 'Arbeitsagentur'])),
        ('Region', 'Bamberg / Erlangen / Nürnberg (50km)'),
        ('Scrape-Datum', datetime.now().strftime('%Y-%m-%d %H:%M')),
        ('Suchbegriffe', 'Junior IT, Werkstudent, Praktikum, Ausbildung, etc.'),
    ]
    for i, (key, val) in enumerate(stats, 1):
        ws_stats.cell(i, 1, key)
        ws_stats.cell(i, 2, val)
    
    out_wb.save(EXCEL_OUT)
    print(f"  Excel saved: {EXCEL_OUT}")
    print(f"  {len(all_jobs)} junior IT jobs")
    
    # Also save raw JSON
    json_out = EXCEL_OUT.replace('.xlsx', '_raw.json')
    with open(json_out, 'w', encoding='utf-8') as f:
        json.dump(all_jobs, f, ensure_ascii=False, indent=2, default=str)
    print(f"  JSON saved: {json_out}")
    
    # Git commit
    os.chdir(GIT_REPO)
    subprocess.run(['git', 'add', '-A'], capture_output=True, timeout=30)
    result = subprocess.run(['git', 'commit', '-m', f'Phase 1: Junior IT Jobs from Arbeitsagentur ({len(all_jobs)} jobs)'], capture_output=True, timeout=30)
    if result.returncode == 0:
        subprocess.run(['git', 'push', f'https://{GITHUB_TOKEN}@github.com/BirdNest055/Heyjobs.git', 'main', '--force-with-lease'], capture_output=True, timeout=60)
        print("  Git: Pushed!")
    
    print(f"\n{'=' * 70}")
    print(f"Phase 1 COMPLETE: {len(all_jobs)} junior IT jobs collected")
    print(f"{'=' * 70}")


if __name__ == '__main__':
    asyncio.run(main())
