#!/usr/bin/env python3
"""
Job Application Scraper v3 - Production version.
- Handles cookie consent dialogs
- Navigates career pages intelligently
- Extracts junior IT job listings with details
- Generates customized application emails
- Saves to Excel + pushes to GitHub every 10 companies
"""

import asyncio
import json
import os
import re
import subprocess
import sys
from datetime import datetime
from urllib.parse import urljoin, urlparse

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# ── CONFIG ────────────────────────────────────────────────────────────────────
EXCEL_OUT = "/home/z/my-project/download/Bewerbungen_Junior_IT.xlsx"
PROGRESS_FILE = "/home/z/my-project/job_scraper_v3_progress.json"
LOG_FILE = "/home/z/my-project/job_scraper_v3_log.txt"
GITHUB_REPO = "/home/z/my-project"
DISPLAY = ":99"
PUSH_EVERY = 10
MAX_COMPANIES = 500
TIMEOUT_PER_COMPANY = 90
MAX_JOBS_PER_COMPANY = 5

# ── EMAIL TEMPLATE ────────────────────────────────────────────────────────────
EMAIL_BODY = """Sehr geehrte{anrede_suffix} {anrede_target},

ich sende Ihnen hiermit meine Bewerbungsunterlagen für die Stelle als {stelle} in Vollzeit.

Ich schließe meine Ausbildung zum Fachinformatiker für Systemintegration bald ab. Das Fachgespräch findet am 11. Juni 2026 statt, ab diesem Zeitpunkt stehe ich für eine Einstellung zur Verfügung.

Während meiner Ausbildung hatte ich Berührungspunkte mit verschiedenen IT-Bereichen: Systemintegration, Webentwicklung und Datenbankentwicklung.

In meiner letzten Praxisphase habe ich mich mit Azure beschäftigt, was direkt in mein IHK-Abschlussprojekt geflossen ist: die automatisierte Anwendungsbereitstellung via App Attach in Azure Virtual Desktop, von der Planung bis zur Dokumentation eigenständig umgesetzt.

Ich arbeite selbstständig und kommuniziere offen. Das haben mir auch meine Betreuer in den Praxisphasen zurückgemeldet. Einen Führerschein der Klasse B besitze ich aktuell nicht, bin aber bereit, diesen zeitnah zu erwerben.

Meine vollständigen Unterlagen finden Sie im Anhang. Ich freue mich auf ein persönliches Gespräch.

Mit freundlichen Grüßen,"""


def log(msg):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line, flush=True)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")


def load_progress():
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"processed": [], "jobs": [], "last_index": 0}


def save_progress(progress):
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump(progress, f, ensure_ascii=False, indent=2)


def git_push(msg):
    try:
        subprocess.run(["git", "add", "-A"], cwd=GITHUB_REPO, capture_output=True, timeout=30)
        subprocess.run(["git", "commit", "-m", msg], cwd=GITHUB_REPO, capture_output=True, timeout=30)
        result = subprocess.run(["git", "push", "origin", "main"], cwd=GITHUB_REPO,
                                capture_output=True, text=True, timeout=60)
        if result.returncode == 0:
            log(f"✅ Pushed: {msg}")
        else:
            log(f"⚠️ Push issue: {result.stderr[:150]}")
    except Exception as e:
        log(f"⚠️ Git error: {e}")


def customize_email(stelle, firma, ansprechpartner=""):
    """Generate customized email from template."""
    if ansprechpartner and ansprechpartner.lower() not in ["n/a", "unbekannt", "", "damen und herren"]:
        if any(t in ansprechpartner.lower() for t in ['frau', 'ms.', 'mrs.']):
            anrede_suffix = ""
            clean = ansprechpartner.replace('Frau', '').replace('Mrs.', '').replace('Ms.', '').strip()
            anrede_target = f"Frau {clean}"
        elif any(t in ansprechpartner.lower() for t in ['herr', 'hr.', 'mr.']):
            anrede_suffix = "r"
            clean = ansprechpartner.replace('Herr', '').replace('Hr.', '').replace('Mr.', '').strip()
            anrede_target = f"Herr {clean}"
        else:
            # Default to "Herr/Frau" when unclear
            anrede_suffix = "r"
            anrede_target = ansprechpartner
    else:
        anrede_suffix = "r"
        anrede_target = "Damen und Herren"

    return EMAIL_BODY.format(anrede_suffix=anrede_suffix, anrede_target=anrede_target, stelle=stelle)


def extract_emails(text):
    if not text:
        return []
    return list(set(re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', text)))


IT_KEYWORDS = [
    'it-', 'it ', 'informatik', 'software', 'systemintegration', 'systemadmin',
    'netzwerk', 'devops', 'cloud', 'data engineer', 'data scien', 'datenbank',
    'cyber', 'security', 'entwickler', 'developer', 'administrator', 'admin',
    'linux', 'sap ', 'azure', 'aws ', 'frontend', 'backend', 'fullstack',
    'full-stack', 'sre ', 'platform engineer', 'infrastruktur', 'infrastructure',
    'it-support', 'helpdesk', 'it-consultant', 'it berater', 'it manager',
    'it specialist', 'system engineer', 'network', 'automation', 'automatisier',
    'test', 'qa ', 'scrum master', 'product owner', 'digital',
    'webentwickler', 'web developer', 'mobile', 'dev ', 'ux ', 'ui ',
    'machine learning', 'ai ', 'künstliche intelligenz', 'projektmanag',
    'anwendungs', 'application', 'desktop', 'rechenzentrum', 'datacenter',
    'technology', 'technolog', 'digitalisier', 'comput', 'hosting',
]

JUNIOR_KEYWORDS = [
    'junior', 'trainee', 'entry', 'graduate', 'start', 'berufseinsteig',
    'jun.', 'young professional', 'werkstudent', 'praktikant', 'azubi',
    'ausbildung', '1st level', 'first level', 'level 1', 'beginner',
    ' apprenticeship', 'dual stud', 'duales studium',
]

SENIOR_KEYWORDS = [
    'senior', 'lead', 'manager', 'director', 'head of', 'vp ',
    'chief', 'c-level', 'principal', 'staff', 'expert',
]


def is_it_job(text):
    tl = text.lower()
    return any(kw in tl for kw in IT_KEYWORDS)


def is_junior_job(text):
    tl = text.lower()
    if any(kw in tl for kw in JUNIOR_KEYWORDS):
        return "junior"
    if any(kw in tl for kw in SENIOR_KEYWORDS):
        return "senior"
    return "open"


def load_companies():
    """Load and merge company data, sort by IT relevance."""
    companies = {}

    # 1. Google Maps
    wb1 = openpyxl.load_workbook("/home/z/my-project/download/Google_Maps_Firmen_Erlangen_50km.xlsx")
    ws1 = wb1["Google Maps Firmen"]
    for r in range(2, ws1.max_row + 1):
        name = (ws1.cell(r, 2).value or "").strip()
        if not name:
            continue
        if name not in companies:
            companies[name] = {
                "name": name,
                "kategorie": ws1.cell(r, 5).value or "",
                "adresse": ws1.cell(r, 6).value or "",
                "plz": ws1.cell(r, 7).value or "",
                "ort": ws1.cell(r, 8).value or "",
                "telefon": ws1.cell(r, 9).value or "",
                "website": ws1.cell(r, 10).value or "",
                "email": "", "ansprechpartner": ""
            }
        elif ws1.cell(r, 10).value and not companies[name]["website"]:
            companies[name]["website"] = ws1.cell(r, 10).value

    # 2. Enriched
    wb2 = openpyxl.load_workbook("/home/z/my-project/download/Firmen_Bamberg_Erlangen_Nuernberg_mit_Website_HTML.xlsx")
    ws2 = wb2["Firmen mit Website & HTML"]
    for r in range(2, ws2.max_row + 1):
        name = (ws2.cell(r, 2).value or "").strip()
        if not name:
            continue
        web_found = ws2.cell(r, 10).value or ""
        email = ws2.cell(r, 11).value or ""
        if name in companies:
            if web_found and not companies[name]["website"]:
                companies[name]["website"] = web_found
            if email and not companies[name]["email"]:
                companies[name]["email"] = email
        else:
            companies[name] = {
                "name": name, "kategorie": ws2.cell(r, 4).value or "",
                "adresse": ws2.cell(r, 5).value or "",
                "plz": ws2.cell(r, 7).value or "",
                "ort": ws2.cell(r, 8).value or "",
                "telefon": ws2.cell(r, 9).value or "",
                "website": web_found, "email": email,
                "ansprechpartner": ""
            }

    # 3. Arbeitsagentur
    wb3 = openpyxl.load_workbook("/home/z/my-project/download/Junior_IT_Jobs_Bamberg_Erlangen_Nuernberg.xlsx")
    ws3 = wb3["Junior IT-Jobs"]
    aa_jobs = {}
    for r in range(2, ws3.max_row + 1):
        employer = (ws3.cell(r, 22).value or "").strip()
        job_title = ws3.cell(r, 2).value or ""
        job_url = ws3.cell(r, 27).value or ""
        if employer:
            aa_jobs.setdefault(employer, []).append({"title": job_title, "url": job_url})

    # Score
    scoring_kws = IT_KEYWORDS + ['maschinenbau', 'industrie', 'elektro', 'energie',
                                  'bank', 'versicher', 'finanz', 'klinik', 'universität',
                                  'forschung', 'research', 'gruppe', 'holding', 'telekommunikat',
                                  'logistik', 'versand', 'handel', 'consult', 'berat']

    scored = []
    for name, data in companies.items():
        if not data["website"]:
            continue
        score = sum(1 for kw in scoring_kws if kw in f"{name} {data['kategorie']}".lower())
        if name in aa_jobs:
            score += 5
        scored.append((score, name, data))

    scored.sort(key=lambda x: -x[0])

    result = []
    for score, name, data in scored:
        data["relevance_score"] = score
        data["arbeitsagentur_jobs"] = aa_jobs.get(name, [])
        result.append(data)

    log(f"Loaded {len(result)} companies with websites")
    return result


class JobScraper:
    def __init__(self, page):
        self.page = page

    async def accept_cookies(self):
        """Try to dismiss cookie consent dialogs."""
        selectors = [
            'button:has-text("Akzeptieren")', 'button:has-text("Alle akzeptieren")',
            'button:has-text("Accept All")', 'button:has-text("Accept")',
            'button:has-text("OK")', 'button:has-text("Verstanden")',
            'button:has-text("Annehmen")', 'button:has-text("Auswahl bestätigen")',
            'button:has-text("Ich stimme zu")', 'button:has-text("Zustimmen")',
            'a:has-text("Akzeptieren")', 'a:has-text("Zustimmen")',
            '#onetrust-accept-btn-handler', '.cc-btn', '.cookie-accept',
            '[data-testid*="accept"]', 'button[id*="accept"]',
            'button[id*="consent"]', '[class*="consent"] button',
            '[class*="cookie"] button:first-child',
            'button.cmplz-btn-accept', 'button#CybotCookiebotDialogBodyLevelButtonLevelOptinAllowAll',
        ]
        for sel in selectors:
            try:
                btn = self.page.locator(sel).first
                if await btn.is_visible(timeout=400):
                    await btn.click()
                    await self.page.wait_for_timeout(600)
                    return True
            except:
                continue
        return False

    async def safe_goto(self, url, timeout=20000):
        """Navigate with error handling."""
        try:
            resp = await self.page.goto(url, wait_until="domcontentloaded", timeout=timeout)
            if resp and resp.status < 500:
                await self.page.wait_for_timeout(2000)
                try:
                    await self.accept_cookies()
                except:
                    pass
                await self.page.wait_for_timeout(500)
                return True
            # Some pages return None response but are still valid
            elif resp is None:
                await self.page.wait_for_timeout(2000)
                try:
                    await self.accept_cookies()
                except:
                    pass
                return True
        except Exception as e:
            # Try with load event instead
            try:
                resp = await self.page.goto(url, wait_until="load", timeout=timeout)
                await self.page.wait_for_timeout(2000)
                return True
            except:
                pass
        return False

    async def get_career_links(self, base_url):
        """Find career page links from the current page."""
        links = await self.page.evaluate("""(baseurl) => {
            const results = [];
            const seen = new Set();
            document.querySelectorAll('a[href]').forEach(a => {
                const text = (a.textContent || '').trim().toLowerCase();
                const href = (a.getAttribute('href') || '').toLowerCase();
                const keywords = ['karriere', 'jobs', 'stellen', 'career', 'bewer',
                                   'offene', 'vacanc', 'position', 'ausschreib',
                                   'unser team', 'arbeit bei', 'work with', 'join',
                                   'now hiring', 'offene positionen', 'jobbörse',
                                   'stellensuche', 'job search'];
                const isCareer = keywords.some(kw => text.includes(kw) || href.includes(kw));
                if (isCareer && a.href && !seen.has(a.href)) {
                    seen.add(a.href);
                    results.push({href: a.href, text: (a.textContent || '').trim().substring(0, 100)});
                }
            });
            return results;
        }""", base_url)

        # Also add standard paths
        parsed = urlparse(base_url)
        domain_base = f"{parsed.scheme}://{parsed.netloc}"
        standard = ["/karriere", "/jobs", "/stellenangebote", "/career", "/careers",
                    "/de/karriere", "/de/jobs", "/de/career", "/en/career",
                    "/offene-stellen", "/bewerbung", "/vacancies", "/unser-team",
                    "/offene-positionen", "/jobboerse", "/stellenangebote.html"]
        for path in standard:
            full = domain_base + path
            if not any(l["href"] == full for l in links):
                links.append({"href": full, "text": path})

        return links

    async def extract_jobs_from_page(self, page_url):
        """Extract IT job listings from a career page."""
        jobs = []

        # Get all text content
        try:
            page_text = await self.page.inner_text("body")
        except:
            try:
                page_text = await self.page.content()
                # Strip HTML tags
                import re as _re
                page_text = _re.sub(r'<[^>]+>', ' ', page_text)
                page_text = _re.sub(r'\s+', ' ', page_text)
            except:
                return jobs

        # Quick check - any IT content?
        if not is_it_job(page_text[:5000]):
            return jobs

        # Method 1: JavaScript extraction - focus on ACTUAL job postings
        raw_jobs = await self.page.evaluate("""() => {
            const results = [];
            const seen = new Set();
            
            // These words strongly indicate an actual JOB POSTING (not a service/category)
            const jobTitlePatterns = [
                /\\(m\\/w\\/d\\)/i, /\\(w\\/m\\/d\\)/i, /\\(m\\/f\\/d\\)/i, /\\(f\\/m\\/d\\)/i,
                /\\(m\\/w\\)/i, /\\(w\\/m\\)/i,
                /entwickler/i, /developer/i, /administrator/i, /engineer/i,
                /analyst/i, /architekt/i, /architect/i, /consultant/i, /berater/i,
                /specialist/i, /spezialist/i, /experte/i, /expert/i,
                /techniker/i, /operator/i, /betreuer/i,
                /trainee/i, /junior/i, /senior/i, /lead/i,
                /praktikant/i, /werkstudent/i, /azubi/i, /auszubildende/i,
                /ingenieur/i, /programmierer/i, /programmer/i, /designer/i,
                /koordinator/i, /coordinator/i, /assistent/i, /assistant/i,
                /referent/i, /sachbearbeiter/i, /disponent/i, /controller/i,
                /scrum\\s*master/i, /product\\s*owner/i, /devops/i, /\\bsre\\b/i,
                /stellenangebot/i, /stellenanzeige/i, /\\bjob\\b/i,
                /vacancy/i, /ausschreibung/i, /opening/i,
                /fachinformatiker/i, /informatiker/i, /mathematiker/i,
                /physiker/i, /naturwissenschaftler/i,
                /kaufmann/i, /kauffrau/i, /kaufleute/i,
                /manager\\s/i, /head\\s+of/i, /director/i,
                /werkstatt/i, /laborant/i, /forscher/i,
            ];
            
            function looksLikeJobPosting(text) {
                // Must match at least one job title pattern
                return jobTitlePatterns.some(p => p.test(text));
            }
            
            const selectors = [
                'a', 'article', 'h2', 'h3', 'h4', 'h5',
                '.job', '.vacancy', '.position', '.stellenanzeige',
                '.job-listing', '.job-item', '.job-entry',
                '.career-item', '.position-item',
                '[data-job]', '[data-vacancy]',
                '.card', '.teaser', '.entry',
                'li', 'tr', '.row', '.item',
                '[class*="job"]', '[class*="career"]', '[class*="position"]',
                '[class*="stellen"]', '[class*="vacanc"]',
            ];
            
            for (const sel of selectors) {
                const elements = document.querySelectorAll(sel);
                for (const el of elements) {
                    const text = (el.textContent || '').trim();
                    if (text.length < 10 || text.length > 1500) continue;
                    
                    const lines = text.split('\\n').map(l => l.trim()).filter(l => l.length > 3);
                    const title = (lines[0] || text.substring(0, 150)).substring(0, 150);
                    
                    // CRITICAL: Must look like a job posting
                    if (!looksLikeJobPosting(title)) continue;
                    
                    const key = title.toLowerCase().substring(0, 50);
                    if (seen.has(key)) continue;
                    seen.add(key);
                    
                    const link = el.closest('a') || el.querySelector('a');
                    const href = link ? link.href : (el.tagName === 'A' ? el.href : '');
                    
                    results.push({title: title, href: href || '', text: text.substring(0, 500)});
                    if (results.length >= 30) break;
                }
                if (results.length >= 30) break;
            }
            return results;
        }""")

        for item in raw_jobs:
            title = item["title"].strip()
            if len(title) < 5:
                continue

            # Check IT relevance
            if not is_it_job(title + " " + item["text"][:200]):
                continue

            # Check level
            level = is_junior_job(title)
            if level == "senior":
                continue

            # Clean up title - remove excessive whitespace/newlines
            title = " ".join(title.split())

            # Build URL
            job_url = item["href"] if item["href"] else ""

            jobs.append({
                "title": title,
                "url": job_url,
                "description": item["text"][:500],
                "is_junior": level == "junior"
            })

        return jobs

    async def find_contact_info(self, base_url):
        """Find email and Ansprechpartner from Impressum/Kontakt."""
        email = ""
        ansprechpartner = ""

        for path in ["/impressum", "/kontakt", "/contact", "/imprint"]:
            try:
                url = base_url + path
                if not await self.safe_goto(url, timeout=12000):
                    continue
                await self.page.wait_for_timeout(1500)
                text = await self.page.inner_text("body")

                # Find email
                emails = extract_emails(text)
                for em in emails:
                    if not any(g in em.lower() for g in ['noreply', 'no-reply', 'spam', 'example', 'test', 'github', 'wordpress']):
                        email = em
                        break
                if not email and emails:
                    email = emails[0]

                # Find Ansprechpartner
                for line in text.split("\n"):
                    line = line.strip()
                    ll = line.lower()
                    if ('herr' in ll or 'frau' in ll) and \
                       any(t in ll for t in ['personal', 'hr', 'bewer', 'kontakt', 'ansprech', 'recruit']):
                        ansprechpartner = line.strip()[:80]
                        break

                if email:
                    break
            except:
                continue

        return email, ansprechpartner

    async def scrape_company(self, company):
        """Main scraping logic for a single company."""
        url = company["website"]
        if not url:
            return [], company.get("email", ""), company.get("ansprechpartner", "")

        # Normalize URL
        url = url.strip()
        if not url.startswith(("http://", "https://")):
            url = "https://" + url
        url = url.rstrip("/")

        parsed = urlparse(url)
        base_url = f"{parsed.scheme}://{parsed.netloc}"

        jobs_found = []
        contact_email = company.get("email", "")
        ansprechpartner = company.get("ansprechpartner", "")

        # Step 1: Visit main page
        log(f"  🌐 Main: {url}")
        if not await self.safe_goto(url, timeout=25000):
            log(f"  ⚠️ Cannot reach main page")
            return [], contact_email, ansprechpartner

        # Step 2: Find career links
        career_links = await self.get_career_links(base_url)
        log(f"  🔗 {len(career_links)} career links")

        # Step 3: Visit career pages and extract jobs
        visited = set()
        for cl in career_links[:8]:
            career_url = cl["href"]
            if career_url in visited:
                continue
            visited.add(career_url)

            log(f"  📄 Career: {career_url[:60]}")
            if not await self.safe_goto(career_url, timeout=18000):
                continue

            await self.page.wait_for_timeout(2500)

            page_jobs = await self.extract_jobs_from_page(career_url)
            if page_jobs:
                log(f"  ✅ {len(page_jobs)} potential IT jobs on this page")
                jobs_found.extend(page_jobs)
                if len(jobs_found) >= MAX_JOBS_PER_COMPANY:
                    break

            # Also check for sub-pages with job listings (click into IT category)
            if not page_jobs:
                # Try clicking IT-specific navigation items
                try:
                    it_nav = await self.page.evaluate("""() => {
                        const links = [];
                        document.querySelectorAll('a').forEach(a => {
                            const text = (a.textContent || '').trim().toLowerCase();
                            const href = (a.getAttribute('href') || '').toLowerCase();
                            if (['it', 'software', 'informatik', 'technik', 'digital', 'technologie', 'data'].some(kw => text === kw || text.startsWith(kw + ' ')) || 
                                href.includes('/it') || href.includes('/software') || href.includes('/digital')) {
                                links.push({href: a.href, text: (a.textContent || '').trim().substring(0, 80)});
                            }
                        });
                        return links.slice(0, 5);
                    }""")

                    for nav in it_nav[:3]:
                        nav_url = nav["href"]
                        if nav_url in visited:
                            continue
                        visited.add(nav_url)
                        log(f"  🔍 IT subpage: {nav['text'][:40]}")
                        if await self.safe_goto(nav_url, timeout=15000):
                            await self.page.wait_for_timeout(2000)
                            sub_jobs = await self.extract_jobs_from_page(nav_url)
                            if sub_jobs:
                                log(f"  ✅ {len(sub_jobs)} jobs on IT subpage")
                                jobs_found.extend(sub_jobs)
                                if len(jobs_found) >= MAX_JOBS_PER_COMPANY:
                                    break
                except:
                    pass

        # Step 4: Get contact info if needed
        if not contact_email or not ansprechpartner:
            found_email, found_ansprech = await self.find_contact_info(base_url)
            if not contact_email and found_email:
                contact_email = found_email
            if not ansprechpartner and found_ansprech:
                ansprechpartner = found_ansprech

        # Step 5: Also include Arbeitsagentur jobs for this company
        for aa_job in company.get("arbeitsagentur_jobs", []):
            title = aa_job["title"]
            level = is_junior_job(title)
            if level == "senior":
                continue
            existing = {j["title"].lower()[:40] for j in jobs_found}
            if title.lower()[:40] not in existing:
                jobs_found.append({
                    "title": title,
                    "url": aa_job.get("url", ""),
                    "description": "Quelle: Arbeitsagentur",
                    "is_junior": level == "junior"
                })

        # Deduplicate
        seen = set()
        unique = []
        for j in jobs_found:
            key = " ".join(j["title"].lower().split())[:50]
            if key not in seen:
                seen.add(key)
                j["email"] = contact_email
                j["ansprechpartner"] = ansprechpartner
                unique.append(j)
            if len(unique) >= MAX_JOBS_PER_COMPANY:
                break

        return unique, contact_email, ansprechpartner


async def main():
    log("=" * 60)
    log("Job Application Scraper v3 - Starting")
    log("=" * 60)

    try:
        companies = load_companies()
        progress = load_progress()
        processed_names = set(progress["processed"])
        jobs_all = progress["jobs"]

        log(f"Previously processed: {len(processed_names)}")
        log(f"Jobs so far: {len(jobs_all)}")

        to_process = [(i, c) for i, c in enumerate(companies)
                      if c["name"] not in processed_names and c["website"]]
        log(f"Remaining: {len(to_process)}")

        if not to_process:
            save_excel(jobs_all)
            return

        to_process = to_process[:MAX_COMPANIES]

        from playwright.async_api import async_playwright

        async with async_playwright() as p:
            browser = await p.chromium.launch(
                headless=False,
                args=["--no-sandbox", "--disable-dev-shm-usage",
                      "--window-size=1920,1080",
                      "--disable-blink-features=AutomationControlled"]
            )
            context = await browser.new_context(
                viewport={"width": 1920, "height": 1080},
                user_agent="Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
                locale="de-DE",
                ignore_https_errors=True
            )
            page = await context.new_page()
            page.set_default_timeout(25000)
            page.set_default_navigation_timeout(25000)

            scraper = JobScraper(page)
            processed_count = 0
            total_new = 0

            for idx, (orig_idx, company) in enumerate(to_process):
                name = company["name"]
                log(f"\n[{idx+1}/{len(to_process)}] {name} (score: {company.get('relevance_score', 0)})")

                try:
                    jobs, email, ansprech = await asyncio.wait_for(
                        scraper.scrape_company(company),
                        timeout=TIMEOUT_PER_COMPANY
                    )
                except asyncio.TimeoutError:
                    log(f"  ⏰ Timeout")
                    jobs, email, ansprech = [], "", ""
                except Exception as e:
                    import traceback
                    log(f"  ❌ Error: {type(e).__name__}: {str(e)[:80]}")
                    log(f"  Traceback: {traceback.format_exc()[:300]}")
                    jobs, email, ansprech = [], "", ""

                if jobs:
                    log(f"  ✅ Found {len(jobs)} IT job(s)!")
                    for job in jobs:
                        email_text = customize_email(
                            stelle=job["title"],
                            firma=name,
                            ansprechpartner=job.get("ansprechpartner", ansprech)
                        )
                        adresse = company.get("adresse", "")
                        if company.get("plz"):
                            adresse += f", {company['plz']}"
                        if company.get("ort"):
                            adresse += f" {company['ort']}"

                        jobs_all.append({
                            "firma": name,
                            "stelle": job["title"],
                            "stellenbeschreibung": job.get("description", ""),
                            "email": job.get("email", email or company.get("email", "")),
                            "ansprechpartner": job.get("ansprechpartner", ansprech),
                            "bewerbung_geschickt": "",
                            "letzter_kontakt": "",
                            "status": "Offen",
                            "website": company["website"],
                            "adresse": adresse.strip(", "),
                            "email_text": email_text,
                            "job_url": job.get("url", ""),
                            "is_junior": job.get("is_junior", False),
                            "scraped_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
                        })
                        total_new += 1
                else:
                    log(f"  ℹ️ No IT jobs found")

                processed_names.add(name)
                processed_count += 1

                progress["processed"] = list(processed_names)
                progress["jobs"] = jobs_all
                progress["last_index"] = orig_idx
                save_progress(progress)

                if processed_count % PUSH_EVERY == 0:
                    log(f"\n📦 Checkpoint: {processed_count} processed, {len(jobs_all)} jobs total")
                    save_excel(jobs_all)
                    git_push(f"Job Scraper v3: {processed_count} co | {len(jobs_all)} jobs | {datetime.now().strftime('%m-%d %H:%M')}")

                await page.wait_for_timeout(500)

            await browser.close()

        log(f"\n{'='*60}")
        log(f"DONE! {processed_count} companies, {total_new} new jobs, {len(jobs_all)} total")
        save_excel(jobs_all)
        git_push(f"Job Scraper v3 DONE: {processed_count} co | {len(jobs_all)} jobs | {datetime.now().strftime('%m-%d')}")

        progress["last_run"] = datetime.now().isoformat()
        save_progress(progress)
    except Exception as e:
        import traceback
        log(f"💥 FATAL ERROR: {type(e).__name__}: {e}")
        log(traceback.format_exc())


def save_excel(jobs):
    """Generate application-ready Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bewerbungen"

    headers = [
        "Firma", "Stelle", "Stellenbeschreibung", "Email", "Ansprechpartner",
        "Bewerbung geschickt am", "Letzter Kontakt am", "Status", "Website",
        "Adresse", "Email Text", "Job-URL", "Junior?", "Scraped At"
    ]

    hdr_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
        c.border = border

    for row, job in enumerate(jobs, 2):
        vals = [
            job.get("firma", ""), job.get("stelle", ""),
            job.get("stellenbeschreibung", ""), job.get("email", ""),
            job.get("ansprechpartner", ""), job.get("bewerbung_geschickt", ""),
            job.get("letzter_kontakt", ""), job.get("status", "Offen"),
            job.get("website", ""), job.get("adresse", ""),
            job.get("email_text", ""), job.get("job_url", ""),
            "Ja" if job.get("is_junior") else "Nein",
            job.get("scraped_at", ""),
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row, col, val)
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {1: 30, 2: 40, 3: 50, 4: 30, 5: 22, 6: 18, 7: 18,
              8: 14, 9: 35, 10: 35, 11: 70, 12: 40, 13: 10, 14: 16}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(jobs)+1}"

    # Status dropdown
    dv = DataValidation(type="list",
        formula1='"Offen,Beworben,Interview eingeplant,Interview absolviert,Zusage,Absage,Kontaktiert,Wiedervorlage,Kein Interesse"',
        allow_blank=True)
    ws.add_data_validation(dv)
    for r in range(2, len(jobs) + 2):
        dv.add(ws.cell(r, 8))

    # Conditional formatting
    lr = len(jobs) + 1
    for status, fill in [
        ("Zusage", PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")),
        ("Absage", PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")),
        ("Beworben", PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")),
        ("Interview eingeplant", PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")),
        ("Interview absolviert", PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")),
    ]:
        ws.conditional_formatting.add(f"H2:H{lr}", CellIsRule(operator="equal", formula=[f'"{status}"'], fill=fill))

    ws.conditional_formatting.add(f"M2:M{lr}",
        CellIsRule(operator="equal", formula=['"Ja"'],
                   fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")))

    # Email Vorlagen sheet
    ws2 = wb.create_sheet("Email Vorlagen")
    for col, h in enumerate(["Firma", "Stelle", "Email an", "Betreff", "Email Text (fertig zum Kopieren)"], 1):
        ws2.cell(1, col, h).font = Font(bold=True)

    for row, job in enumerate(jobs, 2):
        ws2.cell(row, 1, job.get("firma", ""))
        ws2.cell(row, 2, job.get("stelle", ""))
        ws2.cell(row, 3, job.get("email", ""))
        ws2.cell(row, 4, f"Bewerbung als {job.get('stelle', 'Fachinformatiker Systemintegration')} - {job.get('firma', '')}")
        ws2.cell(row, 5, job.get("email_text", ""))

    for col, w in {1: 30, 2: 40, 3: 30, 4: 55, 5: 80}.items():
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.freeze_panes = "A2"

    # Statistics sheet
    ws3 = wb.create_sheet("Statistiken")
    stats = [
        ("Bewerbungen Junior IT - Statistiken", ""),
        ("Stand", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("", ""),
        ("Gesamt Jobs", len(jobs)),
        ("Junior-Jobs (explizit)", sum(1 for j in jobs if j.get("is_junior"))),
        ("Mit Email", sum(1 for j in jobs if j.get("email"))),
        ("Mit Ansprechpartner", sum(1 for j in jobs if j.get("ansprechpartner"))),
        ("", ""),
        ("Nach Status:", ""),
    ]
    for s in ["Offen", "Beworben", "Interview eingeplant", "Interview absolviert", "Zusage", "Absage"]:
        cnt = sum(1 for j in jobs if j.get("status") == s)
        stats.append((f"  {s}", cnt))

    stats.extend([("", ""), ("Nach Ort:", "")])
    ort_counts = {}
    for j in jobs:
        addr = j.get("adresse", "").lower()
        matched = False
        for city in ["Bamberg", "Erlangen", "Nürnberg", "Fürth", "Herzogenaurach",
                      "Bayreuth", "Würzburg", "Schweinfurt", "Coburg", "Forchheim",
                      "Höchstadt", "Lauf", "Roth", "Schwabach"]:
            if city.lower() in addr:
                ort_counts[city] = ort_counts.get(city, 0) + 1
                matched = True
                break
        if not matched:
            ort_counts["Sonstige"] = ort_counts.get("Sonstige", 0) + 1
    for ort, cnt in sorted(ort_counts.items(), key=lambda x: -x[1]):
        stats.append((f"  {ort}", cnt))

    for row, (label, value) in enumerate(stats, 1):
        ws3.cell(row, 1, label).font = Font(bold=(row == 1 or "Nach" in str(label)))
        ws3.cell(row, 2, value)

    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 15

    wb.save(EXCEL_OUT)
    log(f"📊 Saved: {EXCEL_OUT} ({len(jobs)} jobs)")


if __name__ == "__main__":
    os.environ["DISPLAY"] = DISPLAY
    asyncio.run(main())
