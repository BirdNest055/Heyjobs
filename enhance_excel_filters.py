#!/usr/bin/env python3
"""
Enhance Junior IT Jobs Excel with professional filters, data validation,
conditional formatting, derived columns, and statistics.
"""

import json
import re
from datetime import datetime, timedelta
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from copy import copy

# ============================================================
# CONFIG
# ============================================================
EXCEL_IN = "/home/z/my-project/download/Junior_IT_Jobs_Bamberg_Erlangen_Nuernberg.xlsx"
JSON_IN = "/home/z/my-project/download/Junior_IT_Jobs_Bamberg_Erlangen_Nuernberg_raw.json"
EXCEL_OUT = "/home/z/my-project/download/Junior_IT_Jobs_Bamberg_Erlangen_Nuernberg.xlsx"

# ============================================================
# HELPER FUNCTIONS
# ============================================================

def clean_city(city_raw):
    """Clean city name by removing regional suffixes."""
    if not city_raw:
        return ''
    city = city_raw.strip()
    # Remove common suffixes
    for suffix in [', Mittelfranken', ', Oberfranken', ', Unterfranken', 
                   ', Bayern', ', Mittelfranken ', ', Oberfranken ',
                   ', Deutschland', ' bei Coburg', ' an der Pegnitz',
                   ' am Altmühlsee', ' im Markgräflerland',
                   ' im Fichtelgebirge', ', Oberpfalz']:
        city = city.replace(suffix, '')
    return city.strip()


def classify_explicit_junior(junior_kw):
    """Whether the job is explicitly junior (not just 'no senior keyword')."""
    if junior_kw and junior_kw != 'no_senior_keyword':
        return 'Ja'
    return 'Nein'


def days_ago(date_str):
    """Calculate days since a date string."""
    if not date_str:
        return ''
    try:
        dt = datetime.strptime(date_str[:10], '%Y-%m-%d')
        return (datetime.now() - dt).days
    except:
        return ''


def classify_it_bereich(title, hauptberuf=''):
    """Broader IT area classification for filtering."""
    text = f"{title} {hauptberuf}".lower()
    
    if any(kw in text for kw in ['softwareentwickler', 'software developer', 'software engineer',
                                   'programmierer', 'entwickler', 'anwendungsentwicklung',
                                   'frontend', 'backend', 'fullstack', 'full-stack',
                                   'webentwickler', 'app entwickler']):
        return 'Softwareentwicklung'
    if any(kw in text for kw in ['systemintegration', 'systemadministrator', 'sysadmin',
                                   'netzwerkadmin', 'it administrator', 'it-admin',
                                   'systemadministrator', 'infrastruktur']):
        return 'System & Infrastruktur'
    if any(kw in text for kw in ['data scientist', 'data analyst', 'data engineer',
                                   'data analytics', 'business intelligence', 'big data',
                                   'datenbank', 'database']):
        return 'Data & Analytics'
    if any(kw in text for kw in ['devops', 'cloud', 'platform engineer', 'sre',
                                   'site reliability', 'kubernetes', 'container']):
        return 'DevOps & Cloud'
    if any(kw in text for kw in ['security', 'cybersecurity', 'cyber-security',
                                   'it-sicherheit', 'pentest', 'soc', 'iam']):
        return 'IT-Sicherheit'
    if any(kw in text for kw in ['consultant', 'berater', 'beratung']):
        return 'IT-Consulting'
    if any(kw in text for kw in ['support', 'helpdesk', 'service desk', 'first level',
                                   'second level', 'user support', 'help desk']):
        return 'IT-Support'
    if any(kw in text for kw in ['projektmanager', 'project manager', 'scrum master',
                                   'product owner', 'projekt']):
        return 'IT-Projektmanagement'
    if any(kw in text for kw in ['tester', 'qa', 'quality', 'test engineer', 'testautomatisierung']):
        return 'QA & Testing'
    if any(kw in text for kw in ['ux', 'ui', 'user experience', 'user interface', 'design']):
        return 'UX/UI Design'
    if any(kw in text for kw in ['ausbildung', 'duales studium', 'trainee', 'praktikum',
                                   'werkstudent']):
        return 'Einstieg & Ausbildung'
    if any(kw in text for kw in ['sap', 'erp', 'dynamics', 'crm']):
        return 'ERP & CRM'
    if any(kw in text for kw in ['digitalisierung', 'digital', 'e-business']):
        return 'Digitalisierung'
    return 'IT (Sonstige)'


def classify_region(city):
    """Classify city into region for filtering."""
    if not city:
        return 'Unbekannt'
    c = city.lower()
    if any(x in c for x in ['nürnberg', 'nuernberg', 'fürth', 'fuerth', 'erlangen', 'schwabach']):
        return 'Metropolregion Nürnberg (Kern)'
    if any(x in c for x in ['bamberg', 'forchheim', 'coburg', 'bayreuth', 'liche', 'kulmbach']):
        return 'Oberfranken'
    if any(x in c for x in ['ansbach', 'roth', 'weißenburg', 'gunzenhausen', 'feuchtwangen']):
        return 'Mittelfranken (Land)'
    if any(x in c for x in ['herzogenaurach', 'baiersdorf', 'eckental']):
        return 'Erlangen-Umland'
    return 'Weitere Regionen'


# ============================================================
# MAIN: BUILD ENHANCED EXCEL
# ============================================================

def main():
    print("=" * 70)
    print("Junior IT-Jobs Excel Enhancement - Filter & Format")
    print("=" * 70)

    # Load existing Excel data
    print("\n[1] Lade bestehende Excel-Daten...")
    wb_in = openpyxl.load_workbook(EXCEL_IN)
    ws_in = wb_in['Junior IT-Jobs']
    
    # Read all data from existing Excel
    headers_in = [ws_in.cell(1, c).value for c in range(1, ws_in.max_column + 1)]
    print(f"  {ws_in.max_row - 1} Zeilen, {ws_in.max_column} Spalten geladen")
    
    # Load raw JSON for additional fields
    print("\n[2] Lade Rohdaten (JSON)...")
    raw_jobs = json.load(open(JSON_IN, encoding='utf-8'))
    refnr_to_json = {}
    for job in raw_jobs:
        ref = job.get('refnr', '')
        if ref:
            refnr_to_json[ref] = job
    print(f"  {len(refnr_to_json)} Jobs mit Referenznummer")
    
    # ============================================================
    # BUILD NEW EXCEL
    # ============================================================
    print("\n[3] Erstelle neue Excel mit erweiterten Spalten...")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Junior IT-Jobs"
    
    # NEW headers with additional derived columns
    HEADERS = [
        # === Job Identification ===
        'Job-ID',                    # A  (1)
        'Job-Titel',                 # B  (2)
        'Job-Kategorie',             # C  (3)
        'IT-Bereich',                # D  (4)  NEW - broader IT area
        'Erfahrungslevel',           # E  (5)
        'Explizit Junior',           # F  (6)  NEW - explicitly says junior?
        'Junior-Typ',                # G  (7)  NEW - type of junior (Ausbildung/Werkstudent/etc)
        
        # === Job Details ===
        'Beschäftigungsart',         # H  (8)
        'Vollzeit',                  # I  (9)  NEW - Ja/Nein
        'Befristet',                 # J  (10) 
        'Homeoffice',                # K  (11) NEW - Ja/Nein
        'Remote-Option',             # L  (12)
        'Homeoffice-Typ',            # M  (13) NEW - NACH_VEREINBARUNG etc.
        'Technologien',              # N  (14)
        'Quereinstieg',              # O  (15) NEW - Ja/Nein
        'Gehaltsspanne',             # P  (16)
        
        # === Location ===
        'Ort (Original)',            # Q  (17)
        'Stadt',                     # R  (18) NEW - cleaned city
        'Region',                    # S  (19) NEW - region classification
        'PLZ',                       # T  (20)
        'Straße',                    # U  (21)
        
        # === Employer ===
        'Arbeitgeber',               # V  (22)
        'Arbeitgeber (norm.)',       # W  (23)
        'Hauptberuf',                # X  (24)
        'Branche',                   # Y  (25)
        
        # === Source ===
        'Quelle',                    # Z  (26)
        'Job-URL',                   # AA (27)
        'Arbeitsagentur RefNr',      # AB (28)
        'Eintrittsdatum',            # AC (29)
        'Veröffentlicht am',         # AD (30)
        'Veröffentlicht vor (Tage)', # AE (31) NEW - days ago
        'Alter der Ausschreibung',   # AF (32) NEW - fresh/recent/old
        
        # === Classification ===
        'Junior-Keyword Match',      # AG (33)
        'IT-Keyword Match',          # AH (34)
        'Ist Junior-Job',            # AI (35)
        'Ist IT-Job',                # AJ (36)
        
        # === Meta ===
        'Scrape-Datum',              # AK (37)
        'Arbeitgeber-ID',            # AL (38)
    ]
    
    # ============================================================
    # STYLES
    # ============================================================
    header_font = Font(bold=True, color="FFFFFF", size=11, name='Calibri')
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Category colors for conditional formatting
    CATEGORY_COLORS = {
        'Softwareentwicklung': 'DAEEF3',    # light blue
        'System & Infrastruktur': 'E2EFDA', # light green
        'Data & Analytics': 'FCE4D6',       # light orange
        'DevOps & Cloud': 'D9E2F3',         # light blue 2
        'IT-Sicherheit': 'F2DCDB',          # light red
        'IT-Consulting': 'E4DFEC',          # light purple
        'IT-Support': 'FFF2CC',             # light yellow
        'Einstieg & Ausbildung': 'D6DCE4',  # light grey-blue
        'ERP & CRM': 'DEEBF7',              # very light blue
        'Digitalisierung': 'EBF1DE',        # light green 2
        'IT-Projektmanagement': 'F8CBAD',   # light coral
        'QA & Testing': 'C6EFCE',           # light mint
        'UX/UI Design': 'F4CCCC',           # pink
        'IT (Sonstige)': 'F2F2F2',          # very light grey
    }
    
    # Write headers
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Column widths
    widths = {
        'A': 10, 'B': 45, 'C': 22, 'D': 22, 'E': 14, 'F': 14, 'G': 18,
        'H': 16, 'I': 10, 'J': 10, 'K': 12, 'L': 16, 'M': 20, 'N': 35, 'O': 12, 'P': 16,
        'Q': 25, 'R': 18, 'S': 28, 'T': 8, 'U': 25,
        'V': 38, 'W': 28, 'X': 30, 'Y': 22,
        'Z': 16, 'AA': 45, 'AB': 22, 'AC': 14, 'AD': 14, 'AE': 18, 'AF': 18,
        'AG': 22, 'AH': 18, 'AI': 12, 'AJ': 12,
        'AK': 18, 'AL': 16,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # ============================================================
    # MIGRATE DATA
    # ============================================================
    print("\n[4] Migriere Daten mit abgeleiteten Spalten...")
    
    # Map old column names to indices
    col_map = {}
    for c in range(1, ws_in.max_column + 1):
        col_map[ws_in.cell(1, c).value] = c
    
    row_count = 0
    today = datetime.now()
    
    for r in range(2, ws_in.max_row + 1):
        # Read old data
        def get_val(col_name):
            idx = col_map.get(col_name)
            if idx is None:
                return None
            return ws_in.cell(r, idx).value
        
        refnr = get_val('Arbeitsagentur RefNr') or ''
        json_job = refnr_to_json.get(refnr, {})
        
        # Get fields from JSON if available (more complete data)
        title = get_val('Job-Titel') or json_job.get('title', '')
        city_orig = json_job.get('city', '') or get_val('Ort') or ''
        city_clean = clean_city(city_orig)
        region = classify_region(city_clean)
        junior_kw = json_job.get('junior_kw', '') or get_val('Junior-Keyword Match') or ''
        hauptberuf = json_job.get('hauptberuf', '') or get_val('Hauptberuf') or ''
        veroeffdatum = json_job.get('veroeffdatum', '') or get_val('Veroeffentlichungsdatum') or ''
        eintrittsdatum = json_job.get('eintrittsdatum', '') or get_val('Eintrittsdatum') or ''
        homeoffice = json_job.get('homeoffice', False)
        homeoffice_typ = json_job.get('homeofficetyp', '') or get_val('Homeoffice Typ') or ''
        vollzeit = json_job.get('vollzeit', True)
        befristet = json_job.get('befristet', False)
        quereinstieg = json_job.get('quereinstieg', False)
        
        # Derived columns
        it_bereich = classify_it_bereich(title, hauptberuf)
        explicit_junior = classify_explicit_junior(junior_kw)
        
        # Junior type classification
        junior_typ = 'Junior-fähig (kein Senior)'
        if junior_kw == 'ausbildung' or 'ausbildung' in title.lower():
            junior_typ = 'Ausbildung'
        elif junior_kw == 'duales studium' or 'duales studium' in title.lower():
            junior_typ = 'Duales Studium'
        elif junior_kw == 'werkstudent' or 'werkstudent' in title.lower():
            junior_typ = 'Werkstudent'
        elif junior_kw == 'praktikant' or 'praktikum' in title.lower():
            junior_typ = 'Praktikum'
        elif junior_kw == 'junior' or 'junior' in title.lower():
            junior_typ = 'Junior'
        elif junior_kw == 'associate' or 'associate' in title.lower():
            junior_typ = 'Associate'
        elif junior_kw == 'einsteiger' or 'einsteiger' in title.lower():
            junior_typ = 'Berufseinsteiger'
        elif junior_kw == 'absolvent' or 'absolvent' in title.lower():
            junior_typ = 'Absolvent'
        elif junior_kw == 'trainee' or 'trainee' in title.lower():
            junior_typ = 'Trainee'
        
        # Days ago
        days = days_ago(veroeffdatum)
        if days != '' and days >= 0:
            if days <= 7:
                alter = 'Frisch (≤ 7 Tage)'
            elif days <= 30:
                alter = 'Aktuell (8-30 Tage)'
            elif days <= 90:
                alter = 'Älter (31-90 Tage)'
            else:
                alter = 'Veraltet (> 90 Tage)'
        else:
            alter = 'Unbekannt'
        
        # Build row values
        row_data = [
            get_val('Job-ID'),                           # A: Job-ID
            title,                                        # B: Job-Titel
            get_val('Job-Kategorie'),                     # C: Job-Kategorie
            it_bereich,                                   # D: IT-Bereich
            get_val('Erfahrungslevel'),                   # E: Erfahrungslevel
            explicit_junior,                              # F: Explizit Junior
            junior_typ,                                   # G: Junior-Typ
            get_val('Beschäftigungsart'),                 # H: Beschäftigungsart
            'Ja' if vollzeit else 'Nein',                # I: Vollzeit
            'Ja' if befristet else 'Nein',               # J: Befristet
            'Ja' if homeoffice else 'Nein',              # K: Homeoffice
            get_val('Remote-Option'),                     # L: Remote-Option
            homeoffice_typ if homeoffice_typ else '',     # M: Homeoffice-Typ
            get_val('Technologien'),                      # N: Technologien
            'Ja' if quereinstieg else 'Nein',            # O: Quereinstieg
            get_val('Gehaltsspanne'),                     # P: Gehaltsspanne
            city_orig,                                    # Q: Ort (Original)
            city_clean,                                   # R: Stadt
            region,                                       # S: Region
            json_job.get('plz', '') or get_val('PLZ') or '',  # T: PLZ
            json_job.get('street', '') or get_val('Straße') or '',  # U: Straße
            get_val('Arbeitgeber') or json_job.get('employer_name', ''),  # V: Arbeitgeber
            get_val('Arbeitgeber (norm.)'),               # W: Arbeitgeber (norm.)
            hauptberuf,                                   # X: Hauptberuf
            get_val('Branche'),                           # Y: Branche
            get_val('Quelle') or json_job.get('source', ''),  # Z: Quelle
            get_val('Job-URL') or json_job.get('url', ''),  # AA: Job-URL
            refnr,                                        # AB: RefNr
            eintrittsdatum,                               # AC: Eintrittsdatum
            veroeffdatum,                                 # AD: Veröffentlicht am
            days if days != '' else '',                   # AE: Veröffentlicht vor
            alter,                                        # AF: Alter der Ausschreibung
            junior_kw,                                    # AG: Junior-Keyword Match
            get_val('IT-Keyword Match') or json_job.get('it_kw', ''),  # AH: IT-Keyword Match
            get_val('Ist Junior-Job'),                    # AI: Ist Junior-Job
            get_val('Ist IT-Job'),                        # AJ: Ist IT-Job
            get_val('Scrape-Datum'),                      # AK: Scrape-Datum
            get_val('Arbeitgeber-ID') or json_job.get('employer_hash', ''),  # AL: Arbeitgeber-ID
        ]
        
        row_count += 1
        row_idx = row_count + 1  # +1 for header
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = thin_border
            cell.font = Font(name='Calibri', size=10)
    
    total_rows = row_count
    print(f"  {total_rows} Jobs migriert mit {len(HEADERS)} Spalten")
    
    # ============================================================
    # AUTO-FILTER
    # ============================================================
    print("\n[5] Setze Auto-Filter auf alle Spalten...")
    last_col = get_column_letter(len(HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{total_rows + 1}"
    
    # ============================================================
    # EXCEL TABLE (for better filter UI with dropdowns)
    # ============================================================
    print("\n[6] Erstelle Excel-Tabelle mit Filter-Dropdowns...")
    table = Table(
        displayName="JuniorITJobs",
        ref=f"A1:{last_col}{total_rows + 1}"
    )
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)
    
    # ============================================================
    # FROZEN PANES
    # ============================================================
    print("\n[7] Fixiere Kopfzeile und erste 2 Spalten...")
    ws.freeze_panes = 'C2'  # Freeze row 1 and columns A-B
    
    # ============================================================
    # DATA VALIDATION (Dropdown-Filter)
    # ============================================================
    print("\n[8] Füge Datenvalidierung (Dropdowns) hinzu...")
    
    # Job-Kategorie dropdown (Column C)
    job_kategorien = [
        'Softwareentwicklung', 'Systemadministration', 'IT-Consulting',
        'Data Science & Analytics', 'DevOps & Cloud', 'Cyber Security',
        'IT-Support', 'IT-Projektmanagement', 'QA & Testing',
        'UX/UI Design', 'Datenbankadministration', 'IT-Ausbildung & Studium',
        'IT (Sonstige)'
    ]
    dv_kategorie = DataValidation(
        type="list",
        formula1='"' + ','.join(job_kategorien) + '"',
        allow_blank=True,
        showDropDown=False  # In openpyxl, False means SHOW the dropdown
    )
    dv_kategorie.prompt = "Job-Kategorie wählen"
    dv_kategorie.promptTitle = "Job-Kategorie"
    ws.add_data_validation(dv_kategorie)
    dv_kategorie.add(f"C2:C{total_rows + 1}")
    
    # IT-Bereich dropdown (Column D)
    it_bereiche = [
        'Softwareentwicklung', 'System & Infrastruktur', 'Data & Analytics',
        'DevOps & Cloud', 'IT-Sicherheit', 'IT-Consulting', 'IT-Support',
        'IT-Projektmanagement', 'QA & Testing', 'UX/UI Design',
        'Einstieg & Ausbildung', 'ERP & CRM', 'Digitalisierung', 'IT (Sonstige)'
    ]
    dv_bereich = DataValidation(
        type="list",
        formula1='"' + ','.join(it_bereiche) + '"',
        allow_blank=True,
        showDropDown=False
    )
    ws.add_data_validation(dv_bereich)
    dv_bereich.add(f"D2:D{total_rows + 1}")
    
    # Erfahrungslevel (Column E)
    dv_level = DataValidation(
        type="list",
        formula1='"Junior,Mid-Level,Senior,Trainee,Praktikant,Auszubildender"',
        allow_blank=True,
        showDropDown=False
    )
    ws.add_data_validation(dv_level)
    dv_level.add(f"E2:E{total_rows + 1}")
    
    # Explizit Junior (Column F)
    dv_explicit = DataValidation(
        type="list",
        formula1='"Ja,Nein"',
        allow_blank=True,
        showDropDown=False
    )
    ws.add_data_validation(dv_explicit)
    dv_explicit.add(f"F2:F{total_rows + 1}")
    
    # Junior-Typ (Column G)
    junior_typen = [
        'Junior', 'Associate', 'Ausbildung', 'Duales Studium',
        'Werkstudent', 'Praktikum', 'Trainee', 'Berufseinsteiger',
        'Absolvent', 'Junior-fähig (kein Senior)'
    ]
    dv_jtyp = DataValidation(
        type="list",
        formula1='"' + ','.join(junior_typen) + '"',
        allow_blank=True,
        showDropDown=False
    )
    ws.add_data_validation(dv_jtyp)
    dv_jtyp.add(f"G2:G{total_rows + 1}")
    
    # Beschäftigungsart (Column H)
    dv_besch = DataValidation(
        type="list",
        formula1='"Vollzeit,Teilzeit,Werkstudent,Praktikum,Ausbildung,Duales Studium,Minijob"',
        allow_blank=True,
        showDropDown=False
    )
    ws.add_data_validation(dv_besch)
    dv_besch.add(f"H2:H{total_rows + 1}")
    
    # Ja/Nein columns (I, J, K, O)
    for col in ['I', 'J', 'K', 'O']:
        dv_jn = DataValidation(
            type="list",
            formula1='"Ja,Nein"',
            allow_blank=True,
            showDropDown=False
        )
        ws.add_data_validation(dv_jn)
        dv_jn.add(f"{col}2:{col}{total_rows + 1}")
    
    # Remote-Option (Column L)
    dv_remote = DataValidation(
        type="list",
        formula1='"Remote,Hybrid,Vor Ort,Nicht angegeben"',
        allow_blank=True,
        showDropDown=False
    )
    ws.add_data_validation(dv_remote)
    dv_remote.add(f"L2:L{total_rows + 1}")
    
    # Region (Column S)
    regionen = [
        'Metropolregion Nürnberg (Kern)', 'Oberfranken',
        'Mittelfranken (Land)', 'Erlangen-Umland', 'Weitere Regionen'
    ]
    dv_region = DataValidation(
        type="list",
        formula1='"' + ','.join(regionen) + '"',
        allow_blank=True,
        showDropDown=False
    )
    ws.add_data_validation(dv_region)
    dv_region.add(f"S2:S{total_rows + 1}")
    
    # Alter der Ausschreibung (Column AF)
    dv_alter = DataValidation(
        type="list",
        formula1='"Frisch (≤ 7 Tage),Aktuell (8-30 Tage),Älter (31-90 Tage),Veraltet (> 90 Tage),Unbekannt"',
        allow_blank=True,
        showDropDown=False
    )
    ws.add_data_validation(dv_alter)
    dv_alter.add(f"AF2:AF{total_rows + 1}")
    
    # Quelle (Column Z)
    dv_quelle = DataValidation(
        type="list",
        formula1='"Arbeitsagentur,Website,Website (JSON-LD),Stepstone,Indeed"',
        allow_blank=True,
        showDropDown=False
    )
    ws.add_data_validation(dv_quelle)
    dv_quelle.add(f"Z2:Z{total_rows + 1}")
    
    # ============================================================
    # CONDITIONAL FORMATTING
    # ============================================================
    print("\n[9] Füge bedingte Formatierung hinzu...")
    
    # IT-Bereich colors (Column D)
    for bereich, color in CATEGORY_COLORS.items():
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.conditional_formatting.add(
            f"D2:D{total_rows + 1}",
            CellIsRule(operator='equal', formula=[f'"{bereich}"'], fill=fill)
        )
    
    # Explicit Junior = Ja -> green highlight (Column F)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ws.conditional_formatting.add(
        f"F2:F{total_rows + 1}",
        CellIsRule(operator='equal', formula=['"Ja"'], fill=green_fill)
    )
    
    # Homeoffice = Ja -> blue highlight (Column K)
    blue_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    ws.conditional_formatting.add(
        f"K2:K{total_rows + 1}",
        CellIsRule(operator='equal', formula=['"Ja"'], fill=blue_fill)
    )
    
    # Remote -> purple highlight (Column L)
    purple_fill = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")
    ws.conditional_formatting.add(
        f"L2:L{total_rows + 1}",
        CellIsRule(operator='equal', formula=['"Remote"'], fill=purple_fill)
    )
    ws.conditional_formatting.add(
        f"L2:L{total_rows + 1}",
        CellIsRule(operator='equal', formula=['"Hybrid"'], fill=PatternFill(
            start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"
        ))
    )
    
    # Age: Frisch -> green, Veraltet -> red (Column AF)
    ws.conditional_formatting.add(
        f"AF2:AF{total_rows + 1}",
        CellIsRule(operator='equal', formula=['"Frisch (≤ 7 Tage)"'], fill=green_fill)
    )
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ws.conditional_formatting.add(
        f"AF2:AF{total_rows + 1}",
        CellIsRule(operator='equal', formula=['"Veraltet (> 90 Tage)"'], fill=red_fill)
    )
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    ws.conditional_formatting.add(
        f"AF2:AF{total_rows + 1}",
        CellIsRule(operator='equal', formula=['"Aktuell (8-30 Tage)"'], fill=yellow_fill)
    )
    
    # Quereinstieg = Ja -> orange highlight (Column O)
    orange_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    ws.conditional_formatting.add(
        f"O2:O{total_rows + 1}",
        CellIsRule(operator='equal', formula=['"Ja"'], fill=orange_fill)
    )
    
    # ============================================================
    # STATISTICS SHEET
    # ============================================================
    print("\n[10] Erstelle Statistik-Blatt...")
    
    if 'Statistiken' in wb.sheetnames:
        del wb['Statistiken']
    ws_stats = wb.create_sheet('Statistiken')
    
    # Title
    title_font = Font(bold=True, size=16, color="1F4E79", name='Calibri')
    subtitle_font = Font(bold=True, size=12, color="1F4E79", name='Calibri')
    stat_label_font = Font(bold=True, size=10, name='Calibri')
    stat_value_font = Font(size=10, name='Calibri')
    stat_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
    
    ws_stats.column_dimensions['A'].width = 40
    ws_stats.column_dimensions['B'].width = 15
    ws_stats.column_dimensions['C'].width = 5
    ws_stats.column_dimensions['D'].width = 40
    ws_stats.column_dimensions['E'].width = 15
    ws_stats.column_dimensions['F'].width = 5
    ws_stats.column_dimensions['G'].width = 40
    ws_stats.column_dimensions['H'].width = 15
    
    r = 1
    ws_stats.cell(r, 1, "Junior IT-Jobs Bamberg/Erlangen/Nürnberg - Statistiken").font = title_font
    r += 1
    ws_stats.cell(r, 1, f"Stand: {datetime.now().strftime('%d.%m.%Y %H:%M')}").font = Font(size=10, italic=True, color="666666", name='Calibri')
    r += 2
    
    # --- Overview ---
    ws_stats.cell(r, 1, "ÜBERSICHT").font = subtitle_font
    r += 1
    overview = [
        ('Gesamt Junior IT-Jobs', total_rows),
        ('Davon explizit Junior', sum(1 for i in range(2, total_rows + 2) if ws.cell(i, 6).value == 'Ja')),
        ('Davon Junior-fähig (kein Senior)', sum(1 for i in range(2, total_rows + 2) if ws.cell(i, 6).value == 'Nein')),
        ('Mit Homeoffice möglich', sum(1 for i in range(2, total_rows + 2) if ws.cell(i, 11).value == 'Ja')),
        ('Mit Remote möglich', sum(1 for i in range(2, total_rows + 2) if ws.cell(i, 12).value in ['Remote', 'Hybrid'])),
        ('Für Quereinsteiger', sum(1 for i in range(2, total_rows + 2) if ws.cell(i, 15).value == 'Ja')),
        ('Vollzeit', sum(1 for i in range(2, total_rows + 2) if ws.cell(i, 9).value == 'Ja')),
        ('Unbefristet', sum(1 for i in range(2, total_rows + 2) if ws.cell(i, 10).value == 'Nein')),
    ]
    for label, value in overview:
        ws_stats.cell(r, 1, label).font = stat_label_font
        ws_stats.cell(r, 2, value).font = stat_value_font
        ws_stats.cell(r, 2).alignment = Alignment(horizontal='right')
        r += 1
    
    r += 1
    
    # --- By IT-Bereich ---
    ws_stats.cell(r, 1, "NACH IT-BEREICH").font = subtitle_font
    r += 1
    bereich_counter = Counter()
    for i in range(2, total_rows + 2):
        bereich_counter[ws.cell(i, 4).value or 'Unbekannt'] += 1
    for bereich, count in bereich_counter.most_common():
        ws_stats.cell(r, 1, bereich).font = stat_label_font
        ws_stats.cell(r, 2, count).font = stat_value_font
        ws_stats.cell(r, 2).alignment = Alignment(horizontal='right')
        pct = count / total_rows * 100
        ws_stats.cell(r, 3, f"{pct:.1f}%").font = Font(size=9, color="666666", name='Calibri')
        r += 1
    
    r += 1
    
    # --- By Junior-Typ ---
    ws_stats.cell(r, 1, "NACH JUNIOR-TYP").font = subtitle_font
    r += 1
    jtyp_counter = Counter()
    for i in range(2, total_rows + 2):
        jtyp_counter[ws.cell(i, 7).value or 'Unbekannt'] += 1
    for jtyp, count in jtyp_counter.most_common():
        ws_stats.cell(r, 1, jtyp).font = stat_label_font
        ws_stats.cell(r, 2, count).font = stat_value_font
        ws_stats.cell(r, 2).alignment = Alignment(horizontal='right')
        r += 1
    
    r += 1
    
    # --- By Region ---
    start_col = 1
    ws_stats.cell(r, start_col, "NACH REGION").font = subtitle_font
    ws_stats.cell(r, start_col + 3, "TOP 15 STÄDTE").font = subtitle_font
    r += 1
    
    region_counter = Counter()
    for i in range(2, total_rows + 2):
        region_counter[ws.cell(i, 19).value or 'Unbekannt'] += 1
    city_counter = Counter()
    for i in range(2, total_rows + 2):
        city_counter[ws.cell(i, 18).value or 'Unbekannt'] += 1
    
    region_rows = list(region_counter.most_common())
    city_rows = list(city_counter.most_common(15))
    max_rows = max(len(region_rows), len(city_rows))
    
    for idx in range(max_rows):
        if idx < len(region_rows):
            ws_stats.cell(r, 1, region_rows[idx][0]).font = stat_label_font
            ws_stats.cell(r, 2, region_rows[idx][1]).font = stat_value_font
            ws_stats.cell(r, 2).alignment = Alignment(horizontal='right')
        if idx < len(city_rows):
            ws_stats.cell(r, 4, city_rows[idx][0]).font = stat_label_font
            ws_stats.cell(r, 5, city_rows[idx][1]).font = stat_value_font
            ws_stats.cell(r, 5).alignment = Alignment(horizontal='right')
        r += 1
    
    r += 1
    
    # --- By Remote-Option ---
    ws_stats.cell(r, 1, "REMOTE-OPTIONEN").font = subtitle_font
    ws_stats.cell(r, 4, "ALTER DER AUSSCHREIBUNG").font = subtitle_font
    r += 1
    
    remote_counter = Counter()
    for i in range(2, total_rows + 2):
        remote_counter[ws.cell(i, 12).value or 'Unbekannt'] += 1
    age_counter = Counter()
    for i in range(2, total_rows + 2):
        age_counter[ws.cell(i, 32).value or 'Unbekannt'] += 1
    
    remote_rows = list(remote_counter.most_common())
    age_rows_list = [
        ('Frisch (≤ 7 Tage)', age_counter.get('Frisch (≤ 7 Tage)', 0)),
        ('Aktuell (8-30 Tage)', age_counter.get('Aktuell (8-30 Tage)', 0)),
        ('Älter (31-90 Tage)', age_counter.get('Älter (31-90 Tage)', 0)),
        ('Veraltet (> 90 Tage)', age_counter.get('Veraltet (> 90 Tage)', 0)),
        ('Unbekannt', age_counter.get('Unbekannt', 0)),
    ]
    
    max_rows2 = max(len(remote_rows), len(age_rows_list))
    for idx in range(max_rows2):
        if idx < len(remote_rows):
            ws_stats.cell(r, 1, remote_rows[idx][0]).font = stat_label_font
            ws_stats.cell(r, 2, remote_rows[idx][1]).font = stat_value_font
            ws_stats.cell(r, 2).alignment = Alignment(horizontal='right')
        if idx < len(age_rows_list):
            ws_stats.cell(r, 4, age_rows_list[idx][0]).font = stat_label_font
            ws_stats.cell(r, 5, age_rows_list[idx][1]).font = stat_value_font
            ws_stats.cell(r, 5).alignment = Alignment(horizontal='right')
        r += 1
    
    r += 1
    
    # --- Top 20 Arbeitgeber ---
    ws_stats.cell(r, 1, "TOP 20 ARBEITGEBER").font = subtitle_font
    r += 1
    employer_counter = Counter()
    for i in range(2, total_rows + 2):
        emp = ws.cell(i, 22).value or 'Unbekannt'
        employer_counter[emp.strip()] += 1
    
    for emp, count in employer_counter.most_common(20):
        ws_stats.cell(r, 1, emp).font = stat_label_font
        ws_stats.cell(r, 2, count).font = stat_value_font
        ws_stats.cell(r, 2).alignment = Alignment(horizontal='right')
        r += 1
    
    r += 1
    
    # --- Top Technologien ---
    ws_stats.cell(r, 1, "TOP 25 TECHNOLOGIEN").font = subtitle_font
    r += 1
    tech_counter = Counter()
    for i in range(2, total_rows + 2):
        techs = ws.cell(i, 14).value
        if techs:
            for tech in str(techs).split(', '):
                if tech.strip():
                    tech_counter[tech.strip().lower()] += 1
    
    for tech, count in tech_counter.most_common(25):
        ws_stats.cell(r, 1, tech).font = stat_label_font
        ws_stats.cell(r, 2, count).font = stat_value_font
        ws_stats.cell(r, 2).alignment = Alignment(horizontal='right')
        r += 1
    
    # ============================================================
    # FILTER-SCHNELLZUGRIFF SHEET (Quick filter guide)
    # ============================================================
    print("\n[11] Erstelle Filter-Schnellzugriff-Blatt...")
    
    if 'Filter-Guide' in wb.sheetnames:
        del wb['Filter-Guide']
    ws_guide = wb.create_sheet('Filter-Guide')
    
    ws_guide.column_dimensions['A'].width = 25
    ws_guide.column_dimensions['B'].width = 50
    ws_guide.column_dimensions['C'].width = 60
    
    guide_font = Font(bold=True, size=12, color="1F4E79", name='Calibri')
    item_font = Font(size=10, name='Calibri')
    bold_font = Font(bold=True, size=10, name='Calibri')
    
    r = 1
    ws_guide.cell(r, 1, "Filter-Schnellzugriff - Junior IT-Jobs").font = Font(bold=True, size=16, color="1F4E79", name='Calibri')
    r += 2
    
    ws_guide.cell(r, 1, "SPALTE").font = bold_font
    ws_guide.cell(r, 2, "FILTER-OPTION").font = bold_font
    ws_guide.cell(r, 3, "BESCHREIBUNG").font = bold_font
    r += 1
    
    filters = [
        ('IT-Bereich (D)', 'Softwareentwicklung', 'Nur Softwareentwickler-Jobs anzeigen'),
        ('IT-Bereich (D)', 'Einstieg & Ausbildung', 'Ausbildung, Duales Studium, Trainee-Programme'),
        ('IT-Bereich (D)', 'IT-Sicherheit', 'Cyber Security, Penetration Testing, SOC'),
        ('IT-Bereich (D)', 'DevOps & Cloud', 'DevOps, Cloud Engineering, SRE'),
        ('IT-Bereich (D)', 'Data & Analytics', 'Data Science, BI, Big Data'),
        ('IT-Bereich (D)', 'IT-Support', 'Helpdesk, First/Second Level Support'),
        ('', '', ''),
        ('Explizit Junior (F)', 'Ja', 'Nur Jobs mit explizitem Junior-Keyword'),
        ('Explizit Junior (F)', 'Nein', 'Jobs ohne Senior-Keyword (einstiegsfreundlich)'),
        ('', '', ''),
        ('Junior-Typ (G)', 'Junior', 'Explizit als "Junior" ausgeschrieben'),
        ('Junior-Typ (G)', 'Ausbildung', 'Ausbildungsplätze (Fachinformatiker etc.)'),
        ('Junior-Typ (G)', 'Duales Studium', 'Duales Studium Informatik/Wirtschaftsinformatik'),
        ('Junior-Typ (G)', 'Werkstudent', 'Werkstudentenstellen'),
        ('Junior-Typ (G)', 'Praktikum', 'Praktikumsplätze'),
        ('Junior-Typ (G)', 'Trainee', 'Trainee-Programme'),
        ('', '', ''),
        ('Homeoffice (K)', 'Ja', 'Homeoffice möglich'),
        ('Remote-Option (L)', 'Remote', 'Vollständig remote möglich'),
        ('Remote-Option (L)', 'Hybrid', 'Hybrid-Arbeitsmodell'),
        ('', '', ''),
        ('Region (S)', 'Metropolregion Nürnberg (Kern)', 'Nürnberg, Erlangen, Fürth, Schwabach'),
        ('Region (S)', 'Oberfranken', 'Bamberg, Bayreuth, Coburg, Forchheim'),
        ('Region (S)', 'Erlangen-Umland', 'Herzogenaurach, Baiersdorf, Eckental'),
        ('', '', ''),
        ('Quereinstieg (O)', 'Ja', 'Auch für Quereinsteiger geeignet'),
        ('Vollzeit (I)', 'Nein', 'Teilzeit/Werkstudent/Praktikum'),
        ('Befristet (J)', 'Nein', 'Unbefristete Stellen'),
        ('', '', ''),
        ('Alter (AF)', 'Frisch (≤ 7 Tage)', 'In den letzten 7 Tagen veröffentlicht'),
        ('Alter (AF)', 'Aktuell (8-30 Tage)', 'Vor 8-30 Tagen veröffentlicht'),
    ]
    
    for col_name, filter_opt, description in filters:
        if col_name == '' and filter_opt == '':
            r += 1
            continue
        ws_guide.cell(r, 1, col_name).font = bold_font
        ws_guide.cell(r, 2, filter_opt).font = item_font
        ws_guide.cell(r, 3, description).font = item_font
        r += 1
    
    # ============================================================
    # SAVE
    # ============================================================
    print("\n[12] Speichere Excel-Datei...")
    wb.save(EXCEL_OUT)
    
    file_size = Path(EXCEL_OUT).stat().st_size / 1024 / 1024
    print(f"  Gespeichert: {EXCEL_OUT} ({file_size:.1f} MB)")
    
    print("\n" + "=" * 70)
    print("ERGEBNIS:")
    print(f"  {total_rows} Junior IT-Jobs")
    print(f"  {len(HEADERS)} Spalten (inkl. neuer abgeleiteter Filter-Spalten)")
    print("  Auto-Filter auf allen Spalten")
    print("  Excel-Tabelle mit Dropdown-Filtern")
    print("  Datenvalidierung für 10+ kategorische Spalten")
    print("  Bedingte Formatierung (IT-Bereich, Junior, Remote, Alter)")
    print("  Fixierte Kopfzeile + erste 2 Spalten")
    print("  Statistik-Blatt mit Pivot-Übersichten")
    print("  Filter-Guide mit Schnellzugriff-Erklärungen")
    print("=" * 70)


if __name__ == '__main__':
    main()
