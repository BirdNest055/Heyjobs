#!/usr/bin/env python3
"""
Merge enrichment results with original Excel data and create final enriched Excel.
- Loads enrichment data (websites + emails from DDG search + scraping)
- Loads original data from arbeitsagentur_employers.json (addresses, phones)
- Deduplicates by employer name
- Creates comprehensive Excel with all data
"""

import json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from urllib.parse import urlparse

# Load enrichment results
with open('/home/z/my-project/website_search_results.json', 'r') as f:
    enrichment = json.load(f)

# Load original API data 
with open('/home/z/my-project/download/arbeitsagentur_employers.json', 'r') as f:
    api_data = json.load(f)

print(f"Enrichment results: {len(enrichment)}")
print(f"API data entries: {len(api_data)}")

# Build enrichment lookup by name
enrichment_by_name = {}
for key, val in enrichment.items():
    enrichment_by_name[key] = val

# Merge: for each API entry, add enrichment data
# Deduplicate by name - for same name, keep entry with most data
merged = {}
for entry in api_data:
    name = entry.get('Name', '').strip()
    if not name:
        continue
    
    key = name.lower().strip()
    
    # Get enrichment data
    enrich = enrichment_by_name.get(key, {})
    
    # Build merged entry
    current_website = entry.get('Website', '').strip()
    enrich_website = enrich.get('website', '').strip()
    
    # Filter bad websites
    bad_domains = ['duckduckgo.com', 'cylex.de', 'companyhouse.de', 'creditreform', 'kleinanzeigen.de']
    if any(d in enrich_website for d in bad_domains):
        enrich_website = ''
    
    # Choose best website (prefer enrichment, fall back to original)
    website = enrich_website or current_website
    
    # Clean up website (remove trailing paths that are too specific)
    if website:
        # Keep the main URL but preserve useful paths like /kontakt
        pass
    
    # Emails
    enrich_emails = enrich.get('emails', [])
    bewerbung_email = enrich.get('bewerbung_email', '')
    kontakt_email = enrich.get('kontakt_email', '')
    
    # Filter out bad emails
    bad_email_patterns = ['2x.webp', '3x.webp', 'beispiel.de', 'example.com', 'nutzer@', 'email.protected']
    enrich_emails = [e for e in enrich_emails if not any(p in e.lower() for p in bad_email_patterns)]
    if any(p in bewerbung_email.lower() for p in bad_email_patterns):
        bewerbung_email = ''
    if any(p in kontakt_email.lower() for p in bad_email_patterns):
        kontakt_email = ''
    
    # Choose best email
    best_email = bewerbung_email or kontakt_email or (enrich_emails[0] if enrich_emails else '')
    
    merged_entry = {
        'Arbeitgeber_ID': entry.get('Arbeitgeber_ID', ''),
        'Name': name,
        'Straße': entry.get('Straße', ''),
        'PLZ': entry.get('PLZ', ''),
        'Stadt': entry.get('Stadt', ''),
        'Bundesland': entry.get('Bundesland', ''),
        'Telefon': entry.get('Telefon', ''),
        'E-Mail': best_email,
        'Bewerbungs_E-Mail': bewerbung_email,
        'Kontakt_E-Mail': kontakt_email,
        'Alle_E_Mails': ', '.join(enrich_emails[:5]),
        'Website': website,
        'Branche': entry.get('Branche', 'IT, Computer, Telekommunikation'),
        'Quelle': entry.get('Quelle', 'Arbeitsagentur'),
        'Bewerbungsstatus': entry.get('Bewerbungsstatus', 'Offen'),
        'Job_Titel': entry.get('_job_titel', ''),
        'Referenznummer': entry.get('_referenznummer', ''),
    }
    
    # For dedup by name: keep entry with most data
    if key not in merged:
        merged[key] = merged_entry
    else:
        existing = merged[key]
        # Replace if new entry has more data
        existing_data = sum(1 for v in existing.values() if v and str(v).strip() and str(v) != 'None')
        new_data = sum(1 for v in merged_entry.values() if v and str(v).strip() and str(v) != 'None')
        if new_data > existing_data:
            merged[key] = merged_entry

print(f"Merged unique employers: {len(merged)}")

# Sort by name
sorted_employers = sorted(merged.values(), key=lambda x: x.get('Name', '').lower())

# Create Excel
wb = openpyxl.Workbook()

# === Sheet 1: Arbeitgeberverzeichnis ===
ws = wb.active
ws.title = "Arbeitgeberverzeichnis"

# Title row
ws.merge_cells('B1:M1')
ws['B1'] = 'Arbeitgeberverzeichnis - Arbeitsagentur (IT/Telekommunikation) - Angereichert'
ws['B1'].font = Font(name='Calibri', size=16, bold=True, color='1F4E79')

# Headers
headers = ['Arbeitgeber_ID', 'Name', 'Straße', 'PLZ', 'Stadt', 'Bundesland', 'Telefon', 'E-Mail', 'Bewerbungs_E-Mail', 'Website', 'Branche', 'Quelle', 'Bewerbungsstatus']
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

for col, header in enumerate(headers, 2):
    cell = ws.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Data rows
alt_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for row_idx, emp in enumerate(sorted_employers, 5):
    values = [emp.get(h, '') for h in headers]
    for col_idx, val in enumerate(values, 2):
        cell = ws.cell(row=row_idx, column=col_idx, value=val if val and str(val) != 'None' else '')
        cell.font = Font(name='Calibri', size=10)
        cell.border = thin_border
        if row_idx % 2 == 0:
            cell.fill = alt_fill
        
        # Make website clickable
        if col_idx == 11 and val and str(val).startswith('http'):  # Website column
            cell.hyperlink = str(val)
            cell.font = Font(name='Calibri', size=10, color='0563C1', underline='single')
        
        # Make email clickable
        if col_idx in [9, 10] and val and '@' in str(val):  # Email columns
            cell.hyperlink = f'mailto:{val}'
            cell.font = Font(name='Calibri', size=10, color='0563C1', underline='single')

# Column widths
col_widths = {'B': 16, 'C': 42, 'D': 25, 'E': 8, 'F': 25, 'G': 20, 'H': 16, 'I': 30, 'J': 30, 'K': 40, 'L': 30, 'M': 15, 'N': 15}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

# Freeze panes
ws.freeze_panes = 'C5'

# AutoFilter
ws.auto_filter.ref = f'B4:N{4 + len(sorted_employers)}'

# === Sheet 2: Statistik Bundesland ===
ws2 = wb.create_sheet("Statistik Bundesland")
ws2.merge_cells('B1:D1')
ws2['B1'] = 'Statistik nach Bundesland'
ws2['B1'].font = Font(name='Calibri', size=14, bold=True)

# Count by Bundesland
bundesland_count = {}
bundesland_email = {}
for emp in sorted_employers:
    bl = emp.get('Bundesland', 'Unbekannt')
    bundesland_count[bl] = bundesland_count.get(bl, 0) + 1
    if emp.get('E-Mail'):
        bundesland_email[bl] = bundesland_email.get(bl, 0) + 1

ws2.cell(row=3, column=2, value='Bundesland').font = Font(bold=True)
ws2.cell(row=3, column=3, value='Arbeitgeber').font = Font(bold=True)
ws2.cell(row=3, column=4, value='Mit E-Mail').font = Font(bold=True)
ws2.cell(row=3, column=5, value='E-Mail Quote').font = Font(bold=True)

for i, (bl, count) in enumerate(sorted(bundesland_count.items(), key=lambda x: -x[1]), 4):
    ws2.cell(row=i, column=2, value=bl)
    ws2.cell(row=i, column=3, value=count)
    ws2.cell(row=i, column=4, value=bundesland_email.get(bl, 0))
    ws2.cell(row=i, column=5, value=f'{bundesland_email.get(bl, 0)/count*100:.1f}%')

ws2.column_dimensions['B'].width = 25
ws2.column_dimensions['C'].width = 15
ws2.column_dimensions['D'].width = 12
ws2.column_dimensions['E'].width = 12

# === Sheet 3: Bewerbungstracker ===
ws3 = wb.create_sheet("Bewerbungstracker")
ws3.merge_cells('B1:H1')
ws3['B1'] = 'Bewerbungstracker'
ws3['B1'].font = Font(name='Calibri', size=14, bold=True)

tracker_headers = ['Name', 'E-Mail', 'Bewerbungs_E-Mail', 'Website', 'Stadt', 'Bundesland', 'Status', 'Datum', 'Notizen']
for col, h in enumerate(tracker_headers, 2):
    ws3.cell(row=3, column=col, value=h).font = Font(bold=True)

# Pre-fill with employers that have emails
email_employers = [e for e in sorted_employers if e.get('E-Mail')]
email_employers.sort(key=lambda x: (x.get('Bewerbungs_E-Mail', '') != '', x.get('Bundesland', '')))

for i, emp in enumerate(email_employers, 4):
    ws3.cell(row=i, column=2, value=emp.get('Name', ''))
    ws3.cell(row=i, column=3, value=emp.get('E-Mail', ''))
    ws3.cell(row=i, column=4, value=emp.get('Bewerbungs_E-Mail', ''))
    ws3.cell(row=i, column=5, value=emp.get('Website', ''))
    ws3.cell(row=i, column=6, value=emp.get('Stadt', ''))
    ws3.cell(row=i, column=7, value=emp.get('Bundesland', ''))
    ws3.cell(row=i, column=8, value='Offen')
    ws3.cell(row=i, column=9, value='')
    ws3.cell(row=i, column=10, value='')
    
    # Make email clickable
    for c in [3, 4]:
        val = ws3.cell(row=i, column=c).value
        if val and '@' in str(val):
            ws3.cell(row=i, column=c).hyperlink = f'mailto:{val}'
    # Make website clickable
    val = ws3.cell(row=i, column=5).value
    if val and str(val).startswith('http'):
        ws3.cell(row=i, column=5).hyperlink = str(val)

ws3.column_dimensions['B'].width = 40
ws3.column_dimensions['C'].width = 30
ws3.column_dimensions['D'].width = 30
ws3.column_dimensions['E'].width = 35
ws3.column_dimensions['F'].width = 20
ws3.column_dimensions['G'].width = 20
ws3.column_dimensions['H'].width = 12
ws3.column_dimensions['I'].width = 12
ws3.column_dimensions['J'].width = 25

# === Sheet 4: Hinweise ===
ws4 = wb.create_sheet("Hinweise")
ws4['B1'] = 'Hinweise zur Datenanreicherung'
ws4['B1'].font = Font(name='Calibri', size=14, bold=True)

total = len(sorted_employers)
with_website = sum(1 for e in sorted_employers if e.get('Website'))
with_email = sum(1 for e in sorted_employers if e.get('E-Mail'))
with_bewerbung = sum(1 for e in sorted_employers if e.get('Bewerbungs_E-Mail'))

notes = [
    f'Gesamtzahl Arbeitgeber: {total}',
    f'Mit Website: {with_website} ({with_website/total*100:.1f}%)',
    f'Mit E-Mail: {with_email} ({with_email/total*100:.1f}%)',
    f'Mit Bewerbungs-E-Mail: {with_bewerbung} ({with_bewerbung/total*100:.1f}%)',
    '',
    'Datenquellen:',
    '1. Arbeitsagentur REST API (Branche 11: IT/Telekommunikation)',
    '2. DuckDuckGo Websuche für offizielle Unternehmenswebsites',
    '3. HTTP-Scraping von Impressum/Kontaktseiten für E-Mail-Adressen',
    '',
    'Hinweis: E-Mail-Adressen wurden automatisch von Impressum- und Kontaktseiten extrahiert.',
    'Bitte verifizieren Sie die E-Mail-Adressen vor der Nutzung.',
    'Einige E-Mails könnten allgemeine Kontakt-Adressen sein, nicht spezifisch für Bewerbungen.',
    '',
    'Die Arbeitsagentur schützt Kontakt-E-Mails durch CAPTCHA - daher wurden diese extern ergänzt.',
]

for i, note in enumerate(notes, 3):
    ws4.cell(row=i, column=2, value=note)

ws4.column_dimensions['B'].width = 80

# Save
output_path = '/home/z/my-project/download/Arbeitgeber_Arbeitsagentur_Bewerbungskontakte_Angereichert.xlsx'
wb.save(output_path)
print(f"\n=== Excel erstellt: {output_path} ===")
print(f"Arbeitgeber: {total}")
print(f"Mit Website: {with_website} ({with_website/total*100:.1f}%)")
print(f"Mit E-Mail: {with_email} ({with_email/total*100:.1f}%)")
print(f"Mit Bewerbungs-E-Mail: {with_bewerbung} ({with_bewerbung/total*100:.1f}%)")
