#!/usr/bin/env python3
"""
Junior IT Jobs Scraper v2 - Enhanced version
Visits employer websites like a human + uses Arbeitsagentur API
for comprehensive junior-level IT job collection.
"""

import asyncio
import json
import hashlib
import os
import re
import subprocess
import sys
import time
import traceback
from datetime import datetime
from pathlib import Path
from urllib.parse import urljoin, urlparse, quote

import httpx
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright

# ============================================================
# CONFIG
# ============================================================
EXCEL_IN = "/home/z/my-project/download/IT_Arbeitgeber_Bamberg_Erlangen_Nuernberg.xlsx"
EXCEL_FULL = "/home/z/my-project/download/Firmen_Bamberg_Erlangen_Nuernberg_mit_Website_HTML.xlsx"
EXCEL_OUT = "/home/z/my-project/download/Junior_IT_Jobs_Bamberg_Erlangen_Nuernberg.xlsx"
CAREER_HTML_DIR = "/home/z/my-project/download/career_pages_html"
JOB_HTML_DIR = "/home/z/my-project/download/junior_jobs_html"
GIT_REPO = "/home/z/my-project"
GITHUB_TOKEN = "GIT_TOKEN_REDACTED"
COMMIT_EVERY = 10
MAX_CONCURRENT = 4
PAGE_TIMEOUT = 20000
JOB_COMMIT_COUNTER = 0

# Arbeitsagentur API
AA_API_URL = "https://rest.arbeitsagentur.de/jobboerse/jobsuche-service/pc/v6/jobs"
AA_API_KEY = "jobboerse-jobsuche"

# Junior-level keywords
JUNIOR_KEYWORDS = [
    'junior', 'entry level', 'einsteiger', 'anfänger', 'beginner',
    'trainee', 'praktikant', 'intern', 'werkstudent', 'working student',
    'absolvent', 'graduate', 'berufseinsteiger', 'young professional',
    'associate', 'graduate program', 'trainee program',
    'dual student', 'duales studium', 'ausbildung', 'apprentice',
    'werkstudent it', 'werkstudent software', 'praktikum it',
    'praktikum software', 'trainee it', 'trainee software',
    'berufseinsteiger it', 'absolvent informatik',
    'perspektive', 'start career', 'career starter',
    'nachwuchs', 'start-up', 'first job', 'erstjob',
]

# IT-related keywords
IT_KEYWORDS = [
    'software', 'entwickler', 'developer', 'it-', 'informatik',
    'data', 'devops', 'cloud', 'cyber', 'security', 'sicherheit',
    'system', 'admin', 'administrator', 'netzwerk', 'network',
    'frontend', 'backend', 'fullstack', 'full-stack', 'full stack',
    'python', 'java', 'javascript', 'typescript', 'react', 'angular',
    'sap', 'linux', 'sql', 'database', 'datenbank', 'api',
    'machine learning', 'ki', 'künstliche intelligenz', 'ai',
    'scrum', 'agile', 'testing', 'qa', 'quality', 'automation',
    'consultant', 'analyst', 'engineer', 'architekt', 'architect',
    'programmierung', 'programming', 'coding', 'code',
    'web', 'mobile', 'app', 'application', 'anwendung',
    'infrastructure', 'infrastruktur', 'platform', 'plattform',
    'support', 'helpdesk', 'service desk', 'ticket',
    'erp', 'crm', 'dynamics', 'sharepoint',
    'projekt', 'project', 'digitalisierung',
    'technologie', 'technology', 'digital',
]

# Career link text patterns
CAREER_LINK_TEXTS = [
    'karriere', 'career', 'jobs', 'stellenangebote', 'stellen',
    'offene stellen', 'vacancies', 'careers', 'jobbörse',
    'recruiting', 'bewerbung', 'talent', 'opportunities',
    'join us', 'join', 'wir suchen', 'wir stellen ein',
    'offene positionen', 'open positions', 'unser team',
    'arbeiten bei', 'work at', 'mitarbeiter', 'team',
]

# Career URL path patterns
CAREER_URL_PATTERNS = [
    '/karriere', '/career', '/careers', '/jobs', '/stellenangebote',
    '/stellen', '/offene-stellen', '/vacancies', '/job',
    '/jobboerse', '/recruiting', '/bewerbung', '/talent',
    '/opportunities', '/join-us', '/join', '/werde-teil',
    '/unsere-jobs', '/arbeiten-bei', '/offene-positionen',
    '/de/karriere', '/en/career', '/de/career', '/en/careers',
    '/de/jobs', '/en/jobs', '/unternehmen/karriere',
    '/ueber-uns/karriere', '/about/careers',
]


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def sanitize_filename(name, max_len=80):
    name = re.sub(r'[^\w\s.-]', '_', name)
    name = re.sub(r'_{2,}', '_', name)
    name = name.strip('_')
    return name[:max_len]


def is_junior_job(title, description=""):
    text = f"{title} {description}".lower()
    for kw in JUNIOR_KEYWORDS:
        if kw in text:
            return True, kw
    exp_match = re.search(r'(\d+)\s*(?:jahr|year|jahre|years)\s*(?:berufs)?erfahrung', text)
    if exp_match:
        years = int(exp_match.group(1))
        if years <= 2:
            return True, f"erfahrung_{years}jahre"
    if re.search(r'ohne\s*erfahrung|no\s*experience|keine\s*erfahrung|erfahrung\s*nicht', text):
        return True, "no_experience"
    return False, None


def is_it_job(title, description=""):
    text = f"{title} {description}".lower()
    for kw in IT_KEYWORDS:
        if kw in text:
            return True, kw
    return False, None


def is_senior_job(title):
    text = title.lower()
    return any(kw in text for kw in ['senior', 'lead', 'principal', 'head of', 'director', 'manager', 'leitung', 'chef'])


def extract_technologies(text):
    tech_keywords = [
        'python', 'java', 'javascript', 'typescript', 'c#', 'c++',
        'ruby', 'go', 'rust', 'kotlin', 'swift', 'scala', 'php',
        'react', 'angular', 'vue', 'vue.js', 'svelte', 'next.js',
        'node.js', 'nodejs', 'express', 'django', 'flask', 'spring',
        '.net', 'asp.net', 'docker', 'kubernetes', 'k8s', 'aws', 'azure', 'gcp',
        'terraform', 'ansible', 'jenkins', 'gitlab', 'github',
        'sql', 'mysql', 'postgresql', 'postgres', 'mongodb', 'redis',
        'kafka', 'rabbitmq', 'graphql', 'rest', 'api',
        'sap', 'salesforce', 'dynamics', 'sharepoint', 'servicenow',
        'linux', 'windows', 'macos', 'unix',
        'html', 'css', 'scss', 'sass', 'tailwind', 'bootstrap',
        'agile', 'scrum', 'kanban', 'devops', 'ci/cd',
        'machine learning', 'deep learning', 'tensorflow', 'pytorch',
        'nlp', 'computer vision', 'data science', 'big data',
        'jira', 'confluence', 'cybersecurity', 'pentest',
        'powershell', 'bash', 'shell', 'git',
        'azure devops', 'bitbucket', 'intellij', 'vscode',
    ]
    found = []
    text_lower = text.lower()
    for tech in tech_keywords:
        if re.search(r'\b' + re.escape(tech) + r'\b', text_lower):
            found.append(tech)
    return list(set(found))


def classify_job_category(title, description=""):
    text = f"{title} {description}".lower()
    categories = {
        'Softwareentwicklung': ['softwareentwickler', 'software developer', 'software engineer', 'entwickler', 'developer', 'programmierer', 'frontend', 'backend', 'fullstack', 'full-stack'],
        'Data Science & Analytics': ['data scientist', 'data analyst', 'data engineer', 'data analytics', 'business intelligence', 'machine learning', 'ki-entwickler', 'ai developer'],
        'IT-Consulting': ['it consultant', 'it-consultant', 'it berater', 'sap consultant', 'technischer berater', 'digitalisierungsberater'],
        'Systemadministration': ['system administrator', 'systemadministrator', 'sysadmin', 'admin', 'netzwerkadministrator', 'linux admin'],
        'DevOps & Cloud': ['devops', 'cloud engineer', 'cloud architect', 'platform engineer', 'sre', 'infrastruktur'],
        'Cyber Security': ['security', 'cybersecurity', 'information security', 'it-sicherheit', 'pentest', 'soc analyst'],
        'IT-Support': ['it support', 'helpdesk', 'service desk', 'first level', 'second level', 'user support'],
        'IT-Projektmanagement': ['project manager', 'projektmanager', 'scrum master', 'product owner'],
        'QA & Testing': ['quality assurance', 'qa', 'tester', 'test engineer', 'testautomatisierung'],
        'UX/UI Design': ['ux', 'ui', 'user experience', 'user interface', 'design', 'usability'],
        'Datenbankadministration': ['dba', 'datenbankadministrator', 'database administrator'],
        'IT-Ausbildung & Studium': ['ausbildung', 'duales studium', 'trainee', 'praktikum', 'werkstudent'],
    }
    for category, keywords in categories.items():
        for kw in keywords:
            if kw in text:
                return category
    return 'IT (Sonstige)'


def extract_employment_type(title, description=""):
    text = f"{title} {description}".lower()
    if any(kw in text for kw in ['teilzeit', 'part-time', 'part time']):
        return 'Teilzeit'
    if any(kw in text for kw in ['werkstudent', 'working student', 'studentische hilfskraft']):
        return 'Werkstudent'
    if any(kw in text for kw in ['praktikant', 'praktikum', 'intern', 'internship']):
        return 'Praktikum'
    if any(kw in text for kw in ['ausbildung', 'apprentice', 'azubi']):
        return 'Ausbildung'
    if any(kw in text for kw in ['duales studium', 'dual student']):
        return 'Duales Studium'
    if any(kw in text for kw in ['minijob', 'mini-job', 'geringfügig']):
        return 'Minijob'
    return 'Vollzeit'


def extract_remote_option(title, description=""):
    text = f"{title} {description}".lower()
    if any(kw in text for kw in ['hybrid', 'teilweise remote']):
        return 'Hybrid'
    if any(kw in text for kw in ['remote', 'homeoffice', 'home-office', 'home office', 'work from home', 'fully remote']):
        return 'Remote'
    if any(kw in text for kw in ['vor ort', 'on-site', 'onsite']):
        return 'Vor Ort'
    return 'Nicht angegeben'


# ============================================================
# ARBEITSAGENTUR API
# ============================================================

async def query_arbeitsagentur_jobs(client, employer_name, city=None):
    """Query the Arbeitsagentur API for junior IT jobs from a specific employer."""
    jobs = []
    
    # Search for each city
    cities = ['Erlangen', 'Nürnberg', 'Bamberg', 'Fürth', 'Forchheim']
    if city and city not in cities:
        cities.insert(0, city)
    
    for search_city in cities:
        for page in range(2):  # Check first 2 pages
            params = {
                'was': employer_name,
                'wo': search_city,
                'berufsfeld': '11',  # IT
                'page': page,
                'size': 25,
                'arbeitszeit': 'vz,tz',  # Voll- und Teilzeit
            }
            
            headers = {
                'X-API-Key': AA_API_KEY,
                'Accept': 'application/json',
            }
            
            try:
                response = await client.get(AA_API_URL, params=params, headers=headers, timeout=10)
                if response.status_code == 200:
                    data = response.json()
                    stellen = data.get('stellenangebote', [])
                    for stelle in stellen:
                        refnr = stelle.get('refnr', '')
                        titel = stelle.get('beruf', '') or stelle.get('titel', '')
                        arbeitgeber = stelle.get('arbeitgeber', '')
                        ort = stelle.get('arbeitsort', {})
                        ort_name = ort.get('ort', '') if isinstance(ort, dict) else str(ort)
                        
                        # Check if this job is from the target employer
                        employer_lower = employer_name.lower().replace(' ', '')
                        ag_lower = arbeitgeber.lower().replace(' ', '')
                        if employer_lower[:8] in ag_lower or ag_lower[:8] in employer_lower:
                            is_junior, junior_kw = is_junior_job(titel)
                            is_it_flag, it_kw = is_it_job(titel)
                            
                            if is_junior and is_it_flag:
                                jobs.append({
                                    'title': titel,
                                    'employer_name': arbeitgeber,
                                    'city': ort_name,
                                    'source': 'Arbeitsagentur',
                                    'refnr': refnr,
                                    'junior_keyword_match': junior_kw,
                                    'it_keyword_match': it_kw,
                                    'is_junior': True,
                                    'is_it': True,
                                    'url': f'https://www.arbeitsagentur.de/jobsuche/jobdetail/{refnr}',
                                    'description': '',
                                    'vollzeit': stelle.get('vollzeit', True),
                                    'befristet': stelle.get('befristet', False),
                                })
            except Exception as e:
                continue
    
    # Deduplicate by refnr
    seen = set()
    unique = []
    for j in jobs:
        if j.get('refnr') and j['refnr'] not in seen:
            seen.add(j['refnr'])
            unique.append(j)
    
    return unique


# ============================================================
# EXCEL MANAGEMENT
# ============================================================

class JuniorJobsExcel:
    HEADERS = [
        # Job identification
        'Job-ID', 'Job-Titel', 'Job-Titel (Original)', 'Job-Kategorie', 'Erfahrungslevel',
        # Job details
        'Beschäftigungsart', 'Remote-Option', 'Technologien',
        'Ausschreibungstext (Auszug)', 'Job-URL',
        # Employer info
        'Arbeitgeber', 'Arbeitgeber (norm.)', 'Ort', 'PLZ', 'Straße',
        'Telefon', 'E-Mail', 'Branche (Google)',
        # Source info
        'Quelle', 'Karriere-URL', 'Website-URL',
        'Karriere-HTML Pfad', 'Job-HTML Pfad',
        # Classification
        'Junior-Keyword Match', 'IT-Keyword Match', 'Ist Junior-Job', 'Ist IT-Job',
        # Contract details
        'Vollzeit/Teilzeit Detail', 'Befristet', 'Gehaltsspanne', 'Arbeitsagentur RefNr',
        # Metadata
        'Scrape-Datum', 'Scrape-Status', 'Arbeitgeber-ID',
        'Letzte Prüfung', 'Erstellt', 'Aktualisiert',
    ]
    
    def __init__(self, filepath):
        self.filepath = filepath
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Junior IT-Jobs"
        self._write_headers()
        self.row_count = 1
        self.job_counter = 0
    
    def _write_headers(self):
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))
        
        for col, header in enumerate(self.HEADERS, 1):
            cell = self.ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border
        
        widths = [10,40,45,22,14, 16,16,40,60,50, 40,30,15,8,25, 18,30,20, 14,50,50,50,50, 25,25,12,12, 20,12,20,20, 18,14,16,14,18,18]
        for i, w in enumerate(widths):
            self.ws.column_dimensions[get_column_letter(i+1)].width = w
    
    def add_job(self, job_data):
        self.job_counter += 1
        self.row_count += 1
        row = self.row_count
        
        values = [
            job_data.get('job_id', f'J-{self.job_counter:05d}'),
            job_data.get('job_title', ''),
            job_data.get('job_title_original', ''),
            job_data.get('job_category', ''),
            job_data.get('experience_level', ''),
            job_data.get('employment_type', ''),
            job_data.get('remote_option', ''),
            ', '.join(job_data.get('technologies', [])),
            job_data.get('description_excerpt', '')[:500],
            job_data.get('job_url', ''),
            job_data.get('employer_name', ''),
            job_data.get('employer_name_norm', ''),
            job_data.get('city', ''),
            job_data.get('plz', ''),
            job_data.get('street', ''),
            job_data.get('phone', ''),
            job_data.get('email', ''),
            job_data.get('industry', ''),
            job_data.get('source', 'Website'),
            job_data.get('career_url', ''),
            job_data.get('website_url', ''),
            job_data.get('career_html_path', ''),
            job_data.get('job_html_path', ''),
            job_data.get('junior_keyword_match', ''),
            job_data.get('it_keyword_match', ''),
            'Ja' if job_data.get('is_junior', False) else 'Nein',
            'Ja' if job_data.get('is_it', False) else 'Nein',
            job_data.get('employment_detail', ''),
            'Ja' if job_data.get('temporary', False) else 'Nein',
            job_data.get('salary_range', ''),
            job_data.get('refnr', ''),
            job_data.get('scrape_date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            job_data.get('scrape_status', ''),
            job_data.get('employer_id', ''),
            job_data.get('last_check', datetime.now().strftime('%Y-%m-%d')),
            job_data.get('created', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            job_data.get('updated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ]
        
        for col, value in enumerate(values, 1):
            cell = self.ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    def add_stats_sheet(self):
        """Add statistics sheet."""
        if 'Statistiken' in self.wb.sheetnames:
            del self.wb['Statistiken']
        ws_stats = self.wb.create_sheet('Statistiken')
        
        stats = [
            ('Gesamt Junior IT-Jobs', self.job_counter),
            ('Scrape-Datum', datetime.now().strftime('%Y-%m-%d %H:%M')),
        ]
        for i, (key, val) in enumerate(stats, 1):
            ws_stats.cell(i, 1, key)
            ws_stats.cell(i, 2, val)
    
    def save(self):
        self.add_stats_sheet()
        self.wb.save(self.filepath)
    
    def get_job_count(self):
        return self.job_counter


# ============================================================
# WEBSITE SCRAPER
# ============================================================

class WebsiteScraper:
    def __init__(self, excel_manager, semaphore):
        self.excel = excel_manager
        self.semaphore = semaphore
        self.jobs_since_commit = 0
        self.scraped_count = 0
        self.error_count = 0
        self.career_found_count = 0
    
    async def scrape_employer(self, context, employer_data):
        async with self.semaphore:
            website_url = employer_data.get('website_url')
            employer_name = employer_data.get('employer_name', 'Unknown')
            
            if not website_url:
                return []
            
            page = None
            try:
                page = await context.new_page()
                await page.set_extra_http_headers({'Accept-Language': 'de-DE,de;q=0.9,en-US;q=0.8,en;q=0.7'})
                
                # Navigate to homepage
                try:
                    await page.goto(website_url, timeout=PAGE_TIMEOUT, wait_until='domcontentloaded')
                    await asyncio.sleep(1.5)
                except:
                    return []
                
                self.scraped_count += 1
                
                # Step 1: Find career page
                career_url = await self._find_career_link(page, website_url)
                
                if not career_url:
                    career_url = await self._try_career_patterns(page, website_url)
                
                if not career_url:
                    return []
                
                self.career_found_count += 1
                print(f"  [✓] Career: {career_url[:70]}")
                
                # Step 2: Visit career page and scroll to load content
                try:
                    await page.goto(career_url, timeout=PAGE_TIMEOUT, wait_until='domcontentloaded')
                    await asyncio.sleep(2)
                    # Scroll down to trigger lazy loading
                    await page.evaluate('window.scrollTo(0, document.body.scrollHeight / 2)')
                    await asyncio.sleep(1)
                    await page.evaluate('window.scrollTo(0, document.body.scrollHeight)')
                    await asyncio.sleep(1)
                except:
                    pass
                
                # Save career page HTML
                career_html = await page.content()
                career_filename = sanitize_filename(employer_name) + '_career.html'
                career_filepath = os.path.join(CAREER_HTML_DIR, career_filename)
                with open(career_filepath, 'w', encoding='utf-8') as f:
                    f.write(career_html)
                
                # Step 3: Extract job listings
                jobs = await self._extract_jobs(page, career_url, employer_name, career_html)
                
                # Step 4: Visit up to 5 job detail pages for more info
                for job in jobs[:5]:
                    if job.get('url') and job['url'] != career_url:
                        try:
                            detail = await self._visit_job_detail(context, job)
                            if detail:
                                for k, v in detail.items():
                                    if v and not job.get(k):
                                        job[k] = v
                        except:
                            pass
                
                # Step 5: Filter for junior IT jobs
                junior_it_jobs = []
                for job in jobs:
                    is_j, junior_kw = is_junior_job(job.get('title', ''), job.get('description', ''))
                    is_i, it_kw = is_it_job(job.get('title', ''), job.get('description', ''))
                    is_sen = is_senior_job(job.get('title', ''))
                    
                    if is_j and is_i:
                        job['is_junior'] = True
                        job['is_it'] = True
                        job['junior_keyword_match'] = junior_kw
                        job['it_keyword_match'] = it_kw
                        junior_it_jobs.append(job)
                    elif is_i and not is_sen:
                        # IT job without senior keyword - might be junior-friendly
                        job['is_junior'] = True
                        job['is_it'] = True
                        job['junior_keyword_match'] = 'no_senior_keyword'
                        job['it_keyword_match'] = it_kw
                        junior_it_jobs.append(job)
                
                # Step 6: Add to Excel
                for job in junior_it_jobs:
                    job_data = self._build_job_record(job, employer_data, career_url, career_filepath)
                    self.excel.add_job(job_data)
                    self.jobs_since_commit += 1
                    global JOB_COMMIT_COUNTER
                    JOB_COMMIT_COUNTER += 1
                
                if junior_it_jobs:
                    print(f"  [✓✓] {len(junior_it_jobs)} junior IT jobs at {employer_name[:40]}")
                
                # Commit check
                if self.jobs_since_commit >= COMMIT_EVERY:
                    self._commit_and_push()
                    self.jobs_since_commit = 0
                
                return junior_it_jobs
                
            except Exception as e:
                self.error_count += 1
                print(f"  [✗] {employer_name[:40]}: {str(e)[:60]}")
                return []
            finally:
                if page:
                    try:
                        await page.close()
                    except:
                        pass
    
    async def _find_career_link(self, page, base_url):
        """Find career page link on homepage."""
        # Check for cookie consent banner first and dismiss it
        try:
            consent_selectors = [
                'button:has-text("Akzeptieren")', 'button:has-text("Accept")',
                'button:has-text("Alle akzeptieren")', 'button:has-text("Accept All")',
                'button:has-text("OK")', 'button:has-text("Verstanden")',
                '[class*="cookie"] button', '[id*="cookie"] button',
                '#onetrust-accept-btn-handler', '.cc-btn',
            ]
            for sel in consent_selectors:
                try:
                    btn = await page.query_selector(sel)
                    if btn:
                        await btn.click()
                        await asyncio.sleep(0.5)
                        break
                except:
                    continue
        except:
            pass
        
        # Look for career links
        selectors = [
            'a:has-text("Karriere")', 'a:has-text("Career")',
            'a:has-text("Jobs")', 'a:has-text("Stellenangebote")',
            'a:has-text("Stellen")', 'a:has-text("Offene Stellen")',
            'a:has-text("Careers")', 'a:has-text("Wir suchen")',
            'a:has-text("Bewerbung")', 'a:has-text("Werde Teil")',
            'a:has-text("Join us")', 'a:has-text("Join Us")',
            'a:has-text("Jobbörse")', 'a:has-text("Recruiting")',
            'a:has-text("Talent")', 'a:has-text("Opportunities")',
            'a:has-text("Offene Positionen")', 'a:has-text("Open Positions")',
            'a:has-text("Arbeiten bei")', 'a:has-text("Work at")',
        ]
        
        for selector in selectors:
            try:
                elements = await page.query_selector_all(selector)
                for elem in elements:
                    href = await elem.get_attribute('href')
                    if href:
                        full_url = urljoin(base_url, href)
                        url_lower = full_url.lower()
                        if any(kw in url_lower for kw in ['karriere', 'career', 'job', 'stelle', 'vacanc', 'recruit', 'talent', 'bewerbung', 'join', 'work-with', 'opportunit']):
                            return full_url
            except:
                continue
        
        # Check nav/header links
        try:
            nav_links = await page.query_selector_all('nav a, header a, .nav a, .navigation a, .menu a, #menu a, .navbar a')
            for link in nav_links:
                text = (await link.inner_text()).strip().lower()
                href = await link.get_attribute('href')
                if href and any(kw in text for kw in CAREER_LINK_TEXTS):
                    full_url = urljoin(base_url, href)
                    return full_url
        except:
            pass
        
        # Check footer links
        try:
            footer_links = await page.query_selector_all('footer a')
            for link in footer_links:
                text = (await link.inner_text()).strip().lower()
                href = await link.get_attribute('href')
                if href and any(kw in text for kw in CAREER_LINK_TEXTS):
                    full_url = urljoin(base_url, href)
                    return full_url
        except:
            pass
        
        return None
    
    async def _try_career_patterns(self, page, base_url):
        """Try common career URL patterns."""
        parsed = urlparse(base_url)
        base = f"{parsed.scheme}://{parsed.netloc}"
        
        for pattern in CAREER_URL_PATTERNS:
            try:
                test_url = base + pattern
                response = await page.goto(test_url, timeout=8000, wait_until='domcontentloaded')
                if response and response.status < 400:
                    await asyncio.sleep(1)
                    content = await page.content()
                    if any(kw in content.lower() for kw in ['job', 'stelle', 'karriere', 'career', 'vacanc', 'position', 'bewerb']):
                        return test_url
            except:
                continue
        
        try:
            await page.goto(base_url, timeout=PAGE_TIMEOUT, wait_until='domcontentloaded')
        except:
            pass
        
        return None
    
    async def _extract_jobs(self, page, career_url, employer_name, page_html):
        """Extract job listings from career page."""
        jobs = []
        
        # Strategy 1: Look for structured job elements
        job_selectors = [
            '.job-listing a', '.job-list a', '.vacancy a', '.position a',
            '.job-item a', '.job-entry a', '.job-card a', '.job a',
            '[class*="job"] a', '[class*="position"] a', '[class*="vacancy"] a',
            '[class*="stelle"] a', '.listing a', '.opening a',
            'article a', '.card a', 'table tbody tr a',
            '.hiring a', '.open-position a',
            '.list a', '.result a', '.search-result a',
            '[class*="listing"] a', '[class*="teaser"] a',
        ]
        
        for selector in job_selectors:
            try:
                elements = await page.query_selector_all(selector)
                for elem in elements:
                    href = await elem.get_attribute('href')
                    title = (await elem.inner_text()).strip()
                    if title and len(title) > 5 and len(title) < 300:
                        job_url = urljoin(career_url, href) if href else career_url
                        jobs.append({
                            'title': title[:200],
                            'url': job_url,
                            'source': 'Website',
                            'employer_name': employer_name,
                        })
            except:
                continue
        
        # Strategy 2: Parse HTML directly for links with IT-related text
        try:
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(page_html, 'lxml')
            for link in soup.find_all('a', href=True):
                text = link.get_text(strip=True)
                href = link['href']
                if text and 10 < len(text) < 200:
                    text_lower = text.lower()
                    has_it = any(kw in text_lower for kw in IT_KEYWORDS)
                    has_junior = any(kw in text_lower for kw in JUNIOR_KEYWORDS)
                    if has_it or has_junior:
                        job_url = urljoin(career_url, href)
                        if not any(j['url'] == job_url for j in jobs):
                            jobs.append({
                                'title': text[:200],
                                'url': job_url,
                                'source': 'Website',
                                'employer_name': employer_name,
                            })
        except:
            pass
        
        # Strategy 3: Look for JSON-LD JobPosting data
        try:
            json_ld_data = await page.evaluate('''() => {
                const results = [];
                const scripts = document.querySelectorAll('script[type="application/ld+json"]');
                for (const script of scripts) {
                    try {
                        const data = JSON.parse(script.textContent);
                        const items = Array.isArray(data) ? data : [data];
                        for (const item of items) {
                            if (item['@type'] === 'JobPosting') {
                                results.push({
                                    title: item.title || '',
                                    url: item.url || '',
                                    description: item.description || '',
                                    datePosted: item.datePosted || '',
                                    employmentType: item.employmentType || '',
                                });
                            }
                        }
                    } catch(e) {}
                }
                return results;
            }''')
            
            for ld_job in json_ld_data:
                if ld_job.get('title'):
                    jobs.append({
                        'title': ld_job['title'][:200],
                        'url': ld_job.get('url', career_url),
                        'source': 'Website (JSON-LD)',
                        'employer_name': employer_name,
                        'description': ld_job.get('description', ''),
                        'date_posted': ld_job.get('datePosted', ''),
                        'employment_detail': ld_job.get('employmentType', ''),
                    })
        except:
            pass
        
        # Deduplicate by URL
        seen_urls = set()
        unique_jobs = []
        for job in jobs:
            url_key = job.get('url', '') or job.get('title', '')
            if url_key not in seen_urls:
                seen_urls.add(url_key)
                unique_jobs.append(job)
        
        return unique_jobs
    
    async def _visit_job_detail(self, context, job):
        """Visit job detail page for more information."""
        job_url = job.get('url')
        if not job_url:
            return None
        
        detail_page = None
        try:
            detail_page = await context.new_page()
            await detail_page.goto(job_url, timeout=10000, wait_until='domcontentloaded')
            await asyncio.sleep(1.5)
            
            content = await detail_page.content()
            
            # Save HTML
            employer_name = job.get('employer_name', 'unknown')
            title_snippet = sanitize_filename(job.get('title', 'untitled')[:40])
            job_filename = sanitize_filename(employer_name)[:40] + '_' + title_snippet + '.html'
            job_filepath = os.path.join(JOB_HTML_DIR, job_filename)
            with open(job_filepath, 'w', encoding='utf-8') as f:
                f.write(content)
            
            result = {'job_html_path': job_filepath}
            
            # Extract description
            try:
                main_content = await detail_page.query_selector(
                    'main, article, .content, .job-detail, .job-description, '
                    '[class*="job"] [class*="description"], [class*="detail"], #content'
                )
                if main_content:
                    description = (await main_content.inner_text()).strip()
                else:
                    description = await detail_page.evaluate('document.body.innerText')
                    description = (description or '')[:5000]
                
                result['description'] = description
            except:
                result['description'] = ''
            
            # Extract JSON-LD
            try:
                ld = await detail_page.evaluate('''() => {
                    const scripts = document.querySelectorAll('script[type="application/ld+json"]');
                    for (const script of scripts) {
                        try {
                            const data = JSON.parse(script.textContent);
                            const item = Array.isArray(data) ? data[0] : data;
                            if (item['@type'] === 'JobPosting') {
                                return JSON.stringify(item);
                            }
                        } catch(e) {}
                    }
                    return null;
                }''')
                
                if ld:
                    ld_data = json.loads(ld)
                    if not result.get('description'):
                        result['description'] = ld_data.get('description', '')
                    result['employment_detail'] = ld_data.get('employmentType', '')
                    result['date_posted'] = ld_data.get('datePosted', '')
                    
                    salary = ld_data.get('baseSalary', {})
                    if isinstance(salary, dict):
                        val = salary.get('value', {})
                        if isinstance(val, dict):
                            result['salary_range'] = f"{val.get('minValue', '')}-{val.get('maxValue', '')} {salary.get('currency', 'EUR')}"
            except:
                pass
            
            return result
            
        except:
            return None
        finally:
            if detail_page:
                try:
                    await detail_page.close()
                except:
                    pass
    
    def _build_job_record(self, job, employer_data, career_url, career_html_path):
        title = job.get('title', '')
        description = job.get('description', '')
        
        return {
            'job_id': f'J-{self.excel.job_counter + 1:05d}',
            'job_title': title,
            'job_title_original': title,
            'job_category': classify_job_category(title, description),
            'experience_level': 'Junior',
            'employment_type': extract_employment_type(title, description),
            'remote_option': extract_remote_option(title, description),
            'technologies': extract_technologies(f"{title} {description}"),
            'description_excerpt': description[:500] if description else '',
            'job_url': job.get('url', ''),
            'employer_name': employer_data.get('employer_name', ''),
            'employer_name_norm': employer_data.get('employer_name_norm', ''),
            'city': employer_data.get('city', ''),
            'plz': employer_data.get('plz', ''),
            'street': employer_data.get('street', ''),
            'phone': employer_data.get('phone', ''),
            'email': employer_data.get('email', ''),
            'industry': employer_data.get('industry', ''),
            'source': job.get('source', 'Website'),
            'career_url': career_url,
            'website_url': employer_data.get('website_url', ''),
            'career_html_path': career_html_path,
            'job_html_path': job.get('job_html_path', ''),
            'junior_keyword_match': job.get('junior_keyword_match', ''),
            'it_keyword_match': job.get('it_keyword_match', ''),
            'is_junior': job.get('is_junior', True),
            'is_it': job.get('is_it', True),
            'employment_detail': job.get('employment_detail', extract_employment_type(title, description)),
            'temporary': 'befristet' in f"{title} {description}".lower() or job.get('befristet', False),
            'salary_range': job.get('salary_range', ''),
            'refnr': job.get('refnr', ''),
            'scrape_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'scrape_status': 'completed',
            'employer_id': employer_data.get('employer_id', ''),
            'last_check': datetime.now().strftime('%Y-%m-%d'),
            'created': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        }
    
    def _commit_and_push(self):
        try:
            self.excel.save()
            os.chdir(GIT_REPO)
            subprocess.run(['git', 'add', '-A'], capture_output=True, timeout=30)
            commit_msg = f"Junior IT Jobs - {self.excel.get_job_count()} jobs ({datetime.now().strftime('%H:%M')})"
            result = subprocess.run(['git', 'commit', '-m', commit_msg], capture_output=True, timeout=30)
            if result.returncode == 0:
                subprocess.run(
                    ['git', 'push', f'https://{GITHUB_TOKEN}@github.com/BirdNest055/Heyjobs.git', 'main', '--force-with-lease'],
                    capture_output=True, timeout=60
                )
                print(f"  [GIT] Pushed ({self.excel.get_job_count()} jobs)")
        except Exception as e:
            print(f"  [GIT] Error: {str(e)[:80]}")


# ============================================================
# MAIN
# ============================================================

async def main():
    print("=" * 70)
    print("Junior IT Jobs Scraper v2 - B/ER/N")
    print("Visiting websites like a human + Arbeitsagentur API")
    print("=" * 70)
    
    # Load employers with websites
    print("\n[1] Loading employer data...")
    wb = openpyxl.load_workbook(EXCEL_IN)
    ws = wb['Nur IT-Jobs']
    
    employers = []
    seen_urls = set()
    
    for r in range(2, ws.max_row + 1):
        website = ws.cell(r, 9).value
        if website and website not in seen_urls:
            employers.append({
                'employer_id': ws.cell(r, 1).value,
                'employer_name': ws.cell(r, 2).value,
                'employer_name_norm': ws.cell(r, 3).value,
                'city': ws.cell(r, 4).value,
                'plz': ws.cell(r, 5).value,
                'street': ws.cell(r, 6).value,
                'phone': ws.cell(r, 7).value,
                'email': ws.cell(r, 8).value,
                'website_url': website,
                'industry': ws.cell(r, 23).value,
            })
            seen_urls.add(website)
    
    print(f"  {len(employers)} unique employers with websites from IT-Jobs sheet")
    
    # Initialize Excel
    print("\n[2] Initializing output Excel...")
    excel = JuniorJobsExcel(EXCEL_OUT)
    excel.save()
    
    # Step A: Query Arbeitsagentur API for junior IT jobs (parallel, fast)
    print("\n[3] Querying Arbeitsagentur API for junior IT jobs...")
    semaphore = asyncio.Semaphore(MAX_CONCURRENT)
    scraper = WebsiteScraper(excel, semaphore)
    
    async with httpx.AsyncClient() as client:
        aa_jobs_total = 0
        batch_size = 5
        
        for i in range(0, len(employers), batch_size):
            batch = employers[i:i+batch_size]
            tasks = []
            for emp in batch:
                name = emp.get('employer_name', '')
                # Extract core company name for better search
                # Remove location suffixes etc.
                search_name = re.sub(r'\s*(GmbH|AG|e\.K\.|e\.V\.|UG|KG|GbR|OHG).*$', '', name).strip()
                if len(search_name) < 3:
                    search_name = name[:20]
                tasks.append(query_arbeitsagentur_jobs(client, search_name, emp.get('city')))
            
            results = await asyncio.gather(*tasks, return_exceptions=True)
            
            for result in results:
                if isinstance(result, list):
                    for job in result:
                        job_data = scraper._build_job_record(job, 
                            {'employer_name': job.get('employer_name', ''), 'website_url': ''},
                            '', '')
                        job_data['source'] = 'Arbeitsagentur'
                        excel.add_job(job_data)
                        aa_jobs_total += 1
            
            if (i // batch_size + 1) % 5 == 0:
                print(f"  AA API: {i+len(batch)}/{len(employers)} employers checked, {aa_jobs_total} junior IT jobs found")
        
        print(f"  AA API total: {aa_jobs_total} junior IT jobs from {len(employers)} employers")
    
    excel.save()
    scraper._commit_and_push()
    
    # Step B: Visit each employer website
    print(f"\n[4] Visiting {len(employers)} employer websites...")
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=['--no-sandbox', '--disable-dev-shm-usage', '--disable-gpu']
        )
        context = await browser.new_context(
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36',
            viewport={'width': 1920, 'height': 1080},
            locale='de-DE',
        )
        
        # Process in batches
        batch_size = 10
        website_jobs_total = 0
        
        for i in range(0, len(employers), batch_size):
            batch = employers[i:i+batch_size]
            print(f"\n--- Website Batch {i//batch_size + 1}/{(len(employers) + batch_size - 1)//batch_size} ---")
            
            tasks = [scraper.scrape_employer(context, emp) for emp in batch]
            results = await asyncio.gather(*tasks, return_exceptions=True)
            
            batch_jobs = 0
            for result in results:
                if isinstance(result, list):
                    batch_jobs += len(result)
            
            website_jobs_total += batch_jobs
            print(f"  Batch: {batch_jobs} jobs, Total website: {website_jobs_total}, Total all: {excel.get_job_count()}")
            
            excel.save()
        
        await browser.close()
    
    # Final commit
    excel.save()
    scraper._commit_and_push()
    
    print(f"\n{'=' * 70}")
    print(f"COMPLETE!")
    print(f"  Arbeitsagentur junior IT jobs: {aa_jobs_total}")
    print(f"  Website junior IT jobs: {website_jobs_total}")
    print(f"  Total junior IT jobs: {excel.get_job_count()}")
    print(f"  Websites scraped: {scraper.scraped_count}")
    print(f"  Career pages found: {scraper.career_found_count}")
    print(f"  Errors: {scraper.error_count}")
    print(f"  Excel: {EXCEL_OUT}")
    print(f"{'=' * 70}")


if __name__ == '__main__':
    asyncio.run(main())
