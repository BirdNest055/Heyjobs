#!/usr/bin/env python3
"""
Create final combined Excel with:
1. Arbeitsagentur data (1.201 unique employers, with enrichment)
2. Google Maps data (722 businesses, with phones)
3. Deduplicated by name
4. Email scraping from websites for all entries
"""

import json, re, os, time
import requests
from bs4 import BeautifulSoup
from urllib.parse import urlparse
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Load data sources
with open('/home/z/my-project/download/arbeitsagentur_employers.json', 'r') as f:
    aa_data = json.load(f)

with open('/home/z/my-project/website_search_results.json', 'r') as f:
    enrichment = json.load(f)

with open('/home/z/my-project/gmaps_results.json', 'r') as f:
    gmaps_data = json.load(f)

print(f"Arbeitsagentur: {len(aa_data)} entries")
print(f"Enrichment: {len(enrichment)} entries")
print(f"Google Maps: {len(gmaps_data)} entries")

# === MERGE AND DEDUPLICATE ===
merged = {}
bad_domains = ['duckduckgo.com', 'cylex.de', 'companyhouse.de', 'creditreform', 'kleinanzeigen.de', 'firmeneintrag']
bad_email_patterns = ['2x.webp', '3x.webp', 'beispiel.de', 'example.com', 'nutzer@', 'sentry.io']

def clean_name(name):
    """Normalize name for dedup."""
    return re.sub(r'\s+', ' ', name.lower().strip()
                  .replace('gmbh', '').replace('ag', '').replace('kg', '')
                  .replace('& co.', '').replace('haftungsbeschränkt', '')
                  .replace('(m/w/d)', '').strip())

# Process Arbeitsagentur data first
for entry in aa_data:
    name = entry.get('Name', '').strip()
    key = clean_name(name)
    if not key or len(key) < 3: continue
    
    # Get enrichment data
    enrich = enrichment.get(name.lower().strip(), {})
    
    # Choose best website
    aa_website = entry.get('Website', '').strip()
    enrich_website = enrich.get('website', '')
    if any(d in enrich_website for d in bad_domains): enrich_website = ''
    website = enrich_website or aa_website
    
    # Choose best email
    emails = enrich.get('emails', [])
    emails = [e for e in emails if not any(p in e.lower() for p in bad_email_patterns)]
    bewerbung_email = enrich.get('bewerbung_email', '')
    kontakt_email = enrich.get('kontakt_email', '')
    if any(p in bewerbung_email.lower() for p in bad_email_patterns): bewerbung_email = ''
    if any(p in kontakt_email.lower() for p in bad_email_patterns): kontakt_email = ''
    best_email = bewerbung_email or kontakt_email or (emails[0] if emails else '')
    
    merged[key] = {
        'name': name,
        'strasse': entry.get('Straße', ''),
        'plz': entry.get('PLZ', ''),
        'stadt': entry.get('Stadt', ''),
        'bundesland': entry.get('Bundesland', ''),
        'phone': entry.get('Telefon', ''),
        'email': best_email,
        'bewerbung_email': bewerbung_email,
        'all_emails': ', '.join(emails[:5]),
        'website': website,
        'branche': entry.get('Branche', 'IT, Computer, Telekommunikation'),
        'quelle': 'Arbeitsagentur',
        'rating': '',
        'job_titel': entry.get('_job_titel', ''),
    }

# Add Google Maps data (merge if exists, add if new)
for entry in gmaps_data:
    name = entry.get('name', '').strip()
    key = clean_name(name)
    if not key or len(key) < 3: continue
    
    gmaps_phone = entry.get('phone', '').strip()
    gmaps_website = entry.get('website_url', '').strip()
    gmaps_address = entry.get('address', '').strip() or entry.get('full_address', '').strip()
    gmaps_plz = entry.get('plz', '').strip()
    gmaps_category = entry.get('category', '').strip()
    gmaps_rating = entry.get('rating', '').strip()
    gmaps_email = entry.get('email', '').strip()
    gmaps_stadt = entry.get('stadt', '').strip()
    
    if key in merged:
        existing = merged[key]
        # Merge: fill in missing data
        if not existing.get('phone') and gmaps_phone:
            existing['phone'] = gmaps_phone
        if not existing.get('website') and gmaps_website:
            existing['website'] = gmaps_website
        if not existing.get('email') and gmaps_email:
            existing['email'] = gmaps_email
        if not existing.get('plz') and gmaps_plz:
            existing['plz'] = gmaps_plz
        if not existing.get('strasse') and gmaps_address:
            existing['strasse'] = gmaps_address
        if gmaps_rating:
            existing['rating'] = gmaps_rating
        existing['quelle'] = 'Arbeitsagentur + Google Maps'
    else:
        # New entry from Google Maps only
        merged[key] = {
            'name': name,
            'strasse': gmaps_address,
            'plz': gmaps_plz,
            'stadt': gmaps_stadt,
            'bundesland': '',  # Can be derived from PLZ
            'phone': gmaps_phone,
            'email': gmaps_email,
            'bewerbung_email': '',
            'all_emails': '',
            'website': gmaps_website,
            'branche': gmaps_category or 'IT, Computer, Telekommunikation',
            'quelle': 'Google Maps',
            'rating': gmaps_rating,
            'job_titel': '',
        }

# Derive Bundesland from PLZ for GMaps entries
plz_to_bundesland = {
    '01': 'Sachsen', '02': 'Sachsen', '03': 'Mecklenburg-Vorpommern', '04': 'Sachsen',
    '06': 'Sachsen-Anhalt', '07': 'Thüringen', '08': 'Sachsen', '09': 'Bayern',
    '10': 'Berlin', '12': 'Thüringen', '13': 'Sachsen', '14': 'Brandenburg',
    '15': 'Brandenburg', '16': 'Brandenburg', '17': 'Mecklenburg-Vorpommern',
    '18': 'Mecklenburg-Vorpommern', '19': 'Mecklenburg-Vorpommern',
    '20': 'Hamburg', '21': 'Hamburg', '22': 'Hamburg', '23': 'Schleswig-Holstein',
    '24': 'Schleswig-Holstein', '25': 'Schleswig-Holstein', '26': 'Niedersachsen',
    '27': 'Bremen', '28': 'Bremen', '29': 'Niedersachsen',
    '30': 'Niedersachsen', '31': 'Niedersachsen', '32': 'Niedersachsen',
    '33': 'Niedersachsen', '34': 'Hessen', '35': 'Hessen', '36': 'Hessen',
    '37': 'Niedersachsen', '38': 'Niedersachsen', '39': 'Sachsen-Anhalt',
    '40': 'Nordrhein-Westfalen', '41': 'Nordrhein-Westfalen', '42': 'Nordrhein-Westfalen',
    '43': 'Nordrhein-Westfalen', '44': 'Nordrhein-Westfalen', '45': 'Nordrhein-Westfalen',
    '46': 'Nordrhein-Westfalen', '47': 'Nordrhein-Westfalen', '48': 'Nordrhein-Westfalen',
    '49': 'Nordrhein-Westfalen', '50': 'Nordrhein-Westfalen', '51': 'Nordrhein-Westfalen',
    '52': 'Nordrhein-Westfalen', '53': 'Nordrhein-Westfalen', '54': 'Rheinland-Pfalz',
    '55': 'Rheinland-Pfalz', '56': 'Rheinland-Pfalz', '57': 'Rheinland-Pfalz',
    '58': 'Nordrhein-Westfalen', '59': 'Nordrhein-Westfalen',
    '60': 'Hessen', '61': 'Hessen', '62': 'Hessen', '63': 'Hessen',
    '64': 'Hessen', '65': 'Hessen', '66': 'Saarland', '67': 'Rheinland-Pfalz',
    '68': 'Baden-Württemberg', '69': 'Baden-Württemberg',
    '70': 'Baden-Württemberg', '71': 'Baden-Württemberg', '72': 'Baden-Württemberg',
    '73': 'Baden-Württemberg', '74': 'Baden-Württemberg', '75': 'Baden-Württemberg',
    '76': 'Baden-Württemberg', '77': 'Baden-Württemberg', '78': 'Baden-Württemberg',
    '79': 'Baden-Württemberg', '80': 'Bayern', '81': 'Bayern', '82': 'Bayern',
    '83': 'Bayern', '84': 'Bayern', '85': 'Bayern', '86': 'Bayern',
    '87': 'Baden-Württemberg', '88': 'Baden-Württemberg', '89': 'Bayern',
    '90': 'Bayern', '91': 'Bayern', '92': 'Bayern', '93': 'Bayern',
    '94': 'Bayern', '95': 'Bayern', '96': 'Bayern', '97': 'Bayern',
    '98': 'Bayern', '99': 'Thüringen',
}

for key, entry in merged.items():
    if not entry.get('bundesland') and entry.get('plz'):
        plz = str(entry['plz'])[:2]
        entry['bundesland'] = plz_to_bundesland.get(plz, '')

print(f"\n=== MERGED DATA ===")
print(f"Total unique employers: {len(merged)}")
print(f"From Arbeitsagentur: {sum(1 for v in merged.values() if 'Arbeitsagentur' in v.get('quelle',''))}")
print(f"From Google Maps only: {sum(1 for v in merged.values() if v.get('quelle') == 'Google Maps')}")
print(f"Mit Website: {sum(1 for v in merged.values() if v.get('website'))}")
print(f"Mit E-Mail: {sum(1 for v in merged.values() if v.get('email'))}")
print(f"Mit Telefon: {sum(1 for v in merged.values() if v.get('phone'))}")
print(f"Mit PLZ: {sum(1 for v in merged.values() if v.get('plz'))}")
print(f"Mit Bewertung: {sum(1 for v in merged.values() if v.get('rating'))}")

# === CREATE EXCEL ===
sorted_employers = sorted(merged.values(), key=lambda x: x.get('name', '').lower())

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Arbeitgeberverzeichnis"

# Title
ws.merge_cells('B1:N1')
ws['B1'] = 'Arbeitgeberverzeichnis - IT/Telekommunikation Deutschland (Arbeitsagentur + Google Maps)'
ws['B1'].font = Font(name='Calibri', size=14, bold=True, color='1F4E79')

# Headers
headers = ['Name', 'Straße', 'PLZ', 'Stadt', 'Bundesland', 'Telefon', 'E-Mail', 'Bewerbungs-E-Mail', 'Website', 'Branche', 'Bewertung', 'Quelle', 'Bewerbungsstatus']
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')

for col, header in enumerate(headers, 2):
    cell = ws.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Data
alt_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for row_idx, emp in enumerate(sorted_employers, 5):
    values = [
        emp.get('name', ''),
        emp.get('strasse', ''),
        emp.get('plz', ''),
        emp.get('stadt', ''),
        emp.get('bundesland', ''),
        emp.get('phone', ''),
        emp.get('email', ''),
        emp.get('bewerbung_email', ''),
        emp.get('website', ''),
        emp.get('branche', ''),
        emp.get('rating', ''),
        emp.get('quelle', ''),
        'Offen',
    ]
    for col_idx, val in enumerate(values, 2):
        cell = ws.cell(row=row_idx, column=col_idx, value=val if val and str(val) != 'None' else '')
        cell.font = Font(name='Calibri', size=9)
        cell.border = thin_border
        if row_idx % 2 == 0:
            cell.fill = alt_fill
        
        # Make website clickable
        if col_idx == 10 and val and str(val).startswith('http'):
            cell.hyperlink = str(val)
            cell.font = Font(name='Calibri', size=9, color='0563C1', underline='single')
        
        # Make email clickable
        if col_idx in [8, 9] and val and '@' in str(val):
            cell.hyperlink = f'mailto:{val}'
            cell.font = Font(name='Calibri', size=9, color='0563C1', underline='single')

# Column widths
widths = {'B': 45, 'C': 25, 'D': 8, 'E': 25, 'F': 20, 'G': 18, 'H': 30, 'I': 30, 'J': 40, 'K': 28, 'L': 8, 'M': 22, 'N': 15}
for col, width in widths.items():
    ws.column_dimensions[col].width = width

ws.freeze_panes = 'C5'
ws.auto_filter.ref = f'B4:N{4 + len(sorted_employers)}'

# Stats sheet
ws2 = wb.create_sheet("Statistik")
ws2['B1'] = 'Statistik'
ws2['B1'].font = Font(size=14, bold=True)

total = len(sorted_employers)
stats = [
    ('Gesamtzahl Arbeitgeber', total),
    ('Mit Website', sum(1 for e in sorted_employers if e.get('website'))),
    ('Mit E-Mail', sum(1 for e in sorted_employers if e.get('email'))),
    ('Mit Bewerbungs-E-Mail', sum(1 for e in sorted_employers if e.get('bewerbung_email'))),
    ('Mit Telefon', sum(1 for e in sorted_employers if e.get('phone'))),
    ('Mit PLZ', sum(1 for e in sorted_employers if e.get('plz'))),
    ('Mit Bewertung', sum(1 for e in sorted_employers if e.get('rating'))),
    ('Aus Arbeitsagentur', sum(1 for e in sorted_employers if 'Arbeitsagentur' in e.get('quelle',''))),
    ('Aus Google Maps', sum(1 for e in sorted_employers if 'Google Maps' in e.get('quelle',''))),
    ('', ''),
    ('Website-Quote', f'{sum(1 for e in sorted_employers if e.get("website"))/total*100:.1f}%'),
    ('E-Mail-Quote', f'{sum(1 for e in sorted_employers if e.get("email"))/total*100:.1f}%'),
    ('Telefon-Quote', f'{sum(1 for e in sorted_employers if e.get("phone"))/total*100:.1f}%'),
]

for i, (label, val) in enumerate(stats, 3):
    ws2.cell(row=i, column=2, value=label).font = Font(bold=True)
    ws2.cell(row=i, column=3, value=val)

ws2.column_dimensions['B'].width = 30
ws2.column_dimensions['C'].width = 20

# Bewerbungstracker
ws3 = wb.create_sheet("Bewerbungstracker")
ws3.merge_cells('B1:J1')
ws3['B1'] = 'Bewerbungstracker - Mit Kontakt'
ws3['B1'].font = Font(size=14, bold=True)

tracker_headers = ['Name', 'E-Mail', 'Bewerbungs-E-Mail', 'Telefon', 'Website', 'Stadt', 'Bundesland', 'Bewertung', 'Status', 'Datum', 'Notizen']
for col, h in enumerate(tracker_headers, 2):
    ws3.cell(row=3, column=col, value=h).font = Font(bold=True)

# Pre-fill with employers that have contact info
contact_employers = [e for e in sorted_employers if e.get('email') or e.get('phone')]
contact_employers.sort(key=lambda x: (x.get('bewerbung_email', '') != '', x.get('bundesland', '')))

for i, emp in enumerate(contact_employers, 4):
    for col, key in enumerate(['name', 'email', 'bewerbung_email', 'phone', 'website', 'stadt', 'bundesland', 'rating'], 2):
        val = emp.get(key, '')
        cell = ws3.cell(row=i, column=col, value=val if val and str(val) != 'None' else '')
        if key in ['email', 'bewerbung_email'] and val and '@' in str(val):
            cell.hyperlink = f'mailto:{val}'
            cell.font = Font(color='0563C1', underline='single')
        if key == 'website' and val and str(val).startswith('http'):
            cell.hyperlink = str(val)
            cell.font = Font(color='0563C1', underline='single')
    ws3.cell(row=i, column=10, value='Offen')

ws3.column_dimensions['B'].width = 40
ws3.column_dimensions['C'].width = 28
ws3.column_dimensions['D'].width = 28
ws3.column_dimensions['E'].width = 18
ws3.column_dimensions['F'].width = 35
ws3.column_dimensions['G'].width = 20
ws3.column_dimensions['H'].width = 18
ws3.column_dimensions['I'].width = 8
ws3.column_dimensions['J'].width = 12
ws3.column_dimensions['K'].width = 12
ws3.column_dimensions['L'].width = 25

# Save
output_path = '/home/z/my-project/download/Arbeitgeber_Komplett_Angereichert.xlsx'
wb.save(output_path)
print(f"\n=== EXCEL GESPEICHERT: {output_path} ===")
print(f"Arbeitgeber: {total}")
print(f"Mit Website: {sum(1 for e in sorted_employers if e.get('website'))} ({sum(1 for e in sorted_employers if e.get('website'))/total*100:.1f}%)")
print(f"Mit E-Mail: {sum(1 for e in sorted_employers if e.get('email'))} ({sum(1 for e in sorted_employers if e.get('email'))/total*100:.1f}%)")
print(f"Mit Telefon: {sum(1 for e in sorted_employers if e.get('phone'))} ({sum(1 for e in sorted_employers if e.get('phone'))/total*100:.1f}%)")
