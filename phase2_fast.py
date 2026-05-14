#!/usr/bin/env python3
"""
Phase 2 OPTIMIZED: Visit employer websites - faster with stricter timeouts
"""
import asyncio, json, os, re, subprocess
from datetime import datetime
from urllib.parse import urljoin, urlparse
from playwright.async_api import async_playwright
import openpyxl

JSON_IN = "/home/z/my-project/download/Junior_IT_Jobs_Bamberg_Erlangen_Nuernberg_raw.json"
EXCEL_OUT = "/home/z/my-project/download/Junior_IT_Jobs_Bamberg_Erlangen_Nuernberg.xlsx"
CAREER_HTML_DIR = "/home/z/my-project/download/career_pages_html"
GIT_REPO = "/home/z/my-project"
GITHUB_TOKEN = "GIT_TOKEN_REDACTED"
PROGRESS_FILE = "/home/z/my-project/download/phase2_progress.json"

EMPLOYER_EXCEL = "/home/z/my-project/download/IT_Arbeitgeber_Bamberg_Erlangen_Nuernberg.xlsx"
FULL_EMPLOYER_EXCEL = "/home/z/my-project/download/Firmen_Bamberg_Erlangen_Nuernberg_mit_Website_HTML.xlsx"

def sanitize_filename(name, max_len=50):
    name = re.sub(r'[^\w\s.-]', '_', name)
    name = re.sub(r'_{2,}', '_', name)
    return name.strip('_')[:max_len]


def load_employer_websites():
    websites = {}
    try:
        wb = openpyxl.load_workbook(EMPLOYER_EXCEL)
        ws = wb['Nur IT-Jobs']
        for r in range(2, ws.max_row + 1):
            name = ws.cell(r, 2).value or ''
            website = ws.cell(r, 9).value
            if name and website:
                norm = re.sub(r'\s+', '', name.lower())[:15]
                if norm not in websites:
                    websites[norm] = website
        wb.close()
    except:
        pass
    try:
        wb2 = openpyxl.load_workbook(FULL_EMPLOYER_EXCEL)
        ws2 = wb2['Firmen mit Website & HTML']
        for r in range(2, ws2.max_row + 1):
            name = ws2.cell(r, 2).value or ''
            website = ws2.cell(r, 11).value
            if name and website:
                norm = re.sub(r'\s+', '', name.lower())[:15]
                if norm not in websites:
                    websites[norm] = website
        wb2.close()
    except:
        pass
    return websites


def match_employer(aa_name, website_map):
    norm = re.sub(r'\s+', '', aa_name.lower())
    for key, url in website_map.items():
        if key in norm or norm[:15] in key:
            return url
    clean = re.sub(r'(gmbh|ag|ek|ug|kg|gbr|ohg|mbh).*$', '', norm)
    if len(clean) >= 5:
        for key, url in website_map.items():
            ck = re.sub(r'(gmbh|ag|ek|ug|kg|gbr|ohg|mbh).*$', '', key)
            if clean[:8] in ck or ck[:8] in clean:
                return url
    return None


CAREER_SELECTORS = [
    'a:has-text("Karriere")', 'a:has-text("Career")', 'a:has-text("Careers")',
    'a:has-text("Jobs")', 'a:has-text("Stellenangebote")', 'a:has-text("Stellen")',
    'a:has-text("Wir suchen")', 'a:has-text("Bewerbung")', 'a:has-text("Join us")',
    'a:has-text("Recruiting")', 'a:has-text("Offene Stellen")',
    'a:has-text("Offene Positionen")', 'a:has-text("Talent")',
]

CAREER_TEXTS = ['karriere','career','jobs','stellenangebote','stellen','offene stellen',
    'vacancies','careers','recruiting','bewerbung','talent','opportunities',
    'join us','wir suchen','offene positionen']

CAREER_URLS = ['/karriere','/career','/careers','/jobs','/stellenangebote',
    '/stellen','/offene-stellen','/vacancies','/job','/recruiting',
    '/de/karriere','/en/career','/de/career','/en/careers']


async def visit_employer(context, emp_name, website_url):
    """Visit one employer, find career page, save HTML. Fast version."""
    page = None
    result = {
        'employer': emp_name,
        'website': website_url,
        'career_url': None,
        'career_html': None,
        'job_links': [],
        'status': 'ok',
    }
    
    try:
        page = await context.new_page()
        
        # Visit homepage with short timeout
        try:
            await page.goto(website_url, timeout=10000, wait_until='domcontentloaded')
            await asyncio.sleep(1)
        except:
            result['status'] = 'homepage_timeout'
            return result
        
        # Dismiss cookie banner quickly
        try:
            for sel in ['button:has-text("Akzeptieren")', 'button:has-text("Accept")',
                       'button:has-text("OK")', '#onetrust-accept-btn-handler']:
                try:
                    btn = await page.query_selector(sel)
                    if btn: await btn.click(); await asyncio.sleep(0.2); break
                except: continue
        except: pass
        
        # Find career link
        career_url = None
        
        for sel in CAREER_SELECTORS:
            try:
                elems = await page.query_selector_all(sel)
                for elem in elems:
                    href = await elem.get_attribute('href')
                    if href:
                        career_url = urljoin(website_url, href)
                        break
            except: continue
            if career_url: break
        
        if not career_url:
            try:
                for sel in ['nav a', 'header a', 'footer a']:
                    links = await page.query_selector_all(sel)
                    for link in links:
                        text = (await link.inner_text()).strip().lower()
                        href = await link.get_attribute('href')
                        if href and any(kw in text for kw in CAREER_TEXTS):
                            career_url = urljoin(website_url, href)
                            break
                    if career_url: break
            except: pass
        
        if not career_url:
            parsed = urlparse(website_url)
            base = f"{parsed.scheme}://{parsed.netloc}"
            for pattern in CAREER_URLS:
                try:
                    resp = await page.goto(base + pattern, timeout=5000, wait_until='domcontentloaded')
                    if resp and resp.status < 400:
                        content = await page.content()
                        if any(kw in content.lower() for kw in ['job','stelle','karriere','career','position']):
                            career_url = base + pattern
                            break
                except: continue
        
        result['career_url'] = career_url
        
        # Visit career page
        if career_url:
            try:
                await page.goto(career_url, timeout=10000, wait_until='domcontentloaded')
                await asyncio.sleep(1.5)
                await page.evaluate('window.scrollTo(0, document.body.scrollHeight)')
                await asyncio.sleep(0.5)
                
                html = await page.content()
                fn = sanitize_filename(emp_name) + '_career.html'
                fp = os.path.join(CAREER_HTML_DIR, fn)
                with open(fp, 'w', encoding='utf-8') as f:
                    f.write(html)
                result['career_html'] = fp
                
                # Find job links
                links = await page.query_selector_all('a')
                for link in links:
                    text = (await link.inner_text()).strip()
                    href = await link.get_attribute('href')
                    if text and 8 < len(text) < 200 and href:
                        tl = text.lower()
                        if any(kw in tl for kw in ['software','entwickler','developer','it-','data','devops','cloud','system','admin','frontend','backend','fullstack','python','java','sap','consultant','engineer','support','scrum','test','ux','informatik','fachinformatiker','security','architekt','projekt','project']):
                            result['job_links'].append({
                                'title': text[:120],
                                'url': urljoin(career_url, href)
                            })
            except:
                result['status'] = 'career_page_timeout'
        
        return result
        
    except Exception as e:
        result['status'] = f'error: {str(e)[:40]}'
        return result
    finally:
        if page:
            try: await page.close()
            except: pass


async def process_batch(context, batch, semaphore):
    """Process a batch of employers in parallel."""
    async def limited_visit(emp_name, url):
        async with semaphore:
            return await visit_employer(context, emp_name, url)
    
    tasks = [limited_visit(name, url) for name, url in batch]
    return await asyncio.gather(*tasks, return_exceptions=True)


async def main():
    print("="*60)
    print("Phase 2: Visit employer websites (optimized)")
    print("="*60)
    
    os.makedirs(CAREER_HTML_DIR, exist_ok=True)
    
    # Load jobs and employer mapping
    with open(JSON_IN, 'r', encoding='utf-8') as f:
        jobs = json.load(f)
    
    website_map = load_employer_websites()
    print(f"  {len(jobs)} jobs, {len(website_map)} employer websites")
    
    # Get unique employers with websites
    employer_urls = {}
    for job in jobs:
        emp = job.get('employer_name', '')
        if not emp: continue
        norm = re.sub(r'\s+', '', emp.lower())[:20]
        if norm not in employer_urls:
            url = match_employer(emp, website_map)
            if url:
                employer_urls[norm] = (emp, url)
    
    print(f"  {len(employer_urls)} employers with websites to visit")
    
    # Load progress
    done = set()
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE, 'r') as f:
            done = set(json.load(f))
    
    # Filter already done
    todo = {k: v for k, v in employer_urls.items() if k not in done}
    print(f"  {len(todo)} employers remaining ({len(done)} already done)")
    
    all_results = []
    commit_counter = 0
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True, args=['--no-sandbox','--disable-dev-shm-usage'])
        context = await browser.new_context(
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/125.0.0.0',
            viewport={'width': 1920, 'height': 1080},
            locale='de-DE',
        )
        sem = asyncio.Semaphore(3)  # Conservative concurrency
        
        batch_size = 6
        emp_list = list(todo.items())
        
        for i in range(0, len(emp_list), batch_size):
            batch = [(v[0], v[1]) for k, v in emp_list[i:i+batch_size]]
            batch_keys = [k for k, v in emp_list[i:i+batch_size]]
            
            print(f"\n  Batch {i//batch_size+1}/{(len(emp_list)+batch_size-1)//batch_size}")
            
            results = await process_batch(context, batch, sem)
            
            for key, result in zip(batch_keys, results):
                if isinstance(result, dict):
                    all_results.append(result)
                    done.add(key)
                    status = '✓' if result.get('career_url') else '-'
                    print(f"    [{status}] {result.get('employer','?')[:40]} | career: {'found' if result.get('career_url') else 'none'} | jobs: {len(result.get('job_links',[]))}")
            
            # Save progress
            with open(PROGRESS_FILE, 'w') as f:
                json.dump(list(done), f)
            
            # Commit every 10 entries
            commit_counter += len(batch)
            if commit_counter >= COMMIT_EVERY:
                update_excel(all_results)
                git_commit(f"Phase 2: {len(done)} employers visited, {sum(1 for r in all_results if r.get('career_url'))} career pages found")
                commit_counter = 0
        
        await browser.close()
    
    # Final update
    update_excel(all_results)
    git_commit(f"Phase 2 complete: {len(done)} employers, {sum(1 for r in all_results if r.get('career_url'))} career pages")
    
    # Stats
    career_found = sum(1 for r in all_results if r.get('career_url'))
    job_links = sum(len(r.get('job_links', [])) for r in all_results)
    print(f"\n{'='*60}")
    print(f"Phase 2 DONE!")
    print(f"  Visited: {len(all_results)}")
    print(f"  Career pages: {career_found}")
    print(f"  Job links from websites: {job_links}")
    print(f"{'='*60}")


def update_excel(results):
    """Update Excel with career URL and HTML data."""
    try:
        wb = openpyxl.load_workbook(EXCEL_OUT)
        ws = wb['Junior IT-Jobs']
        
        # Build lookup
        lookup = {}
        for r in results:
            emp = r.get('employer', '')
            norm = re.sub(r'\s+', '', emp.lower())[:20]
            lookup[norm] = r
        
        updated = 0
        for row in range(2, ws.max_row + 1):
            emp_name = ws.cell(row, 11).value  # Arbeitgeber
            if not emp_name: continue
            norm = re.sub(r'\s+', '', emp_name.lower())[:20]
            if norm in lookup:
                r = lookup[norm]
                if r.get('career_url'):
                    ws.cell(row, 20, r['career_url'])  # Karriere-URL
                if r.get('career_html'):
                    ws.cell(row, 22, r['career_html'])  # Karriere-HTML Pfad
                if r.get('website'):
                    ws.cell(row, 21, r['website'])  # Website-URL
                if r.get('job_links'):
                    titles = [j['title'] for j in r['job_links'][:5]]
                    if not ws.cell(row, 9).value:
                        ws.cell(row, 9, '; '.join(titles)[:500])
                updated += 1
        
        wb.save(EXCEL_OUT)
        print(f"  Excel: {updated} rows updated")
    except Exception as e:
        print(f"  Excel update error: {e}")


def git_commit(msg):
    try:
        os.chdir(GIT_REPO)
        subprocess.run(['git', 'add', '-A'], capture_output=True, timeout=30)
        r = subprocess.run(['git', 'commit', '-m', msg], capture_output=True, timeout=30)
        if r.returncode == 0:
            subprocess.run(['git', 'push', f'https://{GITHUB_TOKEN}@github.com/BirdNest055/Heyjobs.git', 'main', '--force-with-lease'], capture_output=True, timeout=60)
            print(f"  Git: Pushed!")
    except: pass


if __name__ == '__main__':
    asyncio.run(main())
