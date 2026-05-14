#!/usr/bin/env python3
"""Phase 1 FINAL: Comprehensive Arbeitsagentur API search for junior IT jobs"""
import asyncio, httpx, json, re, os, subprocess
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

AA_API_URL = "https://rest.arbeitsagentur.de/jobboerse/jobsuche-service/pc/v6/jobs"
AA_API_KEY = "jobboerse-jobsuche"
EXCEL_OUT = "/home/z/my-project/download/Junior_IT_Jobs_Bamberg_Erlangen_Nuernberg.xlsx"
GIT_REPO = "/home/z/my-project"
GITHUB_TOKEN = "GIT_TOKEN_REDACTED"

IT_KEYWORDS = ['software','entwickler','developer','it-','informatik','data','devops','cloud','cyber','security','sicherheit','system','admin','netzwerk','frontend','backend','fullstack','full-stack','python','java','javascript','sap','linux','sql','database','datenbank','api','machine learning','scrum','agile','testing','qa','consultant','analyst','engineer','architekt','programmierung','web','mobile','app','infrastructure','support','helpdesk','erp','crm','digitalisierung','technologie','digital','fachinformatiker']

SENIOR_KW = ['senior','lead','principal','head of','director','leitung','chef','manager']

TECH_LIST = ['python','java','javascript','typescript','react','angular','vue','node.js','django','flask','spring','.net','docker','kubernetes','aws','azure','gcp','terraform','ansible','jenkins','sql','mysql','postgresql','mongodb','redis','kafka','graphql','sap','linux','html','css','tailwind','scrum','devops','ci/cd','machine learning','deep learning','tensorflow','pytorch','data science','big data','cybersecurity','pentest','powershell','bash','git']

def is_it_job(t):
    tl = t.lower()
    return any(kw in tl for kw in IT_KEYWORDS)

def is_senior(t):
    tl = t.lower()
    return any(kw in tl for kw in SENIOR_KW)

def is_junior(t):
    tl = t.lower()
    return any(kw in tl for kw in ['junior','einsteiger','trainee','praktikant','intern','werkstudent','working student','absolvent','graduate','berufseinsteiger','associate','duales studium','ausbildung','apprentice','nachwuchs','perspektive'])

def classify_cat(t):
    tl = t.lower()
    if any(k in tl for k in ['softwareentwickler','software developer','entwickler','developer','programmierer','frontend','backend','fullstack','full-stack','fachinformatiker']): return 'Softwareentwicklung'
    if any(k in tl for k in ['data scientist','data analyst','data engineer','business intelligence','machine learning','ki-']): return 'Data Science & Analytics'
    if any(k in tl for k in ['it consultant','it-consultant','it berater','sap consultant','technischer berater']): return 'IT-Consulting'
    if any(k in tl for k in ['system administrator','sysadmin','admin','netzwerkadministrator','fachinformatiker systemintegration']): return 'Systemadministration'
    if any(k in tl for k in ['devops','cloud engineer','platform engineer','sre']): return 'DevOps & Cloud'
    if any(k in tl for k in ['security','cybersecurity','it-sicherheit','pentest']): return 'Cyber Security'
    if any(k in tl for k in ['it support','helpdesk','service desk','first level','second level','user support']): return 'IT-Support'
    if any(k in tl for k in ['project manager','projektmanager','scrum master','product owner']): return 'IT-Projektmanagement'
    if any(k in tl for k in ['qa','tester','test engineer','testautomatisierung']): return 'QA & Testing'
    if any(k in tl for k in ['ux','ui','user experience','user interface','design']): return 'UX/UI Design'
    if any(k in tl for k in ['ausbildung','duales studium','trainee','praktikum','werkstudent']): return 'IT-Ausbildung & Studium'
    return 'IT (Sonstige)'

def extract_techs(text):
    found = []
    tl = text.lower()
    for t in TECH_LIST:
        if re.search(r'\b'+re.escape(t)+r'\b', tl):
            found.append(t)
    return list(set(found))

def get_emp_type(t):
    tl = t.lower()
    if any(k in tl for k in ['werkstudent','working student']): return 'Werkstudent'
    if any(k in tl for k in ['praktikant','praktikum','intern']): return 'Praktikum'
    if any(k in tl for k in ['ausbildung','apprentice','azubi']): return 'Ausbildung'
    if any(k in tl for k in ['duales studium']): return 'Duales Studium'
    if any(k in tl for k in ['teilzeit','part-time']): return 'Teilzeit'
    return 'Vollzeit'

def get_remote(t, ho=False, hot=''):
    tl = t.lower()
    if 'hybrid' in tl: return 'Hybrid'
    if any(k in tl for k in ['remote','homeoffice']): return 'Remote'
    if ho: return 'Hybrid' if 'VEREINBARUNG' in (hot or '') else 'Remote'
    return 'Nicht angegeben'

def parse_item(item):
    title = item.get('stellenangebotsTitel', '')
    refnr = item.get('referenznummer', '') or item.get('chiffrennummer', '')
    
    lokationen = item.get('stellenlokationen', [])
    city = plz = street = ''
    if lokationen:
        addr = lokationen[0].get('adresse', {})
        city = addr.get('ort', '')
        plz = addr.get('plz', '')
        street = (addr.get('strasse', '') + ' ' + addr.get('hausnummer', '')).strip()
    
    employer = item.get('firma', '') or item.get('arbeitgeber', '')
    homeoffice = item.get('homeofficemoeglich', False)
    homeofficetyp = item.get('homeofficetyp', '')
    vollzeit = item.get('arbeitszeitVollzeit', False)
    befristet = item.get('vertragsdauer', '') == 'BEFRISTET'
    eintritt = item.get('eintrittszeitraum', {})
    eintrittsdatum = eintritt.get('von', '') if isinstance(eintritt, dict) else ''
    veroeff = item.get('veroeffentlichungszeitraum', {})
    veroeffdatum = veroeff.get('von', '') if isinstance(veroeff, dict) else ''
    url = item.get('externeURL', '')
    hauptberuf = item.get('hauptberuf', '')
    quereinstieg = item.get('quereinstiegGeeignet', False)
    hash_id = item.get('arbeitgeberKundennummerHash', '')
    
    return {
        'title': title, 'refnr': refnr, 'employer_name': employer,
        'city': city, 'plz': plz, 'street': street,
        'homeoffice': homeoffice, 'homeofficetyp': homeofficetyp,
        'vollzeit': vollzeit, 'befristet': befristet,
        'eintrittsdatum': eintrittsdatum, 'veroeffdatum': veroeffdatum,
        'url': url, 'hauptberuf': hauptberuf,
        'quereinstieg': quereinstieg, 'employer_hash': hash_id,
    }


async def search(client, term, city, page=1, size=100):
    params = {'was': term, 'wo': city, 'umkreis': 50, 'page': page, 'size': size}
    headers = {'X-API-Key': AA_API_KEY, 'Accept': 'application/json'}
    try:
        r = await client.get(AA_API_URL, params=params, headers=headers, timeout=15)
        if r.status_code == 200:
            return r.json()
    except:
        pass
    return {}


async def main():
    print("="*60)
    print("Phase 1 FINAL: Arbeitsagentur API - Junior IT Jobs")
    print("="*60)
    
    all_jobs = {}
    
    # Two sets of searches:
    # 1. Junior-specific searches
    # 2. Broad IT searches (then filter for non-senior)
    
    junior_terms = [
        'Junior IT', 'Junior Software', 'Junior Entwickler',
        'Werkstudent IT', 'Werkstudent Software',
        'Praktikum IT', 'Praktikum Software',
        'Trainee IT', 'Ausbildung Informatik', 'Duales Studium IT',
        'Junior Developer', 'Junior Data', 'Junior DevOps',
        'Einsteiger IT', 'Berufseinsteiger Informatik',
        'Junior Frontend', 'Junior Backend', 'Junior Fullstack',
    ]
    
    broad_terms = [
        'Softwareentwickler', 'Software Developer', 'Frontend Entwickler',
        'Backend Entwickler', 'Fullstack Entwickler', 'IT Administrator',
        'Data Analyst', 'DevOps Engineer', 'Cloud Engineer',
        'IT Support', 'SAP Consultant', 'Software Engineer',
        'System Administrator', 'IT Consultant', 'Web Developer',
        'Informatik', 'IT Fachinformatiker', 'Fachinformatiker',
        'Data Scientist', 'IT Specialist', 'Cyber Security',
        'IT Projektmanager', 'Scrum Master', 'Test Engineer',
        'UX Designer', 'App Entwickler', 'Mobile Developer',
    ]
    
    cities = ['Erlangen', 'Nürnberg', 'Bamberg', 'Fürth', 'Forchheim', 'Herzogenaurach']
    
    all_terms = junior_terms + broad_terms
    all_combos = [(term, city, page) for term in all_terms for city in cities for page in range(1, 6)]
    
    print(f"  {len(all_combos)} search combinations")
    
    async with httpx.AsyncClient() as client:
        sem = asyncio.Semaphore(8)
        
        async def limited(term, city, page):
            async with sem:
                return await search(client, term, city, page)
        
        chunk_size = 20
        for i in range(0, len(all_combos), chunk_size):
            chunk = all_combos[i:i+chunk_size]
            tasks = [limited(t, c, p) for t, c, p in chunk]
            results = await asyncio.gather(*tasks, return_exceptions=True)
            
            for result in results:
                if isinstance(result, dict):
                    for item in result.get('ergebnisliste', []):
                        job = parse_item(item)
                        title = job['title']
                        refnr = job['refnr']
                        
                        if not title or not is_it_job(title):
                            continue
                        if is_senior(title):
                            continue
                        
                        is_j = is_junior(title)
                        jk_list = [kw for kw in ['junior','einsteiger','trainee','werkstudent','praktikant','absolvent','duales studium','ausbildung','berufseinsteiger','associate','nachwuchs','perspektive'] if kw in title.lower()]
                        jk = jk_list[0] if jk_list else 'no_senior_keyword'
                        it_kw = next((kw for kw in IT_KEYWORDS if kw in title.lower()), 'it')
                        
                        job['is_junior'] = is_j
                        job['junior_kw'] = jk
                        job['it_kw'] = it_kw
                        job['remote_option'] = get_remote(title, job.get('homeoffice'), job.get('homeofficetyp',''))
                        job['source'] = 'Arbeitsagentur'
                        
                        key = refnr if refnr else f"{title}_{job.get('employer_hash','')}_{job.get('city','')}"
                        if key not in all_jobs:
                            all_jobs[key] = job
            
            done = min(i+chunk_size, len(all_combos))
            if (i // chunk_size + 1) % 3 == 0 or done >= len(all_combos):
                print(f"  {done}/{len(all_combos)} done | {len(all_jobs)} unique IT jobs (non-senior)")
    
    # Separate junior and non-junior-but-not-senior
    junior_count = len([j for j in all_jobs.values() if j.get('is_junior')])
    non_senior_count = len([j for j in all_jobs.values() if not j.get('is_junior')])
    print(f"\n  Total: {len(all_jobs)} non-senior IT jobs")
    print(f"  Explicitly junior: {junior_count}")
    print(f"  No senior keyword (junior-friendly): {non_senior_count}")
    
    # Create Excel
    print("\n  Creating Excel...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Junior IT-Jobs"
    
    headers = [
        'Job-ID','Job-Titel','Job-Titel (Original)','Job-Kategorie','Erfahrungslevel',
        'Beschäftigungsart','Remote-Option','Technologien',
        'Ausschreibungstext (Auszug)','Job-URL',
        'Arbeitgeber','Arbeitgeber (norm.)','Ort','PLZ','Straße',
        'Telefon','E-Mail','Branche',
        'Quelle','Karriere-URL','Website-URL',
        'Karriere-HTML Pfad','Job-HTML Pfad',
        'Junior-Keyword Match','IT-Keyword Match','Ist Junior-Job','Ist IT-Job',
        'Vollzeit/Teilzeit Detail','Befristet','Gehaltsspanne','Arbeitsagentur RefNr',
        'Hauptberuf','Eintrittsdatum','Veroeffentlichungsdatum',
        'Homeoffice Möglich','Homeoffice Typ','Quereinstieg',
        'Arbeitgeber Hash',
        'Scrape-Datum','Scrape-Status','Arbeitgeber-ID',
        'Letzte Prüfung','Erstellt','Aktualisiert',
    ]
    
    hf = Font(bold=True, color='FFFFFF', size=11)
    hfill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.font = hf; c.fill = hfill
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    
    # Sort: junior first, then by employer
    sorted_jobs = sorted(all_jobs.values(), key=lambda j: (0 if j.get('is_junior') else 1, j.get('employer_name','')))
    
    for idx, job in enumerate(sorted_jobs):
        row = idx + 2
        title = job['title']
        emp = job.get('employer_name', '')
        emp_norm = re.sub(r'\s+', '', re.sub(r'(GmbH|AG|e\.K\.|UG|KG|mbH|Co\.).*$', '', emp)).lower()
        
        values = [
            f'J-{idx+1:05d}',
            title,
            title,
            classify_cat(title),
            'Junior' if job.get('is_junior') else 'Mid-Level',
            get_emp_type(title),
            job.get('remote_option', get_remote(title, job.get('homeoffice'), job.get('homeofficetyp',''))),
            ', '.join(extract_techs(title)),
            '',  # description
            job.get('url', f"https://www.arbeitsagentur.de/jobsuche/jobdetail/{job.get('refnr','')}"),
            emp,
            emp_norm,
            job.get('city',''),
            job.get('plz',''),
            job.get('street',''),
            '',  # phone
            '',  # email
            '',  # branche
            'Arbeitsagentur',
            '',  # karriere url
            '',  # website url
            '',  # career html
            '',  # job html
            job.get('junior_kw',''),
            job.get('it_kw',''),
            'Ja' if job.get('is_junior') else 'Nein',
            'Ja',
            get_emp_type(title),
            'Ja' if job.get('befristet') else 'Nein',
            '',  # salary
            job.get('refnr',''),
            job.get('hauptberuf',''),
            job.get('eintrittsdatum',''),
            job.get('veroeffdatum',''),
            'Ja' if job.get('homeoffice') else 'Nein',
            job.get('homeofficetyp',''),
            'Ja' if job.get('quereinstieg') else 'Nein',
            job.get('employer_hash',''),
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'completed',
            '',
            datetime.now().strftime('%Y-%m-%d'),
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        ]
        for col, v in enumerate(values, 1):
            ws.cell(row, col, v).alignment = Alignment(wrap_text=True, vertical='top')
    
    # Stats sheet
    ws2 = wb.create_sheet('Statistiken')
    stats = [
        ('Gesamt IT-Jobs (Junior & Junior-fähig)', len(all_jobs)),
        ('Davon explizit Junior', junior_count),
        ('Davon Junior-fähig (kein Senior-Keyword)', non_senior_count),
        ('Quelle', 'Arbeitsagentur API'),
        ('Scrape-Datum', datetime.now().strftime('%Y-%m-%d %H:%M')),
    ]
    for i, (k, v) in enumerate(stats, 1):
        ws2.cell(i, 1, k); ws2.cell(i, 2, v)
    
    # Category breakdown
    cats = {}
    for job in all_jobs.values():
        cat = classify_cat(job['title'])
        cats[cat] = cats.get(cat, 0) + 1
    r = len(stats) + 2
    ws2.cell(r, 1, 'Nach Kategorie:')
    for i, (cat, count) in enumerate(sorted(cats.items(), key=lambda x: -x[1])):
        ws2.cell(r+1+i, 1, cat); ws2.cell(r+1+i, 2, count)
    
    # City breakdown
    cities_count = {}
    for job in all_jobs.values():
        c = job.get('city', 'Unbekannt')
        cities_count[c] = cities_count.get(c, 0) + 1
    r2 = r + len(cats) + 3
    ws2.cell(r2, 1, 'Nach Ort:')
    for i, (c, count) in enumerate(sorted(cities_count.items(), key=lambda x: -x[1])):
        ws2.cell(r2+1+i, 1, c); ws2.cell(r2+1+i, 2, count)
    
    # Employer top list
    emps = {}
    for job in all_jobs.values():
        e = job.get('employer_name', 'Unbekannt')
        emps[e] = emps.get(e, 0) + 1
    r3 = r2 + len(cities_count) + 3
    ws2.cell(r3, 1, 'Top Arbeitgeber:')
    for i, (e, count) in enumerate(sorted(emps.items(), key=lambda x: -x[1])[:20]):
        ws2.cell(r3+1+i, 1, e); ws2.cell(r3+1+i, 2, count)
    
    wb.save(EXCEL_OUT)
    print(f"  Excel saved: {EXCEL_OUT}")
    
    # Save JSON
    json_out = EXCEL_OUT.replace('.xlsx', '_raw.json')
    with open(json_out, 'w', encoding='utf-8') as f:
        json.dump(sorted_jobs, f, ensure_ascii=False, indent=2, default=str)
    print(f"  JSON saved: {json_out}")
    
    # Git push
    os.chdir(GIT_REPO)
    subprocess.run(['git', 'add', '-A'], capture_output=True, timeout=30)
    r = subprocess.run(['git', 'commit', '-m', f'Phase 1: Junior IT Jobs ({len(all_jobs)} jobs, {junior_count} junior, {non_senior_count} junior-ready)'], capture_output=True, timeout=30)
    if r.returncode == 0:
        push_r = subprocess.run(['git', 'push', f'https://{GITHUB_TOKEN}@github.com/BirdNest055/Heyjobs.git', 'main', '--force-with-lease'], capture_output=True, timeout=60)
        print(f"  Git: Pushed!")
    else:
        print(f"  Git: No new changes")
    
    print(f"\n{'='*60}")
    print(f"Phase 1 DONE: {len(all_jobs)} IT jobs collected")
    print(f"{'='*60}")

if __name__ == '__main__':
    asyncio.run(main())
