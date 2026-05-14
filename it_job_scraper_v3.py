#!/usr/bin/env python3
"""
IT-Job Scraper v3 - Bamberg, Erlangen, Nürnberg
Strategy:
1. Query Arbeitsagentur API for ALL IT jobs in each city
2. Build set of employer names that have IT jobs
3. Cross-reference with Google Maps employer list
4. For matching employers: collect website, download HTML, extract contacts
5. Also score IT relevance for ALL employers based on name/category
6. Save to database-friendly Excel with atomic columns
7. Commit & push every 10 entries
"""

import json, os, re, sys, time, hashlib, sqlite3, subprocess
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
import requests
from bs4 import BeautifulSoup
import openpyxl

sys.stdout.reconfigure(line_buffering=True)

# ============================================================
# CONFIG
# ============================================================
EXCEL_IN = "/home/z/my-project/download/Google_Maps_Firmen_Erlangen_50km.xlsx"
EXCEL_OUT = "/home/z/my-project/download/IT_Arbeitgeber_Bamberg_Erlangen_Nuernberg.xlsx"
HTML_DIR = "/home/z/my-project/download/it_employer_html"
DB_PATH = "/home/z/my-project/download/it_employers.db"
PROGRESS_FILE = "/home/z/my-project/download/it_scraper_progress.json"
GIT_DIR = "/home/z/my-project"

TARGET_CITIES = {"Bamberg": "96047", "Erlangen": "91052", "Nürnberg": "90402"}
MAX_WORKERS = 8
COMMIT_EVERY = 10

AA_API_URL = "https://rest.arbeitsagentur.de/jobboerse/jobsuche-service/pc/v6/jobs"
AA_API_KEY = "jobboerse-jobsuche"

IT_SEARCH_TERMS = [
    "Software Entwickler", "IT Administrator", "Systemadministrator",
    "DevOps Engineer", "Data Engineer", "Data Scientist",
    "Frontend Entwickler", "Backend Entwickler", "Fullstack Entwickler",
    "IT Projektleiter", "SAP Consultant", "Cloud Engineer",
    "Cyber Security", "IT Support", "Network Engineer",
    "Informatiker", "Softwareentwickler", "IT Specialist",
    "Business Intelligence", "ERP Consultant", "Scrum Master",
    "Product Owner IT", "IT Consultant", "Software Tester",
]

HTTP_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "de-DE,de;q=0.9,en;q=0.5",
}

SESSION = requests.Session()
SESSION.headers.update(HTTP_HEADERS)
SESSION.verify = False

import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ============================================================
# UTILITIES
# ============================================================

def sanitize_filename(name, max_len=80):
    name = re.sub(r'[äöüßÄÖÜ]', lambda m: {'ä':'ae','ö':'oe','ü':'ue','ß':'ss','Ä':'Ae','Ö':'Oe','Ü':'Ue'}[m.group()], name)
    name = re.sub(r'[^a-zA-Z0-9._-]', '_', name)
    name = re.sub(r'_+', '_', name).strip('_')
    return name[:max_len]

def normalize_name(name):
    name = name.lower().strip()
    name = re.sub(r'\s+', ' ', name)
    for suffix in ['gmbh & co. kg', 'gmbh & co kg', 'gmbh co. kg', 'gmbh co kg',
                   'gmbh & co.', 'ag & co. kg', 'ag & co kg',
                   'gmbh', 'ag', 'kg', 'ohg', 'eg', 'se', 'e.k.', 'ek',
                   'e.v.', 'ev', 'mbh', 'ug', 'gbr']:
        name = re.sub(re.escape(suffix), '', name, flags=re.IGNORECASE)
    name = re.sub(r'[^a-z0-9äöüß]', '', name)
    return name

def load_progress():
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE) as f:
            return json.load(f)
    return {"completed": [], "it_jobs_collected": False, "phase": "start"}

def save_progress(progress):
    with open(PROGRESS_FILE, 'w') as f:
        json.dump(progress, f, ensure_ascii=False)

def make_absolute_url(base_url, relative_url):
    if not relative_url:
        return None
    if relative_url.startswith('http'):
        return relative_url
    if relative_url.startswith('//'):
        return 'https:' + relative_url
    if relative_url.startswith('/'):
        from urllib.parse import urlparse
        parsed = urlparse(base_url)
        return f"{parsed.scheme}://{parsed.netloc}{relative_url}"
    return f"{base_url.rstrip('/')}/{relative_url.lstrip('/')}"

# ============================================================
# PHASE 1: COLLECT ALL IT JOBS FROM ARBEITSAGENTUR
# ============================================================

def collect_all_it_jobs():
    """Query AA API for all IT jobs in our 3 cities. Build employer->jobs map."""
    employer_jobs = {}  # normalized_name -> list of jobs
    
    headers = {"X-API-Key": AA_API_KEY, "Accept": "application/json"}
    
    for city, plz in TARGET_CITIES.items():
        print(f"\n--- Sammle IT-Jobs für {city} (PLZ {plz}) ---")
        for term in IT_SEARCH_TERMS:
            page = 1
            total_found = 0
            while True:
                params = {
                    "was": term,
                    "wo": plz,
                    "umkreis": "25",
                    "page": str(page),
                    "size": "100",
                }
                try:
                    resp = requests.get(AA_API_URL, params=params, headers=headers, timeout=15)
                    if resp.status_code != 200:
                        break
                    
                    data = resp.json()
                    ergebnisse = data.get("ergebnisliste", [])
                    max_results = data.get("maxErgebnisse", 0)
                    
                    if not ergebnisse:
                        break
                    
                    for job in ergebnisse:
                        firma = job.get("firma", "").strip()
                        titel = job.get("stellenangebotsTitel", "").strip()
                        refnr = job.get("referenznummer", "")
                        hauptberuf = job.get("hauptberuf", "")
                        berufe = job.get("alleBerufe", [])
                        lokationen = job.get("stellenlokationen", [])
                        ort = ""
                        job_plz = ""
                        if lokationen:
                            addr = lokationen[0].get("adresse", {})
                            ort = addr.get("ort", "")
                            job_plz = addr.get("plz", "")
                        
                        norm_firma = normalize_name(firma)
                        if len(norm_firma) < 3:
                            continue
                        
                        if norm_firma not in employer_jobs:
                            employer_jobs[norm_firma] = {
                                "firma_original": firma,
                                "jobs": [],
                                "stadt": city,
                            }
                        
                        employer_jobs[norm_firma]["jobs"].append({
                            "titel": titel,
                            "hauptberuf": hauptberuf,
                            "berufe": berufe,
                            "referenznummer": refnr,
                            "ort": ort,
                            "plz": job_plz,
                            "eintritt": job.get("eintrittszeitraum", {}).get("von", ""),
                            "art": job.get("stellenangebotsart", ""),
                            "vollzeit": job.get("arbeitszeitVollzeit", False),
                            "unbefristet": job.get("vertragsdauer", "") == "UNBEFRISTET",
                        })
                        total_found += 1
                    
                    # Next page?
                    if len(ergebnisse) < 100 or page * 100 >= max_results:
                        break
                    page += 1
                    time.sleep(0.3)
                    
                except Exception as e:
                    print(f"  Error searching '{term}' in {city}: {e}")
                    break
            
            if total_found > 0:
                print(f"  '{term}': {total_found} Jobs gefunden")
            time.sleep(0.2)
    
    # Deduplicate jobs per employer
    for norm_name, data in employer_jobs.items():
        seen_refs = set()
        unique_jobs = []
        for j in data["jobs"]:
            ref = j.get("referenznummer", "")
            if ref not in seen_refs:
                seen_refs.add(ref)
                unique_jobs.append(j)
        data["jobs"] = unique_jobs
    
    print(f"\nPhase 1 komplett: {len(employer_jobs)} eindeutige Arbeitgeber mit IT-Jobs")
    return employer_jobs

# ============================================================
# PHASE 2: MATCH IT EMPLOYERS WITH GOOGLE MAPS DATA
# ============================================================

def match_employers(employers_list, it_jobs_map):
    """Match Google Maps employers with IT job employers."""
    matched = {}
    
    for emp in employers_list:
        norm = normalize_name(emp['name'])
        score = 0.0
        has_it_jobs = 0
        it_job_data = None
        
        # Direct match
        if norm in it_jobs_map:
            has_it_jobs = 1
            it_job_data = it_jobs_map[norm]
            score = 50
        else:
            # Fuzzy match: check if part of name matches
            for it_norm, it_data in it_jobs_map.items():
                if len(norm) > 4 and len(it_norm) > 4:
                    if norm[:6] in it_norm or it_norm[:6] in norm:
                        has_it_jobs = 1
                        it_job_data = it_data
                        score = 40
                        break
        
        # IT relevance from name/category
        name_lower = emp['name'].lower()
        kat_lower = (emp.get('kategorie') or '').lower()
        
        it_name_words = ['it ', 'it-', 'software', 'tech', 'digital', 'data', 'cloud', 
                         'cyber', 'computer', 'informatik', 'system', 'consulting',
                         'automation', 'iot', 'embedded', 'telekommunikation']
        it_kat_words = ['software', 'it', 'informatik', 'computer', 'technologie', 
                        'digital', 'internet', 'telekommunikation']
        
        name_match = any(w in name_lower for w in it_name_words)
        kat_match = any(w in kat_lower for w in it_kat_words)
        
        if name_match:
            score += 20
        if kat_match:
            score += 15
        
        matched[emp['name']] = {
            'employer': emp,
            'has_it_jobs': has_it_jobs,
            'it_job_data': it_job_data,
            'it_relevanz_score': min(score, 100),
            'name_it_match': name_match,
            'kat_it_match': kat_match,
        }
    
    return matched

# ============================================================
# PHASE 3: ENRICH WITH WEBSITES, HTML, CONTACTS
# ============================================================

def find_website_via_google(name, city):
    """Try to find website using Google search API or fallback methods."""
    # Method 1: Try constructing likely URLs
    clean_name = re.sub(r'(GmbH|AG|KG|OHG|UG|e\.K\.|e\.V\.|mbH|Co\.\s*KG|& Co\.)', '', name, flags=re.IGNORECASE).strip()
    # Remove special chars
    domain_name = re.sub(r'[äöüß]', lambda m: {'ä':'ae','ö':'oe','ü':'ue','ß':'ss'}[m.group()], clean_name)
    domain_name = re.sub(r'[^a-zA-Z0-9\s]', '', domain_name)
    domain_name = re.sub(r'\s+', '', domain_name).lower()
    
    # Try common patterns
    candidates = [
        f"https://www.{domain_name}.de",
        f"https://{domain_name}.de",
    ]
    
    for url in candidates:
        if len(domain_name) < 4:
            break
        try:
            resp = SESSION.get(url, timeout=8, allow_redirects=True)
            if resp.status_code == 200 and 'text/html' in resp.headers.get('content-type', ''):
                return url
        except:
            pass
    
    return None

def download_html(url, label=""):
    try:
        resp = SESSION.get(url, timeout=12, allow_redirects=True)
        if resp.status_code == 200:
            ct = resp.headers.get('content-type', '')
            if 'text/html' in ct or 'text/' in ct:
                return resp.text
    except:
        pass
    return None

def save_html_file(content, company_name, suffix=""):
    if not content:
        return None, None
    safe_name = sanitize_filename(company_name)
    filename = f"{safe_name}_{suffix}.html" if suffix else f"{safe_name}.html"
    filepath = os.path.join(HTML_DIR, filename)
    os.makedirs(HTML_DIR, exist_ok=True)
    with open(filepath, 'w', encoding='utf-8', errors='replace') as f:
        f.write(content)
    html_hash = hashlib.sha256(content.encode()).hexdigest()[:16]
    return filepath, html_hash

def extract_contact_info(html_content):
    info = {}
    impressum_url = None
    karriere_url = None
    
    if not html_content:
        return info, None, None
    
    # Emails
    emails = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', html_content)
    emails = [e for e in emails if not any(x in e.lower() for x in ['example.com', 'sentry', 'webpack', 'noreply', 'wixpress', 'github.com', 'googleapis'])]
    mailtos = re.findall(r'mailto:([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})', html_content)
    emails = list(dict.fromkeys(emails + mailtos))[:5]
    if emails:
        info['email'] = emails[0]
        info['alle_emails'] = "; ".join(emails)
    
    # Geschäftsführer
    gf_match = re.search(r'(?:Geschäftsführer|Geschaeftsfuehrer)[\s:]*([^<\.]{2,100})', html_content, re.IGNORECASE)
    if gf_match:
        gf = BeautifulSoup(gf_match.group(1), 'html.parser').get_text(strip=True)
        if len(gf) > 2:
            info['geschaeftsfuehrer'] = gf[:200]
    
    # Handelsregister
    hr_match = re.search(r'(?:Handelsregister|HRB|HR-Aktenzeichen)[\s:]*([^<\.]{2,100})', html_content, re.IGNORECASE)
    if hr_match:
        hr = BeautifulSoup(hr_match.group(1), 'html.parser').get_text(strip=True)
        info['handelsregister'] = hr[:200]
    
    # Find links
    soup = BeautifulSoup(html_content, 'html.parser')
    
    for a in soup.find_all('a', href=True):
        href = a['href']
        href_lower = href.lower()
        text_lower = a.get_text(strip=True).lower()
        
        if not impressum_url and ('impressum' in href_lower or 'impressum' in text_lower or 'imprint' in href_lower):
            impressum_url = href
        
        if not karriere_url:
            career_words = ['karriere', 'career', 'jobs', 'stellenangebote', 'bewerbung', 'vacancies']
            if any(w in href_lower for w in career_words) or any(w in text_lower for w in career_words):
                karriere_url = href
    
    return info, impressum_url, karriere_url

def check_website_for_it(html_content):
    """Check if website content mentions IT jobs."""
    if not html_content:
        return False, []
    
    text = html_content.lower()
    it_kws = [kw for kw in ['software', 'entwickler', 'developer', 'informatik', 'it-', 
              'data', 'cloud', 'devops', 'administrator', 'programmierer', 'sap', 
              'linux', 'python', 'java ', 'javascript', 'scrum', 'agile', 'docker',
              'kubernetes', 'frontend', 'backend'] if kw in text]
    
    career_kws = [kw for kw in ['karriere', 'career', 'stellenangebote', 'jobs', 'bewerbung'] if kw in text]
    
    has_it = bool(it_kws) and bool(career_kws)
    # Also check specific IT job indicators
    specific = ['entwickler m/w/d', 'developer m/w/d', 'it-jobs', 'softwareentwickler gesucht',
                'it specialist', 'informatiker gesucht', 'devops engineer', 'it-stellen']
    if any(s in text for s in specific):
        has_it = True
    
    return has_it, it_kws[:10]

def enrich_employer(emp_data, matched_info):
    """Enrich single employer with website, HTML, contacts."""
    emp = emp_data
    result = {
        'firmenname': emp['name'],
        'ort': emp['ort'],
        'plz': emp['plz'],
        'strasse': emp['adresse'],
        'telefon': emp.get('telefon', ''),
        'email': '',
        'website_url': '',
        'website_html_path': '',
        'website_html_hash': '',
        'karriere_url': '',
        'karriere_html_path': '',
        'impressum_url': '',
        'impressum_html_path': '',
        'hat_it_jobs': matched_info.get('has_it_jobs', 0),
        'it_job_titel': '',
        'it_job_anzahl': 0,
        'it_job_quelle': '',
        'it_job_details': '',
        'it_keywords_gefunden': '',
        'kategorie_google': emp.get('kategorie', ''),
        'bewertung_google': emp.get('bewertung'),
        'lat': emp.get('lat'),
        'lng': emp.get('lng'),
        'google_maps_link': emp.get('gmaps_link', ''),
        'handelsregister': '',
        'geschaeftsfuehrer': '',
        'it_abteilung_vermutung': 1 if matched_info.get('it_relevanz_score', 0) >= 30 else 0,
        'it_relevanz_score': matched_info.get('it_relevanz_score', 0),
    }
    
    # IT Jobs data
    it_data = matched_info.get('it_job_data')
    if it_data:
        jobs = it_data.get('jobs', [])
        result['it_job_anzahl'] = len(jobs)
        result['it_job_titel'] = "; ".join(j.get('titel', '') for j in jobs[:10])
        result['it_job_quelle'] = 'Arbeitsagentur'
        result['it_job_details'] = json.dumps([{
            'titel': j.get('titel', ''),
            'beruf': j.get('hauptberuf', ''),
            'refnr': j.get('referenznummer', ''),
            'ort': j.get('ort', ''),
            'vollzeit': j.get('vollzeit', False),
            'unbefristet': j.get('unbefristet', False),
        } for j in jobs[:20]], ensure_ascii=False)
    
    # Find website
    website = find_website_via_google(emp['name'], emp['ort'])
    
    if website:
        result['website_url'] = website
        
        # Download main page
        main_html = download_html(website)
        if main_html:
            path, hash_val = save_html_file(main_html, emp['name'])
            result['website_html_path'] = path
            result['website_html_hash'] = hash_val
            
            # Check for IT on website
            has_it, it_kws = check_website_for_it(main_html)
            if has_it:
                if not result['hat_it_jobs']:
                    result['hat_it_jobs'] = 1
                if not result['it_job_quelle']:
                    result['it_job_quelle'] = 'Website'
                result['it_keywords_gefunden'] = ", ".join(it_kws)
                result['it_relevanz_score'] = min(result['it_relevanz_score'] + 15, 100)
            
            # Extract contacts
            contact, imp_url, kar_url = extract_contact_info(main_html)
            if contact.get('email'):
                result['email'] = contact['email']
            if contact.get('geschaeftsfuehrer'):
                result['geschaeftsfuehrer'] = contact['geschaeftsfuehrer']
            if contact.get('handelsregister'):
                result['handelsregister'] = contact['handelsregister']
            
            # Impressum
            if imp_url:
                imp_abs = make_absolute_url(website, imp_url)
                if imp_abs:
                    result['impressum_url'] = imp_abs
                    imp_html = download_html(imp_abs)
                    if imp_html:
                        path, _ = save_html_file(imp_html, emp['name'], "impressum")
                        result['impressum_html_path'] = path
                        imp_c, _, _ = extract_contact_info(imp_html)
                        if imp_c.get('email') and not result['email']:
                            result['email'] = imp_c['email']
                        if imp_c.get('geschaeftsfuehrer') and not result['geschaeftsfuehrer']:
                            result['geschaeftsfuehrer'] = imp_c['geschaeftsfuehrer']
                        if imp_c.get('handelsregister') and not result['handelsregister']:
                            result['handelsregister'] = imp_c['handelsregister']
            
            # Karriere
            if kar_url:
                kar_abs = make_absolute_url(website, kar_url)
                if kar_abs:
                    result['karriere_url'] = kar_abs
                    kar_html = download_html(kar_abs)
                    if kar_html:
                        path, _ = save_html_file(kar_html, emp['name'], "karriere")
                        result['karriere_html_path'] = path
                        has_it_kar, _ = check_website_for_it(kar_html)
                        if has_it_kar and not result['hat_it_jobs']:
                            result['hat_it_jobs'] = 1
    
    result['scrape_status'] = 'completed'
    result['scrape_datum'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    result['letzte_pruefung'] = datetime.now().strftime('%Y-%m-%d')
    
    return result

# ============================================================
# DATABASE
# ============================================================

def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS employers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            firmenname TEXT NOT NULL,
            firmenname_normalisiert TEXT,
            ort TEXT,
            plz TEXT,
            strasse TEXT,
            telefon TEXT,
            email TEXT,
            website_url TEXT,
            website_html_path TEXT,
            website_html_hash TEXT,
            karriere_url TEXT,
            karriere_html_path TEXT,
            impressum_url TEXT,
            impressum_html_path TEXT,
            hat_it_jobs INTEGER DEFAULT 0,
            it_job_titel TEXT,
            it_job_anzahl INTEGER DEFAULT 0,
            it_job_quelle TEXT,
            it_job_details TEXT,
            it_keywords_gefunden TEXT,
            kategorie_google TEXT,
            bewertung_google REAL,
            lat REAL,
            lng REAL,
            google_maps_link TEXT,
            handelsregister TEXT,
            geschaeftsfuehrer TEXT,
            it_abteilung_vermutung INTEGER DEFAULT 0,
            it_relevanz_score REAL DEFAULT 0.0,
            scrape_datum TEXT,
            scrape_status TEXT DEFAULT 'pending',
            letzte_pruefung TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_ort ON employers(ort);
        CREATE INDEX IF NOT EXISTS idx_hat_it ON employers(hat_it_jobs);
        CREATE INDEX IF NOT EXISTS idx_score ON employers(it_relevanz_score);
    """)
    conn.commit()
    return conn

def save_to_db(conn, data):
    cols = ['firmenname','firmenname_normalisiert','ort','plz','strasse',
            'telefon','email','website_url','website_html_path','website_html_hash',
            'karriere_url','karriere_html_path','impressum_url','impressum_html_path',
            'hat_it_jobs','it_job_titel','it_job_anzahl','it_job_quelle',
            'it_job_details','it_keywords_gefunden','kategorie_google',
            'bewertung_google','lat','lng','google_maps_link',
            'handelsregister','geschaeftsfuehrer','it_abteilung_vermutung',
            'it_relevanz_score','scrape_datum','scrape_status','letzte_pruefung']
    vals = [
        data.get('firmenname',''), normalize_name(data.get('firmenname','')),
        data.get('ort',''), data.get('plz',''), data.get('strasse',''),
        data.get('telefon',''), data.get('email',''), data.get('website_url',''),
        data.get('website_html_path',''), data.get('website_html_hash',''),
        data.get('karriere_url',''), data.get('karriere_html_path',''),
        data.get('impressum_url',''), data.get('impressum_html_path',''),
        data.get('hat_it_jobs',0), data.get('it_job_titel',''),
        data.get('it_job_anzahl',0), data.get('it_job_quelle',''),
        data.get('it_job_details',''), data.get('it_keywords_gefunden',''),
        data.get('kategorie_google',''), data.get('bewertung_google'),
        data.get('lat'), data.get('lng'), data.get('google_maps_link',''),
        data.get('handelsregister',''), data.get('geschaeftsfuehrer',''),
        data.get('it_abteilung_vermutung',0), data.get('it_relevanz_score',0.0),
        data.get('scrape_datum',''), data.get('scrape_status',''),
        data.get('letzte_pruefung',''),
    ]
    conn.execute(f"INSERT OR REPLACE INTO employers ({','.join(cols)}) VALUES ({','.join(['?']*len(cols))})", vals)
    conn.commit()

def db_to_excel(conn):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IT Arbeitgeber"
    
    cursor = conn.execute("SELECT * FROM employers ORDER BY it_relevanz_score DESC, hat_it_jobs DESC, firmenname")
    columns = [desc[0] for desc in cursor.description]
    
    header_map = {
        'id': 'ID', 'firmenname': 'Firmenname', 'firmenname_normalisiert': 'Firmenname (normalisiert)',
        'ort': 'Ort', 'plz': 'PLZ', 'strasse': 'Straße',
        'telefon': 'Telefon', 'email': 'E-Mail', 'website_url': 'Website URL',
        'website_html_path': 'Website HTML Pfad', 'website_html_hash': 'Website HTML Hash',
        'karriere_url': 'Karriere-URL', 'karriere_html_path': 'Karriere HTML Pfad',
        'impressum_url': 'Impressum URL', 'impressum_html_path': 'Impressum HTML Pfad',
        'hat_it_jobs': 'IT-Jobs vorhanden', 'it_job_titel': 'IT-Job Titel',
        'it_job_anzahl': 'Anzahl IT-Jobs', 'it_job_quelle': 'IT-Job Quelle',
        'it_job_details': 'IT-Job Details', 'it_keywords_gefunden': 'IT-Keywords gefunden',
        'kategorie_google': 'Kategorie (Google)', 'bewertung_google': 'Bewertung (Google)',
        'lat': 'Breitengrad', 'lng': 'Längengrad', 'google_maps_link': 'Google Maps Link',
        'handelsregister': 'Handelsregister', 'geschaeftsfuehrer': 'Geschäftsführer',
        'it_abteilung_vermutung': 'IT-Abteilung vermutet', 'it_relevanz_score': 'IT-Relevanz Score',
        'scrape_datum': 'Scrape-Datum', 'scrape_status': 'Scrape-Status',
        'letzte_pruefung': 'Letzte Prüfung', 'created_at': 'Erstellt am', 'updated_at': 'Aktualisiert am',
    }
    
    headers = [header_map.get(c, c) for c in columns]
    ws.append(headers)
    for row in cursor:
        ws.append(list(row))
    
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
    
    # Stats
    ws2 = wb.create_sheet("Statistiken")
    stats = [
        ("Gesamt Arbeitgeber", conn.execute("SELECT COUNT(*) FROM employers").fetchone()[0]),
        ("Mit IT-Jobs", conn.execute("SELECT COUNT(*) FROM employers WHERE hat_it_jobs=1").fetchone()[0]),
        ("Mit Website", conn.execute("SELECT COUNT(*) FROM employers WHERE website_url != ''").fetchone()[0]),
        ("Mit E-Mail", conn.execute("SELECT COUNT(*) FROM employers WHERE email != ''").fetchone()[0]),
        ("IT-Abt. vermutet", conn.execute("SELECT COUNT(*) FROM employers WHERE it_abteilung_vermutung=1").fetchone()[0]),
        ("Ø IT-Relevanz", round(conn.execute("SELECT AVG(it_relevanz_score) FROM employers").fetchone()[0] or 0, 1)),
    ]
    for s in stats:
        ws2.append(s)
    
    # Per-city
    ws3 = wb.create_sheet("Nach Ort")
    ws3.append(["Ort", "Gesamt", "Mit IT-Jobs", "Mit Website", "Mit E-Mail", "Ø IT-Relevanz"])
    for row in conn.execute("""
        SELECT ort, COUNT(*),
               SUM(CASE WHEN hat_it_jobs=1 THEN 1 ELSE 0 END),
               SUM(CASE WHEN website_url != '' THEN 1 ELSE 0 END),
               SUM(CASE WHEN email != '' THEN 1 ELSE 0 END),
               ROUND(AVG(it_relevanz_score), 1)
        FROM employers GROUP BY ort ORDER BY COUNT(*) DESC
    """):
        ws3.append(list(row))
    
    wb.save(EXCEL_OUT)
    print(f"Excel saved: {EXCEL_OUT}")

# ============================================================
# MAIN
# ============================================================

def main():
    print("=" * 60)
    print("IT-Job Scraper v3: Bamberg, Erlangen, Nürnberg")
    print("=" * 60)
    
    # Load employers from Excel - EXACT city match
    wb = openpyxl.load_workbook(EXCEL_IN)
    ws = wb['Google Maps Firmen']
    
    employers = []
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        ort = str(row[6] or '').strip()
        name = str(row[1] or '').strip()
        if not name or ort not in TARGET_CITIES:
            continue
        norm = normalize_name(name)
        if norm in seen or len(norm) < 3:
            continue
        seen.add(norm)
        employers.append({
            'name': name, 'ort': ort,
            'plz': str(row[5] or '').strip(),
            'adresse': str(row[4] or '').strip(),
            'telefon': str(row[7] or '').strip(),
            'kategorie': str(row[3] or '').strip(),
            'bewertung': row[2],
            'lat': row[9], 'lng': row[10],
            'gmaps_link': str(row[11] or '').strip(),
        })
    
    print(f"Google Maps: {len(employers)} Arbeitgeber in {list(TARGET_CITIES.keys())}")
    
    progress = load_progress()
    conn = init_db()
    os.makedirs(HTML_DIR, exist_ok=True)
    
    # ========================================
    # PHASE 1: Collect IT jobs from AA API
    # ========================================
    it_jobs_map = {}
    if not progress.get('it_jobs_collected'):
        print("\n=== PHASE 1: Arbeitsagentur IT-Jobs sammeln ===")
        it_jobs_map = collect_all_it_jobs()
        
        # Save it_jobs_map for resume
        with open(os.path.join(os.path.dirname(DB_PATH), 'it_jobs_map.json'), 'w') as f:
            # Convert to serializable
            serializable = {}
            for k, v in it_jobs_map.items():
                serializable[k] = {
                    'firma_original': v['firma_original'],
                    'stadt': v['stadt'],
                    'job_count': len(v['jobs']),
                    'jobs': v['jobs'][:50],  # Limit
                }
            json.dump(serializable, f, ensure_ascii=False, indent=1)
        
        progress['it_jobs_collected'] = True
        progress['it_jobs_count'] = len(it_jobs_map)
        save_progress(progress)
        
        # Commit
        subprocess.run(['git', 'add', '-A'], cwd=GIT_DIR, capture_output=True, timeout=30)
        subprocess.run(['git', 'commit', '-m', f'IT-Scraper Phase 1: {len(it_jobs_map)} IT-Arbeitgeber gefunden'], cwd=GIT_DIR, capture_output=True, timeout=30)
        subprocess.run(['git', 'push', 'origin', 'main'], cwd=GIT_DIR, capture_output=True, timeout=60)
    else:
        print("\nPhase 1 bereits erledigt, lade IT-Jobs...")
        it_jobs_file = os.path.join(os.path.dirname(DB_PATH), 'it_jobs_map.json')
        if os.path.exists(it_jobs_file):
            with open(it_jobs_file) as f:
                loaded = json.load(f)
            for k, v in loaded.items():
                it_jobs_map[k] = {
                    'firma_original': v['firma_original'],
                    'stadt': v.get('stadt', ''),
                    'jobs': v.get('jobs', []),
                }
            print(f"  {len(it_jobs_map)} IT-Arbeitgeber geladen")
    
    # ========================================
    # PHASE 2: Match and score
    # ========================================
    print("\n=== PHASE 2: Arbeitgeber matchen ===")
    matched = match_employers(employers, it_jobs_map)
    it_matched = sum(1 for m in matched.values() if m['has_it_jobs'])
    print(f"  {it_matched} Arbeitgeber mit IT-Jobs gematcht")
    
    # ========================================
    # PHASE 3: Enrich with websites, HTML, contacts
    # ========================================
    print("\n=== PHASE 3: Websites & HTML sammeln ===")
    
    completed_names = set(progress.get('completed', []))
    remaining = [(name, info) for name, info in matched.items() if name not in completed_names]
    total = len(remaining)
    processed = 0
    it_count = 0
    batch_since_commit = 0
    
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {}
        
        for name, info in remaining:
            future = executor.submit(enrich_employer, info['employer'], info)
            futures[future] = name
        
        for future in as_completed(futures):
            name = futures[future]
            try:
                result = future.result()
            except Exception as e:
                print(f"  ERROR: {name}: {e}")
                result = {'firmenname': name, 'scrape_status': f'error: {e}'}
            
            save_to_db(conn, result)
            processed += 1
            batch_since_commit += 1
            
            progress['completed'].append(result.get('firmenname', name))
            if result.get('hat_it_jobs'):
                it_count += 1
            
            status = "✓IT" if result.get('hat_it_jobs') else "   "
            ws_url = result.get('website_url', '')[:35] if result.get('website_url') else '-'
            score = result.get('it_relevanz_score', 0)
            email = "📧" if result.get('email') else "  "
            print(f"  [{processed}/{total}] {status} {email} Score:{score:5.1f} | {result.get('firmenname','')[:42]:42} | {ws_url}")
            
            if batch_since_commit >= COMMIT_EVERY:
                save_progress(progress)
                print(f"\n  --- GIT COMMIT & PUSH ({it_count} IT-Arbeitgeber, {processed}/{total}) ---")
                try:
                    subprocess.run(['git', 'add', '-A'], cwd=GIT_DIR, capture_output=True, timeout=30)
                    msg = f"IT-Scraper: {processed}/{total} enrich, {it_count} IT-Jobs ({datetime.now().strftime('%H:%M')})"
                    r = subprocess.run(['git', 'commit', '-m', msg], cwd=GIT_DIR, capture_output=True, timeout=30)
                    if r.returncode == 0:
                        subprocess.run(['git', 'push', 'origin', 'main'], cwd=GIT_DIR, capture_output=True, timeout=60)
                        print("  --- Push OK ---\n")
                except Exception as e:
                    print(f"  GIT Error: {e}")
                batch_since_commit = 0
    
    # ========================================
    # FINAL: Generate Excel
    # ========================================
    save_progress(progress)
    print("\n=== FINALE EXCEL ERSTELLEN ===")
    db_to_excel(conn)
    
    # Final commit
    try:
        subprocess.run(['git', 'add', '-A'], cwd=GIT_DIR, capture_output=True, timeout=30)
        msg = f"IT-Scraper FERTIG: {it_count} IT-Arbeitgeber aus {total}"
        subprocess.run(['git', 'commit', '-m', msg], cwd=GIT_DIR, capture_output=True, timeout=30)
        subprocess.run(['git', 'push', 'origin', 'main'], cwd=GIT_DIR, capture_output=True, timeout=60)
    except:
        pass
    
    conn.close()
    print(f"\n{'='*60}")
    print(f"FERTIG! {it_count} IT-Arbeitgeber aus {total} gescannt")
    print(f"{'='*60}")

if __name__ == '__main__':
    main()
