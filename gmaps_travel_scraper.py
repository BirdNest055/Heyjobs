#!/usr/bin/env python3
"""
Google Maps Travel Scraper v2 - Production
"Travels" through Google Maps like a human, clicking into detail pages.
Extracts: Direct website URL, full address, PLZ, category, phone, coordinates.
Handles: Missing towns + enrichment of existing companies.
Pushes to GitHub every 10 entries.
"""

import json
import os
import re
import sys
import time
import random
import subprocess
import signal
from datetime import datetime
from pathlib import Path

# --- Configuration ---
CHROME_PATH = os.path.expanduser("~/.cache/ms-playwright/chromium-1223/chrome-linux64/chrome")
RESULTS_FILE = "/home/z/my-project/gmaps_travel_results.json"
PROGRESS_FILE = "/home/z/my-project/gmaps_travel_progress.json"
EXISTING_RESULTS = "/home/z/my-project/gmaps_erlangen_results.json"
TOWNS_FILE = "/home/z/my-project/gmaps_erlangen_towns.json"
OLD_PROGRESS = "/home/z/my-project/gmaps_erlangen_progress.json"
EXCEL_OUTPUT = "/home/z/my-project/download/Google_Maps_Firmen_Erlangen_50km.xlsx"
GITHUB_REPO = "https://github.com/BirdNest055/Heyjobs.git"
GITHUB_TOKEN = "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"
PROJECT_DIR = "/home/z/my-project"
DOWNLOAD_DIR = "/home/z/my-project/download"

# Search terms per population tier
SEARCH_TERMS = {
    1: ["Firma", "Unternehmen", "GmbH", "IT Firma", "Software", "Softwareentwicklung",
        "Handwerk", "Industrie", "Handel", "Dienstleistung", "Gastronomie",
        "Baufirma", "Kfz Werkstatt", "Steuerberater", "Rechtsanwalt",
        "Versicherung", "Immobilien", "Marketing Agentur", "Logistik",
        "Elektro", "Metallbau", "Maschinenbau", "Kunststoff", "Druckerei",
        "Apotheke", "Arztpraxis", "Friseur", "Bäckerei",
        "Ingenieurbüro", "Consulting", "Architekt", "Autohaus",
        "Bank", "Bildung", "Gesundheit", "Pharma"],
    2: ["Firma", "Unternehmen", "GmbH", "IT", "Handwerk", "Dienstleistung",
        "Industrie", "Handel", "Gastronomie", "Steuerberater", "Elektro"],
    3: ["Firma", "Unternehmen", "GmbH", "Handwerk", "Dienstleistung", "Handel", "IT"],
    4: ["Firma", "Unternehmen", "Gewerbe", "Handwerk", "Dienstleistung"],
}


def load_json(path, default=None):
    try:
        with open(path) as f:
            return json.load(f)
    except:
        return default if default is not None else {}


def save_json(path, data):
    with open(path, 'w') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def git_push(message):
    """Push to GitHub"""
    try:
        os.chdir(PROJECT_DIR)
        subprocess.run(["git", "add", "-A"], capture_output=True, timeout=30)
        result = subprocess.run(["git", "commit", "-m", message], capture_output=True, timeout=30)
        if result.returncode != 0 and "nothing to commit" in result.stdout.decode():
            return
        subprocess.run(
            ["git", "push", f"https://{GITHUB_TOKEN}@github.com/BirdNest055/Heyjobs.git", "main"],
            capture_output=True, timeout=60
        )
        print(f"  [GIT] Pushed: {message}")
    except Exception as e:
        print(f"  [GIT] Error: {e}")


def human_delay(min_s=0.5, max_s=2.0):
    time.sleep(random.uniform(min_s, max_s))


def extract_detail(page, town, search_term):
    """Extract all details from a company's Google Maps detail page"""
    result = {
        "name": "", "rating": "", "review_count": "", "category": "",
        "address": "", "plz": "", "city": "", "phone": "", "website": "",
        "lat": "", "lon": "", "gmaps_url": "",
        "search_town": town, "search_term": search_term,
        "scraped_at": datetime.now().isoformat(),
        "source": "google_maps_detail",
    }

    # Wait for detail panel
    try:
        page.wait_for_selector('[role="main"]', timeout=6000)
    except:
        pass

    time.sleep(0.5)

    # 1. Name
    try:
        name_el = page.query_selector('h1.DUwDvf, h1.fontHeadlineLarge')
        if name_el:
            result["name"] = name_el.inner_text().strip()
    except:
        pass

    # 2. Rating
    try:
        rating_el = page.query_selector('div.F7nice')
        if rating_el:
            star_el = rating_el.query_selector('span[aria-label]')
            if star_el:
                aria = star_el.get_attribute('aria-label') or ''
                rating_match = re.search(r'([\d,]+)\s*Sterne', aria)
                if rating_match:
                    result["rating"] = rating_match.group(1).replace(',', '.')
    except:
        pass

    # 3. Review count
    try:
        review_btn = page.query_selector('button[aria-label*="Bewertungen"]')
        if review_btn:
            aria = review_btn.get_attribute('aria-label') or ''
            review_match = re.search(r'([\d.]+)\s*Bewertungen', aria)
            if review_match:
                result["review_count"] = review_match.group(1)
    except:
        pass

    # 4. Category
    try:
        cat_btn = page.query_selector('button[jsaction*="category"]')
        if cat_btn:
            result["category"] = cat_btn.inner_text().strip()
    except:
        pass

    # 5. Website - primary: data-item-id="authority", fallback: aria-label containing "Website"
    try:
        web_link = page.query_selector('a[data-item-id="authority"]')
        if web_link:
            href = web_link.get_attribute('href') or ''
            if href and 'google' not in href.lower():
                result["website"] = href
    except:
        pass

    if not result["website"]:
        try:
            web_link = page.query_selector('a[aria-label*="Website"]')
            if web_link:
                href = web_link.get_attribute('href') or ''
                if href and 'google' not in href.lower() and 'gstatic' not in href.lower() and not href.startswith('tel:'):
                    result["website"] = href
        except:
            pass

    # 6. Address and Phone from data-item-id elements
    try:
        all_els = page.query_selector_all('[data-item-id]')
        for el in all_els:
            item_id = el.get_attribute('data-item-id') or ''
            if 'address' in item_id and not result["address"]:
                text = el.inner_text().strip()
                clean = re.sub(r'[\ue000-\uefff]', '', text).strip()
                # Remove newlines and extra spaces
                clean = re.sub(r'\s+', ' ', clean).strip()
                result["address"] = clean
                plz = re.search(r'(\d{5})', clean)
                if plz:
                    result["plz"] = plz.group(1)
                city_match = re.search(r'\d{5}\s+([A-Za-zäöüßÄÖÜ\-\.]+(?:\s+[A-Za-zäöüßÄÖÜ\-\.]+)*)', clean)
                if city_match:
                    result["city"] = city_match.group(1).strip()

            elif 'phone:tel:' in item_id and not result["phone"]:
                phone_match = re.search(r'phone:tel:(.+)', item_id)
                if phone_match:
                    result["phone"] = phone_match.group(1)
    except:
        pass

    # 7. Coordinates from URL
    try:
        current_url = page.url
        coord_match = re.search(r'@(-?[\d.]+),(-?[\d.]+)', current_url)
        if coord_match:
            result["lat"] = coord_match.group(1)
            result["lon"] = coord_match.group(2)
        result["gmaps_url"] = current_url
    except:
        pass

    return result


def search_and_extract(page, town, search_term, max_scroll=15, max_companies=None):
    """Search Google Maps, scroll to load results, click into each to get details"""
    companies = []
    seen_names = set()

    query = f"{town} {search_term}"
    url = f"https://www.google.com/maps/search/{query.replace(' ', '+')}"

    try:
        page.goto(url, timeout=20000, wait_until="domcontentloaded")
    except:
        try:
            page.goto(url, timeout=15000)
        except:
            print(f"  [WARN] Could not load: {query}")
            return companies

    time.sleep(random.uniform(3, 5))

    # Handle cookie consent
    try:
        accept = page.query_selector('button[aria-label*="Alle ablehnen"], button[aria-label*="Reject"]')
        if accept:
            accept.click()
            time.sleep(1)
    except:
        pass

    # Scroll to load all results
    try:
        scroll_panel = page.query_selector('div[role="feed"]')
        if scroll_panel:
            last_count = 0
            same_rounds = 0
            for i in range(max_scroll):
                page.evaluate('(el) => el.scrollTop = el.scrollHeight', scroll_panel)
                time.sleep(random.uniform(0.8, 1.5))
                items = page.query_selector_all('a[href*="/maps/place/"]')
                if len(items) == last_count:
                    same_rounds += 1
                    if same_rounds >= 3:
                        break
                else:
                    same_rounds = 0
                last_count = len(items)
    except:
        pass

    # Collect result links
    items = page.query_selector_all('a[href*="/maps/place/"]')
    # Deduplicate by href
    seen_hrefs = set()
    unique_items = []
    for item in items:
        href = item.get_attribute('href') or ''
        if href and href not in seen_hrefs:
            seen_hrefs.add(href)
            unique_items.append(item)

    if max_companies:
        unique_items = unique_items[:max_companies]

    print(f"  [{query}] Found {len(unique_items)} unique results")

    for idx, item in enumerate(unique_items):
        try:
            # Get the company name from the link text first
            link_text = ""
            try:
                name_el = item.query_selector('.fontHeadlineSmall, .qBF1Pd')
                if name_el:
                    link_text = name_el.inner_text().strip()
            except:
                pass

            if not link_text:
                try:
                    link_text = item.get_attribute('aria-label') or ''
                except:
                    pass

            if link_text in seen_names:
                continue
            seen_names.add(link_text)

            # Click into detail page
            item.click()
            time.sleep(random.uniform(2, 3.5))

            # Extract details
            detail = extract_detail(page, town, search_term)

            if detail.get("name"):
                companies.append(detail)
                web = detail.get('website', 'N/A')
                plz = detail.get('plz', 'N/A')
                cat = detail.get('category', 'N/A')
                print(f"    [{idx+1}/{len(unique_items)}] {detail['name']} | web={web} | plz={plz} | cat={cat}")
            elif link_text:
                detail["name"] = link_text
                companies.append(detail)
                print(f"    [{idx+1}/{len(unique_items)}] {link_text} (partial)")

            # Go back to results
            try:
                page.go_back(timeout=8000)
                time.sleep(random.uniform(1.5, 2.5))
            except:
                # Re-navigate to search results
                page.goto(url, timeout=15000)
                time.sleep(random.uniform(2, 3))

            # Re-find items after going back
            items = page.query_selector_all('a[href*="/maps/place/"]')
            new_unique = []
            new_seen = set()
            for it in items:
                h = it.get_attribute('href') or ''
                if h and h not in new_seen and h in seen_hrefs:
                    new_seen.add(h)
                    new_unique.append(it)
            unique_items = unique_items  # Keep original list for counting

        except Exception as e:
            print(f"    [{idx+1}] Error: {e}")
            try:
                page.goto(url, timeout=15000)
                time.sleep(2)
            except:
                pass
            continue

    return companies


def create_excel(companies):
    """Create the Excel file with all company data"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import CellIsRule
    except ImportError:
        print("  [WARN] openpyxl not available, skipping Excel creation")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Google Maps Firmen"

    headers = [
        "#", "Firmenname", "Bewertung", "Anzahl Bewertungen", "Kategorie",
        "Adresse", "PLZ", "Ort", "Telefon", "Website",
        "Breitengrad", "Längengrad", "Google Maps Link",
        "Suchort", "Suchbegriff", "Quelle", "Scraped At"
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for idx, c in enumerate(companies, 1):
        row_data = [
            idx, c.get("name", ""), c.get("rating", ""), c.get("review_count", ""),
            c.get("category", ""), c.get("address", ""), c.get("plz", ""),
            c.get("city", ""), c.get("phone", ""), c.get("website", ""),
            c.get("lat", ""), c.get("lon", ""), c.get("gmaps_url", ""),
            c.get("search_town", ""), c.get("search_term", ""),
            c.get("source", ""), c.get("scraped_at", ""),
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=idx + 1, column=col, value=value)
            cell.border = thin_border

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(companies) + 1}"

    # Column widths
    widths = [5, 40, 10, 12, 30, 45, 8, 25, 22, 45, 12, 12, 55, 25, 20, 22, 22]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    # Conditional formatting
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ws.conditional_formatting.add(
        f"J2:J{len(companies)+1}",
        CellIsRule(operator="notEqual", formula=['""'], fill=green_fill)
    )

    # Stats sheet
    ws2 = wb.create_sheet("Statistiken")
    stats = [
        ("Gesamtzahl Firmen", len(companies)),
        ("Mit Website", sum(1 for c in companies if c.get("website"))),
        ("Ohne Website", sum(1 for c in companies if not c.get("website"))),
        ("Mit Telefon", sum(1 for c in companies if c.get("phone"))),
        ("Mit PLZ", sum(1 for c in companies if c.get("plz"))),
        ("Mit Adresse", sum(1 for c in companies if c.get("address"))),
        ("Mit Kategorie", sum(1 for c in companies if c.get("category"))),
        ("Detail-basiert", sum(1 for c in companies if c.get("source") == "google_maps_detail")),
    ]
    ws2.cell(row=1, column=1, value="Statistik").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="Wert").font = Font(bold=True)
    for row, (stat, value) in enumerate(stats, 2):
        ws2.cell(row=row, column=1, value=stat).font = Font(bold=True)
        ws2.cell(row=row, column=2, value=value)
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 15

    # Top Orte
    ws3 = wb.create_sheet("Top Orte")
    city_counts = {}
    for c in companies:
        city = c.get("city", c.get("search_town", "Unbekannt"))
        city_counts[city] = city_counts.get(city, 0) + 1
    ws3.cell(row=1, column=1, value="Ort").font = Font(bold=True)
    ws3.cell(row=1, column=2, value="Anzahl Firmen").font = Font(bold=True)
    for row, (city, count) in enumerate(sorted(city_counts.items(), key=lambda x: -x[1]), 2):
        ws3.cell(row=row, column=1, value=city)
        ws3.cell(row=row, column=2, value=count)
    ws3.column_dimensions['A'].width = 35
    ws3.column_dimensions['B'].width = 15

    # Top Kategorien
    ws4 = wb.create_sheet("Top Kategorien")
    cat_counts = {}
    for c in companies:
        cat = c.get("category", "Ohne Kategorie")
        if cat:
            cat_counts[cat] = cat_counts.get(cat, 0) + 1
    ws4.cell(row=1, column=1, value="Kategorie").font = Font(bold=True)
    ws4.cell(row=1, column=2, value="Anzahl").font = Font(bold=True)
    for row, (cat, count) in enumerate(sorted(cat_counts.items(), key=lambda x: -x[1]), 2):
        ws4.cell(row=row, column=1, value=cat)
        ws4.cell(row=row, column=2, value=count)
    ws4.column_dimensions['A'].width = 40
    ws4.column_dimensions['B'].width = 10

    wb.save(EXCEL_OUTPUT)
    print(f"  [EXCEL] Saved {len(companies)} companies to {EXCEL_OUTPUT}")


def main():
    # Load existing data
    existing = load_json(EXISTING_RESULTS, [])
    progress = load_json(PROGRESS_FILE, {
        "completed_searches": [],
        "enriched_companies": [],
        "phase": "search_new_towns",
        "last_town_idx": 0,
        "last_enrich_idx": 0,
    })
    towns = load_json(TOWNS_FILE, [])
    old_progress = load_json(OLD_PROGRESS, {"completed_searches": []})

    # Merge existing + travel results
    travel_results = load_json(RESULTS_FILE, [])
    all_companies = {}

    for c in existing:
        key = f"{c.get('name', '').lower().strip()}|{c.get('city', c.get('search_town', '')).lower().strip()}"
        if key not in all_companies:
            all_companies[key] = c
        else:
            for k, v in c.items():
                if v and not all_companies[key].get(k):
                    all_companies[key][k] = v

    for c in travel_results:
        key = f"{c.get('name', '').lower().strip()}|{c.get('city', c.get('search_town', '')).lower().strip()}"
        if key not in all_companies:
            all_companies[key] = c
        else:
            # Travel results are higher quality - prefer them for certain fields
            existing_c = all_companies[key]
            for k, v in c.items():
                if v and (not existing_c.get(k) or k in ("website", "category", "plz", "address", "source", "scraped_at")):
                    existing_c[k] = v

    print(f"Loaded {len(existing)} existing + {len(travel_results)} travel = {len(all_companies)} merged companies")

    # Find missing towns
    searched_towns = set()
    for s in old_progress.get("completed_searches", []):
        searched_towns.add(s.split("|")[0])
    for s in progress.get("completed_searches", []):
        searched_towns.add(s.split("|")[0])

    town_names = [t["name"] if isinstance(t, dict) else t for t in towns]
    missing_towns = [t for t in town_names if t not in searched_towns]
    print(f"Missing towns ({len(missing_towns)}): {missing_towns}")

    # Start Xvfb
    xvfb_proc = subprocess.Popen(
        ['Xvfb', ':99', '-screen', '0', '1920x1080x24', '-ac', '+extension', 'GLX', '+render', '-noreset'],
        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL
    )
    time.sleep(2)
    os.environ['DISPLAY'] = ':99'

    from playwright.sync_api import sync_playwright

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=False,
            executable_path=CHROME_PATH,
            args=[
                "--no-sandbox", "--disable-gpu", "--disable-dev-shm-usage",
                "--disable-blink-features=AutomationControlled", "--lang=de-DE",
            ]
        )
        context = browser.new_context(
            viewport={"width": 1920, "height": 1080},
            locale="de-DE",
            user_agent="Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
        )
        page = context.new_page()

        entries_since_push = 0

        # Phase 1: Search missing towns
        if progress.get("phase", "search_new_towns") == "search_new_towns":
            print(f"\n=== PHASE 1: Searching {len(missing_towns)} missing towns ===")

            for town_name in missing_towns:
                town_info = None
                for t in towns:
                    tname = t["name"] if isinstance(t, dict) else t
                    if tname == town_name:
                        town_info = t
                        break

                tier = town_info.get("population_tier", 4) if isinstance(town_info, dict) else 4
                search_terms = SEARCH_TERMS.get(tier, SEARCH_TERMS[4])

                print(f"\n--- {town_name} (tier {tier}, {len(search_terms)} terms) ---")

                for term in search_terms:
                    search_key = f"{town_name}|{term}"

                    if search_key in progress.get("completed_searches", []):
                        continue

                    companies = search_and_extract(page, town_name, term, max_scroll=10)

                    new_count = 0
                    for c in companies:
                        key = f"{c.get('name', '').lower().strip()}|{c.get('city', c.get('search_town', '')).lower().strip()}"
                        if key not in all_companies:
                            all_companies[key] = c
                            new_count += 1
                        else:
                            existing_c = all_companies[key]
                            for k, v in c.items():
                                if v and not existing_c.get(k):
                                    existing_c[k] = v

                    progress.setdefault("completed_searches", []).append(search_key)
                    entries_since_push += new_count

                    print(f"  New: {new_count} | Total: {len(all_companies)} | Since push: {entries_since_push}")

                    # Save progress
                    save_json(RESULTS_FILE, list(all_companies.values()))
                    save_json(PROGRESS_FILE, progress)

                    # Push every 10 entries
                    if entries_since_push >= 10:
                        create_excel(list(all_companies.values()))
                        git_push(f"GMaps Travel: {town_name} | +{entries_since_push} | Total: {len(all_companies)}")
                        entries_since_push = 0

                    human_delay(1, 2)

                print(f"  Town {town_name} done. Total: {len(all_companies)}")

            progress["phase"] = "enrich_existing"
            save_json(PROGRESS_FILE, progress)
            create_excel(list(all_companies.values()))
            git_push(f"GMaps Phase 1 COMPLETE | Total: {len(all_companies)}")
            print(f"\n=== PHASE 1 DONE. Total: {len(all_companies)} ===")

        # Phase 2: Enrich existing companies without website URLs
        if progress.get("phase") == "enrich_existing":
            print(f"\n=== PHASE 2: Enriching companies with missing website URLs ===")

            need_enrichment = []
            for key, c in all_companies.items():
                has_website_flag = c.get("has_website", False)
                has_actual_url = bool(c.get("website") and "google" not in c.get("website", "").lower())
                has_plz = bool(c.get("plz"))
                has_category = bool(c.get("category"))

                if has_website_flag and not has_actual_url:
                    need_enrichment.append((1, key, c))  # Priority 1
                elif not has_plz:
                    need_enrichment.append((2, key, c))
                elif not has_category and has_website_flag:
                    need_enrichment.append((3, key, c))

            need_enrichment.sort(key=lambda x: x[0])
            print(f"Need enrichment: {len(need_enrichment)}")
            print(f"  P1 (need website): {sum(1 for p,_,_ in need_enrichment if p==1)}")
            print(f"  P2 (need PLZ): {sum(1 for p,_,_ in need_enrichment if p==2)}")
            print(f"  P3 (need category): {sum(1 for p,_,_ in need_enrichment if p==3)}")

            enriched_set = set(progress.get("enriched_companies", []))
            start_idx = progress.get("last_enrich_idx", 0)
            enrich_count = 0

            for i in range(start_idx, len(need_enrichment)):
                priority, key, company = need_enrichment[i]
                name = company.get("name", "")
                town = company.get("city", company.get("search_town", ""))

                if key in enriched_set:
                    continue

                # Search for this specific company on Google Maps
                query = f"{name} {town}"
                url = f"https://www.google.com/maps/search/{query.replace(' ', '+')}"

                try:
                    page.goto(url, timeout=15000, wait_until="domcontentloaded")
                    time.sleep(random.uniform(3, 5))

                    # Handle cookies
                    try:
                        accept = page.query_selector('button[aria-label*="Alle ablehnen"], button[aria-label*="Reject"]')
                        if accept:
                            accept.click()
                            time.sleep(0.5)
                    except:
                        pass

                    # Click first result
                    first = page.query_selector('a[href*="/maps/place/"]')
                    if first:
                        first.click()
                        time.sleep(random.uniform(2, 3))
                        detail = extract_detail(page, town, "enrichment")

                        if detail:
                            updated = False
                            if detail.get("website") and "google" not in detail["website"].lower() and not company.get("website"):
                                company["website"] = detail["website"]
                                updated = True
                            if detail.get("phone") and not company.get("phone"):
                                company["phone"] = detail["phone"]
                                updated = True
                            if detail.get("address") and not company.get("address"):
                                company["address"] = detail["address"]
                                updated = True
                            if detail.get("plz") and not company.get("plz"):
                                company["plz"] = detail["plz"]
                                updated = True
                            if detail.get("category") and not company.get("category"):
                                company["category"] = detail["category"]
                                updated = True
                            if detail.get("city") and not company.get("city"):
                                company["city"] = detail["city"]
                                updated = True
                            if detail.get("lat"):
                                company["lat"] = detail["lat"]
                                updated = True
                            if detail.get("lon"):
                                company["lon"] = detail["lon"]
                                updated = True
                            if updated:
                                company["source"] = "google_maps_detail"
                                company["enriched_at"] = datetime.now().isoformat()

                            if company.get("website"):
                                print(f"  [ENRICH] {name}: website={company['website']}")

                except Exception as e:
                    print(f"  [ENRICH-ERR] {name}: {e}")

                enriched_set.add(key)
                enrich_count += 1
                entries_since_push += 1

                # Save progress every 10
                if enrich_count % 10 == 0:
                    save_json(RESULTS_FILE, list(all_companies.values()))
                    progress["enriched_companies"] = list(enriched_set)[-500:]  # Keep last 500
                    progress["last_enrich_idx"] = i + 1
                    save_json(PROGRESS_FILE, progress)
                    create_excel(list(all_companies.values()))
                    git_push(f"GMaps Enrich: {enrich_count} done | Total: {len(all_companies)}")
                    entries_since_push = 0

                human_delay(0.5, 1.5)

            # Final save
            save_json(RESULTS_FILE, list(all_companies.values()))
            progress["enriched_companies"] = list(enriched_set)[-500:]
            progress["last_enrich_idx"] = len(need_enrichment)
            save_json(PROGRESS_FILE, progress)
            create_excel(list(all_companies.values()))
            git_push(f"GMaps COMPLETE: {len(all_companies)} companies enriched")

        browser.close()

    # Cleanup Xvfb
    try:
        os.kill(xvfb_proc.pid, signal.SIGTERM)
    except:
        pass

    print(f"\n=== ALL DONE === Total companies: {len(all_companies)}")


if __name__ == "__main__":
    main()
