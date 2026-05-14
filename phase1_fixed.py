#!/usr/bin/env python3
"""Phase 1: Fast Arbeitsagentur API search for junior IT jobs - FIXED"""
import asyncio, httpx, json, re, os, subprocess
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

AA_API_URL = "https://rest.arbeitsagentur.de/jobboerse/jobsuche-service/pc/v6/jobs"
AA_API_KEY = "jobboerse-jobsuche"
EXCEL_OUT = "/home/z/my-project/download/Junior_IT_Jobs_Bamberg_Erlangen_Nuernberg.xlsx"
GIT_REPO = "/home/z/my-project"
GITHUB_TOKEN = "GIT_TOKEN_REDACTED"

IT_KEYWORDS = ['software','entwickler','developer','it-','informatik','data','devops','cloud','cyber','security','sicherheit','system','admin','administrator','netzwerk','network','frontend','backend','fullstack','full-stack','python','java','javascript','sap','linux','sql','database','datenbank','api','machine learning','scrum','agile','testing','qa','consultant','analyst','engineer','architekt','programmierung','web','mobile','app','infrastructure','support','helpdesk','erp','crm','digitalisierung','technologie','digital']

TECH_LIST = ['python','java','javascript','typescript','react','angular','vue','node.js','django','flask','spring','.net','docker','kubernetes','aws','azure','gcp','terraform','ansible','jenkins','sql','mysql','postgresql','mongodb','redis','kafka','graphql','sap','linux','html','css','scrum','devops','ci/cd','machine learning','deep learning','tensorflow','pytorch','data science','big data','cybersecurity','powershell','bash','git']

def is_it_job(t):
    tl = t.lower()
    return any(kw in tl for kw in IT_KEYWORDS)

def is_senior(t):
    tl = t.lower()
    return any(kw in tl for kw in ['senior','lead','principal','head of','director','leitung','chef','manager'])

def is_junior(t):
    tl = t.lower()
    return any(kw in tl for kw in ['junior','einsteiger','trainee','praktikant','intern','werkstudent','working student','absolvent','graduate','berufseinsteiger','associate','duales studium','ausbildung','apprentice','nachwuchs','perspektive'])

def classify_cat(t):
    tl = t.lower()
    if any(k in tl for k in ['softwareentwickler','software developer','entwickler','developer','programmierer','frontend','backend','fullstack','full-stack']): return 'Softwareentwicklung'
    if any(k in tl for k in ['data scientist','data analyst','data engineer','business intelligence','machine learning']): return 'Data Science & Analytics'
    if any(k in tl for k in ['it consultant','it-consultant','it berater','sap consultant','technischer berater']): return 'IT-Consulting'
    if any(k in tl for k in ['system administrator','sysadmin','admin','netzwerkadministrator']): return 'Systemadministration'
    if any(k in tl for k in ['devops','cloud engineer','platform engineer','sre']): return 'DevOps & Cloud'
    if any(k in tl for k in ['security','cybersecurity','it-sicherheit','pentest']): return 'Cyber Security'
    if any(k in tl for k in ['it support','helpdesk','service desk','first level','second level']): return 'IT-Support'
    if any(k in tl for k in ['project manager','projektmanager','scrum master','product owner']): return 'IT-Projektmanagement'
    if any(k in tl for k in ['qa','tester','test engineer','testautomatisierung']): return 'QA & Testing'
    if any(k in tl for k in ['ux','ui','user experience','user interface']): return 'UX/UI Design'
    if any(k in tl for k in ['ausbildung','duales studium','trainee','praktikum','werkstudent']): return 'IT-Ausbildung & Studium'
    return 'IT (Sonstige)'

def extract_techs(text):
    found = []
    tl = text.lower()
    for t in TECH_LIST:
        if re.search(r'\b'+re.escape(t)+r'\b', tl):
            found.append(t)
    return list(set(found))

def get_emp_type(t):
    tl = t.lower()
    if any(k in tl for k in ['teilzeit','part-time']): return 'Teilzeit'
    if any(k in tl for k in ['werkstudent','working student']): return 'Werkstudent'
    if any(k in tl for k in ['praktikant','praktikum','intern']): return 'Praktikum'
    if any(k in tl for k in ['ausbildung','apprentice','azubi']): return 'Ausbildung'
    if any(k in tl for k in ['duales studium']): return 'Duales Studium'
    return 'Vollzeit'

def get_remote(t, homeoffice=False, homeofficetyp=''):
    tl = t.lower()
    if 'hybrid' in tl: return 'Hybrid'
    if any(k in tl for k in ['remote','homeoffice','home-office']): return 'Remote'
    if homeoffice:
        if homeofficetyp and 'VEREINBARUNG' in (homeofficetyp or ''):
            return 'Hybrid'
        return 'Remote'
    return 'Nicht angegeben'

def parse_job(item):
    """Parse a single job from the API response."""
    title = item.get('stellenangebotsTitel', '')
    refnr = item.get('chiffrenummer', '') or item.get('refnr', '')
    
    # Location
    lokationen = item.get('stellenlokationen', [])
    city = ''
    plz = ''
    if lokationen:
        addr = lokationen[0].get('adresse', {})
        city = addr.get('ort', '')
        plz = addr.get('plz', '')
    
    # Employer - check different field names
    employer = item.get('arbeitgeber', '') or item.get('arbeitgeberName', '')
    
    # Home office
    homeoffice = item.get('homeofficemoeglich', False)
    homeofficetyp = item.get('homeofficetyp', '')
    
    # Contract details
    vollzeit = item.get('arbeitszeitVollzeit', False)
    befristet = item.get('vertragsdauer', '') == 'BEFRISTET'
    
    # Entry date
    eintritt = item.get('eintrittszeitraum', {})
    eintrittsdatum = eintritt.get('von', '') if isinstance(eintritt, dict) else ''
    
    # Publication date
    veroeff = item.get('veroeffentlichungszeitraum', {})
    veroeffdatum = veroeff.get('von', '') if isinstance(veroeff, dict) else ''
    
    # External URL
    url = item.get('externeURL', '') or f'https://www.arbeitsagentur.de/jobsuche/jobdetail/{refnr}'
    
    return {
        'title': title,
        'refnr': refnr,
        'employer_name': employer,
        'city': city,
        'plz': plz,
        'vollzeit': vollzeit,
        'befristet': befristet,
        'eintrittsdatum': eintrittsdatum,
        'veroeffdatum': veroeffdatum,
        'homeoffice': homeoffice,
        'homeofficetyp': homeofficetyp,
        'url': url,
        'quereinstieg': item.get('quereinstiegGeeignet', False),
    }


async def search_jobs(client, search_term, city, umkreis=50, page=1):
    """Single API search"""
    params = {'was': search_term, 'wo': city, 'umkreis': umkreis, 'page': page, 'size': 100}
    headers = {'X-API-Key': AA_API_KEY, 'Accept': 'application/json'}
    try:
        r = await client.get(AA_API_URL, params=params, headers=headers, timeout=15)
        if r.status_code == 200:
            data = r.json()
            return data.get('ergebnisliste', []), data.get('maxErgebnisse', 0)
    except Exception as e:
        pass
    return [], 0


async def main():
    print("="*60)
    print("Phase 1: Arbeitsagentur API - Junior IT Jobs B/ER/N")
    print("="*60)
    
    all_jobs = {}  # refnr -> job dict
    
    cities = ['Erlangen', 'Nürnberg', 'Bamberg', 'Fürth', 'Forchheim', 'Herzogenaurach']
    search_terms = [
        'Junior IT', 'Junior Software', 'Junior Entwickler',
        'Werkstudent IT', 'Werkstudent Software',
        'Praktikum IT', 'Praktikum Software',
        'Trainee IT', 'Ausbildung Informatik', 'Duales Studium IT',
        'Junior Developer', 'Junior Data', 'Junior DevOps',
        'Einsteiger IT', 'Berufseinsteiger Informatik',
        'Junior Frontend', 'Junior Backend', 'Junior Fullstack',
    ]
    
    combos = [(term, city) for term in search_terms for city in cities]
    print(f"  {len(combos)} search combinations")
    
    async with httpx.AsyncClient() as client:
        sem = asyncio.Semaphore(6)
        
        async def limited_search(term, city):
            async with sem:
                return await search_jobs(client, term, city)
        
        # Process in chunks
        chunk_size = 15
        total_results = 0
        for i in range(0, len(combos), chunk_size):
            chunk = combos[i:i+chunk_size]
            tasks = [limited_search(term, city) for term, city in chunk]
            results = await asyncio.gather(*tasks, return_exceptions=True)
            
            for result in results:
                if isinstance(result, tuple):
                    items, max_results = result
                    total_results += max_results
                    for item in items:
                        job = parse_job(item)
                        title = job['title']
                        refnr = job['refnr']
                        
                        if not title or not is_it_job(title):
                            continue
                        if is_senior(title):
                            continue
                        
                        jk = None
                        if is_junior(title):
                            jk = next((kw for kw in ['junior','einsteiger','trainee','werkstudent','praktikant','absolvent','duales studium','ausbildung','berufseinsteiger','associate','nachwuchs','perspektive'] if kw in title.lower()), 'junior_keyword')
                        else:
                            jk = 'no_senior_keyword'
                        
                        job['is_junior'] = is_junior(title)
                        job['junior_kw'] = jk
                        job['it_kw'] = next((kw for kw in IT_KEYWORDS if kw in title.lower()), 'it')
                        job['remote_option'] = get_remote(title, job.get('homeoffice', False), job.get('homeofficetyp', ''))
                        
                        if refnr and refnr not in all_jobs:
                            all_jobs[refnr] = job
            
            done = min(i+chunk_size, len(combos))
            print(f"  {done}/{len(combos)} searches done | {len(all_jobs)} unique junior IT jobs | API total: {total_results}")
        
        # Now do additional broader searches with more pages
        print(f"\n  Fetching more pages for high-yield searches...")
        high_yield = [
            ('Junior Software', 'Erlangen'), ('Junior IT', 'Nürnberg'),
            ('Werkstudent IT', 'Erlangen'), ('Junior Developer', 'Nürnberg'),
            ('Junior Entwickler', 'Erlangen'), ('Praktikum IT', 'Erlangen'),
            ('Ausbildung Informatik', 'Erlangen'), ('Junior Software', 'Nürnberg'),
            ('Werkstudent Software', 'Erlangen'), ('Junior Frontend', 'Erlangen'),
        ]
        
        for term, city in high_yield:
            for page in range(2, 5):  # Pages 2-4
                items, _ = await search_jobs(client, term, city, page=page)
                for item in items:
                    job = parse_job(item)
                    title = job['title']
                    refnr = job['refnr']
                    
                    if not title or not is_it_job(title) or is_senior(title):
                        continue
                    
                    jk = next((kw for kw in ['junior','einsteiger','trainee','werkstudent','praktikant','absolvent','duales studium','ausbildung','berufseinsteiger'] if kw in title.lower()), 'no_senior_keyword')
                    
                    job['is_junior'] = is_junior(title)
                    job['junior_kw'] = jk
                    job['it_kw'] = next((kw for kw in IT_KEYWORDS if kw in title.lower()), 'it')
                    job['remote_option'] = get_remote(title, job.get('homeoffice', False), job.get('homeofficetyp', ''))
                    
                    if refnr and refnr not in all_jobs:
                        all_jobs[refnr] = job
            
            print(f"  Extra pages for '{term}' in {city}: {len(all_jobs)} total")
    
    print(f"\n  Total unique junior IT jobs from API: {len(all_jobs)}")
    
    # Create Excel
    print("\n  Creating Excel...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Junior IT-Jobs"
    
    headers = [
        'Job-ID','Job-Titel','Job-Titel (Original)','Job-Kategorie','Erfahrungslevel',
        'Beschäftigungsart','Remote-Option','Technologien',
        'Ausschreibungstext (Auszug)','Job-URL',
        'Arbeitgeber','Arbeitgeber (norm.)','Ort','PLZ','Straße',
        'Telefon','E-Mail','Branche',
        'Quelle','Karriere-URL','Website-URL',
        'Karriere-HTML Pfad','Job-HTML Pfad',
        'Junior-Keyword Match','IT-Keyword Match','Ist Junior-Job','Ist IT-Job',
        'Vollzeit/Teilzeit Detail','Befristet','Gehaltsspanne','Arbeitsagentur RefNr',
        'Eintrittsdatum','Veroeffentlichungsdatum',
        'Homeoffice Möglich','Homeoffice Typ','Quereinstieg',
        'Scrape-Datum','Scrape-Status','Arbeitgeber-ID',
        'Letzte Prüfung','Erstellt','Aktualisiert',
    ]
    
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.font = hf
        c.fill = hfill
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    
    for idx, (refnr, job) in enumerate(all_jobs.items()):
        row = idx + 2
        title = job['title']
        emp_name = job.get('employer_name', '')
        emp_norm = re.sub(r'\s+', '', re.sub(r'(GmbH|AG|e\.K\.|UG|KG|mbH).*$', '', emp_name)).lower()
        
        values = [
            f'J-{idx+1:05d}', title, title, classify_cat(title),
            'Junior' if job.get('is_junior') else 'Mid-Level',
            get_emp_type(title), job.get('remote_option', get_remote(title)),
            ', '.join(extract_techs(title)),
            '', job.get('url', ''),
            emp_name, emp_norm, job.get('city',''), job.get('plz',''), '',
            '', '', '',
            'Arbeitsagentur', '', '', '', '',
            job.get('junior_kw',''), job.get('it_kw',''),
            'Ja' if job.get('is_junior') else 'Nein', 'Ja',
            get_emp_type(title),
            'Ja' if job.get('befristet') else 'Nein', '', refnr,
            job.get('eintrittsdatum',''), job.get('veroeffdatum',''),
            'Ja' if job.get('homeoffice') else 'Nein',
            job.get('homeofficetyp',''),
            'Ja' if job.get('quereinstieg') else 'Nein',
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'completed', '',
            datetime.now().strftime('%Y-%m-%d'), datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        ]
        
        for col, v in enumerate(values, 1):
            ws.cell(row, col, v).alignment = Alignment(wrap_text=True, vertical='top')
    
    # Stats sheet
    ws2 = wb.create_sheet('Statistiken')
    stats_data = [
        ('Gesamt Junior IT-Jobs', len(all_jobs)),
        ('Quelle', 'Arbeitsagentur API'),
        ('Region', 'Bamberg / Erlangen / Nürnberg (50km)'),
        ('Scrape-Datum', datetime.now().strftime('%Y-%m-%d %H:%M')),
        ('Davon explizit Junior', len([j for j in all_jobs.values() if j.get('is_junior')])),
        ('Davon Homeoffice', len([j for j in all_jobs.values() if j.get('homeoffice')])),
    ]
    for i, (k, v) in enumerate(stats_data, 1):
        ws2.cell(i, 1, k)
        ws2.cell(i, 2, v)
    
    # Category breakdown
    cats = {}
    for job in all_jobs.values():
        cat = classify_cat(job['title'])
        cats[cat] = cats.get(cat, 0) + 1
    ws2.cell(8, 1, 'Nach Kategorie:')
    for i, (cat, count) in enumerate(sorted(cats.items(), key=lambda x: -x[1])):
        ws2.cell(9+i, 1, cat)
        ws2.cell(9+i, 2, count)
    
    # City breakdown
    cities_count = {}
    for job in all_jobs.values():
        c = job.get('city', 'Unbekannt')
        cities_count[c] = cities_count.get(c, 0) + 1
    row_start = 9 + len(cats) + 2
    ws2.cell(row_start, 1, 'Nach Ort:')
    for i, (c, count) in enumerate(sorted(cities_count.items(), key=lambda x: -x[1])):
        ws2.cell(row_start+1+i, 1, c)
        ws2.cell(row_start+1+i, 2, count)
    
    wb.save(EXCEL_OUT)
    print(f"  Excel saved: {EXCEL_OUT}")
    
    # Save raw JSON
    json_out = EXCEL_OUT.replace('.xlsx', '_raw.json')
    with open(json_out, 'w', encoding='utf-8') as f:
        json.dump(list(all_jobs.values()), f, ensure_ascii=False, indent=2, default=str)
    print(f"  JSON saved: {json_out}")
    
    # Git push
    os.chdir(GIT_REPO)
    subprocess.run(['git', 'add', '-A'], capture_output=True, timeout=30)
    r = subprocess.run(['git', 'commit', '-m', f'Phase 1: Junior IT Jobs API ({len(all_jobs)} jobs)'], capture_output=True, timeout=30)
    if r.returncode == 0:
        push_r = subprocess.run(['git', 'push', f'https://{GITHUB_TOKEN}@github.com/BirdNest055/Heyjobs.git', 'main', '--force-with-lease'], capture_output=True, timeout=60)
        print(f"  Git: Pushed! (returncode={push_r.returncode})")
    else:
        print(f"  Git: No changes to commit")
    
    print(f"\n{'='*60}")
    print(f"Phase 1 DONE: {len(all_jobs)} junior IT jobs from API")
    print(f"{'='*60}")

if __name__ == '__main__':
    asyncio.run(main())
