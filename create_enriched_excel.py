#!/usr/bin/env python3
"""Create Excel with Google Maps data + website/email enrichment for Bamberg/Erlangen/Nürnberg."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter

with open('/home/z/my-project/gmaps_erlangen_results.json') as f:
    results = json.load(f)
with open('/home/z/my-project/website_enrichment.json') as f:
    enrichment = json.load(f)

# Filter for target cities
TARGET = ['Bamberg', 'Erlangen', 'Nürnberg']
filtered = [r for r in results if r.get('city') in TARGET]
filtered.sort(key=lambda x: (x.get('city',''), x.get('name','')))

print(f"Total firms in Bamberg/Erlangen/Nürnberg: {len(filtered)}")
print(f"Enrichment entries: {len(enrichment)}")

wb = Workbook()
ws = wb.active
ws.title = "Firmen mit Website & HTML"

headers = ['#', 'Firmenname', 'Bewertung', 'Kategorie', 'Adresse', 'PLZ', 'Ort', 
           'Telefon', 'Website (Google Maps)', 'Website (gefunden)', 
           'E-Mail', 'HTML Datei', 'Impressum URL', 'Google Maps Link']

hfont = Font(name='Arial', bold=True, size=11, color='FFFFFF')
hfill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
halign = Alignment(horizontal='center', vertical='center', wrap_text=True)
bdr = Border(left=Side(style='thin', color='D9E2F3'), right=Side(style='thin', color='D9E2F3'),
             top=Side(style='thin', color='D9E2F3'), bottom=Side(style='thin', color='D9E2F3'))
efill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
gfill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = hfont; c.fill = hfill; c.alignment = halign; c.border = bdr

for i, r in enumerate(filtered, 1):
    name = r.get('name','').strip()
    enrich = enrichment.get(name, {})
    
    row = i + 1
    vals = [
        i, name, r.get('rating',''), r.get('category',''), r.get('address',''),
        r.get('plz',''), r.get('city',''), r.get('phone',''),
        'Ja' if r.get('has_website') else 'Nein',
        enrich.get('website', ''),
        enrich.get('email', ''),
        enrich.get('html_file', ''),
        enrich.get('impressum_url', ''),
        r.get('href', '')
    ]
    
    for col, val in enumerate(vals, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name='Arial', size=10)
        c.alignment = Alignment(vertical='center', wrap_text=(col in [2,5,10,12,13,14]))
        c.border = bdr
        if i % 2 == 0: c.fill = efill
        # Highlight rows with website found
        if enrich.get('website'):
            if col == 10: c.fill = gfill

widths = [5, 35, 8, 22, 35, 8, 15, 18, 12, 40, 30, 35, 40, 45]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(filtered)+1}"
ws.freeze_panes = 'A2'

# Stats sheet
ws2 = wb.create_sheet("Statistiken")
with_web = sum(1 for r in filtered if enrichment.get(r.get('name','').strip(),{}).get('website'))
with_email = sum(1 for r in filtered if enrichment.get(r.get('name','').strip(),{}).get('email'))
with_html = sum(1 for r in filtered if enrichment.get(r.get('name','').strip(),{}).get('html_file'))
stats = [
    ['Statistik', 'Wert'],
    ['Gesamt Firmen', len(filtered)],
    ['Davon Bamberg', sum(1 for r in filtered if r.get('city')=='Bamberg')],
    ['Davon Erlangen', sum(1 for r in filtered if r.get('city')=='Erlangen')],
    ['Davon Nürnberg', sum(1 for r in filtered if r.get('city')=='Nürnberg')],
    ['Website gefunden', with_web],
    ['Website %', f"{with_web/len(filtered)*100:.1f}%"],
    ['E-Mail gefunden', with_email],
    ['E-Mail %', f"{with_email/len(filtered)*100:.1f}%"],
    ['HTML gespeichert', with_html],
    ['HTML %', f"{with_html/len(filtered)*100:.1f}%"],
    ['Bereits angereichert', len(enrichment)],
    ['Noch offen', len(filtered) - len(enrichment)],
]
for ri, rd in enumerate(stats, 1):
    for ci, val in enumerate(rd, 1):
        c = ws2.cell(row=ri, column=ci, value=val)
        if ri == 1: c.font = hfont; c.fill = hfill
        else: c.font = Font(name='Arial', size=10)
ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 15

output = '/home/z/my-project/download/Firmen_Bamberg_Erlangen_Nuernberg_mit_Website_HTML.xlsx'
wb.save(output)
print(f"\nSaved: {output}")
print(f"Firms: {len(filtered)} | Websites: {with_web} | Emails: {with_email} | HTML: {with_html}")
