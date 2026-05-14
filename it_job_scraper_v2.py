#!/usr/bin/env python3
"""
IT-Job Scraper v2 - Bamberg, Erlangen, Nürnberg
- Uses synchronous requests with ThreadPoolExecutor for parallelism
- Better city matching (exact only)
- Arbeitsagentur API + web search for IT jobs
- Downloads HTML, extracts contact info
- Commits every 10 entries
"""

import json, os, re, sys, time, hashlib, sqlite3, subprocess
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
import requests
from bs4 import BeautifulSoup
import openpyxl

# Force unbuffered output
sys.stdout.reconfigure(line_buffering=True)

# ============================================================
# CONFIG
# ============================================================
EXCEL_IN = "/home/z/my-project/download/Google_Maps_Firmen_Erlangen_50km.xlsx"
EXCEL_OUT = "/home/z/my-project/download/IT_Arbeitgeber_Bamberg_Erlangen_Nuernberg.xlsx"
HTML_DIR = "/home/z/my-project/download/it_employer_html"
DB_PATH = "/home/z/my-project/download/it_employers.db"
PROGRESS_FILE = "/home/z/my-project/download/it_scraper_progress.json"
GIT_DIR = "/home/z/my-project"

TARGET_CITIES = ["Bamberg", "Erlangen", "Nürnberg"]  # EXACT matches only
MAX_WORKERS = 12
COMMIT_EVERY = 10

# Arbeitsagentur API
AA_API_URL = "https://rest.arbeitsagentur.de/jobboerse/jobsuche-service/pc/v6/jobs"
AA_API_KEY = "jobboerse-jobsuche"

IT_KEYWORDS = [
    "software", "entwickler", "developer", "it-", "informatik", "data",
    "systemadministrator", "devops", "frontend", "backend", "fullstack",
    "sre", "cloud", "cyber", "security", "netzwerk", "network",
    "administrator", "programmierer", "analyst", "sap", "linux",
    "python", "java ", "javascript", "typescript", "machine learning",
    "ai ", "künstliche intelligenz", "datenbank", "database", "sql",
    "scrum", "agile", "tester", "test engineer", "helpdesk",
    "it consultant", "it berater", "projektmanager it", "product owner",
    "architect", "architekt", "erp", "crm", "bi ", "business intelligence",
    "data engineer", "data scientist", "platform", "infrastruktur",
    "automation", "automatisierung", "iot", "embedded", "microservice",
    "api ", "container", "kubernetes", "docker", "ci/cd", "terraform",
    "ansible", "azure", "aws", "gcp",
]

HTTP_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "de-DE,de;q=0.9,en;q=0.5",
}

# Global session
SESSION = requests.Session()
SESSION.headers.update(HTTP_HEADERS)
SESSION.verify = False

# Suppress SSL warnings
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ============================================================
# UTILITIES
# ============================================================

def sanitize_filename(name, max_len=80):
    name = re.sub(r'[äöüß]', lambda m: {'ä':'ae','ö':'oe','ü':'ue','ß':'ss'}[m.group()], name)
    name = re.sub(r'[^a-zA-Z0-9._-]', '_', name)
    name = re.sub(r'_+', '_', name).strip('_')
    return name[:max_len]

def normalize_name(name):
    name = name.lower().strip()
    name = re.sub(r'\s+', ' ', name)
    for suffix in ['gmbh & co. kg', 'gmbh & co kg', 'gmbh co. kg', 'gmbh co kg',
                   'gmbh & co.', 'ag & co. kg', 'ag & co kg',
                   'gmbh', 'ag', 'kg', 'ohg', 'eg', 'se', 'e.k.', 'ek',
                   'e.v.', 'ev', 'mbh', 'ug', 'gbr']:
        name = re.sub(re.escape(suffix), '', name, flags=re.IGNORECASE)
    name = re.sub(r'[^a-z0-9]', '', name)
    return name

def load_progress():
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE) as f:
            return json.load(f)
    return {"completed": [], "last_index": 0}

def save_progress(progress):
    with open(PROGRESS_FILE, 'w') as f:
        json.dump(progress, f, ensure_ascii=False)

# ============================================================
# ARBEITSAGENTUR API
# ============================================================

def check_aa_it_jobs(company_name, city):
    """Check Arbeitsagentur for IT jobs at this company."""
    results = []
    try:
        params = {
            "was": company_name,
            "wo": city,
            "umkreis": "25",
            "page": "1",
            "size": "25",
        }
        headers = {"X-API-Key": AA_API_KEY, "Accept": "application/json"}
        resp = SESSION.get(AA_API_URL, params=params, headers=headers, timeout=12)
        if resp.status_code == 200:
            data = resp.json()
            jobs = data.get("stellenangebote", [])
            norm_company = normalize_name(company_name)
            for job in jobs:
                title = job.get("beruf", "")
                employer = job.get("arbeitgeber", "")
                norm_employer = normalize_name(employer)
                # Match: company name in employer or vice versa
                if len(norm_company) > 3 and (norm_company[:5] in norm_employer or norm_employer[:5] in norm_company):
                    title_lower = title.lower()
                    is_it = any(kw in title_lower for kw in IT_KEYWORDS)
                    if is_it:
                        results.append({
                            "titel": title,
                            "refnr": job.get("refnr", ""),
                            "ort": job.get("arbeitsort", {}).get("ort", ""),
                            "typ": job.get("arbeitszeit", ""),
                        })
        time.sleep(0.2)
    except Exception as e:
        pass
    return results

# ============================================================
# WEB SEARCH (DuckDuckGo)
# ============================================================

def search_website(company_name, city):
    """Search DuckDuckGo for company website."""
    try:
        query = f'"{company_name}" {city}'
        url = f"https://html.duckduckgo.com/html/?q={query.replace(' ', '+')}"
        resp = SESSION.get(url, timeout=10)
        if resp.status_code == 200:
            soup = BeautifulSoup(resp.text, 'html.parser')
            results = soup.find_all('a', class_='result__a')
            skip = ['facebook.com', 'linkedin.com', 'instagram.com', 'twitter.com',
                    'youtube.com', 'wikipedia.org', 'gelbeseiten.de', 'dasoertliche.de',
                    'meinestadt.de', 'kununu.com', 'yelp.de', 'google.com', 'google.de',
                    'xing.com', 'tiktok.com', 'pinterest.com']
            for r in results[:8]:
                href = r.get('href', '')
                match = re.search(r'uddg=([^&]+)', href)
                if match:
                    actual_url = requests.utils.unquote(match.group(1))
                    if not any(d in actual_url.lower() for d in skip):
                        return actual_url
            # Fallback
            for r in results[:3]:
                href = r.get('href', '')
                match = re.search(r'uddg=([^&]+)', href)
                if match:
                    return requests.utils.unquote(match.group(1))
    except:
        pass
    return None

# ============================================================
# HTML DOWNLOAD & ANALYSIS
# ============================================================

def download_html(url, label=""):
    """Download HTML content."""
    try:
        resp = SESSION.get(url, timeout=12, allow_redirects=True)
        if resp.status_code == 200 and 'text/html' in resp.headers.get('content-type', ''):
            return resp.text
    except:
        pass
    return None

def save_html_file(content, company_name, suffix=""):
    """Save HTML to file."""
    if not content:
        return None, None
    safe_name = sanitize_filename(company_name)
    filename = f"{safe_name}_{suffix}.html" if suffix else f"{safe_name}.html"
    filepath = os.path.join(HTML_DIR, filename)
    os.makedirs(HTML_DIR, exist_ok=True)
    with open(filepath, 'w', encoding='utf-8', errors='replace') as f:
        f.write(content)
    html_hash = hashlib.sha256(content.encode()).hexdigest()[:16]
    return filepath, html_hash

def check_website_for_it_jobs(html_content):
    """Check if website mentions IT jobs."""
    if not html_content:
        return False, [], ""
    
    text = html_content.lower()
    
    career_words = ['karriere', 'career', 'jobs', 'stellenangebote', 'stellenausschreibung',
                    'bewerbung', 'vacancies', 'offene stellen', 'wir suchen']
    career_found = [w for w in career_words if w in text]
    
    it_found = [kw for kw in IT_KEYWORDS if kw in text]
    
    it_job_indicators = [
        'it-stellen', 'it jobs', 'it-jobs', 'softwareentwickler gesucht',
        'entwickler m/w/d', 'developer m/w/d', 'it specialist',
        'informatiker gesucht', 'wir suchen entwickler',
        'it-mitarbeiter', 'systemadministrator gesucht', 'devops engineer',
        'stellenangebot it', 'stellenangebot software',
    ]
    
    specific_matches = [ind for ind in it_job_indicators if ind in text]
    has_it = bool(specific_matches) or (bool(it_found) and bool(career_found))
    
    return has_it, it_found[:10], ", ".join(specific_matches[:3] if specific_matches else career_found[:3])

def extract_contact_info(html_content):
    """Extract email, Geschäftsführer, Handelsregister from HTML."""
    info = {}
    impressum_url = None
    karriere_url = None
    
    if not html_content:
        return info, None, None
    
    # Emails
    emails = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', html_content)
    emails = [e for e in emails if not any(x in e.lower() for x in ['example.com', 'sentry', 'webpack', 'noreply', 'wixpress', 'github.com'])]
    mailtos = re.findall(r'mailto:([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})', html_content)
    emails = list(dict.fromkeys(emails + mailtos))[:5]
    if emails:
        info['email'] = emails[0]
        info['alle_emails'] = "; ".join(emails)
    
    # Geschäftsführer
    gf_match = re.search(r'(?:Geschäftsführer|Geschaeftsfuehrer)[\s:]*(.+?)(?:<br|</|\.|,|$)', html_content, re.IGNORECASE)
    if gf_match:
        gf = BeautifulSoup(gf_match.group(1), 'html.parser').get_text(strip=True)
        if len(gf) > 2:
            info['geschaeftsfuehrer'] = gf[:200]
    
    # Handelsregister
    hr_match = re.search(r'(?:Handelsregister|HRB|HR-Aktenzeichen)[\s:]*(.+?)(?:<br|</|\.|,|$)', html_content, re.IGNORECASE)
    if hr_match:
        hr = BeautifulSoup(hr_match.group(1), 'html.parser').get_text(strip=True)
        info['handelsregister'] = hr[:200]
    
    # Find links
    soup = BeautifulSoup(html_content, 'html.parser')
    
    for a in soup.find_all('a', href=True):
        href = a['href']
        href_lower = href.lower()
        text_lower = a.get_text(strip=True).lower()
        
        if not impressum_url and ('impressum' in href_lower or 'impressum' in text_lower or 'imprint' in href_lower):
            impressum_url = href
        
        if not karriere_url:
            career_words = ['karriere', 'career', 'jobs', 'stellenangebote', 'bewerbung']
            if any(w in href_lower for w in career_words) or any(w in text_lower for w in career_words):
                karriere_url = href
    
    return info, impressum_url, karriere_url

def make_absolute_url(base_url, relative_url):
    """Convert relative URL to absolute."""
    if not relative_url:
        return None
    if relative_url.startswith('http'):
        return relative_url
    if relative_url.startswith('//'):
        return 'https:' + relative_url
    if relative_url.startswith('/'):
        from urllib.parse import urlparse
        parsed = urlparse(base_url)
        return f"{parsed.scheme}://{parsed.netloc}{relative_url}"
    return f"{base_url.rstrip('/')}/{relative_url.lstrip('/')}"

# ============================================================
# IT RELEVANCE SCORE
# ============================================================

def calculate_it_relevance(data):
    score = 0.0
    if data.get('hat_it_jobs'):
        score += 40
        score += min(data.get('it_job_anzahl', 0) * 5, 15)
    if data.get('it_keywords_gefunden'):
        kw_count = len(data['it_keywords_gefunden'].split(', '))
        score += min(kw_count * 3, 20)
    if data.get('karriere_url'):
        score += 10
    kat = (data.get('kategorie_google') or '').lower()
    if any(c in kat for c in ['software', 'it', 'informatik', 'telekommunikation', 'computer', 'technologie', 'digital']):
        score += 15
    name = (data.get('firmenname') or '').lower()
    if any(p in name for p in ['it ', 'software', 'tech', 'digital', 'data', 'cloud', 'cyber', 'computer', 'informatik', 'system']):
        score += 10
    if data.get('it_job_details'):
        score += 10
    return min(score, 100.0)

# ============================================================
# PROCESS SINGLE EMPLOYER
# ============================================================

def process_employer(emp):
    """Process one employer completely."""
    result = {
        'firmenname': emp['name'],
        'ort': emp['ort'],
        'plz': emp['plz'],
        'strasse': emp['adresse'],
        'telefon': emp['telefon'],
        'kategorie_google': emp.get('kategorie', ''),
        'bewertung_google': emp.get('bewertung'),
        'lat': emp.get('lat'),
        'lng': emp.get('lng'),
        'google_maps_link': emp.get('gmaps_link', ''),
        'scrape_status': 'pending',
        'hat_it_jobs': 0,
        'it_job_titel': '',
        'it_job_anzahl': 0,
        'it_job_quelle': '',
        'it_job_details': '',
        'it_keywords_gefunden': '',
        'website_url': '',
        'website_html_path': '',
        'website_html_hash': '',
        'karriere_url': '',
        'karriere_html_path': '',
        'impressum_url': '',
        'impressum_html_path': '',
        'email': '',
        'geschaeftsfuehrer': '',
        'handelsregister': '',
        'it_relevanz_score': 0.0,
        'it_abteilung_vermutung': 0,
    }
    
    try:
        # 1. Check Arbeitsagentur for IT jobs
        aa_jobs = check_aa_it_jobs(emp['name'], emp['ort'])
        if aa_jobs:
            result['hat_it_jobs'] = 1
            result['it_job_anzahl'] = len(aa_jobs)
            result['it_job_titel'] = "; ".join(j['titel'] for j in aa_jobs[:5])
            result['it_job_quelle'] = 'Arbeitsagentur'
            result['it_job_details'] = json.dumps(aa_jobs, ensure_ascii=False)
        
        # 2. Find website
        website = search_website(emp['name'], emp['ort'])
        if not website and emp.get('website_flag', '').lower() == 'ja':
            # Try again with just company name
            website = search_website(emp['name'], '')
        
        if website:
            result['website_url'] = website
            
            # 3. Download main page
            main_html = download_html(website)
            if main_html:
                path, hash_val = save_html_file(main_html, emp['name'])
                result['website_html_path'] = path
                result['website_html_hash'] = hash_val
                
                # 4. Check for IT jobs on website
                has_it, it_kws, it_details = check_website_for_it_jobs(main_html)
                if has_it:
                    result['hat_it_jobs'] = 1
                    if not result['it_job_quelle']:
                        result['it_job_quelle'] = 'Website'
                    result['it_keywords_gefunden'] = ", ".join(it_kws) if it_kws else ""
                    result['it_job_details'] += f" | Website: {it_details}"
                
                # 5. Extract contact info
                contact, impressum_url, karriere_url = extract_contact_info(main_html)
                if contact.get('email'):
                    result['email'] = contact['email']
                if contact.get('geschaeftsfuehrer'):
                    result['geschaeftsfuehrer'] = contact['geschaeftsfuehrer']
                if contact.get('handelsregister'):
                    result['handelsregister'] = contact['handelsregister']
                
                # 6. Download Impressum
                if impressum_url:
                    imp_abs = make_absolute_url(website, impressum_url)
                    if imp_abs:
                        result['impressum_url'] = imp_abs
                        imp_html = download_html(imp_abs)
                        if imp_html:
                            path, _ = save_html_file(imp_html, emp['name'], "impressum")
                            result['impressum_html_path'] = path
                            imp_contact, _, _ = extract_contact_info(imp_html)
                            if imp_contact.get('email') and not result['email']:
                                result['email'] = imp_contact['email']
                            if imp_contact.get('geschaeftsfuehrer') and not result['geschaeftsfuehrer']:
                                result['geschaeftsfuehrer'] = imp_contact['geschaeftsfuehrer']
                            if imp_contact.get('handelsregister') and not result['handelsregister']:
                                result['handelsregister'] = imp_contact['handelsregister']
                
                # 7. Download Karriere page
                if karriere_url:
                    kar_abs = make_absolute_url(website, karriere_url)
                    if kar_abs:
                        result['karriere_url'] = kar_abs
                        kar_html = download_html(kar_abs)
                        if kar_html:
                            path, _ = save_html_file(kar_html, emp['name'], "karriere")
                            result['karriere_html_path'] = path
                            # Check karriere page for IT jobs
                            has_it_kar, it_kws_kar, it_details_kar = check_website_for_it_jobs(kar_html)
                            if has_it_kar:
                                result['hat_it_jobs'] = 1
                                if not result['it_job_quelle']:
                                    result['it_job_quelle'] = 'Website/Karriere'
                                if it_kws_kar and not result['it_keywords_gefunden']:
                                    result['it_keywords_gefunden'] = ", ".join(it_kws_kar)
        
        # 8. Calculate IT relevance
        result['it_relevanz_score'] = calculate_it_relevance(result)
        result['it_abteilung_vermutung'] = 1 if result['it_relevanz_score'] >= 30 else 0
        
        result['scrape_status'] = 'completed'
        result['scrape_datum'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        result['letzte_pruefung'] = datetime.now().strftime('%Y-%m-%d')
        
    except Exception as e:
        result['scrape_status'] = f'error: {str(e)[:100]}'
    
    return result

# ============================================================
# DATABASE
# ============================================================

def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS employers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            firmenname TEXT NOT NULL,
            firmenname_normalisiert TEXT,
            ort TEXT,
            plz TEXT,
            strasse TEXT,
            adresse_vollstaendig TEXT,
            telefon TEXT,
            email TEXT,
            website_url TEXT,
            website_html_path TEXT,
            website_html_hash TEXT,
            karriere_url TEXT,
            karriere_html_path TEXT,
            impressum_url TEXT,
            impressum_html_path TEXT,
            hat_it_jobs INTEGER DEFAULT 0,
            it_job_titel TEXT,
            it_job_anzahl INTEGER DEFAULT 0,
            it_job_quelle TEXT,
            it_job_details TEXT,
            it_keywords_gefunden TEXT,
            kategorie_google TEXT,
            bewertung_google REAL,
            lat REAL,
            lng REAL,
            google_maps_link TEXT,
            handelsregister TEXT,
            geschaeftsfuehrer TEXT,
            it_abteilung_vermutung INTEGER DEFAULT 0,
            it_relevanz_score REAL DEFAULT 0.0,
            scrape_datum TEXT,
            scrape_status TEXT DEFAULT 'pending',
            letzte_pruefung TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_ort ON employers(ort);
        CREATE INDEX IF NOT EXISTS idx_hat_it ON employers(hat_it_jobs);
        CREATE INDEX IF NOT EXISTS idx_score ON employers(it_relevanz_score);
    """)
    conn.commit()
    return conn

def save_to_db(conn, data):
    cols = ['firmenname','firmenname_normalisiert','ort','plz','strasse',
            'telefon','email','website_url','website_html_path','website_html_hash',
            'karriere_url','karriere_html_path','impressum_url','impressum_html_path',
            'hat_it_jobs','it_job_titel','it_job_anzahl','it_job_quelle',
            'it_job_details','it_keywords_gefunden','kategorie_google',
            'bewertung_google','lat','lng','google_maps_link',
            'handelsregister','geschaeftsfuehrer','it_abteilung_vermutung',
            'it_relevanz_score','scrape_datum','scrape_status','letzte_pruefung']
    vals = [
        data.get('firmenname',''), normalize_name(data.get('firmenname','')),
        data.get('ort',''), data.get('plz',''), data.get('strasse',''),
        data.get('telefon',''), data.get('email',''), data.get('website_url',''),
        data.get('website_html_path',''), data.get('website_html_hash',''),
        data.get('karriere_url',''), data.get('karriere_html_path',''),
        data.get('impressum_url',''), data.get('impressum_html_path',''),
        data.get('hat_it_jobs',0), data.get('it_job_titel',''),
        data.get('it_job_anzahl',0), data.get('it_job_quelle',''),
        data.get('it_job_details',''), data.get('it_keywords_gefunden',''),
        data.get('kategorie_google',''), data.get('bewertung_google'),
        data.get('lat'), data.get('lng'), data.get('google_maps_link',''),
        data.get('handelsregister',''), data.get('geschaeftsfuehrer',''),
        data.get('it_abteilung_vermutung',0), data.get('it_relevanz_score',0.0),
        data.get('scrape_datum',''), data.get('scrape_status',''),
        data.get('letzte_pruefung',''),
    ]
    conn.execute(f"INSERT OR REPLACE INTO employers ({','.join(cols)}) VALUES ({','.join(['?']*len(cols))})", vals)
    conn.commit()

def db_to_excel(conn):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IT Arbeitgeber"
    
    cursor = conn.execute("SELECT * FROM employers ORDER BY it_relevanz_score DESC, hat_it_jobs DESC, firmenname")
    columns = [desc[0] for desc in cursor.description]
    
    header_map = {
        'id': 'ID', 'firmenname': 'Firmenname', 'firmenname_normalisiert': 'Firmenname (normalisiert)',
        'ort': 'Ort', 'plz': 'PLZ', 'strasse': 'Straße', 'adresse_vollstaendig': 'Adresse (vollständig)',
        'telefon': 'Telefon', 'email': 'E-Mail', 'website_url': 'Website URL',
        'website_html_path': 'Website HTML Pfad', 'website_html_hash': 'Website HTML Hash',
        'karriere_url': 'Karriere-URL', 'karriere_html_path': 'Karriere HTML Pfad',
        'impressum_url': 'Impressum URL', 'impressum_html_path': 'Impressum HTML Pfad',
        'hat_it_jobs': 'IT-Jobs vorhanden', 'it_job_titel': 'IT-Job Titel',
        'it_job_anzahl': 'Anzahl IT-Jobs', 'it_job_quelle': 'IT-Job Quelle',
        'it_job_details': 'IT-Job Details', 'it_keywords_gefunden': 'IT-Keywords gefunden',
        'kategorie_google': 'Kategorie (Google)', 'bewertung_google': 'Bewertung (Google)',
        'lat': 'Breitengrad', 'lng': 'Längengrad', 'google_maps_link': 'Google Maps Link',
        'handelsregister': 'Handelsregister', 'geschaeftsfuehrer': 'Geschäftsführer',
        'it_abteilung_vermutung': 'IT-Abteilung vermutet', 'it_relevanz_score': 'IT-Relevanz Score',
        'scrape_datum': 'Scrape-Datum', 'scrape_status': 'Scrape-Status',
        'letzte_pruefung': 'Letzte Prüfung', 'created_at': 'Erstellt am', 'updated_at': 'Aktualisiert am',
    }
    
    headers = [header_map.get(c, c) for c in columns]
    ws.append(headers)
    for row in cursor:
        ws.append(list(row))
    
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
    
    # Stats sheet
    ws2 = wb.create_sheet("Statistiken")
    stats = [
        ("Gesamt Arbeitgeber", conn.execute("SELECT COUNT(*) FROM employers").fetchone()[0]),
        ("Mit IT-Jobs", conn.execute("SELECT COUNT(*) FROM employers WHERE hat_it_jobs=1").fetchone()[0]),
        ("Mit Website", conn.execute("SELECT COUNT(*) FROM employers WHERE website_url != ''").fetchone()[0]),
        ("Mit E-Mail", conn.execute("SELECT COUNT(*) FROM employers WHERE email != ''").fetchone()[0]),
        ("IT-Abt. vermutet", conn.execute("SELECT COUNT(*) FROM employers WHERE it_abteilung_vermutung=1").fetchone()[0]),
        ("Ø IT-Relevanz", round(conn.execute("SELECT AVG(it_relevanz_score) FROM employers").fetchone()[0] or 0, 1)),
    ]
    for s in stats:
        ws2.append(s)
    
    # Per-city
    ws3 = wb.create_sheet("Nach Ort")
    ws3.append(["Ort", "Gesamt", "Mit IT-Jobs", "Mit Website", "Mit E-Mail", "Ø IT-Relevanz"])
    for row in conn.execute("""
        SELECT ort, COUNT(*),
               SUM(CASE WHEN hat_it_jobs=1 THEN 1 ELSE 0 END),
               SUM(CASE WHEN website_url != '' THEN 1 ELSE 0 END),
               SUM(CASE WHEN email != '' THEN 1 ELSE 0 END),
               ROUND(AVG(it_relevanz_score), 1)
        FROM employers GROUP BY ort ORDER BY COUNT(*) DESC
    """):
        ws3.append(list(row))
    
    wb.save(EXCEL_OUT)
    print(f"Excel saved: {EXCEL_OUT}")

# ============================================================
# MAIN
# ============================================================

def main():
    print("=" * 60)
    print("IT-Job Scraper v2: Bamberg, Erlangen, Nürnberg")
    print("=" * 60)
    
    # Load employers - EXACT city match only
    wb = openpyxl.load_workbook(EXCEL_IN)
    ws = wb['Google Maps Firmen']
    
    employers = []
    seen = set()
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        ort = str(row[6] or '').strip()
        name = str(row[1] or '').strip()
        if not name:
            continue
        
        # EXACT city match only
        if ort not in TARGET_CITIES:
            continue
        
        norm = normalize_name(name)
        if norm in seen or len(norm) < 3:
            continue
        seen.add(norm)
        
        employers.append({
            'name': name,
            'ort': ort,
            'plz': str(row[5] or '').strip(),
            'adresse': str(row[4] or '').strip(),
            'telefon': str(row[7] or '').strip(),
            'website_flag': str(row[8] or '').strip(),
            'kategorie': str(row[3] or '').strip(),
            'bewertung': row[2],
            'lat': row[9],
            'lng': row[10],
            'gmaps_link': str(row[11] or '').strip(),
        })
    
    print(f"Gefunden: {len(employers)} Arbeitgeber (exakt: {TARGET_CITIES})")
    
    # Load progress
    progress = load_progress()
    completed_names = set(progress.get('completed', []))
    remaining = [e for e in employers if e['name'] not in completed_names]
    print(f"Bereits erledigt: {len(completed_names)}, Verbleibend: {len(remaining)}")
    
    # Init DB
    conn = init_db()
    os.makedirs(HTML_DIR, exist_ok=True)
    
    total = len(remaining)
    processed = 0
    it_count = 0
    batch_since_commit = 0
    
    # Process in parallel with ThreadPoolExecutor
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        # Submit in batches to allow periodic commits
        batch_size = COMMIT_EVERY
        
        for batch_start in range(0, len(remaining), batch_size):
            batch = remaining[batch_start:batch_start + batch_size]
            futures = {executor.submit(process_employer, emp): emp for emp in batch}
            
            for future in as_completed(futures):
                emp = futures[future]
                try:
                    result = future.result()
                except Exception as e:
                    print(f"  ERROR: {emp['name']}: {e}")
                    result = {'firmenname': emp['name'], 'ort': emp['ort'], 'scrape_status': f'error: {e}'}
                
                save_to_db(conn, result)
                processed += 1
                batch_since_commit += 1
                
                progress['completed'].append(result.get('firmenname', emp['name']))
                if result.get('hat_it_jobs'):
                    it_count += 1
                
                status = "✓IT" if result.get('hat_it_jobs') else "   "
                ws_url = result.get('website_url', '')[:40] if result.get('website_url') else '-'
                score = result.get('it_relevanz_score', 0)
                print(f"  [{processed}/{total}] {status} Score:{score:5.1f} | {result.get('firmenname','')[:45]:45} | {ws_url}")
            
            # Save progress
            save_progress(progress)
            
            # Commit & push every batch
            if batch_since_commit >= COMMIT_EVERY:
                print(f"\n  --- GIT COMMIT & PUSH ({it_count} IT-Arbeitgeber bisher) ---")
                try:
                    subprocess.run(['git', 'add', '-A'], cwd=GIT_DIR, capture_output=True, timeout=30)
                    msg = f"IT-Scraper: {processed}/{total} gescannt, {it_count} IT-Jobs ({datetime.now().strftime('%H:%M')})"
                    r = subprocess.run(['git', 'commit', '-m', msg], cwd=GIT_DIR, capture_output=True, timeout=30)
                    if r.returncode == 0:
                        pr = subprocess.run(['git', 'push', 'origin', 'main'], cwd=GIT_DIR, capture_output=True, timeout=60)
                        print(f"  --- Push OK ---\n")
                except Exception as e:
                    print(f"  GIT Error: {e}")
                batch_since_commit = 0
    
    # Final Excel
    print("\nErstelle finale Excel...")
    db_to_excel(conn)
    
    # Final commit & push
    try:
        subprocess.run(['git', 'add', '-A'], cwd=GIT_DIR, capture_output=True, timeout=30)
        msg = f"IT-Scraper FERTIG: {it_count} IT-Arbeitgeber aus {total}"
        subprocess.run(['git', 'commit', '-m', msg], cwd=GIT_DIR, capture_output=True, timeout=30)
        subprocess.run(['git', 'push', 'origin', 'main'], cwd=GIT_DIR, capture_output=True, timeout=60)
    except Exception as e:
        print(f"Final GIT Error: {e}")
    
    conn.close()
    print(f"\n{'='*60}")
    print(f"FERTIG! {it_count} IT-Arbeitgeber aus {total} gescannt")
    print(f"Excel: {EXCEL_OUT}")
    print(f"{'='*60}")

if __name__ == '__main__':
    main()
