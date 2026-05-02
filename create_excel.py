#!/usr/bin/env python3
"""
Create professional Excel file with German employer data.
Uses xlsx skill base template for styling.
"""

import json
import sys
import os

# Add skill templates to path
XLSX_SKILL_DIR = "/home/z/my-project/skills/xlsx"
for sub in [XLSX_SKILL_DIR, os.path.join(XLSX_SKILL_DIR, "templates")]:
    if sub not in sys.path:
        sys.path.insert(0, sub)

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from templates.base import *

# Use professional palette
use_palette_explicit("professional")

# Load employer data
with open('/home/z/my-project/employers_data.json', 'r', encoding='utf-8') as f:
    employers = json.load(f)

# ============================================================
# Create workbook
# ============================================================
wb = Workbook()

# ============================================================
# Sheet 1: Arbeitgeberverzeichnis (Main employer directory)
# ============================================================
ws1 = wb.active
ws1.title = "Arbeitgeberverzeichnis"

# Setup sheet
setup_sheet(ws1, title="Arbeitgeberverzeichnis Deutschland — Bewerbungskontakte", last_col=12)

# Define headers
headers = [
    "Arbeitgeber ID",
    "Firmenname",
    "Straße / Adresse",
    "PLZ",
    "Stadt",
    "Bundesland",
    "Telefon",
    "E-Mail",
    "Website",
    "Branche",
    "Quelle",
    "Bewerbungsstatus",
]

# Column widths
col_widths = [18, 38, 34, 10, 24, 22, 22, 32, 28, 28, 26, 18]

# Write headers at row 4
for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=2):
    cell = ws1.cell(row=4, column=col_idx, value=header)
    ws1.column_dimensions[get_column_letter(col_idx)].width = width

# Style header row
style_header_row(ws1, 4, 2, 13)

# Write data rows
for row_idx, emp in enumerate(employers):
    excel_row = row_idx + 5  # data starts at row 5
    data = [
        emp['arbeitgeber_id'],
        emp['name'],
        emp['adresse'],
        emp['plz'],
        emp['stadt'],
        emp['bundesland'],
        emp['telefon'],
        emp['email'],
        emp['website'],
        emp['branche'],
        emp['quelle'],
        "",  # Bewerbungsstatus - empty for user to fill
    ]
    for col_idx, value in enumerate(data, start=2):
        cell = ws1.cell(row=excel_row, column=col_idx, value=value)
    
    # Style data row
    style_data_row(ws1, excel_row, 2, 13, row_idx)

# Add data validation for Bewerbungsstatus column
status_validation = DataValidation(
    type="list",
    formula1='"Nicht beworben,In Vorbereitung,Beworben,Vorstellungsgespräch,Angebot erhalten,Abgelehnt,Zurückgezogen"',
    allow_blank=True,
)
status_validation.error = "Bitte wählen Sie einen gültigen Status"
status_validation.errorTitle = "Ungültiger Status"
status_validation.prompt = "Wählen Sie den Bewerbungsstatus"
status_validation.promptTitle = "Bewerbungsstatus"
ws1.add_data_validation(status_validation)

# Apply validation to all data rows in status column (col 13 = M)
for row in range(5, 5 + len(employers)):
    status_validation.add(ws1.cell(row=row, column=13))

# Make email and website clickable (hyperlinks)
for row_idx, emp in enumerate(employers):
    excel_row = row_idx + 5
    # Email as mailto link
    if emp['email']:
        email_cell = ws1.cell(row=excel_row, column=9)
        email_cell.hyperlink = f"mailto:{emp['email']}"
        email_cell.font = Font(name=FONT_NAME, size=11, color="0563C1", underline="single")
    
    # Website as hyperlink
    if emp['website']:
        web_cell = ws1.cell(row=excel_row, column=10)
        url = emp['website']
        if not url.startswith('http'):
            url = f"https://{url}"
        web_cell.hyperlink = url
        web_cell.font = Font(name=FONT_NAME, size=11, color="0563C1", underline="single")

# Freeze panes: headers + ID column visible
ws1.freeze_panes = 'D5'

# Add totals/note row
note_row = 5 + len(employers) + 1
ws1.cell(row=note_row, column=2, value=f"Gesamt: {len(employers)} Arbeitgeber")
ws1.cell(row=note_row, column=2).font = font_subheader()
ws1.cell(row=note_row, column=3, value=f"Stand: Mai 2026 | Quelle: berufsstart.de Top 100, DAX 40, Web-Recherche")
ws1.cell(row=note_row, column=3).font = font_caption()

# Print setup
ws1.page_setup.orientation = 'landscape'
ws1.page_setup.fitToWidth = 1
ws1.page_setup.fitToHeight = 0
ws1.print_title_rows = '4:4'


# ============================================================
# Sheet 2: Statistik nach Bundesland
# ============================================================
ws2 = wb.create_sheet("Statistik Bundesland")

setup_sheet(ws2, title="Arbeitgeberstatistik nach Bundesland", last_col=6)

# Count employers by Bundesland
bundesland_counts = {}
for emp in employers:
    bl = emp['bundesland']
    if bl:
        bundesland_counts[bl] = bundesland_counts.get(bl, 0) + 1

# Sort by count descending
sorted_bl = sorted(bundesland_counts.items(), key=lambda x: x[1], reverse=True)

# Headers
bl_headers = ["Bundesland", "Anzahl Arbeitgeber", "Anteil (%)"]
bl_widths = [24, 20, 14]

for col_idx, (header, width) in enumerate(zip(bl_headers, bl_widths), start=2):
    cell = ws2.cell(row=4, column=col_idx, value=header)
    ws2.column_dimensions[get_column_letter(col_idx)].width = width

style_header_row(ws2, 4, 2, 4)

# Data rows
total = len(employers)
for row_idx, (bl_name, count) in enumerate(sorted_bl):
    excel_row = row_idx + 5
    ws2.cell(row=excel_row, column=2, value=bl_name)
    ws2.cell(row=excel_row, column=3, value=count)
    ws2.cell(row=excel_row, column=4, value=round(count / total * 100, 1))
    ws2.cell(row=excel_row, column=4).number_format = '0.0"%"'
    style_data_row(ws2, excel_row, 2, 4, row_idx)

# Total row
total_row = 5 + len(sorted_bl)
ws2.cell(row=total_row, column=2, value="Gesamt")
ws2.cell(row=total_row, column=3, value=total)
ws2.cell(row=total_row, column=4, value="100.0%")
style_total_row(ws2, total_row, 2, 4)

# Add chart
from openpyxl.chart import BarChart, Reference

chart = BarChart()
chart.type = "bar"
chart.style = 10
chart.width = 22
chart.height = 14

data_ref = Reference(ws2, min_col=3, min_row=4, max_row=4 + len(sorted_bl))
cats_ref = Reference(ws2, min_col=2, min_row=5, max_row=4 + len(sorted_bl))
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.shape = 4

# Style chart
setup_chart_titles(chart, title="Arbeitgeber nach Bundesland", y_title="Anzahl", x_title="Bundesland")
apply_chart_colors(chart)

ws2.add_chart(chart, f"F4")


# ============================================================
# Sheet 3: Statistik nach Branche
# ============================================================
ws3 = wb.create_sheet("Statistik Branche")

setup_sheet(ws3, title="Arbeitgeberstatistik nach Branche", last_col=6)

# Count employers by Branche
branche_counts = {}
for emp in employers:
    br = emp['branche']
    if br:
        # Simplify branch name
        br_simple = br.split('/')[0].strip()
        branche_counts[br_simple] = branche_counts.get(br_simple, 0) + 1

sorted_br = sorted(branche_counts.items(), key=lambda x: x[1], reverse=True)

# Headers
br_headers = ["Branche", "Anzahl Arbeitgeber", "Anteil (%)"]
br_widths = [28, 20, 14]

for col_idx, (header, width) in enumerate(zip(br_headers, br_widths), start=2):
    cell = ws3.cell(row=4, column=col_idx, value=header)
    ws3.column_dimensions[get_column_letter(col_idx)].width = width

style_header_row(ws3, 4, 2, 4)

# Data rows
for row_idx, (br_name, count) in enumerate(sorted_br):
    excel_row = row_idx + 5
    ws3.cell(row=excel_row, column=2, value=br_name)
    ws3.cell(row=excel_row, column=3, value=count)
    ws3.cell(row=excel_row, column=4, value=round(count / total * 100, 1))
    ws3.cell(row=excel_row, column=4).number_format = '0.0"%"'
    style_data_row(ws3, excel_row, 2, 4, row_idx)

# Total row
total_row_br = 5 + len(sorted_br)
ws3.cell(row=total_row_br, column=2, value="Gesamt")
ws3.cell(row=total_row_br, column=3, value=total)
ws3.cell(row=total_row_br, column=4, value="100.0%")
style_total_row(ws3, total_row_br, 2, 4)

# Add chart
chart_br = BarChart()
chart_br.type = "bar"
chart_br.style = 10
chart_br.width = 22
chart_br.height = 18

data_ref_br = Reference(ws3, min_col=3, min_row=4, max_row=4 + len(sorted_br))
cats_ref_br = Reference(ws3, min_col=2, min_row=5, max_row=4 + len(sorted_br))
chart_br.add_data(data_ref_br, titles_from_data=True)
chart_br.set_categories(cats_ref_br)
chart_br.shape = 4

setup_chart_titles(chart_br, title="Arbeitgeber nach Branche", y_title="Anzahl", x_title="Branche")
apply_chart_colors(chart_br)

ws3.add_chart(chart_br, f"F4")


# ============================================================
# Sheet 4: Bewerbungstracker
# ============================================================
ws4 = wb.create_sheet("Bewerbungstracker")

setup_sheet(ws4, title="Persönlicher Bewerbungstracker", last_col=10)

# Headers
tracker_headers = [
    "Arbeitgeber ID",
    "Firmenname",
    "Stadt",
    "Branche",
    "Bewerbungsdatum",
    "Status",
    "Kontakt Person",
    "Notizen",
    "Nächster Schritt",
    "Wiedervorlage",
]

tracker_widths = [18, 34, 20, 22, 16, 16, 20, 28, 22, 14]

for col_idx, (header, width) in enumerate(zip(tracker_headers, tracker_widths), start=2):
    cell = ws4.cell(row=4, column=col_idx, value=header)
    ws4.column_dimensions[get_column_letter(col_idx)].width = width

style_header_row(ws4, 4, 2, 11)

# Pre-fill with employer names
for row_idx, emp in enumerate(employers):
    excel_row = row_idx + 5
    ws4.cell(row=excel_row, column=2, value=emp['arbeitgeber_id'])
    ws4.cell(row=excel_row, column=3, value=emp['name'])
    ws4.cell(row=excel_row, column=4, value=emp['stadt'])
    ws4.cell(row=excel_row, column=5, value=emp['branche'])
    style_data_row(ws4, excel_row, 2, 11, row_idx)

# Add status validation for tracker
tracker_status_validation = DataValidation(
    type="list",
    formula1='"Nicht beworben,Recherche,Lebenslauf angepasst,Beworben,Bestätigung erhalten,Vorstellungsgespräch,2. Gespräch,Angebot erhalten,Verhandlung,Angenommen,Abgelehnt,Zurückgezogen"',
    allow_blank=True,
)
tracker_status_validation.prompt = "Wählen Sie den Bewerbungsstatus"
tracker_status_validation.promptTitle = "Status"
ws4.add_data_validation(tracker_status_validation)

for row in range(5, 5 + len(employers)):
    tracker_status_validation.add(ws4.cell(row=row, column=7))

# Date format for Bewerbungsdatum and Wiedervorlage
for row in range(5, 5 + len(employers)):
    ws4.cell(row=row, column=6).number_format = 'DD.MM.YYYY'
    ws4.cell(row=row, column=11).number_format = 'DD.MM.YYYY'

ws4.freeze_panes = 'D5'


# ============================================================
# Sheet 5: Hinweise
# ============================================================
ws5 = wb.create_sheet("Hinweise")

setup_sheet(ws5, title="Hinweise zur Nutzung", last_col=8)

ws5.column_dimensions["B"].width = 20
ws5.column_dimensions["C"].width = 70

notes = [
    ("Datenquelle", "Die Arbeitgeberdaten stammen aus dem berufsstart.de Top 100 Ranking 2025, DAX 40-Unternehmen und ergänzenden Web-Recherchen."),
    ("Kontaktdaten", "Die Kontaktdaten (Adresse, Telefon, E-Mail) wurden aus öffentlich zugänglichen Quellen (Impressum, Kontaktseiten) recherchiert. Bitte überprüfen Sie die Daten vor einer Bewerbung."),
    ("Bewerbungsstatus", "In der Spalte 'Bewerbungsstatus' auf dem Blatt 'Arbeitgeberverzeichnis' können Sie den Status Ihrer Bewerbung über ein Dropdown-Menü einstellen."),
    ("Bewerbungstracker", "Das Blatt 'Bewerbungstracker' bietet detaillierte Möglichkeiten zur Verfolgung Ihrer Bewerbungen inkl. Kontaktperson, Notizen und Wiedervorlagedaten."),
    ("Hyperlinks", "E-Mail-Adressen und Websites sind als klickbare Hyperlinks formatiert."),
    ("Aktualisierung", "Die Daten wurden im Mai 2026 erhoben. Für aktuelle Kontaktdaten besuchen Sie bitte die jeweilige Unternehmenswebsite."),
    ("Deduplizierung", "Doppelte Einträge wurden automatisch entfernt. Jeder Arbeitgeber hat eine eindeutige ID im Format DE-XXXX-NNNN."),
    ("Haftungsausschluss", "Alle Angaben ohne Gewähr. Die Kontaktdaten dienen als Startpunkt für Ihre Bewerbungsrecherche und sollten vorab verifiziert werden."),
]

for row_idx, (key, value) in enumerate(notes):
    excel_row = row_idx + 4
    cell_key = ws5.cell(row=excel_row, column=2, value=key)
    cell_key.font = font_subheader()
    cell_val = ws5.cell(row=excel_row, column=3, value=value)
    cell_val.font = font_body()
    cell_val.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws5.row_dimensions[excel_row].height = 36


# ============================================================
# Save workbook
# ============================================================
output_path = "/home/z/my-project/download/Arbeitgeber_Deutschland_Bewerbungskontakte.xlsx"
os.makedirs(os.path.dirname(output_path), exist_ok=True)
wb.save(output_path)

print(f"✅ Excel file saved to: {output_path}")
print(f"   - Arbeitgeberverzeichnis: {len(employers)} Arbeitgeber mit Kontaktdaten")
print(f"   - Statistik Bundesland: {len(sorted_bl)} Bundesländer")
print(f"   - Statistik Branche: {len(sorted_br)} Branchen")
print(f"   - Bewerbungstracker: Vorformatiert für {len(employers)} Einträge")
print(f"   - Hinweise: Nutzungshinweise und Erklärungen")

