#!/usr/bin/env python3
"""
Robust batch IT-employer scraper.
Processes employers in small batches, commits each batch, can resume.
"""
import json, os, re, sys, time, hashlib, sqlite3, subprocess
from datetime import datetime
import requests
from bs4 import BeautifulSoup
import openpyxl

sys.stdout.reconfigure(line_buffering=True)

# CONFIG
EXCEL_IN = "/home/z/my-project/download/Google_Maps_Firmen_Erlangen_50km.xlsx"
EXCEL_OUT = "/home/z/my-project/download/IT_Arbeitgeber_Bamberg_Erlangen_Nuernberg.xlsx"
HTML_DIR = "/home/z/my-project/download/it_employer_html"
DB_PATH = "/home/z/my-project/download/it_employers.db"
IT_JOBS_MAP = "/home/z/my-project/download/it_jobs_map.json"
PROGRESS_FILE = "/home/z/my-project/download/it_scraper_progress.json"
GIT_DIR = "/home/z/my-project"

TARGET_CITIES = {"Bamberg", "Erlangen", "Nürnberg"}
BATCH_SIZE = 20
COMMIT_EVERY = 10

SESSION = requests.Session()
SESSION.headers.update({"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"})
SESSION.verify = False
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

def sanitize_filename(name, max_len=80):
    name = re.sub(r'[äöüßÄÖÜ]', lambda m: {'ä':'ae','ö':'oe','ü':'ue','ß':'ss','Ä':'Ae','Ö':'Oe','Ü':'Ue'}[m.group()], name)
    name = re.sub(r'[^a-zA-Z0-9._-]', '_', name)
    name = re.sub(r'_+', '_', name).strip('_')
    return name[:max_len]

def normalize_name(name):
    name = name.lower().strip()
    for s in ['gmbh & co. kg','gmbh & co kg','gmbh co. kg','gmbh co kg','gmbh & co.','ag & co. kg','ag & co kg','gmbh','ag','kg','ohg','eg','se','e.k.','ek','e.v.','ev','mbh','ug','gbr']:
        name = re.sub(re.escape(s), '', name, flags=re.IGNORECASE)
    name = re.sub(r'[^a-z0-9äöüß]', '', name)
    return name

def make_abs(base, rel):
    if not rel: return None
    if rel.startswith('http'): return rel
    if rel.startswith('//'): return 'https:' + rel
    if rel.startswith('/'):
        from urllib.parse import urlparse
        p = urlparse(base)
        return f"{p.scheme}://{p.netloc}{rel}"
    return f"{base.rstrip('/')}/{rel.lstrip('/')}"

def find_website(name, city):
    clean = re.sub(r'(GmbH|AG|KG|OHG|UG|e\.K\.|e\.V\.|mbH|Co\.\s*KG|& Co\.|GmbH & Co\. KG)', '', name, flags=re.IGNORECASE).strip()
    domain = re.sub(r'[äöüß]', lambda m: {'ä':'ae','ö':'oe','ü':'ue','ß':'ss'}[m.group()], clean)
    domain = re.sub(r'[^a-zA-Z0-9]', '', domain).lower()
    if len(domain) < 3: return None
    for url in [f"https://www.{domain}.de", f"https://{domain}.de"]:
        try:
            r = SESSION.get(url, timeout=8, allow_redirects=True)
            if r.status_code == 200 and 'text/html' in r.headers.get('content-type',''):
                return url
        except: pass
    return None

def download_html(url):
    try:
        r = SESSION.get(url, timeout=12, allow_redirects=True)
        if r.status_code == 200 and 'text/' in r.headers.get('content-type',''):
            return r.text
    except: pass
    return None

def save_html(content, name, suffix=""):
    if not content: return None, None
    fn = f"{sanitize_filename(name)}_{suffix}.html" if suffix else f"{sanitize_filename(name)}.html"
    fp = os.path.join(HTML_DIR, fn)
    os.makedirs(HTML_DIR, exist_ok=True)
    with open(fp, 'w', encoding='utf-8', errors='replace') as f: f.write(content)
    return fp, hashlib.sha256(content.encode()).hexdigest()[:16]

def extract_info(html):
    info = {}; imp_url = None; kar_url = None
    if not html: return info, None, None
    emails = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', html)
    emails = [e for e in emails if not any(x in e.lower() for x in ['example.com','sentry','webpack','noreply','wixpress','github'])]
    emails = list(dict.fromkeys(emails))[:5]
    if emails: info['email'] = emails[0]
    gf = re.search(r'(?:Geschäftsführer)[\s:]*([^<\.]{2,100})', html, re.IGNORECASE)
    if gf:
        t = BeautifulSoup(gf.group(1), 'html.parser').get_text(strip=True)
        if len(t)>2: info['geschaeftsfuehrer'] = t[:200]
    hr = re.search(r'(?:Handelsregister|HRB)[\s:]*([^<\.]{2,100})', html, re.IGNORECASE)
    if hr:
        t = BeautifulSoup(hr.group(1), 'html.parser').get_text(strip=True)
        info['handelsregister'] = t[:200]
    soup = BeautifulSoup(html, 'html.parser')
    for a in soup.find_all('a', href=True):
        h = a['href'].lower(); t = a.get_text(strip=True).lower()
        if not imp_url and ('impressum' in h or 'impressum' in t): imp_url = a['href']
        if not kar_url and any(w in h for w in ['karriere','career','jobs','stellenangebote','bewerbung']): kar_url = a['href']
    return info, imp_url, kar_url

def check_it(html):
    if not html: return False, []
    text = html.lower()
    it_kws = [kw for kw in ['software','entwickler','developer','informatik','it-','data','cloud','devops','administrator','sap','scrum','docker','frontend','backend'] if kw in text]
    career = any(w in text for w in ['karriere','career','stellenangebote','jobs','bewerbung'])
    specific = ['entwickler m/w/d','developer m/w/d','it-jobs','softwareentwickler gesucht','devops engineer']
    has = (bool(it_kws) and career) or any(s in text for s in specific)
    return has, it_kws[:10]

def process_one(emp, it_data, it_score):
    r = {
        'firmenname': emp['name'], 'ort': emp['ort'], 'plz': emp['plz'],
        'strasse': emp['adresse'], 'telefon': emp.get('telefon',''),
        'email':'', 'website_url':'', 'website_html_path':'', 'website_html_hash':'',
        'karriere_url':'', 'karriere_html_path':'', 'impressum_url':'', 'impressum_html_path':'',
        'hat_it_jobs': 1 if it_data else 0, 'it_job_titel':'', 'it_job_anzahl':0,
        'it_job_quelle':'', 'it_job_details':'', 'it_keywords_gefunden':'',
        'kategorie_google': emp.get('kategorie',''), 'bewertung_google': emp.get('bewertung'),
        'lat': emp.get('lat'), 'lng': emp.get('lng'),
        'google_maps_link': emp.get('gmaps_link',''), 'handelsregister':'', 'geschaeftsfuehrer':'',
        'it_abteilung_vermutung': 1 if it_score >= 30 else 0, 'it_relevanz_score': it_score,
    }
    if it_data:
        jobs = it_data.get('jobs',[])
        r['it_job_anzahl'] = len(jobs)
        r['it_job_titel'] = "; ".join(j.get('titel','') for j in jobs[:10])
        r['it_job_quelle'] = 'Arbeitsagentur'
        r['it_job_details'] = json.dumps([{'titel':j.get('titel',''),'beruf':j.get('hauptberuf',''),'refnr':j.get('referenznummer',''),'ort':j.get('ort','')} for j in jobs[:20]], ensure_ascii=False)
    
    ws = find_website(emp['name'], emp['ort'])
    if ws:
        r['website_url'] = ws
        html = download_html(ws)
        if html:
            p, h = save_html(html, emp['name'])
            r['website_html_path'] = p; r['website_html_hash'] = h
            has_it, it_kws = check_it(html)
            if has_it:
                if not r['hat_it_jobs']: r['hat_it_jobs'] = 1
                if not r['it_job_quelle']: r['it_job_quelle'] = 'Website'
                r['it_keywords_gefunden'] = ", ".join(it_kws)
                r['it_relevanz_score'] = min(r['it_relevanz_score']+15, 100)
            contact, imp_url, kar_url = extract_info(html)
            if contact.get('email'): r['email'] = contact['email']
            if contact.get('geschaeftsfuehrer'): r['geschaeftsfuehrer'] = contact['geschaeftsfuehrer']
            if contact.get('handelsregister'): r['handelsregister'] = contact['handelsregister']
            if imp_url:
                imp_abs = make_abs(ws, imp_url)
                if imp_abs:
                    r['impressum_url'] = imp_abs
                    imp_html = download_html(imp_abs)
                    if imp_html:
                        p2, _ = save_html(imp_html, emp['name'], "impressum")
                        r['impressum_html_path'] = p2
                        ic, _, _ = extract_info(imp_html)
                        if ic.get('email') and not r['email']: r['email'] = ic['email']
                        if ic.get('geschaeftsfuehrer') and not r['geschaeftsfuehrer']: r['geschaeftsfuehrer'] = ic['geschaeftsfuehrer']
                        if ic.get('handelsregister') and not r['handelsregister']: r['handelsregister'] = ic['handelsregister']
            if kar_url:
                kar_abs = make_abs(ws, kar_url)
                if kar_abs:
                    r['karriere_url'] = kar_abs
                    kar_html = download_html(kar_abs)
                    if kar_html:
                        p3, _ = save_html(kar_html, emp['name'], "karriere")
                        r['karriere_html_path'] = p3
    
    r['scrape_status'] = 'completed'
    r['scrape_datum'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    r['letzte_pruefung'] = datetime.now().strftime('%Y-%m-%d')
    return r

def save_to_db(conn, data):
    cols = ['firmenname','firmenname_normalisiert','ort','plz','strasse',
            'telefon','email','website_url','website_html_path','website_html_hash',
            'karriere_url','karriere_html_path','impressum_url','impressum_html_path',
            'hat_it_jobs','it_job_titel','it_job_anzahl','it_job_quelle',
            'it_job_details','it_keywords_gefunden','kategorie_google',
            'bewertung_google','lat','lng','google_maps_link',
            'handelsregister','geschaeftsfuehrer','it_abteilung_vermutung',
            'it_relevanz_score','scrape_datum','scrape_status','letzte_pruefung']
    vals = [data.get('firmenname',''), normalize_name(data.get('firmenname','')),
            data.get('ort',''), data.get('plz',''), data.get('strasse',''),
            data.get('telefon',''), data.get('email',''), data.get('website_url',''),
            data.get('website_html_path',''), data.get('website_html_hash',''),
            data.get('karriere_url',''), data.get('karriere_html_path',''),
            data.get('impressum_url',''), data.get('impressum_html_path',''),
            data.get('hat_it_jobs',0), data.get('it_job_titel',''),
            data.get('it_job_anzahl',0), data.get('it_job_quelle',''),
            data.get('it_job_details',''), data.get('it_keywords_gefunden',''),
            data.get('kategorie_google',''), data.get('bewertung_google'),
            data.get('lat'), data.get('lng'), data.get('google_maps_link',''),
            data.get('handelsregister',''), data.get('geschaeftsfuehrer',''),
            data.get('it_abteilung_vermutung',0), data.get('it_relevanz_score',0.0),
            data.get('scrape_datum',''), data.get('scrape_status',''),
            data.get('letzte_pruefung','')]
    conn.execute(f"INSERT OR REPLACE INTO employers ({','.join(cols)}) VALUES ({','.join(['?']*len(cols))})", vals)
    conn.commit()

def main():
    print("=== IT-Scraper Batch ===")
    
    # Load IT jobs
    with open(IT_JOBS_MAP) as f: it_jobs = json.load(f)
    print(f"IT-Jobs: {len(it_jobs)} Arbeitgeber")
    
    # Load Google Maps
    wb = openpyxl.load_workbook(EXCEL_IN)
    ws = wb['Google Maps Firmen']
    employers = []; seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        ort = str(row[6] or '').strip()
        name = str(row[1] or '').strip()
        if not name or ort not in TARGET_CITIES: continue
        norm = normalize_name(name)
        if norm in seen or len(norm) < 3: continue
        seen.add(norm)
        employers.append({'name':name,'ort':ort,'plz':str(row[5] or '').strip(),
                          'adresse':str(row[4] or '').strip(),'telefon':str(row[7] or '').strip(),
                          'kategorie':str(row[3] or '').strip(),'bewertung':row[2],
                          'lat':row[9],'lng':row[10],'gmaps_link':str(row[11] or '').strip()})
    print(f"Google Maps: {len(employers)} Arbeitgeber")
    
    # Match
    matched = []
    for emp in employers:
        norm = normalize_name(emp['name'])
        it_data = None; score = 0.0
        if norm in it_jobs: it_data = it_jobs[norm]; score = 50
        else:
            for it_norm, it_d in it_jobs.items():
                if len(norm)>4 and len(it_norm)>4 and (norm[:6] in it_norm or it_norm[:6] in norm):
                    it_data = it_d; score = 40; break
        name_lower = emp['name'].lower()
        kat_lower = (emp.get('kategorie') or '').lower()
        if any(w in name_lower for w in ['it ','it-','software','tech','digital','data','cloud','cyber','computer','informatik','system']): score += 20
        if any(w in kat_lower for w in ['software','it','informatik','computer','technologie','digital']): score += 15
        matched.append({'employer': emp, 'it_data': it_data, 'it_score': min(score, 100)})
    
    matched.sort(key=lambda x: (-x['it_score'], x['employer']['name']))
    
    # Progress
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE) as f: progress = json.load(f)
    else:
        progress = {"completed": []}
    
    completed = set(progress['completed'])
    remaining = [m for m in matched if m['employer']['name'] not in completed]
    print(f"Verbleibend: {len(remaining)} (erledigt: {len(completed)})")
    
    # Init DB
    conn = sqlite3.connect(DB_PATH)
    conn.executescript("""CREATE TABLE IF NOT EXISTS employers (
        id INTEGER PRIMARY KEY AUTOINCREMENT, firmenname TEXT NOT NULL,
        firmenname_normalisiert TEXT, ort TEXT, plz TEXT, strasse TEXT,
        telefon TEXT, email TEXT, website_url TEXT, website_html_path TEXT,
        website_html_hash TEXT, karriere_url TEXT, karriere_html_path TEXT,
        impressum_url TEXT, impressum_html_path TEXT, hat_it_jobs INTEGER DEFAULT 0,
        it_job_titel TEXT, it_job_anzahl INTEGER DEFAULT 0, it_job_quelle TEXT,
        it_job_details TEXT, it_keywords_gefunden TEXT, kategorie_google TEXT,
        bewertung_google REAL, lat REAL, lng REAL, google_maps_link TEXT,
        handelsregister TEXT, geschaeftsfuehrer TEXT,
        it_abteilung_vermutung INTEGER DEFAULT 0, it_relevanz_score REAL DEFAULT 0.0,
        scrape_datum TEXT, scrape_status TEXT, letzte_pruefung TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP, updated_at TEXT DEFAULT CURRENT_TIMESTAMP
    ); CREATE INDEX IF NOT EXISTS idx_ort ON employers(ort);
    CREATE INDEX IF NOT EXISTS idx_hat_it ON employers(hat_it_jobs);
    CREATE INDEX IF NOT EXISTS idx_score ON employers(it_relevanz_score);""")
    conn.commit()
    os.makedirs(HTML_DIR, exist_ok=True)
    
    # Process sequentially (more reliable than ThreadPool)
    total = len(remaining)
    processed = 0; it_count = 0; batch_commit = 0
    
    for m in remaining:
        result = process_one(m['employer'], m['it_data'], m['it_score'])
        save_to_db(conn, result)
        processed += 1; batch_commit += 1
        progress['completed'].append(result.get('firmenname', ''))
        if result.get('hat_it_jobs'): it_count += 1
        
        s = "✓IT" if result.get('hat_it_jobs') else "   "
        ws = result.get('website_url','')[:35] if result.get('website_url') else '-'
        sc = result.get('it_relevanz_score',0)
        em = "📧" if result.get('email') else "  "
        print(f"  [{processed}/{total}] {s} {em} Sc:{sc:5.1f} | {result.get('firmenname','')[:42]:42} | {ws}")
        
        if batch_commit >= COMMIT_EVERY:
            with open(PROGRESS_FILE, 'w') as f: json.dump(progress, f, ensure_ascii=False)
            try:
                subprocess.run(['git','add','-A'], cwd=GIT_DIR, capture_output=True, timeout=30)
                msg = f"IT-Scraper: {processed}/{total} | {it_count} IT-Jobs ({datetime.now().strftime('%H:%M')})"
                r = subprocess.run(['git','commit','-m',msg], cwd=GIT_DIR, capture_output=True, timeout=30)
                if r.returncode == 0:
                    subprocess.run(['git','push','origin','main'], cwd=GIT_DIR, capture_output=True, timeout=60)
                    print(f"  >>> GIT PUSH ({it_count} IT) <<<")
            except Exception as e: print(f"  GIT: {e}")
            batch_commit = 0
    
    # Final
    with open(PROGRESS_FILE, 'w') as f: json.dump(progress, f, ensure_ascii=False)
    
    # Generate Excel
    wb_out = openpyxl.Workbook()
    ws_main = wb_out.active
    ws_main.title = "IT Arbeitgeber"
    cursor = conn.execute("SELECT * FROM employers ORDER BY it_relevanz_score DESC, hat_it_jobs DESC, firmenname")
    columns = [d[0] for d in cursor.description]
    hm = {'id':'ID','firmenname':'Firmenname','firmenname_normalisiert':'Firmenname (norm.)',
          'ort':'Ort','plz':'PLZ','strasse':'Straße','telefon':'Telefon','email':'E-Mail',
          'website_url':'Website URL','website_html_path':'HTML Pfad','website_html_hash':'HTML Hash',
          'karriere_url':'Karriere-URL','karriere_html_path':'Karriere HTML','impressum_url':'Impressum URL',
          'impressum_html_path':'Impressum HTML','hat_it_jobs':'IT-Jobs','it_job_titel':'IT-Job Titel',
          'it_job_anzahl':'Anzahl IT-Jobs','it_job_quelle':'IT-Job Quelle','it_job_details':'IT-Job Details',
          'it_keywords_gefunden':'IT-Keywords','kategorie_google':'Kategorie (Google)',
          'bewertung_google':'Bewertung','lat':'Breitengrad','lng':'Längengrad',
          'google_maps_link':'Google Maps Link','handelsregister':'Handelsregister',
          'geschaeftsfuehrer':'Geschäftsführer','it_abteilung_vermutung':'IT-Abt. vermutet',
          'it_relevanz_score':'IT-Relevanz Score','scrape_datum':'Scrape-Datum',
          'scrape_status':'Scrape-Status','letzte_pruefung':'Letzte Prüfung',
          'created_at':'Erstellt','updated_at':'Aktualisiert'}
    ws_main.append([hm.get(c,c) for c in columns])
    for row in cursor: ws_main.append(list(row))
    for col in ws_main.columns:
        ml = max(len(str(cell.value or '')) for cell in col)
        ws_main.column_dimensions[col[0].column_letter].width = min(ml+2, 50)
    
    # Stats
    ws2 = wb_out.create_sheet("Statistiken")
    for s in [("Gesamt",conn.execute("SELECT COUNT(*) FROM employers").fetchone()[0]),
              ("IT-Jobs",conn.execute("SELECT COUNT(*) FROM employers WHERE hat_it_jobs=1").fetchone()[0]),
              ("Website",conn.execute("SELECT COUNT(*) FROM employers WHERE website_url != ''").fetchone()[0]),
              ("E-Mail",conn.execute("SELECT COUNT(*) FROM employers WHERE email != ''").fetchone()[0])]:
        ws2.append(s)
    
    # Per city
    ws3 = wb_out.create_sheet("Nach Ort")
    ws3.append(["Ort","Gesamt","IT-Jobs","Website","E-Mail","Ø Score"])
    for row in conn.execute("""SELECT ort,COUNT(*),SUM(CASE WHEN hat_it_jobs=1 THEN 1 ELSE 0 END),
        SUM(CASE WHEN website_url != '' THEN 1 ELSE 0 END),SUM(CASE WHEN email != '' THEN 1 ELSE 0 END),
        ROUND(AVG(it_relevanz_score),1) FROM employers GROUP BY ort ORDER BY COUNT(*) DESC"""):
        ws3.append(list(row))
    
    wb_out.save(EXCEL_OUT)
    print(f"Excel: {EXCEL_OUT}")
    
    try:
        subprocess.run(['git','add','-A'], cwd=GIT_DIR, capture_output=True, timeout=30)
        subprocess.run(['git','commit','-m',f"IT-Scraper FERTIG: {it_count} IT-Arbeitgeber"], cwd=GIT_DIR, capture_output=True, timeout=30)
        subprocess.run(['git','push','origin','main'], cwd=GIT_DIR, capture_output=True, timeout=60)
    except: pass
    conn.close()
    print(f"FERTIG! {it_count}/{total}")

if __name__ == '__main__':
    main()
