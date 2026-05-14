#!/usr/bin/env python3
"""
Parallel IT-Job Scraper for Bamberg, Erlangen, Nürnberg employers.
- Reads employers from Google Maps Excel
- For each employer: checks if they have current IT jobs (Arbeitsagentur API + web search)
- Collects website URLs, downloads HTML
- Saves to database-friendly Excel with atomic columns
- Commits & pushes every 10 entries
"""

import asyncio
import aiohttp
import json
import os
import re
import time
import hashlib
import sqlite3
import subprocess
import sys
import traceback
from datetime import datetime
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor
from bs4 import BeautifulSoup
import openpyxl

# ============================================================
# CONFIG
# ============================================================
EXCEL_IN = "/home/z/my-project/download/Google_Maps_Firmen_Erlangen_50km.xlsx"
EXCEL_OUT = "/home/z/my-project/download/IT_Arbeitgeber_Bamberg_Erlangen_Nuernberg.xlsx"
HTML_DIR = "/home/z/my-project/download/it_employer_html"
DB_PATH = "/home/z/my-project/download/it_employers.db"
PROGRESS_FILE = "/home/z/my-project/download/it_scraper_progress.json"
GIT_DIR = "/home/z/my-project"

TARGET_CITIES = ["Bamberg", "Erlangen", "Nürnberg"]
MAX_CONCURRENT = 15  # parallel requests
COMMIT_EVERY = 10

# Arbeitsagentur API
AA_API_URL = "https://rest.arbeitsagentur.de/jobboerse/jobsuche-service/pc/v6/jobs"
AA_API_KEY = "jobboerse-jobsuche"

# IT-Keywords for job matching
IT_KEYWORDS = [
    "software", "entwickler", "developer", "it-", "informatik", "data",
    "systemadministrator", "devops", "frontend", "backend", "fullstack",
    "sre", "cloud", "cyber", "security", "netzwerk", "network",
    "administrator", "programmierer", "programmiererin", "analyst",
    "sap", "linux", "python", "java", "javascript", "typescript",
    "machine learning", "ai", "künstliche intelligenz", "datenbank",
    "database", "sql", "scrum", "agile", "tester", "test engineer",
    "support", "helpdesk", "it consultant", "it berater", "projektmanager it",
    "product owner", "architect", "architekt", "ciso", "cto",
    "erp", "crm", "bi ", "business intelligence", "data engineer",
    "data scientist", "platform", "infrastruktur", "infrastructure",
    "automation", "automatisierung", "robotic", "iot", "embedded",
    "microservice", "api", "container", "kubernetes", "docker",
    "ci/cd", "terraform", "ansible", "azure", "aws", "gcp",
]

IT_CATEGORIES = [
    "Informatik", "IT", "Softwareentwicklung", "Datenverarbeitung",
    "Telekommunikation", "Elektrotechnik",  # sometimes IT-adjacent
]

# ============================================================
# DATABASE SCHEMA (atomic columns - database friendly)
# ============================================================
DB_SCHEMA = """
CREATE TABLE IF NOT EXISTS employers (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    firmenname TEXT NOT NULL,
    firmenname_normalisiert TEXT,
    ort TEXT,
    plz TEXT,
    strasse TEXT,
    adresse_vollstaendig TEXT,
    bundesland TEXT,
    land TEXT DEFAULT 'Deutschland',
    telefon TEXT,
    email TEXT,
    website_url TEXT,
    website_html_path TEXT,
    website_html_hash TEXT,
    karriere_url TEXT,
    karriere_html_path TEXT,
    impressum_url TEXT,
    impressum_html_path TEXT,
    hat_it_jobs INTEGER DEFAULT 0,
    it_job_titel TEXT,
    it_job_anzahl INTEGER DEFAULT 0,
    it_job_quelle TEXT,
    it_job_details TEXT,
    it_keywords_gefunden TEXT,
    branche TEXT,
    kategorie_google TEXT,
    bewertung_google REAL,
    anzahl_bewertungen INTEGER,
    lat REAL,
    lng REAL,
    google_maps_link TEXT,
    arbeitsagentur_id TEXT,
    arbeitsagentur_name TEXT,
    firmengroesse TEXT,
    rechtsform TEXT,
    gruendungsjahr TEXT,
    ust_id TEXT,
    handelsregister TEXT,
    geschaeftsfuehrer TEXT,
    angestelltenzahl TEXT,
    it_abteilung_vermutung INTEGER DEFAULT 0,
    it_relevanz_score REAL DEFAULT 0.0,
    scrape_datum TEXT,
    scrape_status TEXT DEFAULT 'pending',
    letzte_pruefung TEXT,
    notizen TEXT,
    created_at TEXT DEFAULT CURRENT_TIMESTAMP,
    updated_at TEXT DEFAULT CURRENT_TIMESTAMP
);

CREATE INDEX IF NOT EXISTS idx_ort ON employers(ort);
CREATE INDEX IF NOT EXISTS idx_hat_it_jobs ON employers(hat_it_jobs);
CREATE INDEX IF NOT EXISTS idx_it_relevanz ON employers(it_relevanz_score);
CREATE INDEX IF NOT EXISTS idx_firmenname ON employers(firmenname_normalisiert);
"""

# ============================================================
# UTILITIES
# ============================================================

def sanitize_filename(name, max_len=80):
    """Create filesystem-safe filename from company name."""
    name = re.sub(r'[äöüß]', lambda m: {'ä':'ae','ö':'oe','ü':'ue','ß':'ss'}[m.group()], name)
    name = re.sub(r'[^a-zA-Z0-9._-]', '_', name)
    name = re.sub(r'_+', '_', name).strip('_')
    return name[:max_len]

def normalize_name(name):
    """Normalize company name for dedup."""
    name = name.lower().strip()
    name = re.sub(r'\s+', ' ', name)
    # Remove common suffixes
    for suffix in ['gmbh & co. kg', 'gmbh & co kg', 'gmbh co. kg', 'gmbh co kg',
                   'gmbh & co.', 'ag & co. kg', 'ag & co kg',
                   'gmbh', 'ag', 'kg', 'ohg', 'eg', 'se', 'e.k.', 'ek',
                   'e.v.', 'ev', 'mbh', 'ug', 'gbr']:
        name = re.sub(re.escape(suffix), '', name, flags=re.IGNORECASE)
    name = re.sub(r'[^a-z0-9]', '', name)
    return name

def load_progress():
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE) as f:
            return json.load(f)
    return {"completed": [], "it_employers": [], "last_index": 0}

def save_progress(progress):
    with open(PROGRESS_FILE, 'w') as f:
        json.dump(progress, f, ensure_ascii=False, indent=1)

def git_commit_push(count, total):
    """Commit and push every COMMIT_EVERY entries."""
    try:
        subprocess.run(['git', 'add', '-A'], cwd=GIT_DIR, capture_output=True, timeout=30)
        msg = f"IT-Job Scraper: {count}/{total} Arbeitgeber gescannt ({datetime.now().strftime('%H:%M:%S')})"
        result = subprocess.run(['git', 'commit', '-m', msg], cwd=GIT_DIR, capture_output=True, timeout=30)
        if result.returncode == 0 or 'nothing to commit' not in result.stdout.decode():
            push_result = subprocess.run(['git', 'push', 'origin', 'main'], cwd=GIT_DIR, capture_output=True, timeout=60)
            print(f"  [GIT] Committed & pushed ({count}/{total})")
    except Exception as e:
        print(f"  [GIT] Error: {e}")

# ============================================================
# ARBEITSAGENTUR API - Check for IT jobs by employer name
# ============================================================

async def check_aa_it_jobs(session, company_name, city):
    """Check Arbeitsagentur API for IT jobs at this company."""
    results = []
    try:
        # Search for jobs at this company
        params = {
            "was": company_name,
            "wo": city,
            "umkreis": "25",
            "page": "1",
            "size": "10",
            "arbeitszeit": "vz,tz,mj",
        }
        headers = {
            "X-API-Key": AA_API_KEY,
            "Accept": "application/json",
        }
        async with session.get(AA_API_URL, params=params, headers=headers, timeout=aiohttp.ClientTimeout(total=15)) as resp:
            if resp.status == 200:
                data = await resp.json()
                jobs = data.get("stellenangebote", [])
                for job in jobs:
                    title = job.get("beruf", "").lower()
                    employer = job.get("arbeitgeber", "").lower()
                    # Check if this job is actually from this employer
                    norm_company = normalize_name(company_name)
                    norm_employer = normalize_name(employer)
                    if norm_company[:5] in norm_employer or norm_employer[:5] in norm_company:
                        # Check if IT-related
                        is_it = any(kw in title for kw in IT_KEYWORDS)
                        if is_it:
                            results.append({
                                "titel": job.get("beruf", ""),
                                "refnr": job.get("refnr", ""),
                                "ort": job.get("arbeitsort", {}).get("ort", ""),
                                "typ": job.get("arbeitszeit", ""),
                                "url": f"https://www.arbeitsagentur.de/jobsuche/jobdetail/{job.get('refnr', '')}",
                            })
            else:
                pass  # API rate limit or error
    except Exception as e:
        pass  # Timeout, connection error etc.
    
    await asyncio.sleep(0.3)  # Rate limiting
    return results

# ============================================================
# WEB SEARCH - Find website and check for IT jobs
# ============================================================

async def search_company_website(session, company_name, city):
    """Search DuckDuckGo for company website."""
    try:
        query = f"{company_name} {city} website"
        url = f"https://html.duckduckgo.com/html/?q={query.replace(' ', '+')}"
        async with session.get(url, timeout=aiohttp.ClientTimeout(total=12)) as resp:
            if resp.status == 200:
                html = await resp.text()
                soup = BeautifulSoup(html, 'html.parser')
                results = soup.find_all('a', class_='result__a')
                for r in results[:5]:
                    href = r.get('href', '')
                    # Extract actual URL from DDG redirect
                    match = re.search(r'uddg=([^&]+)', href)
                    if match:
                        actual_url = match.group(1)
                        # Skip social media and directories
                        skip_domains = ['facebook.com', 'linkedin.com', 'instagram.com', 
                                       'twitter.com', 'youtube.com', 'wikipedia.org',
                                       'gelbeseiten.de', 'dasoertliche.de', 'meinestadt.de',
                                       'kununu.com', 'yelp.de', 'google.com']
                        if not any(d in actual_url.lower() for d in skip_domains):
                            return actual_url
                # Fallback: try result URLs directly
                for r in results[:3]:
                    href = r.get('href', '')
                    if href.startswith('http'):
                        return href
    except:
        pass
    return None

async def download_html_content(session, url, label=""):
    """Download HTML content from URL."""
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "de-DE,de;q=0.9,en;q=0.5",
        }
        async with session.get(url, headers=headers, timeout=aiohttp.ClientTimeout(total=15), allow_redirects=True, ssl=False) as resp:
            if resp.status == 200:
                content = await resp.text()
                return content
    except:
        pass
    return None

def save_html(content, company_name, suffix=""):
    """Save HTML to file and return path."""
    if not content:
        return None, None
    safe_name = sanitize_filename(company_name)
    if suffix:
        filename = f"{safe_name}_{suffix}.html"
    else:
        filename = f"{safe_name}.html"
    filepath = os.path.join(HTML_DIR, filename)
    os.makedirs(HTML_DIR, exist_ok=True)
    with open(filepath, 'w', encoding='utf-8', errors='replace') as f:
        f.write(content)
    html_hash = hashlib.sha256(content.encode()).hexdigest()[:16]
    return filepath, html_hash

# ============================================================
# WEBSITE IT-JOB CHECK
# ============================================================

def check_website_for_it_jobs(html_content):
    """Check if website content mentions IT jobs / career pages."""
    if not html_content:
        return False, [], ""
    
    text = html_content.lower()
    
    # Check for career/karriere page links
    career_patterns = ['karriere', 'career', 'jobs', 'stellenangebote', 'stellenausschreibung',
                       'bewerbung', 'vacancies', 'offene stellen', 'wir suchen']
    career_found = []
    for p in career_patterns:
        if p in text:
            career_found.append(p)
    
    # Check for IT-specific job mentions
    it_found = []
    for kw in IT_KEYWORDS:
        if kw in text:
            it_found.append(kw)
    
    # More specific IT job indicators
    it_job_indicators = [
        'it-stellen', 'it jobs', 'it-jobs', 'softwareentwickler gesucht',
        'entwickler m/w/d', 'developer m/w/d', 'it specialist',
        'informatiker gesucht', 'wir suchen entwickler', 'wir suchen developer',
        'it-mitarbeiter', 'systemadministrator gesucht', 'devops engineer',
        'stellenangebot it', 'stellenangebot software', 'job it', 'job software',
    ]
    
    has_it_jobs = False
    specific_matches = []
    for indicator in it_job_indicators:
        if indicator in text:
            has_it_jobs = True
            specific_matches.append(indicator)
    
    # Also strong signal: IT keywords + career page
    if it_found and career_found:
        has_it_jobs = True
    
    return has_it_jobs, it_found[:10], ", ".join(specific_matches[:5] if specific_matches else career_found[:3])

# ============================================================
# IMPRESSUM / KONTAKT SCRAPING
# ============================================================

def extract_contact_info(html_content):
    """Extract email, phone, address from Impressum."""
    if not html_content:
        return {}, None
    
    info = {}
    text = html_content
    
    # Email
    emails = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', text)
    # Filter out common non-business emails
    emails = [e for e in emails if not any(x in e.lower() for x in ['example.com', 'sentry', 'webpack', 'noreply', 'wixpress'])]
    # Also check mailto: links
    mailtos = re.findall(r'mailto:([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})', text)
    emails = list(set(emails + mailtos))
    if emails:
        info['email'] = emails[0]
        info['alle_emails'] = "; ".join(emails[:5])
    
    # Phone
    phones = re.findall(r'(\+49[\s-/]*[\d\s-/]{8,})', text)
    if phones:
        info['telefon'] = phones[0].strip()
    
    # Geschäftsführer
    gf_match = re.search(r'(?:Geschäftsführer|Geschaeftsfuehrer|GF|Vorstand)[\s:]*(.+?)(?:<|\.|,|$)', text)
    if gf_match:
        gf = BeautifulSoup(gf_match.group(1), 'html.parser').get_text(strip=True)
        if len(gf) > 2:
            info['geschaeftsfuehrer'] = gf[:200]
    
    # Handelsregister
    hr_match = re.search(r'(?:Handelsregister|HRB|HR-Aktenzeichen|Amtsgericht)[\s:]*(.+?)(?:<|\.|,|$)', text)
    if hr_match:
        hr = BeautifulSoup(hr_match.group(1), 'html.parser').get_text(strip=True)
        info['handelsregister'] = hr[:200]
    
    # Impressum URL
    soup = BeautifulSoup(text, 'html.parser')
    impressum_link = None
    for a in soup.find_all('a', href=True):
        href = a['href'].lower()
        text_link = a.get_text(strip=True).lower()
        if 'impressum' in href or 'imprint' in href or 'impressum' in text_link:
            impressum_link = a['href']
            break
    
    # Karriere URL
    karriere_link = None
    for a in soup.find_all('a', href=True):
        href = a['href'].lower()
        text_link = a.get_text(strip=True).lower()
        career_words = ['karriere', 'career', 'jobs', 'stellenangebote', 'bewerbung']
        if any(w in href for w in career_words) or any(w in text_link for w in career_words):
            karriere_link = a['href']
            break
    
    return info, impressum_link, karriere_link

# ============================================================
# IT RELEVANCE SCORING
# ============================================================

def calculate_it_relevance(employer_data):
    """Calculate IT relevance score 0-100."""
    score = 0.0
    
    # Has IT jobs via Arbeitsagentur (strong signal)
    if employer_data.get('hat_it_jobs'):
        score += 40
        anzahl = employer_data.get('it_job_anzahl', 0)
        score += min(anzahl * 5, 15)  # up to +15 for multiple IT jobs
    
    # IT keywords on website
    if employer_data.get('it_keywords_gefunden'):
        kws = employer_data.get('it_keywords_gefunden', '')
        kw_count = len(kws.split(', ')) if kws else 0
        score += min(kw_count * 3, 20)
    
    # Career page with IT mentions
    if employer_data.get('karriere_url'):
        score += 10
    
    # Company category IT-related
    kat = (employer_data.get('kategorie_google') or '').lower()
    it_cats = ['software', 'it', 'informatik', 'telekommunikation', 'computer', 'technologie', 'digital', 'internet']
    if any(c in kat for c in it_cats):
        score += 15
    
    # Company name IT-related
    name = (employer_data.get('firmenname') or '').lower()
    it_name_parts = ['it ', 'software', 'tech', 'digital', 'data', 'cloud', 'cyber', 'computer', 'informatik', 'system']
    if any(p in name for p in it_name_parts):
        score += 10
    
    # Website has IT job indicators
    if employer_data.get('it_job_details'):
        score += 10
    
    return min(score, 100.0)

# ============================================================
# MAIN SCRAPING LOGIC
# ============================================================

async def process_employer(session, emp, semaphore):
    """Process a single employer: check for IT jobs, collect website, HTML."""
    async with semaphore:
        result = {
            **emp,
            'scrape_status': 'pending',
            'hat_it_jobs': 0,
            'it_job_titel': '',
            'it_job_anzahl': 0,
            'it_job_quelle': '',
            'it_job_details': '',
            'it_keywords_gefunden': '',
            'website_url': '',
            'website_html_path': '',
            'website_html_hash': '',
            'karriere_url': '',
            'karriere_html_path': '',
            'impressum_url': '',
            'impressum_html_path': '',
            'email': '',
            'geschaeftsfuehrer': '',
            'handelsregister': '',
            'it_relevanz_score': 0.0,
            'it_abteilung_vermutung': 0,
        }
        
        try:
            # 1. Check Arbeitsagentur for IT jobs
            aa_jobs = await check_aa_it_jobs(session, emp['name'], emp['ort'])
            if aa_jobs:
                result['hat_it_jobs'] = 1
                result['it_job_anzahl'] = len(aa_jobs)
                result['it_job_titel'] = "; ".join(j['titel'] for j in aa_jobs[:5])
                result['it_job_quelle'] = 'Arbeitsagentur'
                result['it_job_details'] = json.dumps(aa_jobs, ensure_ascii=False)
            
            # 2. Find website
            website = None
            if emp.get('website') and emp['website'].lower() == 'ja':
                # Need to search for actual URL
                website = await search_company_website(session, emp['name'], emp['ort'])
            elif emp.get('website') and emp['website'].startswith('http'):
                website = emp['website']
            else:
                website = await search_company_website(session, emp['name'], emp['ort'])
            
            if website:
                result['website_url'] = website
                
                # 3. Download main page HTML
                main_html = await download_html_content(session, website, "main")
                if main_html:
                    path, hash_val = save_html(main_html, emp['name'])
                    result['website_html_path'] = path
                    result['website_html_hash'] = hash_val
                    
                    # 4. Check website for IT jobs
                    has_it, it_kws, it_details = check_website_for_it_jobs(main_html)
                    if has_it:
                        result['hat_it_jobs'] = 1
                        if not result['it_job_quelle']:
                            result['it_job_quelle'] = 'Website'
                        result['it_keywords_gefunden'] = ", ".join(it_kws) if it_kws else ""
                        result['it_job_details'] = result.get('it_job_details', '') + f" | Website: {it_details}"
                    
                    # 5. Extract contact info
                    contact, impressum_url, karriere_url = extract_contact_info(main_html)
                    if contact.get('email'):
                        result['email'] = contact['email']
                    if contact.get('geschaeftsfuehrer'):
                        result['geschaeftsfuehrer'] = contact['geschaeftsfuehrer']
                    if contact.get('handelsregister'):
                        result['handelsregister'] = contact['handelsregister']
                    
                    # 6. Download Impressum page
                    if impressum_url:
                        # Make absolute URL
                        if impressum_url.startswith('/'):
                            from urllib.parse import urlparse
                            parsed = urlparse(website)
                            impressum_url = f"{parsed.scheme}://{parsed.netloc}{impressum_url}"
                        elif not impressum_url.startswith('http'):
                            impressum_url = f"{website.rstrip('/')}/{impressum_url.lstrip('/')}"
                        
                        result['impressum_url'] = impressum_url
                        imp_html = await download_html_content(session, impressum_url, "impressum")
                        if imp_html:
                            path, _ = save_html(imp_html, emp['name'], "impressum")
                            result['impressum_html_path'] = path
                            # Extract more contact info from Impressum
                            imp_contact, _, _ = extract_contact_info(imp_html)
                            if imp_contact.get('email') and not result['email']:
                                result['email'] = imp_contact['email']
                            if imp_contact.get('geschaeftsfuehrer') and not result['geschaeftsfuehrer']:
                                result['geschaeftsfuehrer'] = imp_contact['geschaeftsfuehrer']
                            if imp_contact.get('handelsregister') and not result['handelsregister']:
                                result['handelsregister'] = imp_contact['handelsregister']
                    
                    # 7. Download Karriere page
                    if karriere_url:
                        if karriere_url.startswith('/'):
                            from urllib.parse import urlparse
                            parsed = urlparse(website)
                            karriere_url = f"{parsed.scheme}://{parsed.netloc}{karriere_url}"
                        elif not karriere_url.startswith('http'):
                            karriere_url = f"{website.rstrip('/')}/{karriere_url.lstrip('/')}"
                        
                        result['karriere_url'] = karriere_url
                        kar_html = await download_html_content(session, karriere_url, "karriere")
                        if kar_html:
                            path, _ = save_html(kar_html, emp['name'], "karriere")
                            result['karriere_html_path'] = path
                            # Check karriere page for IT jobs too
                            has_it_kar, it_kws_kar, it_details_kar = check_website_for_it_jobs(kar_html)
                            if has_it_kar:
                                result['hat_it_jobs'] = 1
                                if not result['it_job_quelle']:
                                    result['it_job_quelle'] = 'Website/Karriere'
                                if it_kws_kar and not result['it_keywords_gefunden']:
                                    result['it_keywords_gefunden'] = ", ".join(it_kws_kar)
            
            # 8. Calculate IT relevance score
            result['it_relevanz_score'] = calculate_it_relevance(result)
            result['it_abteilung_vermutung'] = 1 if result['it_relevanz_score'] >= 30 else 0
            
            result['scrape_status'] = 'completed'
            result['scrape_datum'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            result['letzte_pruefung'] = datetime.now().strftime('%Y-%m-%d')
            
        except Exception as e:
            result['scrape_status'] = f'error: {str(e)[:100]}'
            traceback.print_exc()
        
        return result

# ============================================================
# DATABASE OPERATIONS
# ============================================================

def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.executescript(DB_SCHEMA)
    conn.commit()
    return conn

def save_to_db(conn, data):
    """Save employer data to SQLite."""
    columns = [
        'firmenname', 'firmenname_normalisiert', 'ort', 'plz', 'strasse',
        'adresse_vollstaendig', 'telefon', 'email', 'website_url',
        'website_html_path', 'website_html_hash', 'karriere_url',
        'karriere_html_path', 'impressum_url', 'impressum_html_path',
        'hat_it_jobs', 'it_job_titel', 'it_job_anzahl', 'it_job_quelle',
        'it_job_details', 'it_keywords_gefunden', 'kategorie_google',
        'bewertung_google', 'lat', 'lng', 'google_maps_link',
        'handelsregister', 'geschaeftsfuehrer', 'it_abteilung_vermutung',
        'it_relevanz_score', 'scrape_datum', 'scrape_status', 'letzte_pruefung'
    ]
    values = [
        data.get('firmenname', data.get('name', '')),
        normalize_name(data.get('firmenname', data.get('name', ''))),
        data.get('ort', ''),
        data.get('plz', ''),
        data.get('adresse', ''),
        data.get('adresse_vollstaendig', ''),
        data.get('telefon', ''),
        data.get('email', ''),
        data.get('website_url', ''),
        data.get('website_html_path', ''),
        data.get('website_html_hash', ''),
        data.get('karriere_url', ''),
        data.get('karriere_html_path', ''),
        data.get('impressum_url', ''),
        data.get('impressum_html_path', ''),
        data.get('hat_it_jobs', 0),
        data.get('it_job_titel', ''),
        data.get('it_job_anzahl', 0),
        data.get('it_job_quelle', ''),
        data.get('it_job_details', ''),
        data.get('it_keywords_gefunden', ''),
        data.get('kategorie_google', data.get('kategorie', '')),
        data.get('bewertung_google', None),
        data.get('lat', None),
        data.get('lng', None),
        data.get('google_maps_link', data.get('gmaps_link', '')),
        data.get('handelsregister', ''),
        data.get('geschaeftsfuehrer', ''),
        data.get('it_abteilung_vermutung', 0),
        data.get('it_relevanz_score', 0.0),
        data.get('scrape_datum', ''),
        data.get('scrape_status', ''),
        data.get('letzte_pruefung', ''),
    ]
    
    placeholders = ','.join(['?' for _ in columns])
    sql = f"INSERT OR REPLACE INTO employers ({','.join(columns)}) VALUES ({placeholders})"
    conn.execute(sql, values)
    conn.commit()

def db_to_excel(conn):
    """Export SQLite database to Excel."""
    wb = openpyxl.Workbook()
    
    # Main sheet
    ws = wb.active
    ws.title = "IT Arbeitgeber"
    
    # Get all data
    cursor = conn.execute("SELECT * FROM employers ORDER BY it_relevanz_score DESC, hat_it_jobs DESC, firmenname")
    columns = [desc[0] for desc in cursor.description]
    
    # Header
    header_map = {
        'id': 'ID', 'firmenname': 'Firmenname', 'firmenname_normalisiert': 'Firmenname (normalisiert)',
        'ort': 'Ort', 'plz': 'PLZ', 'strasse': 'Straße', 'adresse_vollstaendig': 'Adresse (vollständig)',
        'bundesland': 'Bundesland', 'land': 'Land', 'telefon': 'Telefon', 'email': 'E-Mail',
        'website_url': 'Website URL', 'website_html_path': 'Website HTML Pfad',
        'website_html_hash': 'Website HTML Hash', 'karriere_url': 'Karriere-URL',
        'karriere_html_path': 'Karriere HTML Pfad', 'impressum_url': 'Impressum URL',
        'impressum_html_path': 'Impressum HTML Pfad', 'hat_it_jobs': 'IT-Jobs vorhanden',
        'it_job_titel': 'IT-Job Titel', 'it_job_anzahl': 'Anzahl IT-Jobs',
        'it_job_quelle': 'IT-Job Quelle', 'it_job_details': 'IT-Job Details',
        'it_keywords_gefunden': 'IT-Keywords gefunden', 'branche': 'Branche',
        'kategorie_google': 'Kategorie (Google)', 'bewertung_google': 'Bewertung (Google)',
        'anzahl_bewertungen': 'Anzahl Bewertungen', 'lat': 'Breitengrad', 'lng': 'Längengrad',
        'google_maps_link': 'Google Maps Link', 'arbeitsagentur_id': 'Arbeitsagentur ID',
        'arbeitsagentur_name': 'Arbeitsagentur Name', 'firmengroesse': 'Firmengröße',
        'rechtsform': 'Rechtsform', 'gruendungsjahr': 'Gründungsjahr',
        'ust_id': 'USt-ID', 'handelsregister': 'Handelsregister',
        'geschaeftsfuehrer': 'Geschäftsführer', 'angestelltenzahl': 'Angestelltenzahl',
        'it_abteilung_vermutung': 'IT-Abteilung vermutet', 'it_relevanz_score': 'IT-Relevanz Score',
        'scrape_datum': 'Scrape-Datum', 'scrape_status': 'Scrape-Status',
        'letzte_pruefung': 'Letzte Prüfung', 'notizen': 'Notizen',
        'created_at': 'Erstellt am', 'updated_at': 'Aktualisiert am',
    }
    
    headers = [header_map.get(c, c) for c in columns]
    ws.append(headers)
    
    # Data rows
    for row in cursor:
        ws.append(list(row))
    
    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
    
    # Stats sheet
    ws2 = wb.create_sheet("Statistiken")
    stats = [
        ("Gesamt Arbeitgeber", conn.execute("SELECT COUNT(*) FROM employers").fetchone()[0]),
        ("Mit IT-Jobs", conn.execute("SELECT COUNT(*) FROM employers WHERE hat_it_jobs=1").fetchone()[0]),
        ("Mit Website", conn.execute("SELECT COUNT(*) FROM employers WHERE website_url != ''").fetchone()[0]),
        ("Mit E-Mail", conn.execute("SELECT COUNT(*) FROM employers WHERE email != ''").fetchone()[0]),
        ("IT-Abteilung vermutet", conn.execute("SELECT COUNT(*) FROM employers WHERE it_abteilung_vermutung=1").fetchone()[0]),
        ("Durchschn. IT-Relevanz", round(conn.execute("SELECT AVG(it_relevanz_score) FROM employers").fetchone()[0] or 0, 1)),
    ]
    for s in stats:
        ws2.append(s)
    
    # Per-city stats
    ws3 = wb.create_sheet("Nach Ort")
    ws3.append(["Ort", "Gesamt", "Mit IT-Jobs", "Mit Website", "Mit E-Mail", "Ø IT-Relevanz"])
    cursor2 = conn.execute("""
        SELECT ort, COUNT(*), 
               SUM(CASE WHEN hat_it_jobs=1 THEN 1 ELSE 0 END),
               SUM(CASE WHEN website_url != '' THEN 1 ELSE 0 END),
               SUM(CASE WHEN email != '' THEN 1 ELSE 0 END),
               ROUND(AVG(it_relevanz_score), 1)
        FROM employers GROUP BY ort ORDER BY COUNT(*) DESC
    """)
    for row in cursor2:
        ws3.append(list(row))
    
    wb.save(EXCEL_OUT)
    print(f"Excel saved to {EXCEL_OUT}")

# ============================================================
# MAIN
# ============================================================

async def main():
    print("=" * 60)
    print("IT-Job Scraper: Bamberg, Erlangen, Nürnberg")
    print("=" * 60)
    
    # Load employers from Excel
    wb = openpyxl.load_workbook(EXCEL_IN)
    ws = wb['Google Maps Firmen']
    
    employers = []
    seen = set()
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        ort = str(row[6] or '').strip()
        name = str(row[1] or '').strip()
        if not name:
            continue
        
        # Filter for target cities
        matched_city = None
        for city in TARGET_CITIES:
            if city.lower() in ort.lower():
                matched_city = city
                break
        if not matched_city:
            continue
        
        # Dedup
        norm = normalize_name(name)
        if norm in seen or len(norm) < 3:
            continue
        seen.add(norm)
        
        employers.append({
            'name': name,
            'ort': ort,
            'plz': str(row[5] or '').strip(),
            'adresse': str(row[4] or '').strip(),
            'telefon': str(row[7] or '').strip(),
            'website_flag': str(row[8] or '').strip(),  # "Ja" / "Nein"
            'kategorie': str(row[3] or '').strip(),
            'bewertung': row[2],
            'lat': row[9],
            'lng': row[10],
            'gmaps_link': str(row[11] or '').strip(),
            'suchort': str(row[12] or '').strip(),
        })
    
    print(f"Gefunden: {len(employers)} eindeutige Arbeitgeber in {TARGET_CITIES}")
    
    # Load progress
    progress = load_progress()
    completed_names = set(progress.get('completed', []))
    remaining = [e for e in employers if e['name'] not in completed_names]
    print(f"Bereits gescannt: {len(completed_names)}, Verbleibend: {len(remaining)}")
    
    # Init DB
    conn = init_db()
    
    # Create HTML dir
    os.makedirs(HTML_DIR, exist_ok=True)
    
    # Process in batches with concurrency
    semaphore = asyncio.Semaphore(MAX_CONCURRENT)
    batch_size = COMMIT_EVERY
    total = len(remaining)
    processed = 0
    it_count = 0
    
    connector = aiohttp.TCPConnector(limit=MAX_CONCURRENT, limit_per_host=2, ssl=False)
    async with aiohttp.ClientSession(connector=connector) as session:
        # Process in chunks
        for chunk_start in range(0, len(remaining), batch_size):
            chunk = remaining[chunk_start:chunk_start + batch_size]
            tasks = [process_employer(session, emp, semaphore) for emp in chunk]
            results = await asyncio.gather(*tasks, return_exceptions=True)
            
            for i, res in enumerate(results):
                if isinstance(res, Exception):
                    print(f"  ERROR: {chunk[i]['name']}: {res}")
                    continue
                
                save_to_db(conn, res)
                processed += 1
                
                # Update progress
                progress['completed'].append(res.get('firmenname', res.get('name', '')))
                if res.get('hat_it_jobs'):
                    it_count += 1
                
                status = "✓ IT-JOBS" if res.get('hat_it_jobs') else "  no IT"
                website = res.get('website_url', '')[:40] if res.get('website_url') else '-'
                print(f"  [{processed}/{total}] {status} | {res.get('firmenname', res.get('name', ''))[:50]} | {website}")
            
            # Save progress
            save_progress(progress)
            
            # Commit & push every batch
            print(f"\n  --- Commit & Push (bisher {it_count} IT-Arbeitgeber) ---")
            git_commit_push(processed, total)
    
    # Generate final Excel
    print("\nErstelle finale Excel...")
    db_to_excel(conn)
    
    # Final commit & push
    subprocess.run(['git', 'add', '-A'], cwd=GIT_DIR, capture_output=True)
    msg = f"IT-Job Scraper ABGESCHLOSSEN: {it_count} IT-Arbeitgeber aus {total} gescannt"
    subprocess.run(['git', 'commit', '-m', msg], cwd=GIT_DIR, capture_output=True)
    subprocess.run(['git', 'push', 'origin', 'main'], cwd=GIT_DIR, capture_output=True, timeout=60)
    
    conn.close()
    print(f"\n{'=' * 60}")
    print(f"FERTIG! {it_count} IT-Arbeitgeber gefunden aus {total} gescannten")
    print(f"Excel: {EXCEL_OUT}")
    print(f"HTML: {HTML_DIR}/ ({len(os.listdir(HTML_DIR))} Dateien)")
    print(f"DB: {DB_PATH}")
    print(f"{'=' * 60}")

if __name__ == '__main__':
    asyncio.run(main())
