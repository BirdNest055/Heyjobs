#!/usr/bin/env python3
"""
Job Application Scraper v5 - Batch mode.
Processes a small batch of companies, saves progress, exits.
Designed to be called repeatedly: python3 job_application_scraper_v5.py [BATCH_SIZE]
Resumes from saved progress automatically.
"""

import asyncio
import json
import os
import re
import subprocess
import sys
from datetime import datetime
from urllib.parse import urlparse

import openpyxl

# ── CONFIG ────────────────────────────────────────────────────────────────────
EXCEL_OUT = "/home/z/my-project/download/Bewerbungen_Junior_IT.xlsx"
PROGRESS_FILE = "/home/z/my-project/job_scraper_v5_progress.json"
LOG_FILE = "/home/z/my-project/job_scraper_v5_log.txt"
GITHUB_REPO = "/home/z/my-project"
DISPLAY = ":99"
DEFAULT_BATCH = 3
MAX_JOBS_PER_COMPANY = 5
TIMEOUT_PER_COMPANY = 90

EMAIL_BODY = """Sehr geehrte{anrede_suffix} {anrede_target},

ich sende Ihnen hiermit meine Bewerbungsunterlagen für die Stelle als {stelle} in Vollzeit.

Ich schließe meine Ausbildung zum Fachinformatiker für Systemintegration bald ab. Das Fachgespräch findet am 11. Juni 2026 statt, ab diesem Zeitpunkt stehe ich für eine Einstellung zur Verfügung.

Während meiner Ausbildung hatte ich Berührungspunkte mit verschiedenen IT-Bereichen: Systemintegration, Webentwicklung und Datenbankentwicklung.

In meiner letzten Praxisphase habe ich mich mit Azure beschäftigt, was direkt in mein IHK-Abschlussprojekt geflossen ist: die automatisierte Anwendungsbereitstellung via App Attach in Azure Virtual Desktop, von der Planung bis zur Dokumentation eigenständig umgesetzt.

Ich arbeite selbstständig und kommuniziere offen. Das haben mir auch meine Betreuer in den Praxisphasen zurückgemeldet. Einen Führerschein der Klasse B besitze ich aktuell nicht, bin aber bereit, diesen zeitnah zu erwerben.

Meine vollständigen Unterlagen finden Sie im Anhang. Ich freue mich auf ein persönliches Gespräch.

Mit freundlichen Grüßen,"""


def log(msg):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line, flush=True)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")


def load_progress():
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"processed": [], "jobs": [], "last_index": 0}


def save_progress(progress):
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump(progress, f, ensure_ascii=False, indent=2)


def git_push(msg):
    try:
        subprocess.run(["git", "add", "-A"], cwd=GITHUB_REPO, capture_output=True, timeout=30)
        subprocess.run(["git", "commit", "-m", msg], cwd=GITHUB_REPO, capture_output=True, timeout=30)
        result = subprocess.run(["git", "push", "origin", "main"], cwd=GITHUB_REPO,
                                capture_output=True, text=True, timeout=60)
        if result.returncode == 0:
            log(f"✅ Pushed: {msg}")
        else:
            log(f"⚠️ Push issue: {result.stderr[:150]}")
    except Exception as e:
        log(f"⚠️ Git error: {e}")


def customize_email(stelle, firma, ansprechpartner=""):
    if ansprechpartner and ansprechpartner.lower() not in ["n/a", "unbekannt", "", "damen und herren"]:
        if any(t in ansprechpartner.lower() for t in ['frau', 'ms.', 'mrs.']):
            anrede_suffix = ""
            clean = ansprechpartner.replace('Frau', '').replace('Mrs.', '').replace('Ms.', '').strip()
            anrede_target = f"Frau {clean}"
        elif any(t in ansprechpartner.lower() for t in ['herr', 'hr.', 'mr.']):
            anrede_suffix = "r"
            clean = ansprechpartner.replace('Herr', '').replace('Hr.', '').replace('Mr.', '').strip()
            anrede_target = f"Herr {clean}"
        else:
            anrede_suffix = "r"
            anrede_target = ansprechpartner
    else:
        anrede_suffix = ""
        anrede_target = "Damen und Herren"
    return EMAIL_BODY.format(anrede_suffix=anrede_suffix, anrede_target=anrede_target, stelle=stelle)


def extract_emails(text):
    if not text:
        return []
    return list(set(re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', text)))


IT_KEYWORDS = ['it-', 'it ', 'informatik', 'software', 'systemintegration', 'netzwerk',
    'devops', 'cloud', 'data ', 'datenbank', 'cyber', 'security', 'entwickler',
    'developer', 'administrator', 'admin', 'linux', 'azure', 'aws ',
    'frontend', 'backend', 'fullstack', 'full-stack', 'infrastruktur', 'support',
    'consultant', 'engineer', 'digital', 'scrum', 'projektmanag', 'technology',
    'technolog', 'digitalisier', 'comput', 'automation', 'automatisier']

JUNIOR_KEYWORDS = ['junior', 'trainee', 'entry', 'graduate', 'berufseinsteig',
    'werkstudent', 'praktikant', 'azubi', 'ausbildung', 'duales studium']

SENIOR_KEYWORDS = ['senior', 'lead', 'director', 'head of', 'vp ', 'chief', 'principal']


def is_it_job(text):
    return any(kw in text.lower() for kw in IT_KEYWORDS)

def is_junior_job(text):
    tl = text.lower()
    if any(kw in tl for kw in JUNIOR_KEYWORDS): return "junior"
    if any(kw in tl for kw in SENIOR_KEYWORDS): return "senior"
    return "open"


def load_companies():
    companies = {}
    wb1 = openpyxl.load_workbook("/home/z/my-project/download/Google_Maps_Firmen_Erlangen_50km.xlsx")
    ws1 = wb1["Google Maps Firmen"]
    for r in range(2, ws1.max_row + 1):
        name = (ws1.cell(r, 2).value or "").strip()
        if not name: continue
        if name not in companies:
            companies[name] = {"name": name, "kategorie": ws1.cell(r, 5).value or "",
                "adresse": ws1.cell(r, 6).value or "", "plz": ws1.cell(r, 7).value or "",
                "ort": ws1.cell(r, 8).value or "", "telefon": ws1.cell(r, 9).value or "",
                "website": ws1.cell(r, 10).value or "", "email": "", "ansprechpartner": ""}
        elif ws1.cell(r, 10).value and not companies[name]["website"]:
            companies[name]["website"] = ws1.cell(r, 10).value

    wb2 = openpyxl.load_workbook("/home/z/my-project/download/Firmen_Bamberg_Erlangen_Nuernberg_mit_Website_HTML.xlsx")
    ws2 = wb2["Firmen mit Website & HTML"]
    for r in range(2, ws2.max_row + 1):
        name = (ws2.cell(r, 2).value or "").strip()
        if not name: continue
        web_found = ws2.cell(r, 10).value or ""
        email = ws2.cell(r, 11).value or ""
        if name in companies:
            if web_found and not companies[name]["website"]: companies[name]["website"] = web_found
            if email and not companies[name]["email"]: companies[name]["email"] = email
        else:
            companies[name] = {"name": name, "kategorie": ws2.cell(r, 4).value or "",
                "adresse": ws2.cell(r, 5).value or "", "plz": ws2.cell(r, 7).value or "",
                "ort": ws2.cell(r, 8).value or "", "telefon": ws2.cell(r, 9).value or "",
                "website": web_found, "email": email, "ansprechpartner": ""}

    wb3 = openpyxl.load_workbook("/home/z/my-project/download/Junior_IT_Jobs_Bamberg_Erlangen_Nuernberg.xlsx")
    ws3 = wb3["Junior IT-Jobs"]
    aa_jobs = {}
    for r in range(2, ws3.max_row + 1):
        employer = (ws3.cell(r, 22).value or "").strip()
        job_title = ws3.cell(r, 2).value or ""
        job_url = ws3.cell(r, 27).value or ""
        if employer: aa_jobs.setdefault(employer, []).append({"title": job_title, "url": job_url})

    scoring_kws = IT_KEYWORDS + ['maschinenbau', 'industrie', 'elektro', 'bank', 'versicher',
        'finanz', 'klinik', 'universität', 'forschung', 'gruppe', 'holding',
        'telekommunikat', 'logistik', 'handel', 'consult', 'berat']

    scored = []
    for name, data in companies.items():
        if not data["website"]: continue
        score = sum(1 for kw in scoring_kws if kw in f"{name} {data['kategorie']}".lower())
        if name in aa_jobs: score += 5
        scored.append((score, name, data))

    scored.sort(key=lambda x: -x[0])
    result = []
    for score, name, data in scored:
        data["relevance_score"] = score
        data["arbeitsagentur_jobs"] = aa_jobs.get(name, [])
        result.append(data)

    return result


class JobScraper:
    def __init__(self, page):
        self.page = page

    async def accept_cookies(self):
        for sel in ['button:has-text("Akzeptieren")', 'button:has-text("Alle akzeptieren")',
            'button:has-text("Accept")', 'button:has-text("OK")',
            'button:has-text("Verstanden")', 'button:has-text("Zustimmen")',
            '#onetrust-accept-btn-handler', '.cc-btn']:
            try:
                btn = self.page.locator(sel).first
                if await btn.is_visible(timeout=400):
                    await btn.click()
                    await self.page.wait_for_timeout(600)
                    return True
            except: continue
        return False

    async def safe_goto(self, url, timeout=20000):
        try:
            await self.page.goto(url, wait_until="domcontentloaded", timeout=timeout)
            await self.page.wait_for_timeout(2000)
            try: await self.accept_cookies()
            except: pass
            return True
        except: return False

    async def inject_and_extract_jobs(self):
        """Inject JS and extract jobs in one call to minimize evaluate calls."""
        try:
            return await self.page.evaluate("""() => {
                const jobWords = ['entwickler', 'developer', 'administrator', 'engineer',
                    'berater', 'consultant', 'trainee', 'junior', 'senior', 'azubi',
                    'praktikant', 'werkstudent', 'ingenieur', 'fachinformatiker',
                    'informatiker', 'programmierer', 'sachbearbeiter', 'assistent',
                    'techniker', 'disponent', 'referent', 'controller', 'experte',
                    'spezialist', 'specialist', 'scrum master', 'product owner', 'devops',
                    'm/w/d', 'w/m/d', 'm/f/d', 'f/m/d', 'm/w', 'w/m', 'architekt',
                    'architect', 'analyst', 'operator', 'manager', 'head', 'director',
                    'koordinator', 'coordinator', 'projektmanager', 'project manager',
                    'kaufmann', 'kauffrau', 'laborant', 'forscher', 'researcher'];
                function looksLikeJob(text) {
                    const tl = text.toLowerCase();
                    return jobWords.some(kw => tl.includes(kw));
                }
                const els = document.querySelectorAll('a, h2, h3, h4, h5, li, article, tr');
                const results = [];
                const seen = new Set();
                for (const el of els) {
                    const text = (el.textContent || '').trim();
                    if (text.length < 10 || text.length > 1500) continue;
                    const lines = text.split(String.fromCharCode(10)).map(l => l.trim()).filter(l => l.length > 3);
                    const title = (lines[0] || text.substring(0, 150)).substring(0, 150);
                    if (!looksLikeJob(title)) continue;
                    const key = title.toLowerCase().substring(0, 50);
                    if (seen.has(key)) continue;
                    seen.add(key);
                    const link = el.closest('a') || el.querySelector('a');
                    const href = link ? link.href : (el.tagName === 'A' ? el.href : '');
                    results.push({title: title, href: href || '', text: text.substring(0, 500)});
                }
                return results;
            }""")
        except:
            return []

    async def get_career_links(self, base_url):
        try:
            return await self.page.evaluate("""() => {
                const keywords = ['karriere', 'jobs', 'stellen', 'career', 'bewer', 'offene', 'vacanc', 'position'];
                const results = [];
                const seen = new Set();
                document.querySelectorAll('a[href]').forEach(a => {
                    const text = (a.textContent || '').trim().toLowerCase();
                    const href = (a.getAttribute('href') || '').toLowerCase();
                    const isMatch = keywords.some(kw => text.includes(kw) || href.includes(kw));
                    if (isMatch && a.href && !seen.has(a.href)) {
                        seen.add(a.href);
                        results.push({href: a.href, text: (a.textContent || '').trim().substring(0, 100)});
                    }
                });
                return results;
            }""")
        except:
            return []

    async def extract_jobs_from_page(self):
        raw_jobs = await self.inject_and_extract_jobs()
        jobs = []
        for item in raw_jobs:
            title = item.get("title", "").strip()
            if len(title) < 5: continue
            if not is_it_job(title + " " + item.get("text", "")[:200]): continue
            level = is_junior_job(title)
            if level == "senior": continue
            title = " ".join(title.split())
            jobs.append({"title": title, "url": item.get("href", ""),
                "description": item.get("text", "")[:500], "is_junior": level == "junior"})
        return jobs

    async def find_contact_info(self, base_url):
        email, ansprechpartner = "", ""
        for path in ["/impressum", "/kontakt", "/contact"]:
            try:
                if not await self.safe_goto(base_url + path, timeout=12000): continue
                await self.page.wait_for_timeout(1500)
                text = await self.page.inner_text("body")
                emails = extract_emails(text)
                for em in emails:
                    if not any(g in em.lower() for g in ['noreply', 'no-reply', 'spam', 'example', 'github']):
                        email = em; break
                if not email and emails: email = emails[0]
                for line in text.split("\n"):
                    line = line.strip()
                    ll = line.lower()
                    if ('herr' in ll or 'frau' in ll) and any(t in ll for t in ['personal', 'hr', 'bewer', 'kontakt', 'ansprech']):
                        ansprechpartner = line[:80]; break
                if email: break
            except: continue
        return email, ansprechpartner

    async def scrape_company(self, company):
        url = company["website"]
        if not url: return [], company.get("email", ""), company.get("ansprechpartner", "")
        url = url.strip()
        if not url.startswith(("http://", "https://")): url = "https://" + url
        url = url.rstrip("/")
        base_url = f"{urlparse(url).scheme}://{urlparse(url).netloc}"
        jobs_found = []
        contact_email = company.get("email", "")
        ansprechpartner = company.get("ansprechpartner", "")

        # Visit main page
        log(f"  🌐 Main: {url}")
        if not await self.safe_goto(url, timeout=25000):
            return [], contact_email, ansprechpartner

        # Find career links
        career_links = await self.get_career_links(base_url)
        parsed = urlparse(base_url)
        domain_base = f"{parsed.scheme}://{parsed.netloc}"
        for path in ["/karriere", "/jobs", "/stellenangebote", "/career", "/careers",
                     "/de/karriere", "/offene-stellen", "/bewerbung", "/vacancies"]:
            full = domain_base + path
            if not any(l["href"] == full for l in career_links):
                career_links.append({"href": full, "text": path})
        log(f"  🔗 {len(career_links)} career links")

        # Visit career pages
        visited = set()
        for cl in career_links[:6]:
            career_url = cl["href"]
            if career_url in visited: continue
            visited.add(career_url)
            log(f"  📄 {career_url[:60]}")
            if not await self.safe_goto(career_url, timeout=18000): continue
            await self.page.wait_for_timeout(2500)
            page_jobs = await self.extract_jobs_from_page()
            if page_jobs:
                log(f"  ✅ {len(page_jobs)} IT jobs found")
                jobs_found.extend(page_jobs)
                if len(jobs_found) >= MAX_JOBS_PER_COMPANY: break

        # Contact info
        if not contact_email or not ansprechpartner:
            found_email, found_ansprech = await self.find_contact_info(base_url)
            if not contact_email and found_email: contact_email = found_email
            if not ansprechpartner and found_ansprech: ansprechpartner = found_ansprech

        # Arbeitsagentur jobs
        for aa_job in company.get("arbeitsagentur_jobs", []):
            title = aa_job["title"]
            level = is_junior_job(title)
            if level == "senior": continue
            existing = {j["title"].lower()[:40] for j in jobs_found}
            if title.lower()[:40] not in existing:
                jobs_found.append({"title": title, "url": aa_job.get("url", ""),
                    "description": "Quelle: Arbeitsagentur", "is_junior": level == "junior"})

        # Deduplicate
        seen = set()
        unique = []
        for j in jobs_found:
            key = " ".join(j["title"].lower().split())[:50]
            if key not in seen:
                seen.add(key)
                j["email"] = contact_email
                j["ansprechpartner"] = ansprechpartner
                unique.append(j)
            if len(unique) >= MAX_JOBS_PER_COMPANY: break
        return unique, contact_email, ansprechpartner


async def run_batch(batch_size):
    log(f"Starting batch of {batch_size} companies")
    companies = load_companies()
    progress = load_progress()
    processed_names = set(progress["processed"])
    jobs_all = progress["jobs"]

    to_process = [(i, c) for i, c in enumerate(companies)
                  if c["name"] not in processed_names and c["website"]]
    to_process = to_process[:batch_size]

    if not to_process:
        log("Nothing to process!")
        save_excel(jobs_all)
        return

    from playwright.async_api import async_playwright
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False,
            args=["--no-sandbox", "--disable-dev-shm-usage", "--window-size=1920,1080"])
        context = await browser.new_context(viewport={"width": 1920, "height": 1080},
            user_agent="Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
            locale="de-DE", ignore_https_errors=True)
        page = await context.new_page()
        page.set_default_timeout(25000)
        page.set_default_navigation_timeout(25000)

        scraper = JobScraper(page)
        processed_count = 0
        total_new = 0

        for idx, (orig_idx, company) in enumerate(to_process):
            name = company["name"]
            log(f"\n[{idx+1}/{len(to_process)}] {name} (score: {company.get('relevance_score', 0)})")
            try:
                jobs, email, ansprech = await asyncio.wait_for(
                    scraper.scrape_company(company), timeout=TIMEOUT_PER_COMPANY)
            except asyncio.TimeoutError:
                log(f"  ⏰ Timeout"); jobs, email, ansprech = [], "", ""
            except Exception as e:
                import traceback
                log(f"  ❌ Error: {type(e).__name__}: {str(e)[:60]}")
                log(traceback.format_exc()[:300])
                jobs, email, ansprech = [], "", ""

            if jobs:
                log(f"  ✅ Found {len(jobs)} IT job(s)!")
                for job in jobs:
                    email_text = customize_email(stelle=job["title"], firma=name,
                        ansprechpartner=job.get("ansprechpartner", ansprech))
                    adresse = company.get("adresse", "")
                    if company.get("plz"): adresse += f", {company['plz']}"
                    if company.get("ort"): adresse += f" {company['ort']}"
                    jobs_all.append({
                        "firma": name, "stelle": job["title"],
                        "stellenbeschreibung": job.get("description", ""),
                        "email": job.get("email", email or company.get("email", "")),
                        "ansprechpartner": job.get("ansprechpartner", ansprech),
                        "bewerbung_geschickt": "", "letzter_kontakt": "",
                        "status": "Offen", "website": company["website"],
                        "adresse": adresse.strip(", "), "email_text": email_text,
                        "job_url": job.get("url", ""),
                        "is_junior": job.get("is_junior", False),
                        "scraped_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    })
                    total_new += 1
            else:
                log(f"  ℹ️ No IT jobs found")

            processed_names.add(name)
            processed_count += 1
            progress["processed"] = list(processed_names)
            progress["jobs"] = jobs_all
            progress["last_index"] = orig_idx
            save_progress(progress)
            await page.wait_for_timeout(500)

        await browser.close()

    log(f"Batch done! {processed_count} companies, {total_new} new jobs, {len(jobs_all)} total")
    save_excel(jobs_all)
    git_push(f"Job Scraper v5: {processed_count} co | {len(jobs_all)} jobs | {datetime.now().strftime('%m-%d %H:%M')}")


def save_excel(jobs):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bewerbungen"
    headers = ["Firma", "Stelle", "Stellenbeschreibung", "Email", "Ansprechpartner",
        "Bewerbung geschickt am", "Letzter Kontakt am", "Status", "Website",
        "Adresse", "Email Text", "Job-URL", "Junior?", "Scraped At"]

    hdr_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h); c.font = hdr_font; c.fill = hdr_fill; c.alignment = hdr_align; c.border = border
    for row, job in enumerate(jobs, 2):
        vals = [job.get("firma", ""), job.get("stelle", ""), job.get("stellenbeschreibung", ""),
            job.get("email", ""), job.get("ansprechpartner", ""), job.get("bewerbung_geschickt", ""),
            job.get("letzter_kontakt", ""), job.get("status", "Offen"), job.get("website", ""),
            job.get("adresse", ""), job.get("email_text", ""), job.get("job_url", ""),
            "Ja" if job.get("is_junior") else "Nein", job.get("scraped_at", "")]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row, col, val); c.border = border; c.alignment = Alignment(vertical="top", wrap_text=True)

    for col, w in {1:30, 2:40, 3:50, 4:30, 5:22, 6:18, 7:18, 8:14, 9:35, 10:35, 11:70, 12:40, 13:10, 14:16}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(jobs)+1}"

    dv = DataValidation(type="list", formula1='"Offen,Beworben,Interview eingeplant,Interview absolviert,Zusage,Absage,Kontaktiert,Wiedervorlage,Kein Interesse"', allow_blank=True)
    ws.add_data_validation(dv)
    for r in range(2, len(jobs)+2): dv.add(ws.cell(r, 8))

    lr = len(jobs) + 1
    for s, f in [("Zusage","C6EFCE"),("Absage","FFC7CE"),("Beworben","BDD7EE"),("Interview eingeplant","FFEB9C"),("Interview absolviert","F4B084")]:
        ws.conditional_formatting.add(f"H2:H{lr}", CellIsRule(operator="equal", formula=[f'"{s}"'],
            fill=PatternFill(start_color=f, end_color=f, fill_type="solid")))
    ws.conditional_formatting.add(f"M2:M{lr}", CellIsRule(operator="equal", formula=['"Ja"'],
        fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")))

    # Email Vorlagen
    ws2 = wb.create_sheet("Email Vorlagen")
    for col, h in enumerate(["Firma", "Stelle", "Email an", "Betreff", "Email Text"], 1):
        ws2.cell(1, col, h).font = Font(bold=True)
    for row, job in enumerate(jobs, 2):
        ws2.cell(row, 1, job.get("firma", "")); ws2.cell(row, 2, job.get("stelle", ""))
        ws2.cell(row, 3, job.get("email", ""))
        ws2.cell(row, 4, f"Bewerbung als {job.get('stelle', 'Fachinformatiker Systemintegration')} - {job.get('firma', '')}")
        ws2.cell(row, 5, job.get("email_text", ""))
    for col, w in {1:30, 2:40, 3:30, 4:55, 5:80}.items():
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.freeze_panes = "A2"

    # Statistics
    ws3 = wb.create_sheet("Statistiken")
    stats = [("Bewerbungen Junior IT", ""), ("Stand", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Gesamt", len(jobs)), ("Junior", sum(1 for j in jobs if j.get("is_junior"))),
        ("Mit Email", sum(1 for j in jobs if j.get("email"))), ("Mit Ansprechpartner", sum(1 for j in jobs if j.get("ansprechpartner")))]
    for row, (l, v) in enumerate(stats, 1):
        ws3.cell(row, 1, l).font = Font(bold=(row<=1)); ws3.cell(row, 2, v)
    ws3.column_dimensions["A"].width = 30; ws3.column_dimensions["B"].width = 15

    wb.save(EXCEL_OUT)
    log(f"📊 Saved: {EXCEL_OUT} ({len(jobs)} jobs)")


if __name__ == "__main__":
    batch_size = int(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_BATCH
    os.environ["DISPLAY"] = DISPLAY
    asyncio.run(run_batch(batch_size))
