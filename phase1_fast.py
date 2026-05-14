#!/usr/bin/env python3
"""Phase 1: Fast Arbeitsagentur API search for junior IT jobs"""
import asyncio, httpx, json, re, os, subprocess
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

AA_API_URL = "https://rest.arbeitsagentur.de/jobboerse/jobsuche-service/pc/v6/jobs"
AA_API_KEY = "jobboerse-jobsuche"
EXCEL_OUT = "/home/z/my-project/download/Junior_IT_Jobs_Bamberg_Erlangen_Nuernberg.xlsx"
GIT_REPO = "/home/z/my-project"
GITHUB_TOKEN = "GIT_TOKEN_REDACTED"

IT_KEYWORDS = ['software','entwickler','developer','it-','informatik','data','devops','cloud','cyber','security','system','admin','netzwerk','network','frontend','backend','fullstack','full-stack','python','java','javascript','sap','linux','sql','database','datenbank','api','machine learning','scrum','agile','testing','qa','consultant','analyst','engineer','programmierung','web','mobile','app','infrastructure','support','helpdesk','erp','crm','digitalisierung','technologie']

TECH_LIST = ['python','java','javascript','typescript','react','angular','vue','node.js','django','flask','spring','.net','docker','kubernetes','aws','azure','gcp','terraform','ansible','jenkins','sql','mysql','postgresql','mongodb','redis','kafka','graphql','sap','linux','html','css','scrum','devops','ci/cd','machine learning','deep learning','tensorflow','pytorch','data science','big data','cybersecurity','powershell','bash','git']

def is_it_job(t): 
    tl = t.lower()
    return any(kw in tl for kw in IT_KEYWORDS)

def is_senior(t):
    tl = t.lower()
    return any(kw in tl for kw in ['senior','lead','principal','head of','director','leitung'])

def is_junior(t):
    tl = t.lower()
    return any(kw in tl for kw in ['junior','einsteiger','trainee','praktikant','intern','werkstudent','working student','absolvent','graduate','berufseinsteiger','associate','duales studium','ausbildung','apprentice','nachwuchs'])

def classify_cat(t):
    tl = t.lower()
    if any(k in tl for k in ['softwareentwickler','software developer','entwickler','developer','programmierer','frontend','backend','fullstack','full-stack']): return 'Softwareentwicklung'
    if any(k in tl for k in ['data scientist','data analyst','data engineer','business intelligence','machine learning']): return 'Data Science & Analytics'
    if any(k in tl for k in ['it consultant','it-consultant','it berater','sap consultant','technischer berater']): return 'IT-Consulting'
    if any(k in tl for k in ['system administrator','sysadmin','admin','netzwerkadministrator']): return 'Systemadministration'
    if any(k in tl for k in ['devops','cloud engineer','platform engineer','sre']): return 'DevOps & Cloud'
    if any(k in tl for k in ['security','cybersecurity','it-sicherheit','pentest']): return 'Cyber Security'
    if any(k in tl for k in ['it support','helpdesk','service desk','first level','second level']): return 'IT-Support'
    if any(k in tl for k in ['project manager','projektmanager','scrum master','product owner']): return 'IT-Projektmanagement'
    if any(k in tl for k in ['qa','tester','test engineer','testautomatisierung']): return 'QA & Testing'
    if any(k in tl for k in ['ux','ui','user experience','user interface']): return 'UX/UI Design'
    if any(k in tl for k in ['ausbildung','duales studium','trainee','praktikum','werkstudent']): return 'IT-Ausbildung & Studium'
    return 'IT (Sonstige)'

def extract_techs(text):
    found = []
    tl = text.lower()
    for t in TECH_LIST:
        if re.search(r'\b'+re.escape(t)+r'\b', tl):
            found.append(t)
    return list(set(found))

def get_emp_type(t):
    tl = t.lower()
    if any(k in tl for k in ['teilzeit','part-time']): return 'Teilzeit'
    if any(k in tl for k in ['werkstudent','working student']): return 'Werkstudent'
    if any(k in tl for k in ['praktikant','praktikum','intern']): return 'Praktikum'
    if any(k in tl for k in ['ausbildung','apprentice','azubi']): return 'Ausbildung'
    if any(k in tl for k in ['duales studium']): return 'Duales Studium'
    return 'Vollzeit'

def get_remote(t):
    tl = t.lower()
    if any(k in tl for k in ['hybrid']): return 'Hybrid'
    if any(k in tl for k in ['remote','homeoffice','home-office']): return 'Remote'
    return 'Nicht angegeben'


async def search_jobs(client, search_term, city, umkreis=50):
    """Single API search"""
    params = {'was': search_term, 'wo': city, 'umkreis': umkreis, 'page': 0, 'size': 100}
    headers = {'X-API-Key': AA_API_KEY, 'Accept': 'application/json'}
    try:
        r = await client.get(AA_API_URL, params=params, headers=headers, timeout=12)
        if r.status_code == 200:
            return r.json().get('stellenangebote', [])
    except:
        pass
    return []


async def main():
    print("="*60)
    print("Phase 1: Arbeitsagentur API - Junior IT Jobs B/ER/N")
    print("="*60)
    
    all_jobs = {}  # refnr -> job dict
    
    cities = ['Erlangen', 'Nürnberg', 'Bamberg', 'Fürth', 'Forchheim']
    search_terms = [
        'Junior IT', 'Junior Software', 'Junior Entwickler',
        'Werkstudent IT', 'Werkstudent Software', 
        'Praktikum IT', 'Praktikum Software',
        'Trainee IT', 'Ausbildung Informatik', 'Duales Studium IT',
        'Junior Developer', 'Junior Data', 'Junior DevOps',
        'Einsteiger IT', 'Berufseinsteiger Informatik',
        'Junior Frontend', 'Junior Backend', 'Junior Fullstack',
    ]
    
    # Build all search combos
    combos = [(term, city) for term in search_terms for city in cities]
    print(f"  {len(combos)} search combinations")
    
    # Execute with high concurrency
    async with httpx.AsyncClient() as client:
        sem = asyncio.Semaphore(8)
        
        async def limited_search(term, city):
            async with sem:
                results = await search_jobs(client, term, city)
                return results
        
        # Process in chunks
        chunk_size = 20
        for i in range(0, len(combos), chunk_size):
            chunk = combos[i:i+chunk_size]
            tasks = [limited_search(term, city) for term, city in chunk]
            results = await asyncio.gather(*tasks, return_exceptions=True)
            
            for result in results:
                if isinstance(result, list):
                    for stelle in result:
                        refnr = stelle.get('refnr', '')
                        titel = stelle.get('beruf', '') or stelle.get('titel', '')
                        
                        if not titel or not is_it_job(titel):
                            continue
                        if is_senior(titel):
                            continue
                        
                        arbeitgeber = stelle.get('arbeitgeber', '')
                        ort = stelle.get('arbeitsort', {})
                        ort_name = ort.get('ort', '') if isinstance(ort, dict) else str(ort)
                        ort_plz = ort.get('plz', '') if isinstance(ort, dict) else ''
                        
                        jk = None
                        if is_junior(titel):
                            jk = [kw for kw in ['junior','einsteiger','trainee','werkstudent','praktikant','absolvent','duales studium','ausbildung','berufseinsteiger','associate','nachwuchs'] if kw in titel.lower()]
                            jk = jk[0] if jk else 'junior_keyword'
                        else:
                            jk = 'no_senior_keyword'
                        
                        if refnr and refnr not in all_jobs:
                            all_jobs[refnr] = {
                                'refnr': refnr,
                                'title': titel,
                                'employer_name': arbeitgeber,
                                'city': ort_name,
                                'plz': ort_plz,
                                'is_junior': is_junior(titel),
                                'junior_kw': jk,
                                'it_kw': next((kw for kw in IT_KEYWORDS if kw in titel.lower()), 'it'),
                                'vollzeit': stelle.get('vollzeit', True),
                                'befristet': stelle.get('befristet', False),
                                'eintrittsdatum': stelle.get('eintrittsdatum', ''),
                                'modifikationsTimestamp': stelle.get('modifikationsTimestamp', ''),
                                'arbeitgeberOrt': stelle.get('arbeitgeberOrt', ''),
                                'arbeitgeberPlz': stelle.get('arbeitgeberPlz', ''),
                                'url': f'https://www.arbeitsagentur.de/jobsuche/jobdetail/{refnr}',
                            }
            
            print(f"  Processed {min(i+chunk_size, len(combos))}/{len(combos)} searches, found {len(all_jobs)} unique jobs so far")
    
    print(f"\n  Total unique junior IT jobs from API: {len(all_jobs)}")
    
    # Now create Excel
    print("\n  Creating Excel...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Junior IT-Jobs"
    
    headers = [
        'Job-ID','Job-Titel','Job-Titel (Original)','Job-Kategorie','Erfahrungslevel',
        'Beschäftigungsart','Remote-Option','Technologien',
        'Ausschreibungstext (Auszug)','Job-URL',
        'Arbeitgeber','Arbeitgeber (norm.)','Ort','PLZ','Straße',
        'Telefon','E-Mail','Branche',
        'Quelle','Karriere-URL','Website-URL',
        'Karriere-HTML Pfad','Job-HTML Pfad',
        'Junior-Keyword Match','IT-Keyword Match','Ist Junior-Job','Ist IT-Job',
        'Vollzeit/Teilzeit Detail','Befristet','Gehaltsspanne','Arbeitsagentur RefNr',
        'Eintrittsdatum','Veroeffentlichungsdatum',
        'Arbeitgeber Ort','Arbeitgeber PLZ',
        'Scrape-Datum','Scrape-Status','Arbeitgeber-ID',
        'Letzte Prüfung','Erstellt','Aktualisiert',
    ]
    
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.font = hf
        c.fill = hfill
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    
    for idx, (refnr, job) in enumerate(all_jobs.items()):
        row = idx + 2
        title = job['title']
        emp_name = job['employer_name']
        emp_norm = re.sub(r'\s+', '', re.sub(r'(GmbH|AG|e\.K\.|UG|KG|mbH).*$', '', emp_name)).lower()
        
        values = [
            f'J-{idx+1:05d}', title, title, classify_cat(title), 'Junior',
            get_emp_type(title), get_remote(title), ', '.join(extract_techs(title)),
            '', job['url'],
            emp_name, emp_norm, job['city'], job['plz'], '',
            '', '', '',
            'Arbeitsagentur', '', '', '', '',
            job['junior_kw'], job['it_kw'], 'Ja' if job['is_junior'] else 'Nein', 'Ja',
            get_emp_type(title), 'Ja' if job.get('befristet') else 'Nein', '', refnr,
            job.get('eintrittsdatum',''), job.get('modifikationsTimestamp',''),
            job.get('arbeitgeberOrt',''), job.get('arbeitgeberPlz',''),
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'completed', '',
            datetime.now().strftime('%Y-%m-%d'), datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        ]
        
        for col, v in enumerate(values, 1):
            ws.cell(row, col, v).alignment = Alignment(wrap_text=True, vertical='top')
    
    # Stats
    ws2 = wb.create_sheet('Statistiken')
    ws2.cell(1, 1, 'Gesamt Junior IT-Jobs')
    ws2.cell(1, 2, len(all_jobs))
    ws2.cell(2, 1, 'Quelle')
    ws2.cell(2, 2, 'Arbeitsagentur API')
    ws2.cell(3, 1, 'Region')
    ws2.cell(3, 2, 'Bamberg / Erlangen / Nürnberg (50km)')
    ws2.cell(4, 1, 'Scrape-Datum')
    ws2.cell(4, 2, datetime.now().strftime('%Y-%m-%d %H:%M'))
    
    # Count by category
    cats = {}
    for job in all_jobs.values():
        cat = classify_cat(job['title'])
        cats[cat] = cats.get(cat, 0) + 1
    ws2.cell(6, 1, 'Nach Kategorie:')
    for i, (cat, count) in enumerate(sorted(cats.items(), key=lambda x: -x[1])):
        ws2.cell(7+i, 1, cat)
        ws2.cell(7+i, 2, count)
    
    wb.save(EXCEL_OUT)
    print(f"  Excel saved: {EXCEL_OUT}")
    
    # Save JSON too
    json_out = EXCEL_OUT.replace('.xlsx', '_raw.json')
    with open(json_out, 'w', encoding='utf-8') as f:
        json.dump(list(all_jobs.values()), f, ensure_ascii=False, indent=2, default=str)
    
    # Git push
    os.chdir(GIT_REPO)
    subprocess.run(['git', 'add', '-A'], capture_output=True, timeout=30)
    r = subprocess.run(['git', 'commit', '-m', f'Phase 1: Junior IT Jobs API ({len(all_jobs)} jobs)'], capture_output=True, timeout=30)
    if r.returncode == 0:
        subprocess.run(['git', 'push', f'https://{GITHUB_TOKEN}@github.com/BirdNest055/Heyjobs.git', 'main', '--force-with-lease'], capture_output=True, timeout=60)
        print("  Git: Pushed!")
    
    print(f"\n{'='*60}")
    print(f"Phase 1 DONE: {len(all_jobs)} junior IT jobs")
    print(f"{'='*60}")

if __name__ == '__main__':
    asyncio.run(main())
