import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load data
with open('/home/z/my-project/gmaps_erlangen_results.json') as f:
    results = json.load(f)

print(f"Loaded {len(results)} results")

# Deduplicate by name+address combination
seen = set()
unique = []
for r in results:
    key = (r.get('name','').strip().lower(), r.get('address','').strip().lower())
    if key not in seen and r.get('name','').strip():
        seen.add(key)
        unique.append(r)

print(f"After dedup: {len(unique)} unique companies")

# Sort by city, then name
unique.sort(key=lambda x: (x.get('city',''), x.get('name','')))

wb = Workbook()
ws = wb.active
ws.title = "Google Maps Firmen"

# Headers
headers = ['#', 'Firmenname', 'Bewertung', 'Adresse', 'PLZ', 'Ort', 'Telefon', 'Website verfügbar', 'Breitengrad', 'Längengrad', 'Google Maps Link', 'Suchort', 'Quelle']

# Styling
header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='D9E2F3'),
    right=Side(style='thin', color='D9E2F3'),
    top=Side(style='thin', color='D9E2F3'),
    bottom=Side(style='thin', color='D9E2F3')
)
even_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# Data rows
for i, r in enumerate(unique, 1):
    row = i + 1
    values = [
        i,
        r.get('name', ''),
        r.get('rating', ''),
        r.get('address', ''),
        r.get('plz', ''),
        r.get('city', ''),
        r.get('phone', ''),
        'Ja' if r.get('has_website') else 'Nein',
        r.get('lat', ''),
        r.get('lon', ''),
        r.get('href', ''),
        r.get('search_town', ''),
        r.get('source', '')
    ]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(name='Arial', size=10)
        cell.alignment = Alignment(vertical='center', wrap_text=(col in [2,4,11]))
        cell.border = thin_border
        if i % 2 == 0:
            cell.fill = even_fill

# Column widths
widths = [5, 35, 8, 40, 8, 20, 18, 10, 10, 10, 45, 20, 12]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Auto-filter
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(unique)+1}"

# Freeze top row
ws.freeze_panes = 'A2'

# Stats sheet
ws2 = wb.create_sheet("Statistiken")
stats = [
    ['Statistik', 'Wert'],
    ['Gesamt Einträge (roh)', len(results)],
    ['Einzigartige Firmen', len(unique)],
    ['Städte/Orte abgedeckt', len(set(r.get('city','') for r in unique if r.get('city')))],
    ['Mit Website', sum(1 for r in unique if r.get('has_website'))],
    ['Mit Telefon', sum(1 for r in unique if r.get('phone'))],
    ['Mit Adresse', sum(1 for r in unique if r.get('address'))],
    ['Mit PLZ', sum(1 for r in unique if r.get('plz'))],
    ['Website %', f"{sum(1 for r in unique if r.get('has_website'))/len(unique)*100:.1f}%"],
    ['Telefon %', f"{sum(1 for r in unique if r.get('phone'))/len(unique)*100:.1f}%"],
]
for row_idx, row_data in enumerate(stats, 1):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        if row_idx == 1:
            cell.font = header_font
            cell.fill = header_fill
        cell.font = Font(name="Arial", size=10)
ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 15

# Top cities sheet
ws3 = wb.create_sheet("Top Orte")
from collections import Counter
city_counts = Counter(r.get('city','') for r in unique if r.get('city'))
ws3.cell(row=1, column=1, value='Ort').font = header_font
ws3.cell(row=1, column=1).fill = header_fill
ws3.cell(row=1, column=2, value='Anzahl Firmen').font = header_font
ws3.cell(row=1, column=2).fill = header_fill
for i, (city, count) in enumerate(city_counts.most_common(50), 2):
    ws3.cell(row=i, column=1, value=city)
    ws3.cell(row=i, column=2, value=count)
ws3.column_dimensions['A'].width = 25
ws3.column_dimensions['B'].width = 15

# Save
output = '/home/z/my-project/download/Google_Maps_Firmen_Erlangen_50km.xlsx'
wb.save(output)
print(f"\nSaved to: {output}")
print(f"Total unique firms: {len(unique)}")
